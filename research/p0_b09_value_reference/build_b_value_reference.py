"""B-09 Phase 3 — isolated candidate B value reference (pure cross-sectional).

Authorized scope (user ruling 2026-08-17): 資料可行性 / 隔離 candidate reference 建立 /
feature graph freeze / B-02 dependency-L 機械推導. NO performance, IC, selection, A0-A3.

Implements the frozen F-E(a) ruling:
  Value = 純當期產業內橫斷面估值 percentile   (+ PEG, built elsewhere)
  - NO expanding self-history window   (removes V3-style path dependence)
  - NO MIN_PCT_SAMPLES=60 sample gate
  - NO 2019 compatibility anchor       (F-D: anchor's only rationale was v4.5
                                        production compatibility, retired by Frozen A)

Two deviations from the retired production reference, both forced and disclosed:

  D1. Industry level = TSE 產業 (not TEJ 產業).
      Reason: 歷史產業類別.xlsx carries change history ONLY for TSE產業 and
      TEJ子產業 — there is no 前N次TEJ產業變更 column. A PIT-correct industry
      timeline is therefore not reconstructible at the TEJ產業 level that the
      retired reference used. TSE產業 is also the official exchange
      classification (external standard, not vendor-proprietary).

  D2. Industry assignment is POINT-IN-TIME, not the static current snapshot.
      Reason: 49.4% of stocks have >=1 TSE industry change (2,760 events dated
      >=2004). Using ~/tej_cache/industry_map.parquet (no date column, current
      snapshot) would assign post-change industries backwards through history —
      look-ahead affecting roughly half the universe.

Zero free parameters. Group-size floor is 2, the mathematical minimum for a
within-group rank to be defined — not a tuned threshold.

Writes ONLY to this directory. Does not touch ~/tej_cache, ~/market_cache,
cloud_cache, or any production path.
"""

import json
import os
import re
import sys

import duckdb
import openpyxl
import pandas as pd

HERE = os.path.dirname(os.path.abspath(__file__))
REPO = os.path.dirname(os.path.dirname(HERE))
TEJ_CACHE = os.path.expanduser(os.environ.get("TEJ_CACHE", "~/tej_cache"))
IND_XLSX = os.path.join(REPO, "tej_exports", "DataExport0806", "產業類別", "歷史產業類別.xlsx")
FROZEN = os.path.join(REPO, "research", "p0_r2_identity_collector", "a_leg_parity_result.json")

OUT_PARQUET = os.path.join(HERE, "b_value_reference_candidate.parquet")
OUT_INDMAP = os.path.join(HERE, "pit_industry_timeline.parquet")
OUT_JSON = os.path.join(HERE, "build_report.json")

MIN_GROUP = 2   # mathematical minimum for a within-group rank, not a tuned value
_DATE = re.compile(r"^(\d{4})/(\d{1,2})/(\d{1,2})$")


def _norm_date(v):
    if v is None:
        return None
    s = str(v).strip()
    if not s or s == ".":
        return None
    m = _DATE.match(s)
    if m:
        return "%04d-%02d-%02d" % (int(m.group(1)), int(m.group(2)), int(m.group(3)))
    return None


def _norm_code(v):
    """'M1300 塑膠工業' -> 'M1300'."""
    if v is None:
        return None
    s = str(v).strip()
    if not s or s == ".":
        return None
    return s.split()[0]


def build_pit_industry_timeline():
    """(stock_id, effective_from, tse_ind_code) step function, ascending.

    Semantics verified against 1316 上曜 / 1319 東陽 / 1229 聯華:
    each 前N次 record is (industry it BECAME, effective date); 前一次 is most
    recent and always agrees with the current TSE產業 column.
    """
    wb = openpyxl.load_workbook(IND_XLSX, read_only=True, data_only=True)
    try:
        ws = wb[wb.sheetnames[0]]
        it = ws.iter_rows(values_only=True)
        hdr = list(next(it))
        I = {c: i for i, c in enumerate(hdr)}
        chg_pairs = [("前三次TSE產業變更", "前三次TSE產業變更日"),
                     ("前二次TSE產業變更", "前二次TSE產業變更日"),
                     ("前一次TSE產業變更", "前一次TSE產業變更日")]
        rows = []
        stats = {"stocks": 0, "with_changes": 0, "records": 0,
                 "listing_only": 0, "current_agrees_with_latest": 0,
                 "current_disagrees": 0}
        for r in it:
            if r[0] is None:
                continue
            sid = str(r[I["代號"]]).strip()
            if not sid:
                continue
            stats["stocks"] += 1
            recs = []
            for c_ind, c_date in chg_pairs:
                code, d = _norm_code(r[I[c_ind]]), _norm_date(r[I[c_date]])
                if code and d:
                    recs.append((d, code))
            listed = _norm_date(r[I["首次掛牌日期"]])
            first_ind = _norm_code(r[I["首次掛牌TSE產業"]])
            if listed and first_ind:
                recs.append((listed, first_ind))
            if not recs:
                cur = _norm_code(r[I["TSE產業_代碼"]])
                if cur:
                    recs.append(("1900-01-01", cur))
                    stats["listing_only"] += 1
            recs = sorted(set(recs))
            if len(recs) > 1:
                stats["with_changes"] += 1
            cur = _norm_code(r[I["TSE產業_代碼"]])
            if cur and recs:
                if recs[-1][1] == cur:
                    stats["current_agrees_with_latest"] += 1
                else:
                    stats["current_disagrees"] += 1
                    # current column is authoritative for "now": append it with
                    # no known effective date only if it adds information.
            for d, code in recs:
                rows.append({"stock_id": sid, "effective_from": d, "tse_ind_code": code})
                stats["records"] += 1
    finally:
        wb.close()
    df = pd.DataFrame(rows).sort_values(["stock_id", "effective_from"]).reset_index(drop=True)
    return df, stats


def frozen_decision_dates():
    with open(FROZEN, encoding="utf-8") as fh:
        per_date = json.load(fh)["a_leg_parity_result"]["per_date"]
    return sorted(per_date.keys())


def industry_as_of(timeline: pd.DataFrame, date: str) -> pd.Series:
    """stock_id -> tse_ind_code effective at `date` (latest record <= date)."""
    sub = timeline[timeline["effective_from"] <= date]
    if sub.empty:
        return pd.Series(dtype=object)
    return sub.groupby("stock_id")["tse_ind_code"].last()


def main():
    print("building PIT industry timeline ...")
    timeline, tstats = build_pit_industry_timeline()
    print("  stocks=%d records=%d with_changes=%d current_agrees=%d disagrees=%d"
          % (tstats["stocks"], tstats["records"], tstats["with_changes"],
             tstats["current_agrees_with_latest"], tstats["current_disagrees"]))
    timeline.to_parquet(OUT_INDMAP, index=False)

    dates = frozen_decision_dates()
    print("frozen decision dates:", len(dates), dates[0], "..", dates[-1])

    con = duckdb.connect()
    dl = ",".join("'%s'" % d for d in dates)
    pv = con.execute(f"""
        SELECT stock_id, CAST(date AS VARCHAR) AS date, PER_TSE
        FROM read_parquet('{TEJ_CACHE}/price_valuation/*.parquet', union_by_name=true)
        WHERE CAST(date AS VARCHAR) IN ({dl})
    """).fetchdf()
    print("price_valuation rows on decision dates:", len(pv))

    out = []
    diag = []
    for d in dates:
        day = pv[pv["date"] == d]
        n_quote = len(day)
        if n_quote == 0:
            diag.append({"date": d, "quoted": 0, "pe_positive": 0, "with_industry": 0,
                         "scored": 0, "groups": 0})
            continue
        ind = industry_as_of(timeline, d)
        day = day.copy()
        day["tse_ind_code"] = day["stock_id"].map(ind)
        day["PER_TSE"] = pd.to_numeric(day["PER_TSE"], errors="coerce")
        n_ind = int(day["tse_ind_code"].notna().sum())
        elig = day[(day["PER_TSE"] > 0) & day["tse_ind_code"].notna()].copy()
        n_pos = int((day["PER_TSE"] > 0).sum())
        if elig.empty:
            diag.append({"date": d, "quoted": n_quote, "pe_positive": n_pos,
                         "with_industry": n_ind, "scored": 0, "groups": 0})
            continue
        g = elig.groupby("tse_ind_code")["PER_TSE"]
        sizes = g.transform("size")
        elig = elig[sizes >= MIN_GROUP]
        if elig.empty:
            diag.append({"date": d, "quoted": n_quote, "pe_positive": n_pos,
                         "with_industry": n_ind, "scored": 0, "groups": 0})
            continue
        # cheaper = better -> invert the PE rank
        pct = elig.groupby("tse_ind_code")["PER_TSE"].rank(pct=True)
        elig["value_ind_pct_b"] = (1.0 - pct) * 100.0
        out.append(elig[["stock_id", "date", "tse_ind_code", "PER_TSE", "value_ind_pct_b"]])
        diag.append({"date": d, "quoted": n_quote, "pe_positive": n_pos,
                     "with_industry": n_ind, "scored": len(elig),
                     "groups": int(elig["tse_ind_code"].nunique())})

    ref = pd.concat(out, ignore_index=True) if out else pd.DataFrame()
    ref.to_parquet(OUT_PARQUET, index=False)
    dg = pd.DataFrame(diag)

    cov = (dg["scored"] / dg["quoted"].replace(0, pd.NA)).astype(float)
    payload = {
        "study": "B-09 Phase 3 — isolated candidate B value reference",
        "read_only_inputs": True,
        "writes": [os.path.relpath(OUT_PARQUET, REPO).replace("\\", "/"),
                   os.path.relpath(OUT_INDMAP, REPO).replace("\\", "/"),
                   os.path.relpath(OUT_JSON, REPO).replace("\\", "/")],
        "touched_production_paths": "NONE — no write to ~/tej_cache, ~/market_cache, cloud_cache, data/runtime_cache",
        "definition": {
            "value_ind_pct_b": "(1 - within-TSE-industry percentile rank of PER_TSE) * 100; higher = cheaper",
            "expanding_window": "NONE (removes V3-style path dependence)",
            "min_pct_samples_gate": "NONE (60-sample gate removed)",
            "history_anchor": "NONE (2019 compatibility anchor removed per F-D)",
            "industry_level": "TSE 產業 (D1: only level with reconstructible PIT change history)",
            "industry_assignment": "POINT-IN-TIME step function (D2)",
            "min_group_size": MIN_GROUP,
            "free_parameters": 0,
        },
        "industry_timeline": tstats,
        "coverage": {
            "decision_dates": len(dates),
            "date_range": [dates[0], dates[-1]],
            "rows_written": int(len(ref)),
            "scored_per_date_min": int(dg["scored"].min()),
            "scored_per_date_median": float(dg["scored"].median()),
            "scored_per_date_max": int(dg["scored"].max()),
            "coverage_vs_quoted_min": float(cov.min()),
            "coverage_vs_quoted_median": float(cov.median()),
            "groups_per_date_median": float(dg["groups"].median()),
        },
        "per_date_diagnostics": diag,
    }
    with open(OUT_JSON, "w", encoding="utf-8") as fh:
        json.dump(payload, fh, ensure_ascii=False, indent=1)

    print("rows written:", len(ref))
    print("scored/date  min=%d median=%.0f max=%d"
          % (dg["scored"].min(), dg["scored"].median(), dg["scored"].max()))
    print("coverage vs quoted  min=%.3f median=%.3f" % (cov.min(), cov.median()))
    print("groups/date median=%.0f" % dg["groups"].median())
    print("wrote", os.path.relpath(OUT_JSON, REPO))
    return 0


if __name__ == "__main__":
    sys.exit(main())
