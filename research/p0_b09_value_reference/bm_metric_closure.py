"""B-09 Phase 3b — Ruling A (PIT-UNRESOLVED) + Ruling B (B/M Metric Closure).

Authorized scope (user ruling 2026-08-17):
  A. 92 disagreeing stocks -> PIT-UNRESOLVED. After the last DATED classification,
     do NOT backfill from the current snapshot and do NOT assume the old class
     persists. Value = NA over the unknown interval; B-15 complete-case drops it.
     Coverage loss may be quantified but MUST NOT be used to change the rule.
  B. B/M Metric Closure, pure data only: per-month PBR coverage, root cause of the
     34.8% minimum, null / <=0 / source-gap decomposition, and lineage of
     1/PBR_TSE vs canonical Book Equity / Market Equity.

PROHIBITED and not performed: returns, IC, Sharpe, CAGR/MDD, selection lists, A0-A3.

Writes only into this directory. No production path is touched.
"""

import json
import os
import re
import sys

import duckdb
import openpyxl
import pandas as pd

HERE = os.path.dirname(os.path.abspath(__file__))
REPO = os.path.dirname(os.path.dirname(HERE))
TEJ_CACHE = os.path.expanduser(os.environ.get("TEJ_CACHE", "~/tej_cache"))
IND_XLSX = os.path.join(REPO, "tej_exports", "DataExport0806", "產業類別", "歷史產業類別.xlsx")
FROZEN = os.path.join(REPO, "research", "p0_r2_identity_collector", "a_leg_parity_result.json")

OUT_JSON = os.path.join(HERE, "bm_metric_closure.json")
OUT_TIMELINE = os.path.join(HERE, "pit_industry_timeline_v2.parquet")
OUT_REF = os.path.join(HERE, "b_value_reference_candidate_v2.parquet")

MIN_GROUP = 2
WINDOW_START = "2014-07"   # frozen per ruling C (L=18, first eligible 2014-07)
_DATE = re.compile(r"^(\d{4})/(\d{1,2})/(\d{1,2})$")


def _nd(v):
    if v is None:
        return None
    s = str(v).strip()
    if not s or s == ".":
        return None
    m = _DATE.match(s)
    return "%04d-%02d-%02d" % (int(m.group(1)), int(m.group(2)), int(m.group(3))) if m else None


def _nc(v):
    if v is None:
        return None
    s = str(v).strip()
    return s.split()[0] if s and s != "." else None


def build_timeline_v2():
    """Timeline + per-stock `unresolved_from` (Ruling A)."""
    wb = openpyxl.load_workbook(IND_XLSX, read_only=True, data_only=True)
    try:
        ws = wb[wb.sheetnames[0]]
        it = ws.iter_rows(values_only=True)
        hdr = list(next(it))
        I = {c: i for i, c in enumerate(hdr)}
        pairs = [("前三次TSE產業變更", "前三次TSE產業變更日"),
                 ("前二次TSE產業變更", "前二次TSE產業變更日"),
                 ("前一次TSE產業變更", "前一次TSE產業變更日")]
        rows, unresolved = [], {}
        n_stocks = n_unres = 0
        for r in it:
            if r[0] is None:
                continue
            sid = str(r[I["代號"]]).strip()
            if not sid:
                continue
            n_stocks += 1
            recs = [(_nd(r[I[cd]]), _nc(r[I[ci]])) for ci, cd in pairs]
            recs = [(d, c) for d, c in recs if d and c]
            listed, first_ind = _nd(r[I["首次掛牌日期"]]), _nc(r[I["首次掛牌TSE產業"]])
            if listed and first_ind:
                recs.append((listed, first_ind))
            cur = _nc(r[I["TSE產業_代碼"]])
            if not recs:
                if cur:
                    recs.append(("1900-01-01", cur))
            recs = sorted(set(recs))
            if cur and recs and recs[-1][1] != cur:
                # Ruling A: an undated change exists after the last dated record.
                # The whole interval from that record onward is UNRESOLVED.
                unresolved[sid] = recs[-1][0]
                n_unres += 1
            for d, c in recs:
                rows.append({"stock_id": sid, "effective_from": d, "tse_ind_code": c})
    finally:
        wb.close()
    tl = pd.DataFrame(rows).sort_values(["stock_id", "effective_from"]).reset_index(drop=True)
    tl["unresolved_from"] = tl["stock_id"].map(unresolved)
    return tl, unresolved, {"stocks": n_stocks, "unresolved_stocks": n_unres}


def industry_as_of(tl, unresolved, date):
    sub = tl[tl["effective_from"] <= date]
    ser = sub.groupby("stock_id")["tse_ind_code"].last() if not sub.empty else pd.Series(dtype=object)
    if not ser.empty and unresolved:
        bad = [s for s, d in unresolved.items() if date >= d and s in ser.index]
        if bad:
            ser = ser.drop(index=bad)      # Ruling A: NA over the unknown interval
    return ser


def main():
    print("A · rebuilding timeline with PIT-UNRESOLVED ...")
    tl, unresolved, tstats = build_timeline_v2()
    tl.to_parquet(OUT_TIMELINE, index=False)
    print("  stocks=%d  unresolved=%d" % (tstats["stocks"], tstats["unresolved_stocks"]))

    dates = sorted(json.load(open(FROZEN, encoding="utf-8"))["a_leg_parity_result"]["per_date"])
    win = [d for d in dates if d[:7] >= WINDOW_START]
    print("  frozen window: %d months %s .. %s" % (len(win), win[0], win[-1]))

    con = duckdb.connect()
    dl = ",".join("'%s'" % d for d in dates)
    pv = con.execute(f"""
        SELECT stock_id, CAST(date AS VARCHAR) AS date, PER_TSE, PBR_TSE
        FROM read_parquet('{TEJ_CACHE}/price_valuation/*.parquet', union_by_name=true)
        WHERE CAST(date AS VARCHAR) IN ({dl})
    """).fetchdf()
    pv["PER_TSE"] = pd.to_numeric(pv["PER_TSE"], errors="coerce")
    pv["PBR_TSE"] = pd.to_numeric(pv["PBR_TSE"], errors="coerce")

    print("B · per-date coverage decomposition ...")
    diag, out = [], []
    for d in dates:
        day = pv[pv["date"] == d]
        q = len(day)
        if q == 0:
            continue
        ind = industry_as_of(tl, unresolved, d)
        day = day.copy()
        day["tse_ind_code"] = day["stock_id"].map(ind)
        rec = {
            "date": d,
            "quoted": q,
            "ind_ok": int(day["tse_ind_code"].notna().sum()),
            "ind_na_unresolved": int(day["stock_id"].isin(
                [s for s, dd in unresolved.items() if d >= dd]).sum()),
            "pe_null": int(day["PER_TSE"].isna().sum()),
            "pe_le0": int((day["PER_TSE"] <= 0).sum()),
            "pe_pos": int((day["PER_TSE"] > 0).sum()),
            "pb_null": int(day["PBR_TSE"].isna().sum()),
            "pb_le0": int((day["PBR_TSE"] <= 0).sum()),
            "pb_pos": int((day["PBR_TSE"] > 0).sum()),
        }
        # canonical B/M candidate reference (Ruling B direction), under Ruling A
        elig = day[(day["PBR_TSE"] > 0) & day["tse_ind_code"].notna()].copy()
        if not elig.empty:
            sizes = elig.groupby("tse_ind_code")["PBR_TSE"].transform("size")
            elig = elig[sizes >= MIN_GROUP].copy()
        if not elig.empty:
            elig["bm"] = 1.0 / elig["PBR_TSE"]          # Book/Market
            pct = elig.groupby("tse_ind_code")["bm"].rank(pct=True)
            elig["value_ind_pct_bm"] = pct * 100.0       # higher B/M = cheaper
            out.append(elig[["stock_id", "date", "tse_ind_code", "PBR_TSE",
                             "bm", "value_ind_pct_bm"]])
            rec["scored_bm"] = len(elig)
        else:
            rec["scored_bm"] = 0
        diag.append(rec)

    ref = pd.concat(out, ignore_index=True) if out else pd.DataFrame()
    ref.to_parquet(OUT_REF, index=False)
    dg = pd.DataFrame(diag)
    dgw = dg[dg["date"].str[:7] >= WINDOW_START]

    def med(df, c):
        return float(df[c].median())

    worst = dg.assign(cov=dg.pb_pos / dg.quoted).nsmallest(5, "cov")
    worst_w = dgw.assign(cov=dgw.pb_pos / dgw.quoted).nsmallest(5, "cov")

    print("C · lineage check: PBR/PER should equal ROE_ttm ...")
    fs = con.execute(f"""
        SELECT stock_id, CAST(date AS VARCHAR) AS date, net_income, equity
        FROM read_parquet('{TEJ_CACHE}/financial_statements/*.parquet', union_by_name=true)
        WHERE net_income IS NOT NULL AND equity > 0
    """).fetchdf().sort_values(["stock_id", "date"])
    fs["ni_ttm"] = fs.groupby("stock_id")["net_income"].transform(
        lambda s: s.rolling(4, min_periods=4).sum())
    fs["roe_ttm"] = fs["ni_ttm"] / fs["equity"]
    fs = fs.dropna(subset=["roe_ttm"])
    # asof-join each decision date to the latest quarter strictly before it
    lineage = []
    fs_idx = {sid: g for sid, g in fs.groupby("stock_id")}
    sample_dates = [d for d in win][::12]      # one per year inside frozen window
    for d in sample_dates:
        day = pv[(pv["date"] == d) & (pv["PER_TSE"] > 0) & (pv["PBR_TSE"] > 0)]
        vals = []
        for sid, per, pbr in zip(day["stock_id"], day["PER_TSE"], day["PBR_TSE"]):
            g = fs_idx.get(sid)
            if g is None:
                continue
            gg = g[g["date"] < d]
            if gg.empty:
                continue
            roe = gg.iloc[-1]["roe_ttm"]
            if roe and roe > 0:
                vals.append((pbr / per) / roe)
        if vals:
            s = pd.Series(vals)
            lineage.append({"date": d, "n": len(s),
                            "ratio_median": float(s.median()),
                            "ratio_p10": float(s.quantile(0.10)),
                            "ratio_p90": float(s.quantile(0.90)),
                            "within_20pct_of_1": float(((s - 1).abs() <= 0.20).mean())})

    payload = {
        "study": "B-09 Phase 3b — Ruling A (PIT-UNRESOLVED) + Ruling B (B/M Metric Closure)",
        "prohibited_not_performed": ["returns", "IC", "Sharpe", "CAGR/MDD", "selection lists", "A0-A3"],
        "frozen_window": {"first_eligible": WINDOW_START, "months": len(win),
                          "range": [win[0], win[-1]], "L": 18},
        "ruling_A_pit_unresolved": {
            "rule": "For stocks whose current TSE class disagrees with the last dated record, the interval from that record onward is UNRESOLVED: no current-snapshot backfill, no persistence assumption. Industry = NA -> Value = NA -> dropped by B-15 complete-case.",
            "stocks_total": tstats["stocks"],
            "stocks_unresolved": tstats["unresolved_stocks"],
            "coverage_loss_median_stocks_per_date_in_window": med(dgw, "ind_na_unresolved"),
            "coverage_loss_median_pct_of_quoted": round(
                100 * med(dgw, "ind_na_unresolved") / med(dgw, "quoted"), 3),
            "note": "Quantified only. Per ruling, this number must NOT be used to change the rule.",
        },
        "ruling_B_bm_metric_closure": {
            "full_grid_255": {
                "pe_pos_coverage_median": round(med(dg, "pe_pos") / med(dg, "quoted"), 4),
                "pb_pos_coverage_median": round(med(dg, "pb_pos") / med(dg, "quoted"), 4),
                "worst_dates": worst[["date", "quoted", "pb_null", "pb_le0", "pb_pos"]].to_dict("records"),
            },
            "frozen_window_141": {
                "quoted_median": med(dgw, "quoted"),
                "pe_null_median": med(dgw, "pe_null"),
                "pe_le0_median": med(dgw, "pe_le0"),
                "pe_pos_coverage_median": round(med(dgw, "pe_pos") / med(dgw, "quoted"), 4),
                "pb_null_median": med(dgw, "pb_null"),
                "pb_le0_median": med(dgw, "pb_le0"),
                "pb_pos_coverage_median": round(med(dgw, "pb_pos") / med(dgw, "quoted"), 4),
                "pb_pos_coverage_min": round(float((dgw.pb_pos / dgw.quoted).min()), 4),
                "worst_dates": worst_w[["date", "quoted", "pb_null", "pb_le0", "pb_pos"]].to_dict("records"),
                "scored_bm_median": med(dgw, "scored_bm"),
                "scored_bm_coverage_median": round(med(dgw, "scored_bm") / med(dgw, "quoted"), 4),
            },
            "lineage_1_over_PBR_vs_BE_over_ME": {
                "identity_tested": "ME = PER x Earnings and ME = PBR x BookEquity  =>  PBR/PER = Earnings/BookEquity = ROE_ttm. A ratio of (PBR/PER)/ROE_ttm ~= 1 confirms PER_TSE and PBR_TSE share one market-cap and one share basis, i.e. 1/PBR_TSE is BE/ME.",
                "per_year_samples": lineage,
            },
        },
        "per_date_diagnostics": diag,
    }
    with open(OUT_JSON, "w", encoding="utf-8") as fh:
        json.dump(payload, fh, ensure_ascii=False, indent=1)

    print()
    print("A · unresolved stocks dropped per date (window): median %.0f (%.3f%% of quoted)"
          % (med(dgw, "ind_na_unresolved"),
             100 * med(dgw, "ind_na_unresolved") / med(dgw, "quoted")))
    print("B · window medians: quoted %.0f | PE>0 %.3f | PBR>0 %.3f | scored_bm %.3f"
          % (med(dgw, "quoted"), med(dgw, "pe_pos") / med(dgw, "quoted"),
             med(dgw, "pb_pos") / med(dgw, "quoted"),
             med(dgw, "scored_bm") / med(dgw, "quoted")))
    print("  PBR null median %.0f | PBR<=0 median %.0f" % (med(dgw, "pb_null"), med(dgw, "pb_le0")))
    print("  worst dates on FULL 255 grid:")
    for r in worst[["date", "quoted", "pb_null", "pb_le0", "pb_pos"]].to_dict("records"):
        print("   ", r)
    print("  worst dates INSIDE frozen window:")
    for r in worst_w[["date", "quoted", "pb_null", "pb_le0", "pb_pos"]].to_dict("records"):
        print("   ", r)
    print("C · lineage (PBR/PER)/ROE_ttm:")
    for r in lineage:
        print("    %s n=%4d median=%.3f p10=%.3f p90=%.3f within20%%=%.2f"
              % (r["date"], r["n"], r["ratio_median"], r["ratio_p10"],
                 r["ratio_p90"], r["within_20pct_of_1"]))
    print("wrote", os.path.relpath(OUT_JSON, REPO))
    return 0


if __name__ == "__main__":
    sys.exit(main())
