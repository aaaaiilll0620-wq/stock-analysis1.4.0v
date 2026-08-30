# -*- coding: utf-8 -*-
"""W6b-2 · readers that turn VERIFIED leaves into raw frames.

One reader per locator form. Each one takes a leaf that has already passed the
manifest engine's checks — declared, present, hash-matched, member-inventoried —
and reads only the entries that leaf marked `consumed`. Nothing here globs, and
nothing here decides what to read: the manifest decided, and this obeys it.

HOW CORRECTNESS IS ESTABLISHED
------------------------------
Not by assertion. P2-3 requires a field to mean ONE thing across both routes, so
an L3 reader is right exactly when it reproduces what L2's materializer produced
from the same bytes. `research/b0_l3/verify_reader_parity.py` does that against
the sealed panels, and the tests run it.

That matters more than it sounds. The failure mode for a second implementation
of a parsing rule is never a crash — it is a slightly different number. TEJ
publishes 成交量(千股); the frozen adv20 lineage (C-25) is
`close x Trading_Volume` against an ABSOLUTE NTD floor (§4.2). An L3 reader that
forgot the x1000 would not raise. It would quietly make every security illiquid.

WHERE THE LINE BETWEEN TRANSCRIPTION AND IMPORT FALLS
-----------------------------------------------------
Two kinds of knowledge live in an L2 builder, and they are treated differently
here:

    vendor dialect      column names, encodings, separators, date shapes,
                        which spreadsheet cell holds what
                        -> TRANSCRIBED into this module

    canonical semantics what a reason code MEANS, which status a row may
                        become, how a multiplier is formed, whether an event is
                        reconstructible
                        -> IMPORTED from `core.*`

`build_market_state.py` already states the rule for its own family: "All
vendor-specific column names live here; `core.b0_market_state` never sees a
Chinese column name." An importer per route, one set of semantics for both.

So this module imports `core.b0_market_state`, `core.b0_corporate_actions`,
`core.b0_bonus_share_source` and `core.b0_features` — all normative, all inside
the A2 route closure, all bound by the seal — and it imports NEITHER
`tej_importer` NOR any `research/b0_materializer/build_*.py`. Two reasons, and
the second is the one that actually decides it:

  1. `tej_importer` is a live importer serving a dozen datasets and sits outside
     the normative set. An L3 run whose answer moved because an unrelated
     dataset's spec was edited would be a run whose inputs nobody bound.
  2. A reader that CALLS L2's parser cannot be checked against L2's output —
     parity would be a tautology. Transcription is what makes the comparison
     mean something, and the comparison is what makes transcription safe.

⚠ These readers do NOT touch `data/b0/`. Those panels are L2's, frozen in place
under R-W1-1. The parity checker reads them, because comparing against them is
the whole point — but it reads them and nothing else.

⚠ Nor do they apply the frozen window. L2's fundamentals panels stop at
`window_end` (2026-03-31) and its bonus panel at the 141-period lookback union;
an L3 reader that inherited those bounds would be prospective in name only. The
readers return everything the sources hold, and the parity checker restricts to
the overlap before comparing — the same shape as the calendar's prefix rule.

EVALUATION ORDER
----------------
The leaf dependency graph is a real order, not a formality:

    security_status  ->  corporate_actions  ->  bonus_shares
    calendar         ->  bonus_shares

`corporate_actions` needs the status table for its holder-side reorganization
exits, and `bonus_shares` needs the ledger it produces plus the session list.
"""
from __future__ import annotations

import io
import os
import sys
import zipfile

HERE = os.path.dirname(os.path.abspath(__file__))
REPO = os.path.dirname(os.path.dirname(HERE))
for _p in (REPO, os.path.join(REPO, "research", "b0_materializer")):
    if _p not in sys.path:
        sys.path.insert(0, _p)

from core.b0_canonical_hash import file_sha256                  # noqa: E402
from source_ownership_manifest import (                          # noqa: E402
    LEAF_FILENAME, SELF_HASH_FIELD, ManifestError,
    assert_archive_members_match, load_leaf,
)

# The frozen 2019+ price columns and the unit restoration. Named rather than
# inlined so a diff shows a change to them.
PRICE_COLUMNS = ("證券代碼", "年月日", "開盤價(元)", "收盤價(元)", "成交量(千股)")
VOLUME_THOUSANDS_TO_SHARES = 1000.0
VINTAGE_BOUNDARY = "2019-01-01"


class ReaderError(SystemExit):
    """Fail-loud: a declared source could not be read as declared."""


def _leaf_and_landing(run_dir: str, dataset: str):
    leaf = load_leaf(os.path.join(run_dir, LEAF_FILENAME % dataset))
    landing = leaf["landing_directory"]
    if not os.path.isabs(landing):
        landing = os.path.join(REPO, landing)
    return leaf, landing


def consumed_entries(leaf: dict) -> list:
    """Only what the manifest said to read. Order is the leaf's, not the disk's."""
    return [e for e in leaf["entries"] if e["disposition"] == "consumed"]


def _verified_path(landing: str, entry: dict) -> str:
    """Re-check the bytes at read time, not only at declare time.

    The manifest engine checks these when the leaf is verified; checking again
    here closes the window between verification and use. It is cheap next to
    reading a 270 MB member.

    An entry may name its own landing directory. `prices` needs that: §2.8.3
    splits its lineage at 2019-01-01 and the two halves live in different trees,
    so one landing per leaf cannot address both.
    """
    landing = entry.get("landing_directory") or landing
    if not os.path.isabs(landing):
        landing = os.path.join(REPO, landing)
    path = os.path.join(landing, entry["locator"])
    if not os.path.isfile(path):
        raise ReaderError("abort: declared source %s is not present"
                          % entry["locator"])
    got = file_sha256(path)
    if got != entry["raw_sha256"]:
        raise ReaderError(
            "abort: %s changed between declaration and read.\n"
            "  declared: %s\n  on disk:  %s" % (entry["locator"],
                                                entry["raw_sha256"], got))
    return path


# --- declared-format dispatch ---------------------------------------------------
#
# WHICH PARSER opens a file is a property of the DECLARATION, never of the name.
# Two files in this corpus lie by extension and neither lies loudly:
#
#     2026 0826 2385家.csv    BOM ff fe, zero commas, TAB separated
#     月營收7月完整.zip        a zip wrapping exactly that dialect
#
# `pd.read_excel` on the archive raises, which is the lucky case. The unlucky
# case is the one this dispatch exists for: a csv opened with the wrong
# encoding/separator pair returns a SINGLE-COLUMN frame and raises nothing, and
# every guard downstream — required columns, period algebra, duplicate keys —
# then reports on a frame that was never read.
#
# So the leaf's `format` string selects the reader, and a format with no
# registered reader ABORTS naming both the format and the file. It never falls
# back to a guess: a guess that happens to be wrong is indistinguishable from a
# guess that happens to be right until a number moves.
#
# The vocabulary is the one the producers write:
#   `build_financials_leaf.DECLARATION`      "xlsx", "csv:utf-16:tab"
#   `build_flat_leaves.FLAT_FAMILIES`        extension-derived, plus the
#                                            per-file `declarations` override
#                                            "zip:csv:utf-16:tab"
# and `source_ownership_manifest` keys its archive rules off the `zip` PREFIX
# (`_assert_archive_inventory`), which is why the qualified form keeps it.

CSV_UTF16_TAB_ENCODING = "utf-16"
CSV_UTF16_TAB_SEPARATOR = "\t"


def _assert_declared_format(entry: dict, accepted: tuple, reader_name: str):
    """A reader may only open what its own transcription declares it can.

    Not decoration. `read_revenue` read every consumed entry with
    `pd.read_excel` for as long as the family held one workbook, and the day a
    second format was declared the assumption became a crash on the L3
    prospective path and stayed invisible on the sealed L2 path (which clips at
    `window_end` and never reaches the new month).
    """
    fmt = str(entry.get("format", ""))
    if fmt not in accepted:
        raise ReaderError(
            "abort: %s is declared format %r, and %s has no reader for it (it "
            "transcribes %s). A reader that guessed a dialect here would not "
            "raise — a wrong encoding/separator pair yields a single-column "
            "frame — so an unhandled declared format aborts instead."
            % (entry.get("locator", "<entry>"), fmt, reader_name,
               ", ".join(repr(a) for a in accepted)))
    return fmt


def _table_xlsx(path: str, entry: dict):
    import pandas as pd

    return pd.read_excel(path, engine="openpyxl")


def _table_csv_utf16_tab(path: str, entry: dict):
    import pandas as pd

    return pd.read_csv(path, encoding=CSV_UTF16_TAB_ENCODING,
                       sep=CSV_UTF16_TAB_SEPARATOR)


def _table_zip_csv_utf16_tab(path: str, entry: dict):
    """The same dialect one container down: one csv per DECLARED member.

    The member inventory is checked first and the members are read in the
    inventory's order — not `namelist()`'s. A member that appeared inside a
    declared archive is as invisible as a file that appeared in a declared
    directory, and `assert_archive_members_match` is where that is caught.
    """
    import pandas as pd

    declared = [m["name"] for m in (entry.get("members") or ())]
    if not declared:
        raise ReaderError(
            "abort: %s is declared %r but carries no member inventory. A "
            "qualified archive format is a promise about what is INSIDE the "
            "container, and it cannot be honoured against an unlisted member "
            "set." % (entry.get("locator", "<entry>"), entry.get("format")))
    assert_archive_members_match(path, entry)

    frames = []
    with zipfile.ZipFile(path) as zf:
        for name in declared:
            txt = zf.read(name).decode(CSV_UTF16_TAB_ENCODING)
            frames.append(pd.read_csv(io.StringIO(txt),
                                      sep=CSV_UTF16_TAB_SEPARATOR))
    return pd.concat(frames, ignore_index=True) if len(frames) > 1 else frames[0]


# The registry IS the contract. A format may not be declared without an entry
# here, and an entry here is a statement that the dialect was transcribed.
DECLARED_TABLE_READERS = {
    "xlsx": _table_xlsx,
    "csv:utf-16:tab": _table_csv_utf16_tab,
    "zip:csv:utf-16:tab": _table_zip_csv_utf16_tab,
}


def _read_declared_table(entry: dict, path: str):
    """Open one consumed entry as a frame, chosen by its DECLARED format."""
    fmt = str(entry.get("format", ""))
    reader = DECLARED_TABLE_READERS.get(fmt)
    if reader is None:
        raise ReaderError(
            "abort: no reader is registered for declared format %r (%s). A "
            "format may not be declared without a reader that can honour it, "
            "and guessing one from the extension is how a UTF-16/TAB export "
            "becomes a silent single-column frame. Register the dialect in "
            "DECLARED_TABLE_READERS or do not declare it."
            % (fmt, entry.get("locator", os.path.basename(path))))
    return reader(path, entry)


# --- shared TEJ dialect ---------------------------------------------------------
#
# THREE different "parse a number" rules live in this module, and they are three
# rules rather than one because three sources mean different things by a blank:
#
#     _num       a valuation ratio; 0.0 is UNDEFINED, not a value
#     _ca_num    a share quantity; "." is absent, 0 is a real zero
#     _bonus_num an exchange figure that may carry a unit suffix
#
# Collapsing them would be the quiet kind of wrong.

def _zip_tsv_rows(path: str, expected_columns: tuple) -> list:
    """UTF-16 TSV inside a zip, one member per era.

    The encoding/separator pair is pinned rather than sniffed for the same
    reason it is pinned for prices: reading this with the default separator
    yields a one-column frame and raises nothing.
    """
    rows = []
    with zipfile.ZipFile(path) as zf:
        for name in zf.namelist():
            lines = [ln for ln in zf.read(name).decode("utf-16").splitlines()
                     if ln.strip()]
            header = lines[0].split("\t")
            if header != list(expected_columns):
                raise ReaderError(
                    "abort: %s:%s has schema %s, not the declared %s. A schema "
                    "change is a source change, not something a reader absorbs."
                    % (os.path.basename(path), name, header,
                       list(expected_columns)))
            rows += [dict(zip(header, ln.split("\t"))) for ln in lines[1:]]
    return rows


def _d8(v):
    """TEJ's 8-digit date stamp -> ISO, or None. `.` means absent."""
    s = str(v).strip().split(".")[0]
    return "%s-%s-%s" % (s[:4], s[4:6], s[6:]) if len(s) == 8 and s.isdigit() else None


def _sid(value) -> str:
    """`1101 台泥` -> `1101`. The merged id/name column TEJ ships."""
    return str(value).split()[0]


# Transcribed from `tej_importer`. The alias table is deliberately tiny and
# exact: the 2026-08-10 Q2 re-export ships 年/月 where the older template ships
# 年月, and nothing else is guessed.
SOURCE_COLUMN_ALIASES = {"年月": ("年/月",)}
ID_SPLIT_COLS = ("證券代碼", "股票代號")


def _normalize_aliases(df, source_name: str = ""):
    """Alias -> canonical, fail-closed when both are present and disagree."""
    out = df.copy()
    for canonical, aliases in SOURCE_COLUMN_ALIASES.items():
        present = [a for a in aliases if a in out.columns]
        if canonical in out.columns:
            for alias in present:
                left = out[canonical].astype(str).str.strip()
                right = out[alias].astype(str).str.strip()
                bad = ~(left.eq(right)
                        | (out[canonical].isna() & out[alias].isna()))
                if bool(bad.any()):
                    raise ReaderError(
                        "abort: %s has %d row(s) where canonical %r and alias "
                        "%r disagree. A silent overwrite here would change what "
                        "period a statement belongs to."
                        % (source_name or "<source>", int(bad.sum()),
                           canonical, alias))
        elif present:
            if len(present) > 1:
                raise ReaderError(
                    "abort: %s carries %d aliases for %r (%s); which one is the "
                    "column is not decidable here."
                    % (source_name or "<source>", len(present), canonical,
                       present))
            out[canonical] = out[present[0]]
    return out


def _split_id_name(df):
    """`代號`/`名稱` split, or the merged `1101 台泥` form. Both, nothing else."""
    df = df.copy()
    if "代號" in df.columns and "名稱" in df.columns:
        df["stock_id"] = df["代號"].astype(str).str.strip()
        df["stock_name"] = df["名稱"].astype(str).str.strip()
        return df
    id_col = next((c for c in ID_SPLIT_COLS if c in df.columns), None)
    if id_col is None:
        raise ReaderError(
            "abort: no identifier column (代號 or one of %s); columns are %s"
            % (list(ID_SPLIT_COLS), list(df.columns)))
    parts = df[id_col].astype(str).str.strip().str.split(n=1, expand=True)
    df["stock_id"] = parts[0].str.strip()
    df["stock_name"] = parts[1].str.strip() if 1 in parts.columns else ""
    return df


# --- archive_with_member_inventory: prices --------------------------------------
#
# The two legs are two FORMATS as well as two eras, and the reader hardcodes a
# dialect per leg. Naming the formats it transcribes is what makes a third one
# an abort instead of a misparse (`build_prices_leaf` writes "zip" for the
# 2019+ archives and "parquet" for the pre-2019 cache).
PRICES_ARCHIVE_FORMATS = ("zip",)
PRICES_CACHE_FORMATS = ("parquet",)


def read_prices(run_dir: str, date_min: str, date_max: str):
    """Both price legs, from the sources the leaf declares.

    §2.8.3 splits the lineage at 2019-01-01 and the two halves disagree about
    what a volume number means, which is a difference that does not raise:

        <= 2018   per-security parquet, `Trading_Volume` ALREADY in shares
        >= 2019   UTF-16/TAB inside archives, 成交量(千股), restored by x1000

    Mixing the conventions moves every security 1000x across §4.2's absolute
    NTD liquidity floor. The wrong encoding/separator pair silently yields one
    column. Both are pinned per leg rather than sniffed.
    """
    import pandas as pd

    leaf, landing = _leaf_and_landing(run_dir, "prices")
    entries = consumed_entries(leaf)
    if not entries:
        raise ReaderError("abort: the prices leaf consumes nothing")

    frames = [_read_pre_2019_leg(landing, entries, date_min, date_max)]
    for entry in entries:
        if entry.get("leg") == "pre-2019":
            continue
        _assert_declared_format(entry, PRICES_ARCHIVE_FORMATS,
                                "the 2019+ price leg")
        path = _verified_path(landing, entry)
        # The member inventory is the contract one level down.
        assert_archive_members_match(path, entry)
        declared = [m["name"] for m in entry["members"]]
        if len(declared) != 1:
            raise ReaderError(
                "abort: %s declares %d members; the price leg expects exactly "
                "one CSV per archive." % (entry["locator"], len(declared)))

        with zipfile.ZipFile(path) as zf:
            with zf.open(declared[0]) as fh:
                txt = fh.read().decode("utf-16")

        d = pd.read_csv(io.StringIO(txt), sep="\t", dtype=str,
                        usecols=list(PRICE_COLUMNS))
        d["date"] = pd.to_datetime(d["年月日"], errors="coerce").dt.strftime(
            "%Y-%m-%d")
        d = d[d["date"].notna()]
        d = d[(d["date"] >= date_min) & (d["date"] <= date_max)]
        d["stock_id"] = d["證券代碼"].astype(str).str.split().str[0].str.strip()
        d["open"] = pd.to_numeric(d["開盤價(元)"], errors="coerce")
        d["close"] = pd.to_numeric(d["收盤價(元)"], errors="coerce")
        d["volume_shares"] = pd.to_numeric(
            d["成交量(千股)"], errors="coerce") * VOLUME_THOUSANDS_TO_SHARES
        frames.append(d[["stock_id", "date", "open", "close", "volume_shares"]])

    archive = pd.concat(frames[1:], ignore_index=True)
    early = archive[archive["date"] < VINTAGE_BOUNDARY]
    if len(early):
        raise ReaderError(
            "abort: %d rows from the archive leg are dated before %s; the "
            "pre-2019 era belongs to a different lineage (§2.8.3)."
            % (len(early), VINTAGE_BOUNDARY))

    legs = [f for f in (frames[0], archive) if len(f)]
    out = pd.concat(legs, ignore_index=True) if legs else archive
    dupes = int(out.duplicated(subset=["stock_id", "date"]).sum())
    if dupes:
        raise ReaderError(
            "abort: %d (stock_id, date) key(s) appear in BOTH legs. The two "
            "meet at %s, not overlap — an overlapping row would mean two "
            "different unit conventions describing the same session."
            % (dupes, VINTAGE_BOUNDARY))
    return out.sort_values(["stock_id", "date"]).reset_index(drop=True)


PRE_2019_COLUMNS = ("stock_id", "date", "open", "close", "Trading_Volume")


def _read_pre_2019_leg(landing: str, entries, date_min: str, date_max: str):
    """The <= 2018 leg, one declared parquet per security.

    ⚠ The D-1 quarantine is on the 2019+ ERA of this cache, not on the cache.
    The same file holds admissible pre-2019 rows and quarantined 2019+ rows, so
    the restriction cannot be expressed by which files are declared: it is
    enforced here, on every file, and a row that survives it aborts.
    """
    import pandas as pd

    declared = [e for e in entries if e.get("leg") == "pre-2019"]
    if not declared:
        raise ReaderError(
            "abort: the prices leaf declares no pre-2019 leg. Without it every "
            "listing spell for a security already listed in 2018 opens at the "
            "first 2019 session — a fabricated listing date, not a missing one "
            "(O-G has no `opened_by` value for 'the corpus stops here').")

    frames = []
    for entry in declared:
        _assert_declared_format(entry, PRICES_CACHE_FORMATS,
                                "the pre-2019 price leg")
        path = _verified_path(landing, entry)
        d = pd.read_parquet(path, columns=list(PRE_2019_COLUMNS))
        if d.empty:
            continue
        d["date"] = d["date"].astype(str)
        d = d[d["date"] < VINTAGE_BOUNDARY]
        d = d[(d["date"] >= date_min) & (d["date"] <= date_max)]
        if d.empty:
            continue
        d["stock_id"] = d["stock_id"].astype(str)
        # ALREADY shares on this leg. No x1000 — see LEG_UNIT_CONVENTIONS.
        d["volume_shares"] = pd.to_numeric(d["Trading_Volume"], errors="coerce")
        frames.append(d[["stock_id", "date", "open", "close", "volume_shares"]])

    if not frames:
        return pd.DataFrame(columns=["stock_id", "date", "open", "close",
                                     "volume_shares"])
    out = pd.concat(frames, ignore_index=True)
    late = out[out["date"] >= VINTAGE_BOUNDARY]
    if len(late):
        raise ReaderError(
            "abort: %d rows from the cache leg are dated %s or later, which is "
            "the D-1 quarantined era of this cache."
            % (len(late), VINTAGE_BOUNDARY))
    return out


# --- board_date_payload_key: valuation ------------------------------------------

VALUATION_FORMATS = ("json:exchange_payload",)


def read_valuation(run_dir: str) -> dict:
    """{stock_id: {'per_tse', 'pbr_tse', 'board'}} for the leaf's session.

    Resolves columns BY THE DECLARED NAME. The harvester's own parser resolves
    them with `idx()`, which returns None on a rename and then skips every row —
    turning a renamed column into an all-NA session rather than an error. Here a
    rename has already aborted at declare time, and this reads the index the
    leaf recorded.
    """
    import json

    leaf, landing = _leaf_and_landing(run_dir, "valuation")
    out = {}
    for entry in consumed_entries(leaf):
        _assert_declared_format(entry, VALUATION_FORMATS, "the valuation reader")
        path = _verified_path(landing, entry)
        with open(path, encoding="utf-8") as fh:
            doc = json.load(fh)

        cur = doc
        for step in entry["rows_path"]:
            cur = cur[step] if isinstance(step, int) else cur.get(step)
        rows = cur or []
        fields = entry["resolved_fields"]
        i_id = fields.index(entry["id_field"])
        i_pbr = fields.index("股價淨值比")
        i_pe = fields.index("本益比")

        for r in rows:
            sid = str(r[i_id]).strip()
            out[sid] = {"board": entry["board"],
                        "pbr_tse": _num(r[i_pbr]), "per_tse": _num(r[i_pe])}
    if not out:
        raise ReaderError("abort: the valuation leaf yielded no rows")
    return out


def _num(value):
    """The frozen sentinel rule: 0.0 is UNDEFINED for a ratio, not a value."""
    try:
        v = float(str(value).replace(",", "").strip())
    except (TypeError, ValueError):
        return None
    if v == 0.0:
        return None                      # `valuation_sentinel_zero_is_undefined`
    return v


# --- flat_directory_filename: calendar ------------------------------------------

CALENDAR_FORMATS = ("parquet",)


def read_calendar(run_dir: str) -> tuple:
    """Sessions from the declared TAIEX series — never `data/b0/`."""
    import pandas as pd

    leaf, landing = _leaf_and_landing(run_dir, "calendar")
    entries = consumed_entries(leaf)
    if len(entries) != 1:
        raise ReaderError(
            "abort: the calendar leaf declares %d consumed sources; exactly one "
            "series defines the sessions." % len(entries))
    _assert_declared_format(entries[0], CALENDAR_FORMATS, "the calendar reader")
    path = _verified_path(landing, entries[0])
    df = pd.read_parquet(path)
    return tuple(sorted({str(d) for d in df["date"]}))


# --- archive set: security_status -----------------------------------------------

SUSPENSION_COLUMNS = ("證券代碼", "年月日", "恢復交易日", "暫停交易原因")
STATUS_SOURCE_LABEL = "TEJ 暫停交易"
# `_zip_tsv_rows` hardcodes UTF-16 + TAB one level inside the container. The
# plain `zip` this family declares carries that dialect; a QUALIFIED zip
# (`zip:csv:big5:comma`, say) would not, and must not be read by this path.
STATUS_ARCHIVE_FORMATS = ("zip",)

# O-E-1, declared rather than assumed: available_from = 年月日, and the guard
# then lets a record explain only sessions STRICTLY AFTER that date. This export
# carries no filing timestamp at all (O-F finding A-1), so the convention has to
# be stated somewhere — and it is stated in exactly one place per route.
STATUS_AVAILABILITY_CONVENTION = "available_from = 年月日 (O-E-1)"


def read_security_status(run_dir: str) -> list:
    """The canonical status records: one per filed change, resumptions included.

    The reason -> status mapping is NOT decided here. `core.b0_market_state`
    carries it because O-F ruling 4 made it normative: a row in this export is
    not automatically a suspension, and a 停止過戶 window or an uninterpretable
    reason must fail closed rather than be promoted. This reader reports what
    the normative mapping produced and nothing else.
    """
    from core.b0_market_state import (
        STATUS_DELISTED, STATUS_LISTED, assert_not_promoted_to_suspended,
        status_for_event,
    )

    leaf, landing = _leaf_and_landing(run_dir, "security_status")
    entries = consumed_entries(leaf)
    if not entries:
        raise ReaderError("abort: the security_status leaf consumes nothing")

    rows = []
    for entry in entries:
        _assert_declared_format(entry, STATUS_ARCHIVE_FORMATS,
                                "the security_status reader")
        path = _verified_path(landing, entry)
        assert_archive_members_match(path, entry)
        for r in _zip_tsv_rows(path, SUSPENSION_COLUMNS):
            sid = _sid(r["證券代碼"])
            start, resume = _d8(r["年月日"]), _d8(r["恢復交易日"])
            reason = str(r["暫停交易原因"]).strip()
            if not start:
                continue
            status = status_for_event(reason)
            if status is None:
                # Fail closed: no record, so it can never stand over a session
                # as an explanation for a missing price.
                continue
            assert_not_promoted_to_suspended(reason, status)
            rows.append({"stock_id": sid, "status": status,
                         "effective_from": start, "available_from": start,
                         "reason": reason, "source": STATUS_SOURCE_LABEL})
            # A resumption is a separate filed fact, knowable on the day it
            # happens. Emitting it lets a later `listed` record cancel an
            # earlier suspension instead of the suspension explaining gaps
            # forever.
            if resume and status != STATUS_DELISTED:
                rows.append({"stock_id": sid, "status": STATUS_LISTED,
                             "effective_from": resume, "available_from": resume,
                             "reason": "resume", "source": STATUS_SOURCE_LABEL})
    if not rows:
        raise ReaderError("abort: the security_status leaf yielded no records")
    rows.sort(key=lambda r: (r["stock_id"], r["effective_from"], r["status"]))
    return rows


STATUS_COLUMNS = ("stock_id", "status", "effective_from", "available_from",
                  "reason", "source")


# --- archive set + leaf dependency: corporate_actions ---------------------------

CA_KIND_BY_COLUMN = {
    "減資(仟股)": "capital_reduction",
    # B0.3 R2/R4: keyed on the SOURCE COLUMN, which is the immutable provenance.
    # This export is a per-security share-formation table, so 合併(仟股) on a row
    # is that security's OWN issuance — issuer-side, never the holder leg.
    "合併(仟股)": "issuer_side_merger_share_issuance",
    "股份轉換(仟股": "issuer_side_share_conversion_issuance",
    "變更股票面額股數(仟股)": "par_value_change",
    "現金增資(仟股)": "cash_capital_increase",
    "証券轉換_可轉債(仟股)": "convertible_bond_conversion",
    "庫藏股註銷(仟股)": "treasury_cancellation",
    "員工分紅(仟股)": "employee_bonus",
    "受讓(仟股)": "transfer_in",
    "其它(仟股)": "other_share_change",
}
# Columns whose non-zero value on the same row would contaminate a share-count
# identity, so a reduction rate must not be derived from counts when present.
CA_SHARE_MOVING_COLUMNS = tuple(CA_KIND_BY_COLUMN) + ("盈餘增資(仟股)",
                                                      "公積增資(仟股)")

# B0.4: the second admitted lineage. `security_status` states, authoritatively,
# that a listed security ceased trading and why — the ONLY source in the corpus
# that establishes the disappearing side of a reorganization at all. These are
# the rows carrying `holder_side_reorganization_exit`, which is what B0.7
# terminates on at period 67.
REORGANIZATION_EXIT_REASONS = {
    "合併下市": "MERGER",
    "併入控股公司下市": "HOLDING_COMPANY_CONVERSION",
}

# Same UTF-16 archive dialect as security_status, and the same reason to name it.
CA_ARCHIVE_FORMATS = ("zip",)

LEDGER_COLUMNS = ("stock_id", "kind", "source_field", "ex_or_effective_date",
                  "reconstructibility", "reason", "credit_tradable_date",
                  "new_shares_thousands", "share_multiplier", "cash_per_share",
                  "cash_payment_date", "zero_day_receivable")


def _ca_norm_date(v):
    """Both shapes the 配股相關 export uses: `20140731` and `2014/07/31`."""
    s = str(v).strip()
    if not s or s == ".":
        return None
    if len(s) == 8 and s.isdigit():
        return "%s-%s-%s" % (s[:4], s[4:6], s[6:])
    p = s.replace("/", "-").split("-")
    if len(p) == 3 and all(x.isdigit() for x in p):
        return "%04d-%02d-%02d" % (int(p[0]), int(p[1]), int(p[2]))
    return None


def _ca_num(v):
    """A share quantity. `.` is absent; 0 is a real zero and stays one."""
    s = str(v).strip()
    if not s or s == ".":
        return None
    try:
        return float(s)
    except ValueError:
        return None


def _ca_load_rows(run_dir: str) -> list:
    """Every row of the seven declared 配股相關 archives, header-checked.

    Unlike the other archive families this export's header is NOT pinned to a
    constant here: it is wide, it differs across eras, and `CA_KIND_BY_COLUMN`
    already addresses it by name. A column that vanishes therefore produces
    absent quantities rather than an abort — which is the L2 behaviour, and
    changing it would be changing the ledger, not reading it.
    """
    leaf, landing = _leaf_and_landing(run_dir, "corporate_actions")
    entries = consumed_entries(leaf)
    if not entries:
        raise ReaderError("abort: the corporate_actions leaf consumes nothing")

    rows = []
    for entry in entries:
        _assert_declared_format(entry, CA_ARCHIVE_FORMATS,
                                "the corporate_actions reader")
        path = _verified_path(landing, entry)
        assert_archive_members_match(path, entry)
        with zipfile.ZipFile(path) as zf:
            for name in zf.namelist():
                lines = zf.read(name).decode("utf-16").split("\n")
                hdr = lines[0].rstrip("\r").split("\t")
                for line in lines[1:]:
                    if not line.strip():
                        continue
                    f = line.rstrip("\r").split("\t")
                    if len(f) < len(hdr):
                        continue
                    rows.append(dict(zip(hdr, f)))
    return rows


def _ca_build_records(rows):
    """Yield (kind, normalised record) for every share-changing row."""
    import collections

    # Previous known par value per security, for par-value-change events.
    seq = collections.defaultdict(list)
    for r in rows:
        d = _ca_norm_date(r.get("年月日"))
        if d:
            seq[_sid(r["證券代碼"])].append((d, r))
    for k in seq:
        seq[k].sort(key=lambda t: t[0])
    prev_par = {}
    for k, items in seq.items():
        running = None
        for d, r in items:
            prev_par[(k, d)] = running
            p = _ca_num(r.get("面額"))
            if p:
                running = p

    for r in rows:
        ex = _ca_norm_date(r.get("年月日"))
        sid = _sid(r["證券代碼"])
        tot = _ca_num(r.get("總股數(仟股)"))

        new_shares = ((_ca_num(r.get("盈餘增資(仟股)")) or 0.0)
                      + (_ca_num(r.get("公積增資(仟股)")) or 0.0))
        if new_shares != 0.0:
            li = _ca_norm_date(r.get("股票股利上市日"))
            pa = _ca_norm_date(r.get("股票股利發放日"))
            credit = max([d for d in (li, pa) if d], default=None)
            rate = ((_ca_num(r.get("盈餘配股率 %")) or 0.0)
                    + (_ca_num(r.get("公積配股率 %")) or 0.0))
            yield "stock_dividend", {
                "stock_id": sid, "ex_right_date": ex,
                "new_shares_thousands": new_shares,
                "distribution_ratio_pct": rate or None,
                "credit_tradable_date": credit,
                "is_ex_right_event": str(r.get("配股(Y/N)", "")).strip() == "Y",
            }

        for col, kind in CA_KIND_BY_COLUMN.items():
            q = _ca_num(r.get(col))
            if not q:
                continue
            rec = {"stock_id": sid, "ex_right_date": ex, "effective_date": ex,
                   "quantity_thousands": q, "total_shares_thousands": tot}
            if kind == "capital_reduction":
                rate = _ca_num(r.get("減資率 %"))
                if rate is None and tot:
                    # Arithmetic identity, not a model: reduction / pre-event
                    # shares. Only usable when nothing else moved the count on
                    # the same row, otherwise the identity is contaminated.
                    contaminated = any(
                        (_ca_num(r.get(c)) or 0.0) != 0.0
                        for c in CA_SHARE_MOVING_COLUMNS if c != col)
                    if not contaminated:
                        pre = tot + q
                        if pre > 0:
                            rate = q / pre * 100.0
                            rec["reduction_rate_derived"] = True
                rec["reduction_rate_pct"] = rate
                rec["effective_date"] = _ca_norm_date(r.get("除權減資基準日")) or ex
                rec["cash_per_share"] = _ca_num(r.get("減資每股退還現金"))
                rec["cash_payment_date"] = _ca_norm_date(r.get("減資現金退款日"))
            elif kind == "par_value_change":
                rec["new_par"] = _ca_num(r.get("面額"))
                rec["old_par"] = prev_par.get((sid, ex))
                rec["changed_shares_thousands"] = q
            yield kind, rec


def assert_status_dependency_holds(run_dir: str) -> str:
    """The CA leaf binds THIS run's security_status leaf by payload hash.

    The aggregate validator checks this when the source set is assembled;
    checking it again at read time is the same argument as re-hashing a file:
    the binding is what makes "the status source this ledger was built from"
    a fact rather than a coincidence of which files happen to be on disk.
    """
    leaf, _ = _leaf_and_landing(run_dir, "corporate_actions")
    dep = (leaf.get("derived_dependencies") or {}).get("security_status")
    if not dep:
        raise ReaderError(
            "abort: the corporate_actions leaf declares no security_status "
            "dependency, but its holder-side reorganization exits come from "
            "nowhere else. A ledger built without it is missing the "
            "disappearing side of every merger.")
    status_leaf = load_leaf(os.path.join(run_dir, dep["leaf"]))
    if status_leaf[SELF_HASH_FIELD] != dep["payload_sha256"]:
        raise ReaderError(
            "abort: the security_status leaf in this run is not the one the "
            "corporate_actions leaf was built against.\n"
            "  bound:   %s\n  present: %s"
            % (dep["payload_sha256"], status_leaf[SELF_HASH_FIELD]))
    return dep["payload_sha256"]


def read_corporate_actions(run_dir: str, status_rows=None) -> list:
    """The canonical ledger rows: 配股相關 events PLUS holder-side exits.

    Reconstructibility is not decided here. `core.b0_corporate_actions.classify`
    decides it, and a NOT_RECONSTRUCTIBLE row is part of the source rather than
    an absence to be filtered out — B0.7 terminates on exactly those rows, and a
    reader that dropped them would make the run look like it had more history
    than it does.
    """
    from core.b0_corporate_actions import (
        EVENT_KIND_BY_KEY, assert_every_holder_affecting_kind_has_a_handler,
        classify,
    )

    assert_every_holder_affecting_kind_has_a_handler()
    assert_status_dependency_holds(run_dir)

    rows = _ca_load_rows(run_dir)
    events = [classify(kind, rec) for kind, rec in _ca_build_records(rows)]

    # B0.4 · holder-side coverage, materialized for EVERY status-defined
    # disappearance — no filter on price-universe membership, window or
    # holdings. The coverage invariant is about the corpus, not about what a
    # particular run happened to touch.
    if status_rows is None:
        status_rows = read_security_status(run_dir)
    for r in status_rows:
        if r["status"] != "delisted":
            continue
        if r["reason"] not in REORGANIZATION_EXIT_REASONS:
            continue
        events.append(classify("holder_side_reorganization_exit", {
            "stock_id": r["stock_id"],
            "effective_date": r["effective_from"],
            "status_reason": r["reason"]}))

    out = []
    for e in sorted(events, key=lambda x: (x.ex_or_effective_date, x.stock_id,
                                           x.kind)):
        row = {c: getattr(e, c) for c in LEDGER_COLUMNS if c != "source_field"}
        spec = EVENT_KIND_BY_KEY.get(e.kind)
        # B0.3 R4: the source column travels with the event, so a later audit
        # can be run on PROVENANCE rather than on the canonical kind name —
        # which is precisely the shortcut that produced the conflation.
        row["source_field"] = spec.source_column if spec else ""
        out.append({c: row[c] for c in LEDGER_COLUMNS})
    return out


# --- harvested_payload_key: bonus_shares ----------------------------------------

def _bonus_num(v):
    """An exchange figure, which may arrive with a unit suffix attached."""
    s = str(v).replace(",", "").strip()
    for tail in ("元／股", "元/股", "股", "元"):
        if s.endswith(tail):
            s = s[: -len(tail)].strip()
    if s in ("", "-", "--", ".", "N/A", "nan", "None"):
        return None
    try:
        f = float(s)
    except ValueError:
        return None
    return None if f != f else f


BONUS_FORMATS = ("json:harvested_envelope",)


def _bonus_envelopes(run_dir: str, layer_prefix: str):
    """Declared envelopes of one layer, hash-checked, in leaf order."""
    import json

    leaf, landing = _leaf_and_landing(run_dir, "bonus_shares")
    for entry in consumed_entries(leaf):
        if not entry["payload_key"].startswith(layer_prefix):
            continue
        _assert_declared_format(entry, BONUS_FORMATS, "the bonus_shares reader")
        path = _verified_path(landing, entry)
        with open(path, encoding="utf-8") as fh:
            yield entry, json.load(fh)


def _bonus_official_index(run_dir: str):
    """(stock_id, scheduled_date) -> row, from the two range reports.

    The range reports establish contemporaneous BOARD membership: a security is
    on TWSE for a date because the TWSE report for that date carries it. TPEx
    carries the bonus allotment in the range table itself; TWSE needs the
    per-event detail, joined separately.
    """
    from core.b0_bonus_share_source import OFFICIAL_BONUS_FIELD

    tpex_b = OFFICIAL_BONUS_FIELD["TPEx"]
    idx, upstream = {}, {}

    for _entry, rec in _bonus_envelopes(run_dir, "twse_range_"):
        upstream[rec["key"]] = rec["sha256"]
        pay = rec["payload"]
        fields = pay.get("fields") or []
        if not fields:
            continue
        ix = {c: i for i, c in enumerate(fields)}
        for r in pay.get("data") or []:
            s = str(r[ix["資料日期"]])
            d = "%04d-%02d-%02d" % (int(s.split("年")[0]) + 1911,
                                    int(s.split("年")[1].split("月")[0]),
                                    int(s.split("月")[1].split("日")[0]))
            idx[(str(r[ix["股票代號"]]).strip(), d)] = {
                "board": "TWSE", "bonus_per_1000": None,
                "payload_key": rec["key"], "payload_sha256": rec["sha256"]}

    for _entry, rec in _bonus_envelopes(run_dir, "tpex_range_"):
        upstream[rec["key"]] = rec["sha256"]
        for tb in rec["payload"].get("tables") or []:
            fields = tb.get("fields") or []
            if not fields:
                continue
            if tpex_b not in fields:
                raise ReaderError(
                    "abort: TPEx schema in %s has no %r column. The disclosure "
                    "changed; the parser version must change with it rather "
                    "than this reader guessing a replacement."
                    % (rec["key"], tpex_b))
            ix = {c: i for i, c in enumerate(fields)}
            for r in tb.get("data") or []:
                q = str(r[ix["除權息日期"]]).split("/")
                if len(q) != 3 or not q[0].isdigit():
                    continue
                d = "%04d-%02d-%02d" % (int(q[0]) + 1911, int(q[1]), int(q[2]))
                idx[(str(r[ix["代號"]]).strip(), d)] = {
                    "board": "TPEx", "bonus_per_1000": _bonus_num(r[ix[tpex_b]]),
                    "payload_key": rec["key"], "payload_sha256": rec["sha256"]}
    return idx, upstream


def _bonus_detail_index(run_dir: str):
    from core.b0_bonus_share_source import OFFICIAL_BONUS_FIELD

    twse_a = OFFICIAL_BONUS_FIELD["TWSE"]
    out, upstream = {}, {}
    for _entry, rec in _bonus_envelopes(run_dir, "twse_detail_"):
        upstream[rec["key"]] = rec["sha256"]
        stk, ymd = rec["key"].split("twse_detail_")[1].rsplit("_", 1)
        iso = "%s-%s-%s" % (ymd[:4], ymd[4:6], ymd[6:8])
        pay = rec["payload"]
        fields, data = pay.get("fields") or [], pay.get("data") or []
        if not fields or not data:
            out[(stk, iso)] = {"bonus_per_1000": None,
                               "payload_key": rec["key"],
                               "payload_sha256": rec["sha256"]}
            continue
        hits = [i for i, c in enumerate(fields) if c.strip() == twse_a.strip()]
        if len(hits) != 1:
            raise ReaderError(
                "abort: TWSE detail schema in %s resolves %r to %d columns. "
                "The report changed; the parser version must change with it."
                % (rec["key"], twse_a, len(hits)))
        out[(stk, iso)] = {"bonus_per_1000": _bonus_num(data[0][hits[0]]),
                           "payload_key": rec["key"],
                           "payload_sha256": rec["sha256"]}
    return out, upstream


def read_bonus_shares(run_dir: str, window_from: str, window_to: str,
                      ledger=None, sessions=None):
    """The C-51 holder-multiplier panel for the stock_dividend events in a window.

    The window is an ARGUMENT, not a constant. L2's is the union every
    141-period momentum_12_1 / sigma20d lookback reaches, which is a property of
    L2's frozen window; an L3 run's is a property of its own. Baking L2's in
    would make a prospective panel that silently stops in March.

    `m = 1 + 每千股無償配股/1000` and the dispositions come from
    `core.b0_bonus_share_source`. An UNRESOLVED event never carries a number,
    and that is enforced there rather than promised here.
    """
    import collections

    import pandas as pd

    from core.b0_bonus_share_source import (
        BONUS_PARSER_VERSION, MATCHED_DISPOSITION, OFFICIAL_ENDPOINT,
        assert_no_inferred_multiplier, assert_same_market_effective_event,
        holder_multiplier_from_bonus, is_pre_listing, market_effective_session,
        resolve_disposition,
    )

    if sessions is None:
        sessions = list(read_calendar(run_dir))
    sessions = sorted(sessions)
    sset = set(sessions)
    if ledger is None:
        ledger = read_corporate_actions(run_dir)

    events = [r for r in ledger
              if r["kind"] == "stock_dividend"
              and r["ex_or_effective_date"]
              and window_from <= r["ex_or_effective_date"] <= window_to]

    idx, up_range = _bonus_official_index(run_dir)
    det, up_det = _bonus_detail_index(run_dir)
    by_sec = collections.defaultdict(set)
    for sid, d in idx:
        by_sec[sid].add(d)

    def official_for(sid, ex):
        """Which official row, if any, describes this ledger event.

        Exact key first, then the closed-market normalisation — which is
        checked by the normative module rather than re-implemented here.
        """
        if (sid, ex) in idx:
            return ex
        for sched in sorted(by_sec.get(sid, ())):
            if sched in sset or sched >= ex:
                continue
            try:
                assert_same_market_effective_event(sched, ex, sessions)
            except Exception:
                continue
            if market_effective_session(sched, sessions) == ex:
                return sched
        return None

    have_detail = {(a, b.replace("-", "")) for a, b in det}
    missing = []
    for ev in events:
        sid, ex = ev["stock_id"], ev["ex_or_effective_date"]
        sched = official_for(sid, ex)
        if (sched is not None and idx[(sid, sched)]["board"] == "TWSE"
                and (sid, sched.replace("-", "")) not in have_detail):
            missing.append([sid, sched.replace("-", "")])
    if missing:
        raise ReaderError(
            "abort: %d TWSE detail payload(s) this panel needs are not declared "
            "in the bonus_shares leaf, e.g. %s. Harvest them and build a NEW "
            "run manifest — a run may not acquire a source after its sources "
            "were bound." % (len(missing), missing[:5]))

    rows = []
    for ev in events:
        sid, ex = ev["stock_id"], ev["ex_or_effective_date"]
        sched = official_for(sid, ex)
        board = bonus = pkey = psha = None
        if sched is not None:
            row = idx[(sid, sched)]
            board = row["board"]
            pkey, psha = row["payload_key"], row["payload_sha256"]
            if board == "TWSE":
                d = det[(sid, sched)]
                bonus = d["bonus_per_1000"]
                pkey, psha = d["payload_key"], d["payload_sha256"]
            else:
                bonus = row["bonus_per_1000"]
            if bonus is not None and bonus <= 0:
                # An official row that says "no bonus" is not a multiplier of 1
                # waiting to be applied; it contradicts the ledger's
                # classification of the event and must not become one.
                bonus = None
        prior = [d for d in by_sec.get(sid, ()) if d < ex]
        pre = is_pre_listing(ex, prior)
        disp = resolve_disposition(official_bonus_per_1000=bonus,
                                   pre_listing=pre)
        mult = (holder_multiplier_from_bonus(bonus)
                if disp == MATCHED_DISPOSITION else None)
        assert_no_inferred_multiplier(disp, mult)
        rows.append({
            "stock_id": sid,
            "official_scheduled_ex_right_date": sched,
            "market_effective_session": ex,
            "board": board,
            "source_endpoint": OFFICIAL_ENDPOINT[board] if board else None,
            "payload_key": pkey,
            "payload_sha256": psha,
            "bonus_shares_per_1000": bonus,
            "holder_multiplier": mult,
            "disposition": disp,
            # Carried, not folded into the disposition. R2's NOT_APPLICABLE is
            # about B0 having no market history to adjust; the ledger's own
            # NOT_RECONSTRUCTIBLE verdict is a different reason for arriving
            # there, and collapsing the two would make the panel claim more
            # than it knows.
            "ledger_reconstructibility": ev["reconstructibility"],
            "parser_version": BONUS_PARSER_VERSION,
        })
    if not rows:
        raise ReaderError(
            "abort: no stock_dividend events in %s..%s; a bonus panel with no "
            "rows would silently disable the share-unit adjustment."
            % (window_from, window_to))
    panel = pd.DataFrame(rows).sort_values(
        ["stock_id", "market_effective_session"]).reset_index(drop=True)
    panel.attrs["upstream_sha256"] = {**up_range, **up_det}
    return panel


# --- flat_directory_filename: financials ----------------------------------------
#
# Transcribed from `tej_importer.DATASETS["financial_statements"]`, not imported
# from it. See the module docstring for why.

FINANCIALS_PERIOD_COL = "年月"
FINANCIALS_ID_COL = "證券代碼"
FINANCIALS_CONSOLIDATION_COL = "合併(Y/N)"
FINANCIALS_PERIOD_FORMATS = ("%Y%m", "%Y/%m")
FINANCIALS_RELEASE_COL = "財報發布日"
FINANCIALS_RELEASE_FORMAT = "%Y/%m/%d"
FINANCIALS_REQUIRED_COLUMNS = (
    "證券代碼", "年月", "季別", "財報發布日", "營業收入淨額", "歸屬母公司淨利（損）",
    "每股盈餘", "資產總額", "負債總額", "股東權益總額", "來自營運之現金流量")
FINANCIALS_RENAME = {
    "季別": "quarter",
    "財報發布日": "release_date",
    "營業收入淨額": "_revenue_thousand",
    "營業毛利": "_gross_profit_thousand",
    "營業利益": "_operating_income_thousand",
    "歸屬母公司淨利（損）": "_net_income_thousand",
    "每股盈餘": "eps",
    "資產總額": "_total_assets_thousand",
    "負債總額": "_total_liab_thousand",
    "流動資產": "_current_assets_thousand",
    "流動負債": "_current_liab_thousand",
    "股東權益總額": "_equity_thousand",
    "來自營運之現金流量": "_ocf_thousand",
    "  購置不動產廠房設備（含預付）－CFI": "_capex_thousand",
}
FINANCIALS_THOUSAND_COLS = {
    "_revenue_thousand": "revenue",
    "_gross_profit_thousand": "gross_profit",
    "_operating_income_thousand": "operating_income",
    "_net_income_thousand": "net_income",
    "_total_assets_thousand": "total_assets",
    "_total_liab_thousand": "total_liabilities",
    "_current_assets_thousand": "current_assets",
    "_current_liab_thousand": "current_liabilities",
    "_equity_thousand": "equity",
    "_ocf_thousand": "operating_cash_flow",
    "_capex_thousand": "capex",
}
FINANCIALS_NUMERIC_COLS = ("eps", "quarter")
THOUSANDS = 1000.0


def _norm_period(value) -> str:
    """`年月` -> canonical 'YYYYMM'. Both frozen formats, nothing else."""
    s = str(value).strip().replace("/", "")
    if len(s) != 6 or not s.isdigit():
        raise ReaderError(
            "abort: period value %r is not one of the frozen formats %s. A "
            "third format is a source change, not something a reader absorbs."
            % (value, list(FINANCIALS_PERIOD_FORMATS)))
    return s


def _owns_predicate(declaration):
    """`['202606']` or `'<= 202603'` -> a predicate over canonical periods."""
    if isinstance(declaration, (tuple, list)):
        owned = {_norm_period(p) for p in declaration}
        return lambda p: p in owned
    if isinstance(declaration, str) and declaration.startswith("<="):
        bound = _norm_period(declaration[2:])
        return lambda p: p <= bound
    raise ReaderError(
        "abort: ownership declaration %r is not a supported form. Use a list of "
        "periods or '<= YYYYMM'. An unparsed declaration is an undeclared one."
        % (declaration,))


def read_financials(run_dir: str):
    """Quarterly fundamentals with their real announcement date (§2.2).

    Two sources overlap on 2026-06, so which one owns a period is DECLARED in
    the leaf and applied here. Nothing is dropped silently: a period a file
    contains but neither owns nor yields aborts.

    No window filter. §2.2's availability rule (`release_date <= decision_date`)
    belongs to the decision, not to the reader.
    """
    import pandas as pd

    leaf, landing = _leaf_and_landing(run_dir, "financials")
    entries = consumed_entries(leaf)
    if not entries:
        raise ReaderError("abort: the financials leaf consumes nothing")

    frames, owned_by = [], {}
    for entry in entries:
        name = entry["locator"]
        path = _verified_path(landing, entry)
        df = _read_declared_table(entry, path)
        df = _normalize_aliases(df, name)
        df = _split_id_name(df)

        missing = [c for c in FINANCIALS_REQUIRED_COLUMNS
                   if c != FINANCIALS_ID_COL and c not in df.columns]
        if missing:
            raise ReaderError(
                "abort: %s is missing frozen required column(s) %s. A source "
                "that cannot supply them cannot be imported under the frozen "
                "spec." % (name, missing))

        periods = df[FINANCIALS_PERIOD_COL].map(_norm_period)
        owns = _owns_predicate(entry["owns"])
        yields_ = {_norm_period(p) for p in (entry.get("yields") or ())}
        stray = sorted({p for p in periods.unique()
                        if not owns(p) and p not in yields_})
        if stray:
            raise ReaderError(
                "abort: %s contains period(s) %s that it neither owns nor "
                "yields. Dropping them would be the silent skip this contract "
                "replaces; keeping them would make two files canonical for one "
                "period. Declare which it is." % (name, ", ".join(stray)))

        keep = periods.map(owns)
        for p in periods[keep].unique():
            if owned_by.get(p, name) != name:
                raise ReaderError(
                    "abort: period %s was contributed by both %s and %s; "
                    "ownership did not partition the sources."
                    % (p, owned_by[p], name))
            owned_by[p] = name
        frames.append(df[keep].copy())

    raw = pd.concat(frames, ignore_index=True)
    key = raw.groupby([raw[FINANCIALS_PERIOD_COL].map(_norm_period),
                       raw["stock_id"].astype(str).str.strip()]).size()
    collided = key[key > 1]
    if len(collided):
        raise ReaderError(
            "abort: %d (period, security) key(s) appear more than once after "
            "ownership filtering, e.g. %s."
            % (len(collided), list(collided.index[:5])))

    df = raw.rename(columns=FINANCIALS_RENAME).copy()
    df["stock_id"] = raw["stock_id"].astype(str).str.strip()
    df["date"] = _parse_period(raw[FINANCIALS_PERIOD_COL])
    df["release_date"] = pd.to_datetime(raw[FINANCIALS_RELEASE_COL],
                                        format=FINANCIALS_RELEASE_FORMAT,
                                        errors="coerce")
    # §2.2: a row whose announcement date cannot be read has no availability,
    # and a fixed-lag substitute is forbidden. Dropped, never back-filled.
    df = df[df["release_date"].notna()].copy()

    for src, dst in FINANCIALS_THOUSAND_COLS.items():
        df[dst] = pd.to_numeric(df[src], errors="coerce") * THOUSANDS
    for c in FINANCIALS_NUMERIC_COLS:
        df[c] = pd.to_numeric(df[c], errors="coerce")
    df["consolidated"] = raw.loc[df.index, FINANCIALS_CONSOLIDATION_COL] \
        .astype(str).str.strip()

    keep_cols = (["stock_id", "date", "release_date", "quarter", "consolidated",
                  "eps"] + sorted(FINANCIALS_THOUSAND_COLS.values()))
    df = df[keep_cols].sort_values(["stock_id", "date"]).reset_index(drop=True)

    dupes = int(df.duplicated(subset=["stock_id", "date"]).sum())
    if dupes:
        raise ReaderError(
            "abort: %d duplicate (stock_id, period) rows. Which of two "
            "competing statements for one security-period wins is NOT specified "
            "by the master preregistration and must not be decided here (M-3)."
            % dupes)
    return df


def _parse_period(series):
    """Both formats the frozen spec accepts, and nothing else."""
    import pandas as pd

    out = pd.Series(pd.NaT, index=series.index, dtype="datetime64[ns]")
    raw = series.astype(str).str.strip()
    for fmt in FINANCIALS_PERIOD_FORMATS:
        need = out.isna()
        if not need.any():
            break
        out.loc[need] = pd.to_datetime(raw[need], format=fmt, errors="coerce")
    if out.isna().any():
        bad = sorted(raw[out.isna()].unique())[:5]
        raise ReaderError(
            "abort: %d period value(s) match neither frozen format %s: %s."
            % (int(out.isna().sum()), list(FINANCIALS_PERIOD_FORMATS), bad))
    return out


# --- flat_directory_filename: revenue -------------------------------------------

REVENUE_PERIOD_COL = "年月"
REVENUE_PERIOD_FORMAT = "%Y%m"
REVENUE_RELEASE_COL = "營收發布日"
REVENUE_RELEASE_FORMAT = "%Y%m%d"
REVENUE_AMOUNT_COL = "單月營收(千元)"
REVENUE_REQUIRED_COLUMNS = ("證券代碼", "年月", "營收發布日", "單月營收成長率％",
                            "單月營收(千元)")
REVENUE_CARRY = ("stock_id", "date", "release_date", "revenue")

# §3.5 derives revenue_yoy — and revenue_accel as a difference of exactly that
# quantity — through `core.b0_features.compute_revenue_yoy`. The vendor's own
# 單月營收成長率％ is required to be PRESENT (its absence would mean a different
# export) but is deliberately NOT carried: shipping it would put a second growth
# definition within reach of the decision layer.
REVENUE_EXCLUDED_BY_LINEAGE = ("revenue_yoy_pct", "cum_revenue",
                               "cum_revenue_last_year", "revenue_last_year")


def read_revenue(run_dir: str):
    """Monthly revenue with its REAL announcement date (§2.2).

    §2.2, quoted: 月營收：讀真實 `release_date`，不得使用固定 lag 代理. A month
    without one is dropped, never given a proxy — and §2.1 makes where those
    dates begin (2013-01) the binding constraint on the whole window.

    MIXED FORMATS, one family. Since 2026-08-30 this family declares two
    sources and they are not the same kind of file: a workbook and a zip
    wrapping a UTF-16/TAB csv (`zip:csv:utf-16:tab`). The parser therefore comes
    from the DECLARED format, exactly as it does for financials — reading every
    entry with `pd.read_excel` was what crashed the prospective path the day
    July's completed export was declared.

    OWNERSHIP, not de-duplication. The two sources OVERLAP on 202607: the
    workbook was exported 2026-08-06 and carries a PARTIAL month (406 of 2,002
    securities), the archive carries the completed one and OWNS it. On the 406
    they share, the archive is not merely wider — it is REVISED (3003:
    658,000 -> 657,875 千元). Which value survives may therefore not be decided
    by concat order, `drop_duplicates` or the duplicate-key guard below: it is
    decided by the leaf's `owns`/`yields` declaration, applied here.
    """
    import pandas as pd

    leaf, landing = _leaf_and_landing(run_dir, "revenue")
    entries = consumed_entries(leaf)
    if not entries:
        raise ReaderError("abort: the revenue leaf consumes nothing")

    # Ownership is a family-wide property or it is nothing — the same rule
    # `build_flat_leaves.build` enforces at declare time. An entry without
    # `owns` beside entries that have one is a claimant no overlap check can
    # see, so a partial declaration aborts rather than being half-applied.
    owning = [e for e in entries if "owns" in e]
    if owning and len(owning) != len(entries):
        raise ReaderError(
            "abort: the revenue leaf declares period ownership for %s but not "
            "for %s. Within one family ownership is declared for every consumed "
            "source or for none; an entry without `owns` is an undeclared "
            "claimant, and the row it wins would be decided by concat order."
            % ([e["locator"] for e in owning],
               [e["locator"] for e in entries if "owns" not in e]))

    frames, owned_by = [], {}
    for entry in entries:
        name = entry["locator"]
        path = _verified_path(landing, entry)
        raw = _read_declared_table(entry, path)
        raw = _normalize_aliases(raw, name)
        raw = _split_id_name(raw)
        missing = [c for c in REVENUE_REQUIRED_COLUMNS
                   if c != "證券代碼" and c not in raw.columns]
        if missing:
            raise ReaderError(
                "abort: %s lacks frozen required column(s) %s; without a real "
                "release_date §2.2 leaves no admissible way to date these rows."
                % (name, missing))

        df = pd.DataFrame(index=raw.index)
        df["stock_id"] = raw["stock_id"].astype(str).str.strip()
        df["date"] = pd.to_datetime(
            raw[REVENUE_PERIOD_COL].astype(str).str.strip(),
            format=REVENUE_PERIOD_FORMAT, errors="coerce")
        # TEJ writes "." for an absent announcement date (all pre-2013 rows).
        # Coerced to NaT and dropped below — never given a lag proxy (§2.2).
        rel = pd.to_numeric(raw[REVENUE_RELEASE_COL], errors="coerce")
        df["release_date"] = pd.to_datetime(
            rel.astype("Int64").astype(str).replace("<NA>", ""),
            format=REVENUE_RELEASE_FORMAT, errors="coerce")
        df["revenue"] = pd.to_numeric(raw[REVENUE_AMOUNT_COL],
                                      errors="coerce") * THOUSANDS

        if owning:
            # `_norm_period` on BOTH sides of the comparison, deliberately: two
            # normalisers for the two halves of one predicate is how a declared
            # period stops matching the period it names.
            periods = raw[REVENUE_PERIOD_COL].map(_norm_period)
            owns = _owns_predicate(entry["owns"])
            yields_ = {_norm_period(p) for p in (entry.get("yields") or ())}
            stray = sorted({p for p in periods.unique()
                            if not owns(p) and p not in yields_})
            if stray:
                raise ReaderError(
                    "abort: %s contains month(s) %s that it neither owns nor "
                    "yields. Dropping them would be a silent skip; keeping them "
                    "would make two exports canonical for one month. Declare "
                    "which it is." % (name, ", ".join(stray)))
            keep = periods.map(owns)
            for p in periods[keep].unique():
                if owned_by.get(p, name) != name:
                    raise ReaderError(
                        "abort: month %s was contributed by both %s and %s; "
                        "ownership did not partition the sources."
                        % (p, owned_by[p], name))
                owned_by[p] = name
            df = df[keep]

        frames.append(df[list(REVENUE_CARRY)])

    panel = pd.concat(frames, ignore_index=True)
    panel = panel[panel["release_date"].notna()].copy()

    dupes = int(panel.duplicated(subset=["stock_id", "date"]).sum())
    if dupes:
        raise ReaderError(
            "abort: %d duplicate (stock_id, month) revenue rows; which one is "
            "the month's revenue is not specified (M-3)." % dupes)

    panel = panel.sort_values(["stock_id", "date"]).reset_index(drop=True)
    panel["date"] = panel["date"].dt.strftime("%Y-%m-%d")
    panel["release_date"] = panel["release_date"].dt.strftime("%Y-%m-%d")
    return panel


# --- flat_directory_filename: industry ------------------------------------------

INDUSTRY_CHANGE_PAIRS = (("前三次TSE產業變更", "前三次TSE產業變更日"),
                         ("前二次TSE產業變更", "前二次TSE產業變更日"),
                         ("前一次TSE產業變更", "前一次TSE產業變更日"))
INDUSTRY_NO_HISTORY_EFFECTIVE_FROM = "1900-01-01"
INDUSTRY_CURRENT_COL = "TSE產業_代碼"
# Cell-addressed through openpyxl rather than through a frame, so this reader
# does not go via DECLARED_TABLE_READERS — but it is bound by the same rule.
INDUSTRY_FORMATS = ("xlsx",)


def _ind_code(v):
    if v is None:
        return None
    s = str(v).strip()
    return s.split()[0] if s and s != "." else None


def _ind_date(v):
    if v is None:
        return None
    s = str(v).strip()
    if not s or s == ".":
        return None
    s = s.replace("/", "-")
    return s[:10] if len(s) >= 10 else None


def read_industry(run_dir: str):
    """The §2.3 point-in-time TSE industry STEP FUNCTION, not a snapshot.

    A current-state lookup is industry look-ahead for about half the universe:
    1,203 securities (49.4%) have changed TSE industry at least once, so O-E
    ruled the live `industry_map` NOT_PIT_SAFE.

    Each 前N次 record is `(the industry it BECAME, effective date)`. Where the
    current column disagrees with the latest dated record, the interval from
    that record onward is UNRESOLVED — NOT back-filled from the snapshot, and
    NOT assumed to have kept the old classification. UNRESOLVED means industry
    NA, which means Value NA, which means §4.1 complete-case drops it.
    """
    import openpyxl
    import pandas as pd

    leaf, landing = _leaf_and_landing(run_dir, "industry")
    entries = consumed_entries(leaf)
    if len(entries) != 1:
        raise ReaderError(
            "abort: the industry leaf declares %d consumed sources; exactly one "
            "dated table defines the timeline." % len(entries))
    path = _verified_path(landing, entries[0])

    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    try:
        ws = wb[wb.sheetnames[0]]
        it = ws.iter_rows(values_only=True)
        idx = {c: i for i, c in enumerate(list(next(it)))}
        rows, unresolved = [], {}
        for r in it:
            if r[0] is None:
                continue
            sid = str(r[idx["代號"]]).strip()
            if not sid:
                continue
            recs = []
            for c_ind, c_date in INDUSTRY_CHANGE_PAIRS:
                code, d = _ind_code(r[idx[c_ind]]), _ind_date(r[idx[c_date]])
                if code and d:
                    recs.append((d, code))
            listed = _ind_date(r[idx["首次掛牌日期"]])
            first_ind = _ind_code(r[idx["首次掛牌TSE產業"]])
            if listed and first_ind:
                recs.append((listed, first_ind))
            cur = _ind_code(r[idx[INDUSTRY_CURRENT_COL]])
            if not recs and cur:
                # The one place a current value is used, and safe only because
                # there is no change history to contradict it.
                recs.append((INDUSTRY_NO_HISTORY_EFFECTIVE_FROM, cur))
            recs = sorted(set(recs))
            if cur and recs and recs[-1][1] != cur:
                unresolved[sid] = recs[-1][0]
            for d, code in recs:
                rows.append({"stock_id": sid, "effective_from": d,
                             "tse_ind_code": code})
    finally:
        wb.close()

    if not rows:
        raise ReaderError("abort: the industry leaf yielded no timeline rows")
    tl = pd.DataFrame(rows).sort_values(
        ["stock_id", "effective_from"]).reset_index(drop=True)
    tl["unresolved_from"] = tl["stock_id"].map(unresolved)
    return tl


def industry_as_of(timeline, as_of: str) -> dict:
    """{stock_id: tse_ind_code or UNRESOLVED} at a session. The only resolver.

    Returning the sentinel rather than dropping the security is deliberate:
    `SecurityPitInputs.pit_industry` defaults to UNRESOLVED and §3.2 turns that
    into Value = NA. A silently absent security and an UNRESOLVED one are the
    same outcome for Value but not the same fact, and the state should carry
    the fact.
    """
    from core.b0_features import INDUSTRY_UNRESOLVED

    sub = timeline[timeline["effective_from"] <= as_of]
    if sub.empty:
        return {}
    latest = sub.groupby("stock_id").last()
    out = {}
    for sid, row in latest.iterrows():
        unres = row["unresolved_from"]
        if isinstance(unres, str) and as_of >= unres:
            out[str(sid)] = INDUSTRY_UNRESOLVED
        else:
            out[str(sid)] = str(row["tse_ind_code"])
    return out


READERS = {
    "prices": read_prices,
    "valuation": read_valuation,
    "calendar": read_calendar,
    "security_status": read_security_status,
    "corporate_actions": read_corporate_actions,
    "bonus_shares": read_bonus_shares,
    "financials": read_financials,
    "revenue": read_revenue,
    "industry": read_industry,
}

# Every required family now has a reader. The mapping is kept (rather than
# deleted) because "not implemented" has to remain a DECLARED state: the test
# that `READERS | READERS_NOT_IMPLEMENTED == REQUIRED_DATASETS` is what would
# catch a tenth family arriving with no reader and nobody noticing.
READERS_NOT_IMPLEMENTED: dict = {}
