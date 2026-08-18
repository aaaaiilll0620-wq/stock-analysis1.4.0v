"""P0-R5 Phase A — Stage 2: monthly_revenue + financial_statements (READ-ONLY).

Approved scope only (docs/prereg_P0_R5_..._2026-08-16.md, Phase A). Reads the two
raw .xlsx files with openpyxl read_only=True. Calls no production module, writes
no cache, judges no tdcc_weekly (blocked by §3a-2).

SPECIFICATION GAP DISCLOSED HERE — §3 requires classifying every "必要列"
(required row) of the frozen decision-time stock-month population, but the prereg
never defines which monthly_revenue rows are "required" for a given
(decision month, ticker). That definition is made HERE, during Phase A, and is
therefore NOT pre-frozen. To stop the definition from driving the verdict, the
result is reported under three lookbacks (L = 3, 6, 12 months) and the verdict is
only treated as stable if it is identical under all three.

Definition used, for decision month m and ticker t in the frozen population at m:
  required rows = monthly_revenue rows for t whose 年月 falls in
                  [m_yyyymm - L, m_yyyymm] inclusive.
  RECONSTRUCTIBLE      -> every required row has a genuine parseable 營收發布日
  NOT_RECONSTRUCTIBLE  -> at least one required row carries '.' / blank / unparseable
  NO_ROWS              -> no rows exist in the window (reported separately; NOT
                          silently counted as either pass or fail)

Rationale for including m itself: if the current month's row is present with a
genuine date we can determine it was not yet public at m; if it carries a
placeholder we cannot. Including it is the conservative choice.
"""

import json
import os
import re
import sys

import openpyxl

REPO = os.path.dirname(os.path.dirname(os.path.dirname(os.path.abspath(__file__))))
CORPUS = os.path.join(REPO, "tej_exports", "DataExport0806")
HERE = os.path.dirname(os.path.abspath(__file__))
OUT = os.path.join(HERE, "phase_a_stage2_revenue_financials.json")

POP_DIR = os.path.join(REPO, "research", "p0_u1_canonical_universe",
                       "canonical_universe_by_date")
FROZEN_MONTHS = os.path.join(REPO, "research", "p0_r2_identity_collector",
                             "a_leg_parity_result.json")

MR_PATH = os.path.join(CORPUS, "月營收2004-202608", "20260806091706.xlsx")
FS_PATH = os.path.join(CORPUS, "財報2004~202606", "20260806090633.xlsx")

WINDOW_START = "2013-01-01"
LOOKBACKS = [3, 6, 12]

# §3 escalation thresholds (revision 3, AC-R5-1b). Partial preregistration:
# set with knowledge of P0-R4's observed values. Must not be adjusted after
# seeing Phase A results.
THRESH_FULL_WINDOW = 0.005   # 0.5%
THRESH_PER_MONTH = 0.020     # 2.0%

DOT = "."
_YYYYMMDD = re.compile(r"^\d{8}$")
_SLASHED = re.compile(r"^(\d{4})/(\d{1,2})/(\d{1,2})$")


def ticker_of(cell):
    if cell is None:
        return None
    s = str(cell).strip()
    if not s:
        return None
    return s.split()[0]


def ym_of(cell):
    if cell is None:
        return None
    s = str(cell).strip()
    return s if re.match(r"^\d{6}$", s) else None


def classify_yyyymmdd(v):
    """genuine | placeholder | blank | unparseable"""
    if v is None:
        return "blank"
    s = str(v).strip()
    if not s:
        return "blank"
    if s == DOT:
        return "placeholder"
    if _YYYYMMDD.match(s):
        mm, dd = int(s[4:6]), int(s[6:8])
        return "genuine" if 1 <= mm <= 12 and 1 <= dd <= 31 else "unparseable"
    if s.endswith(".0") and _YYYYMMDD.match(s[:-2]):
        return "genuine"
    return "unparseable"


def classify_slashed(v):
    if v is None:
        return "blank"
    s = str(v).strip()
    if not s:
        return "blank"
    if s == DOT:
        return "placeholder"
    m = _SLASHED.match(s)
    if m and 1 <= int(m.group(2)) <= 12 and 1 <= int(m.group(3)) <= 31:
        return "genuine"
    if _YYYYMMDD.match(s):
        return "genuine"
    return "unparseable"


def ym_shift(ym, back):
    y, m = int(ym[:4]), int(ym[4:])
    total = y * 12 + (m - 1) - back
    return "%04d%02d" % (total // 12, total % 12 + 1)


def load_window_and_population():
    with open(FROZEN_MONTHS, encoding="utf-8") as fh:
        per_date = json.load(fh)["a_leg_parity_result"]["per_date"]
    window = sorted(m for m in per_date if m >= WINDOW_START)
    pop = {}
    for m in window:
        path = os.path.join(POP_DIR, "canonical_universe_%s.csv" % m)
        tickers = []
        with open(path, encoding="utf-8-sig") as fh:
            next(fh)
            for line in fh:
                parts = line.split(",")
                if len(parts) > 1:
                    tickers.append(parts[1].strip())
        pop[m] = tickers
    return window, pop


def scan_monthly_revenue():
    """(ticker, ym) -> classification, plus per-year tallies over the whole file."""
    table = {}
    per_year = {}
    wb = openpyxl.load_workbook(MR_PATH, read_only=True, data_only=True)
    try:
        ws = wb[wb.sheetnames[0]]
        first = True
        for row in ws.iter_rows(values_only=True):
            if first:
                first = False
                continue
            t = ticker_of(row[0])
            ym = ym_of(row[1])
            if not t or not ym:
                continue
            cls = classify_yyyymmdd(row[2])
            table[(t, ym)] = cls
            y = ym[:4]
            per_year.setdefault(y, {}).setdefault(cls, 0)
            per_year[y][cls] += 1
    finally:
        wb.close()
    return table, per_year


def evaluate_monthly_revenue(window, pop, table, lookback):
    per_month = []
    tot_pairs = tot_bad = tot_norows = 0
    worst = []
    for m in window:
        m_ym = m[:4] + m[5:7]
        yms = [ym_shift(m_ym, k) for k in range(lookback + 1)]
        n = bad = norows = 0
        for t in pop[m]:
            n += 1
            seen = False
            broken = False
            for ym in yms:
                cls = table.get((t, ym))
                if cls is None:
                    continue
                seen = True
                if cls != "genuine":
                    broken = True
            if not seen:
                norows += 1
            elif broken:
                bad += 1
        rate = (bad / n) if n else 0.0
        per_month.append({"month": m, "stock_months": n,
                          "not_reconstructible": bad, "no_rows": norows,
                          "rate": rate})
        tot_pairs += n
        tot_bad += bad
        tot_norows += norows
        if rate > THRESH_PER_MONTH:
            worst.append({"month": m, "rate": rate, "not_reconstructible": bad,
                          "stock_months": n})
    full_rate = (tot_bad / tot_pairs) if tot_pairs else 0.0
    max_month = max(per_month, key=lambda r: r["rate"]) if per_month else None
    passes = (full_rate <= THRESH_FULL_WINDOW) and not worst
    return {
        "lookback_months": lookback,
        "total_stock_months": tot_pairs,
        "total_not_reconstructible": tot_bad,
        "total_no_rows_in_window": tot_norows,
        "full_window_rate": full_rate,
        "full_window_threshold": THRESH_FULL_WINDOW,
        "full_window_pass": full_rate <= THRESH_FULL_WINDOW,
        "months_exceeding_per_month_threshold": len(worst),
        "per_month_threshold": THRESH_PER_MONTH,
        "per_month_pass": not worst,
        "worst_months": sorted(worst, key=lambda r: -r["rate"])[:15],
        "max_rate_month": max_month,
        "item_status": "PASS" if passes else "FAIL",
        "per_month": per_month,
    }


def scan_financial_statements(window, pop):
    tickers = set()
    for m in window:
        tickers.update(pop[m])
    lo = ym_shift(window[0][:4] + window[0][5:7], 24)   # 2-year margin before window
    hi = window[-1][:4] + window[-1][5:7]
    per_year = {}
    tot = bad = 0
    samples = []
    wb = openpyxl.load_workbook(FS_PATH, read_only=True, data_only=True)
    try:
        ws = wb[wb.sheetnames[0]]
        first = True
        for row in ws.iter_rows(values_only=True):
            if first:
                first = False
                continue
            t = ticker_of(row[0])
            ym = ym_of(row[1])
            if not t or not ym or t not in tickers:
                continue
            if ym < lo or ym > hi:
                continue
            cls = classify_slashed(row[51])
            y = ym[:4]
            per_year.setdefault(y, {}).setdefault(cls, 0)
            per_year[y][cls] += 1
            tot += 1
            if cls != "genuine":
                bad += 1
                if len(samples) < 20:
                    samples.append({"ticker": t, "ym": ym, "value": repr(row[51]),
                                    "class": cls})
    finally:
        wb.close()
    rate = (bad / tot) if tot else 0.0
    return {
        "scope": "rows whose ticker appears in the frozen population and whose 年月 is in [%s, %s] (24-month margin before window start)" % (lo, hi),
        "rows_in_scope": tot,
        "non_genuine_rows": bad,
        "non_genuine_rate": rate,
        "non_genuine_samples": samples,
        "per_year": per_year,
        "item_status": "PASS" if bad == 0 else ("PASS_WITH_DISCLOSURE" if rate <= THRESH_FULL_WINDOW else "FAIL"),
    }


def main():
    window, pop = load_window_and_population()
    print("window months:", len(window), window[0], "..", window[-1])
    print("scanning monthly_revenue ...")
    table, per_year = scan_monthly_revenue()
    print("  rows indexed:", len(table))

    mr_results = {}
    for L in LOOKBACKS:
        r = evaluate_monthly_revenue(window, pop, table, L)
        mr_results["lookback_%d" % L] = r
        print("  L=%2d  stock-months=%d  not_reconstructible=%d (%.4f%%)  "
              "months>2%%=%d  -> %s"
              % (L, r["total_stock_months"], r["total_not_reconstructible"],
                 r["full_window_rate"] * 100,
                 r["months_exceeding_per_month_threshold"], r["item_status"]))

    verdicts = {r["item_status"] for r in mr_results.values()}
    stable = len(verdicts) == 1
    print("  verdict stable across lookbacks:", stable, verdicts)

    print("scanning financial_statements ...")
    fs = scan_financial_statements(window, pop)
    print("  rows in scope=%d  non_genuine=%d (%.6f%%)  -> %s"
          % (fs["rows_in_scope"], fs["non_genuine_rows"],
             fs["non_genuine_rate"] * 100, fs["item_status"]))

    payload = {
        "study": "P0-R5 Phase A — Stage 2 (monthly_revenue, financial_statements)",
        "read_only": True,
        "tdcc_weekly": "NOT_JUDGED — blocked by §3a-2 pending user ruling",
        "specification_gap_disclosed": {
            "gap": "§3 never defines which monthly_revenue rows are the '必要列' for a given (decision month, ticker).",
            "resolution": "Defined in this script during Phase A, therefore NOT pre-frozen.",
            "mitigation": "Reported under lookbacks 3/6/12; verdict treated as stable only if identical under all three.",
            "verdict_stable_across_lookbacks": stable,
        },
        "thresholds": {
            "full_window": THRESH_FULL_WINDOW,
            "per_month": THRESH_PER_MONTH,
            "purity": "Partial preregistration — set with knowledge of P0-R4 observed values; not adjusted after seeing these results.",
        },
        "monthly_revenue": {
            "source": os.path.relpath(MR_PATH, REPO).replace("\\", "/"),
            "whole_file_per_year_release_date_classes": per_year,
            "window_evaluation": mr_results,
        },
        "financial_statements": dict(fs, source=os.path.relpath(FS_PATH, REPO).replace("\\", "/")),
    }
    with open(OUT, "w", encoding="utf-8") as fh:
        json.dump(payload, fh, ensure_ascii=False, indent=1)
    print("wrote", os.path.relpath(OUT, REPO))
    return 0


if __name__ == "__main__":
    sys.exit(main())
