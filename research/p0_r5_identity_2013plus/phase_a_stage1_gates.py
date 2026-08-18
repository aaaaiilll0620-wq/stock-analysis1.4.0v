"""P0-R5 Phase A — Stage 1 gates (READ-ONLY).

Scope is strictly the approved Phase A (docs/prereg_P0_R5_..._2026-08-16.md,
approved 2026-08-16, receipt docs/prereg_P0_R5_approval_receipt_2026-08-16.json):

  1. §1 mechanical gate — derive the 159-month window from the frozen 255
     decision months and assert the count is exactly 159.
  2. Frozen decision-time stock-month population — resolve which artifact holds
     the per-month ticker lists and verify it across ALL 159 months (not a
     sample), per the full-corpus principle of AC-R5-2.
  3. §3 raw-file SHA256 rules — recompute hashes for every source file this
     study touches; match against P0-R4-recorded hashes where they exist,
     record source_pre_sha256 where they do not.
  4. Header presence — all 16 institutional_gross files (AC-R5-2 full corpus),
     plus monthly_revenue / financial_statements main files.

This script does NOT: build any candidate cache, call tej_importer.py or any
core.* production module, write to ~/tej_cache / market_cache /
data/runtime_cache, compute parity/performance, or judge tdcc_weekly (blocked
by §3a-2 until the user rules between option 2 and option 3).
"""

import hashlib
import json
import os
import sys

import openpyxl

REPO = os.path.dirname(os.path.dirname(os.path.dirname(os.path.abspath(__file__))))
CORPUS = os.path.join(REPO, "tej_exports", "DataExport0806")
OUT = os.path.join(os.path.dirname(os.path.abspath(__file__)), "phase_a_stage1_gates.json")

FROZEN_MONTHS = os.path.join(
    REPO, "research", "p0_r2_identity_collector", "a_leg_parity_result.json"
)
POP_DIR = os.path.join(
    REPO, "research", "p0_u1_canonical_universe", "canonical_universe_by_date"
)
R4_PREFLIGHT = os.path.join(
    REPO, "research", "p0_r4_datamigration_feasibility",
    "dataexport0806_4dataset_preflight.json",
)

WINDOW_START = "2013-01-01"
EXPECTED_MONTHS = 159

INSTITUTIONAL_REQUIRED_COLS = [
    "外資買進張數", "外資賣出張數",
    "投信買進張數", "投信賣出張數",
    "外資總投資股率%", "投信持股率%",
]


def sha256(path):
    h = hashlib.sha256()
    with open(path, "rb") as fh:
        for chunk in iter(lambda: fh.read(1 << 20), b""):
            h.update(chunk)
    return h.hexdigest()


def header_of(path):
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    try:
        ws = wb[wb.sheetnames[0]]
        for row in ws.iter_rows(min_row=1, max_row=1, values_only=True):
            return [("" if c is None else str(c).strip()) for c in row]
        return []
    finally:
        wb.close()


def gate_1_months():
    with open(FROZEN_MONTHS, encoding="utf-8") as fh:
        per_date = json.load(fh)["a_leg_parity_result"]["per_date"]
    frozen = sorted(per_date.keys())
    window = [m for m in frozen if m >= WINDOW_START]
    return {
        "source": os.path.relpath(FROZEN_MONTHS, REPO).replace("\\", "/"),
        "derivation": "sorted(per_date.keys()) filtered >= %s" % WINDOW_START,
        "frozen_month_count": len(frozen),
        "frozen_month_range": [frozen[0], frozen[-1]] if frozen else None,
        "derived_month_count": len(window),
        "derived_month_range": [window[0], window[-1]] if window else None,
        "expected": EXPECTED_MONTHS,
        "status": "PASS" if len(window) == EXPECTED_MONTHS else "FAIL",
    }, window, per_date


def gate_2_population(window, per_date):
    """Verify the population artifact across ALL 159 months, not a sample."""
    rows = []
    mismatches = []
    missing = []
    for m in window:
        path = os.path.join(POP_DIR, "canonical_universe_%s.csv" % m)
        if not os.path.exists(path):
            missing.append(m)
            continue
        with open(path, encoding="utf-8-sig") as fh:
            n = sum(1 for _ in fh) - 1
        oracle = per_date[m]["population_diagnostics"]["oracle_population"]
        rows.append({"month": m, "csv_rows": n, "oracle_population": oracle})
        if n != oracle:
            mismatches.append({"month": m, "csv_rows": n, "oracle_population": oracle})
    status = "PASS" if not mismatches and not missing else "FAIL"
    return {
        "artifact": os.path.relpath(POP_DIR, REPO).replace("\\", "/") + "/canonical_universe_{month}.csv",
        "cross_check": "csv data-row count == a_leg_parity_result.per_date[month].population_diagnostics.oracle_population",
        "months_checked": len(rows),
        "months_missing_file": missing,
        "mismatch_count": len(mismatches),
        "mismatches": mismatches[:20],
        "total_stock_months": sum(r["csv_rows"] for r in rows),
        "status": status,
        "note": (
            "Full-corpus over all 159 window months, not sampled. A FAIL here means the "
            "frozen decision-time population is UNRESOLVED and Phase A stops rather than "
            "substituting a different population."
        ),
    }


def gate_3_hashes():
    with open(R4_PREFLIGHT, encoding="utf-8") as fh:
        r4_blob = json.dumps(json.load(fh), ensure_ascii=False)
    import re
    r4_hashes = set(re.findall(r"[0-9a-f]{64}", r4_blob))

    targets = []
    mr = os.path.join(CORPUS, "月營收2004-202608", "20260806091706.xlsx")
    fs = os.path.join(CORPUS, "財報2004~202606", "20260806090633.xlsx")
    targets.append(("monthly_revenue", mr))
    targets.append(("financial_statements", fs))
    ig_dir = os.path.join(CORPUS, "法人回測2004-20260806")
    for name in sorted(os.listdir(ig_dir)):
        if name.lower().endswith(".xlsx"):
            targets.append(("institutional_gross", os.path.join(ig_dir, name)))

    out = []
    for dataset, path in targets:
        if not os.path.exists(path):
            out.append({"dataset": dataset, "path": path, "status": "FAIL",
                        "reason": "FILE_NOT_FOUND"})
            continue
        digest = sha256(path)
        known = digest in r4_hashes
        out.append({
            "dataset": dataset,
            "file": os.path.basename(path),
            "path": os.path.relpath(path, REPO).replace("\\", "/"),
            "size_bytes": os.path.getsize(path),
            "source_pre_sha256": digest,
            "recorded_by_p0_r4": known,
            "match_vs_p0_r4": "MATCH" if known else "NOT_PREVIOUSLY_RECORDED",
        })
    return out


def gate_4_headers(hash_rows):
    results = []
    for row in hash_rows:
        if row.get("status") == "FAIL":
            continue
        path = os.path.join(REPO, row["path"])
        try:
            hdr = header_of(path)
        except Exception as exc:  # noqa: BLE001 - recorded, not swallowed
            results.append({"file": row["file"], "dataset": row["dataset"],
                            "status": "FAIL", "reason": "HEADER_READ_ERROR: %s" % exc})
            continue
        entry = {"dataset": row["dataset"], "file": row["file"],
                 "column_count": len(hdr)}
        if row["dataset"] == "institutional_gross":
            missing = [c for c in INSTITUTIONAL_REQUIRED_COLS if c not in hdr]
            entry["required_cols_present"] = not missing
            entry["missing_required_cols"] = missing
            entry["status"] = "PASS" if not missing else "FAIL"
        else:
            entry["header"] = hdr
            entry["status"] = "PASS"
        results.append(entry)
    return results


def main():
    g1, window, per_date = gate_1_months()
    print("GATE 1 (159-month derivation):", g1["status"],
          g1["derived_month_count"], g1["derived_month_range"])
    if g1["status"] != "PASS":
        print("STOP: §1 gate FAIL — not proceeding to any further judgement.")
        json.dump({"gate_1_month_window": g1}, open(OUT, "w", encoding="utf-8"),
                  ensure_ascii=False, indent=1)
        return 1

    g2 = gate_2_population(window, per_date)
    print("GATE 2 (frozen population, all 159 months):", g2["status"],
          "months=%d mismatches=%d stock_months=%d"
          % (g2["months_checked"], g2["mismatch_count"], g2["total_stock_months"]))

    g3 = gate_3_hashes()
    n_match = sum(1 for r in g3 if r.get("match_vs_p0_r4") == "MATCH")
    n_new = sum(1 for r in g3 if r.get("match_vs_p0_r4") == "NOT_PREVIOUSLY_RECORDED")
    print("GATE 3 (SHA256): files=%d matched_p0_r4=%d newly_recorded=%d"
          % (len(g3), n_match, n_new))
    for r in g3:
        print("   %-22s %-46s %s %s" % (r["dataset"], r["file"][:46],
                                        r["source_pre_sha256"][:12],
                                        r["match_vs_p0_r4"]))

    g4 = gate_4_headers(g3)
    ig = [r for r in g4 if r["dataset"] == "institutional_gross"]
    ig_fail = [r for r in ig if r["status"] != "PASS"]
    print("GATE 4 (headers): institutional_gross %d/%d files carry all 6 required cols"
          % (len(ig) - len(ig_fail), len(ig)))
    for r in ig_fail:
        print("   FAIL", r["file"], r.get("missing_required_cols"))

    payload = {
        "study": "P0-R5 Phase A — Stage 1 gates",
        "read_only": True,
        "tdcc_weekly": "NOT_JUDGED — blocked by §3a-2 pending user ruling (option 2 vs option 3)",
        "gate_1_month_window": g1,
        "gate_2_frozen_population": g2,
        "gate_3_source_hashes": g3,
        "gate_4_headers": g4,
    }
    with open(OUT, "w", encoding="utf-8") as fh:
        json.dump(payload, fh, ensure_ascii=False, indent=1)
    print("\nwrote", os.path.relpath(OUT, REPO))
    return 0


if __name__ == "__main__":
    sys.exit(main())
