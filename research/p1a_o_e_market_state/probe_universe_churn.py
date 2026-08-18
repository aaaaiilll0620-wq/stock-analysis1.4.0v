"""O-E · is the per-year TEJ price export universe-stable?

90 securities share the terminal date 2018-12-28. Because the export is one file
per year, "last seen 2018-12-28" only means "present through 2018, absent from
2019" — the shared date is an artifact of the file layout, not of a shared event.

The question that matters is whether year-over-year disappearance is ordinary
delisting churn or a break in the export's own universe definition. If the 2019
vintage silently dropped names that were still trading, every downstream
population count in the window is affected, which is far larger than O-E.

READ-ONLY. No return, IC, Sharpe, ranking or selection quantity is computed.
"""

import json
import os
import sys
from collections import defaultdict

from openpyxl import load_workbook

HERE = os.path.dirname(os.path.abspath(__file__))
REPO = os.path.dirname(os.path.dirname(HERE))
D = os.path.join(REPO, "tej_exports", "DataExport0806", "個股股價、本益比2004-20260806")
OUT = os.path.join(HERE, "universe_churn.json")

YEARS = [str(y) for y in range(2012, 2027)]


def year_ids(year):
    p = os.path.join(D, f"{year}DataExport.xlsx")
    if not os.path.exists(p):
        return None, None
    wb = load_workbook(p, read_only=True, data_only=True)
    ws = wb[wb.sheetnames[0]]
    it = ws.iter_rows(values_only=True)
    hdr = next(it)
    i_id = next((j for j, c in enumerate(hdr) if "代號" in str(c)), 0)
    i_d = next((j for j, c in enumerate(hdr) if "年月日" in str(c)), 2)
    ids = set()
    last = defaultdict(str)
    for row in it:
        v = row[i_id]
        if v is None:
            continue
        sid = str(v).split()[0]
        ids.add(sid)
        d = str(row[i_d])[:10]
        if d > last[sid]:
            last[sid] = d
    wb.close()
    return ids, last


def main():
    per_year, per_year_last = {}, {}
    for y in YEARS:
        ids, last = year_ids(y)
        if ids is None:
            print(f"{y}: <no file>")
            continue
        per_year[y] = ids
        per_year_last[y] = last
        print(f"{y}: {len(ids):,} securities")

    ys = sorted(per_year)
    report = {}
    print("\nyear-over-year churn:")
    print("  year   present  dropped  added   dropped-but-traded-to-year-end")
    for a, b in zip(ys, ys[1:]):
        dropped = per_year[a] - per_year[b]
        added = per_year[b] - per_year[a]
        # A name that traded right up to the last session of year `a` and then
        # vanishes is the suspicious kind: a genuine delisting usually stops
        # mid-year.
        year_end = max(per_year_last[a].values()) if per_year_last[a] else ""
        to_end = sum(1 for s in dropped if per_year_last[a].get(s, "") == year_end)
        report[f"{a}->{b}"] = {
            "present": len(per_year[a]), "dropped": len(dropped),
            "added": len(added), "dropped_but_traded_to_year_end": to_end,
            "year_end_session": year_end,
            "examples": sorted(s for s in dropped
                               if per_year_last[a].get(s, "") == year_end)[:15],
        }
        print("  %s   %6d   %6d   %5d   %5d" %
              (a, len(per_year[a]), len(dropped), len(added), to_end))

    payload = {"study": "O-E per-year export universe churn",
               "read_only": True, "performance_computed": False,
               "source": os.path.relpath(D, REPO).replace("\\", "/"),
               "securities_per_year": {y: len(v) for y, v in per_year.items()},
               "churn": report}
    with open(OUT, "w", encoding="utf-8") as fh:
        json.dump(payload, fh, ensure_ascii=False, indent=1)
    print("\nwrote", os.path.relpath(OUT, REPO))
    return 0


if __name__ == "__main__":
    sys.exit(main())
