"""O-E · confirm or refute survivorship truncation in the price export.

The churn probe found: 11-20 securities drop out per year through 2018, none of
them trading to year-end; then 110 drop at 2018->2019 (90 of them trading right
to the last session of 2018); then ZERO drop in every year 2019-2024.

Zero delistings across six years is not a market fact. The hypothesis is that
the 2019+ vintage of this export was pulled with a "currently listed" universe
filter, so every security that delisted after 2018 is absent from those files
entirely — which would make its last appearance 2018-12-28 regardless of when it
actually left.

Two independent tests, neither of which relies on reasoning about the vendor:

  T1  Do the 90 appear in OTHER exports (配股相關, 暫停交易) with events dated
      after 2018? If a security has a 2021 corporate action, it existed in 2021.
  T2  Do the 90 have delisting-type suspension reasons, and when?

READ-ONLY. No return, IC, Sharpe, ranking or selection quantity is computed.
"""

import glob
import json
import os
import sys
import zipfile
from collections import Counter, defaultdict

import pandas as pd
from openpyxl import load_workbook

HERE = os.path.dirname(os.path.abspath(__file__))
REPO = os.path.dirname(os.path.dirname(HERE))
EXPORT = os.path.join(REPO, "tej_exports", "DataExport0806")
PRICE = os.path.join(os.path.expanduser("~"), "tej_cache", "price_valuation")
OUT = os.path.join(HERE, "survivorship_check.json")


def d8(v):
    s = str(v).strip().split(".")[0]
    return f"{s[:4]}-{s[4:6]}-{s[6:]}" if len(s) == 8 and s.isdigit() else None


def cohort_from_prices():
    files = sorted(glob.glob(os.path.join(PRICE, "*.parquet")))
    last = {}
    for f in files:
        df = pd.read_parquet(f, columns=["stock_id", "date"])
        if df.empty:
            continue
        sid = str(df["stock_id"].iloc[0])
        last[sid] = str(df["date"].max())
    px_max = max(last.values())
    return {s for s, d in last.items() if d == "2018-12-28"}, last, px_max


def corporate_action_dates():
    """stock_id -> latest 年月日 seen in the 配股相關 corpus (covers 上下市)."""
    latest = defaultdict(str)
    src = os.path.join(EXPORT, "配股相關2004-20260817")
    for z in sorted(glob.glob(os.path.join(src, "*.zip"))):
        zf = zipfile.ZipFile(z)
        for name in zf.namelist():
            txt = zf.read(name).decode("utf-16")
            lines = txt.split("\n")
            hdr = lines[0].rstrip("\r").split("\t")
            i_id, i_d = hdr.index("證券代碼"), hdr.index("年月日")
            for line in lines[1:]:
                if not line.strip():
                    continue
                f = line.rstrip("\r").split("\t")
                if len(f) <= max(i_id, i_d):
                    continue
                sid = f[i_id].split()[0]
                d = d8(f[i_d])
                if d and d > latest[sid]:
                    latest[sid] = d
    return latest


def suspensions():
    rows = []
    d = os.path.join(EXPORT, "暫停交易2004-20260806")
    for f in sorted(glob.glob(os.path.join(d, "**", "*.xlsx"), recursive=True)):
        for r in load_workbook(f, read_only=True, data_only=True)[
                load_workbook(f, read_only=True).sheetnames[0]].iter_rows(values_only=True):
            pass
    # simpler: pandas
    rows = []
    for f in sorted(glob.glob(os.path.join(d, "**", "*.xlsx"), recursive=True)):
        df = pd.read_excel(f)
        for r in df.to_dict("records"):
            rows.append({"stock_id": str(r["證券代碼"]).split()[0],
                         "start": d8(r["年月日"]),
                         "reason": str(r["暫停交易原因"]).strip()})
    return rows


def main():
    cohort, last_price, px_max = cohort_from_prices()
    print(f"price data max: {px_max}   cohort (last price 2018-12-28): {len(cohort)}")

    ca = corporate_action_dates()
    after_2018 = {s: ca[s] for s in cohort if ca.get(s, "") > "2018-12-31"}
    print(f"\nT1 · cohort members with a 配股相關 event AFTER 2018: "
          f"{len(after_2018)}/{len(cohort)}")
    for s, d in sorted(after_2018.items())[:15]:
        print(f"     {s}  latest corporate-action row {d}")

    susp = suspensions()
    by_id = defaultdict(list)
    for r in susp:
        by_id[r["stock_id"]].append(r)
    delist_words = ("下市", "終止", "合併", "轉換", "併入")
    t2 = {}
    for s in cohort:
        rows = [r for r in by_id.get(s, []) if r["start"]]
        d_rows = [r for r in rows if any(w in r["reason"] for w in delist_words)]
        if d_rows:
            t2[s] = sorted((r["start"], r["reason"][:24]) for r in d_rows)[-1]
    print(f"\nT2 · cohort members with a delisting-type suspension: {len(t2)}/{len(cohort)}")
    for s, (d, why) in sorted(t2.items())[:15]:
        print(f"     {s}  {d}  {why}")
    after = {s: v for s, v in t2.items() if v[0] > "2018-12-31"}
    print(f"     of which dated AFTER 2018: {len(after)}")

    # control: do NON-cohort securities that are still priced have events too?
    still = [s for s, d in last_price.items() if d == px_max]
    ctrl = sum(1 for s in still[:300] if ca.get(s, "") > "2018-12-31")
    print(f"\ncontrol · of 300 still-priced securities, {ctrl} have a post-2018 "
          f"corporate-action row (the corpus does cover post-2018 events)")

    evidence = sorted(set(after_2018) | set(after))
    payload = {
        "study": "O-E survivorship truncation check",
        "read_only": True, "performance_computed": False,
        "price_data_max": px_max,
        "cohort_size": len(cohort),
        "cohort": sorted(cohort),
        "T1_corporate_action_after_2018": {s: after_2018[s] for s in sorted(after_2018)},
        "T2_delisting_suspension": {s: list(t2[s]) for s in sorted(t2)},
        "T2_delisting_after_2018": {s: list(after[s]) for s in sorted(after)},
        "cohort_with_post_2018_evidence": evidence,
        "cohort_with_post_2018_evidence_n": len(evidence),
    }
    with open(OUT, "w", encoding="utf-8") as fh:
        json.dump(payload, fh, ensure_ascii=False, indent=1)
    print(f"\n=> {len(evidence)}/{len(cohort)} cohort members have independent "
          f"evidence of existing after 2018.")
    print("wrote", os.path.relpath(OUT, REPO))
    return 0


if __name__ == "__main__":
    sys.exit(main())
