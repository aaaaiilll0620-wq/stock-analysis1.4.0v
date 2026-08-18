"""D-1 · audit the price corpus against an INDEPENDENT security master.

The previous D-1 evidence was self-referential: the price corpus was judged by
its own churn pattern. That detects the contamination but cannot size it, and it
cannot say which securities are absent without reading the answer off the very
data that is broken.

`基本資料/公司資料.xlsx` provides an independent reference. It carries 上市日 and
下市日期 for securities including ones long delisted (row 1 is 000116 日盛證券,
delisted 2002-02-05), so "which securities were listed during year Y" can be
answered without consulting the price files at all.

TWO HARD CONSTRAINTS ON THAT FILE, both enforced elsewhere:

  1. It is a CURRENT SNAPSHOT (目前狀態, no as-of). Under O-E it is therefore
     NOT_PIT_SAFE and must never become a B0 runtime input — knowing a delisting
     date in advance is exactly the look-ahead O-B exists to prevent. It is used
     here as an AUDIT reference only.
  2. The output is diagnostic, NOT a patch list. The D-1 remedy remains a full
     re-export; nothing here may be used to top up the contaminated corpus with
     the specific securities it happens to be missing.

Pre-2019 years act as the control: whatever mismatch is structural (market-type
scope, non-equity securities, master staleness) shows up there too, so the
survivorship signature is the DIFFERENCE between the eras, not the level.

READ-ONLY. No return, IC, Sharpe, ranking or selection quantity is computed.
"""

import glob
import json
import os
import sys
from collections import Counter, defaultdict

import pandas as pd
from openpyxl import load_workbook

HERE = os.path.dirname(os.path.abspath(__file__))
REPO = os.path.dirname(os.path.dirname(HERE))
MASTER = os.path.join(REPO, "tej_exports", "DataExport0806", "基本資料", "公司資料.xlsx")
PRICE = os.path.join(os.path.expanduser("~"), "tej_cache", "price_valuation")
OUT = os.path.join(HERE, "universe_vs_master.json")

YEARS = [str(y) for y in range(2012, 2027)]

# Scope is taken from the HISTORICAL listing-date columns, never from 上市別.
# 上市別 is a current-snapshot label and it is REWRITTEN on delisting: every one
# of the 90 securities whose price series stops at 2018-12-28 now reads UNPUB or
# PUB, so filtering on it silently excludes precisely the delisted names the
# audit exists to find. `TSE上市日` / `OTC上市日` are historical facts and are
# immune to that rewrite. (This is the same class of defect as industry_map —
# a current label applied to history — and it is why the file is audit-only.)
EXCHANGE_LISTING_DATE_COLUMNS = ("TSE上市日", "OTC上市日")


def norm(v):
    s = str(v).strip() if v is not None else ""
    if not s or s in (".", "None", "nan"):
        return None
    s = s.replace("/", "-")
    p = s.split("-")
    if len(p) == 3 and all(x.strip().isdigit() for x in p):
        return "%04d-%02d-%02d" % (int(p[0]), int(p[1]), int(p[2]))
    return None


def load_master():
    wb = load_workbook(MASTER, read_only=True, data_only=True)
    ws = wb[wb.sheetnames[0]]
    it = ws.iter_rows(values_only=True)
    hdr = [str(c).strip() for c in next(it)]
    idx = {c: i for i, c in enumerate(hdr)}
    rows = []
    for r in it:
        if r[idx["代號"]] is None:
            continue
        sid = str(r[idx["代號"]]).strip()
        exchange_from = None
        for col in EXCHANGE_LISTING_DATE_COLUMNS:
            d = norm(r[idx[col]]) if col in idx else None
            if d and (exchange_from is None or d < exchange_from):
                exchange_from = d
        any_listed = exchange_from
        for col in ("最近上市日", "REG上市日", "創新版上市日", "首次掛牌日期"):
            d = norm(r[idx[col]]) if col in idx else None
            if d and (any_listed is None or d < any_listed):
                any_listed = d
        rows.append({
            "stock_id": sid,
            "market": str(r[idx["上市別"]]).strip() if r[idx["上市別"]] else "",
            "status": str(r[idx["目前狀態"]]).strip() if r[idx["目前狀態"]] else "",
            "exchange_listed_from": exchange_from,   # TSE/OTC only, historical
            "listed_from": any_listed,
            "delisted_on": norm(r[idx["下市日期"]]),
        })
    wb.close()
    return rows


def price_years():
    """symbol -> set of years with at least one observed price."""
    out = {}
    last = {}
    for f in sorted(glob.glob(os.path.join(PRICE, "*.parquet"))):
        df = pd.read_parquet(f, columns=["stock_id", "date"])
        if df.empty:
            continue
        sid = str(df["stock_id"].iloc[0])
        d = df["date"].astype(str)
        out[sid] = {x[:4] for x in d}
        last[sid] = d.max()
    return out, last


def _load_status_spans():
    """stock_id -> ((effective_from, status), ...) from the registered O-E source."""
    import csv
    path = os.path.join(REPO, "data", "b0", "security_status.csv")
    spans = defaultdict(list)
    if not os.path.exists(path):
        return spans
    with open(path, encoding="utf-8") as fh:
        for r in csv.DictReader(fh):
            spans[r["stock_id"]].append((r["effective_from"], r["status"]))
    for k in spans:
        spans[k].sort()
    return spans


def _covered_by_status(spans, y0, y1):
    """True if a non-trading status was in force for the whole of [y0, y1]."""
    if not spans:
        return False
    in_force = None
    for eff, status in spans:
        if eff <= y0:
            in_force = status
        elif eff <= y1:
            # a status change inside the year means it traded part of it
            return False
    return in_force in ("suspended", "delisted")


def _emit_csv(report, clusters):
    """Flat schema the D-1 blocking verifier reads."""
    import csv
    out = os.path.join(REPO, "data", "b0", "price_universe_audit.csv")
    os.makedirs(os.path.dirname(out), exist_ok=True)
    cols = ["year", "expected_from_reference", "observed_in_corpus", "missing",
            "missing_though_listed_after_year_end",
            "unexplained_missing_though_listed",
            "exits_observed", "exits_expected_from_reference"]
    with open(out, "w", encoding="utf-8", newline="") as fh:
        w = csv.DictWriter(fh, fieldnames=cols)
        w.writeheader()
        for y, v in sorted(report.items()):
            w.writerow({**{c: v.get(c) for c in cols}, "year": y})
    out2 = os.path.join(REPO, "data", "b0", "price_universe_clusters.csv")
    with open(out2, "w", encoding="utf-8", newline="") as fh:
        w = csv.DictWriter(fh, fieldnames=["date", "corpus_terminations",
                                           "reference_delistings_on_date"])
        w.writeheader()
        for d, v in sorted(clusters.items()):
            w.writerow({"date": d, "corpus_terminations": v["corpus_terminations"],
                        "reference_delistings_on_date": v["master_delistings_that_day"]})
    return out, out2


def main():
    master = load_master()
    print(f"security master rows: {len(master):,}")
    print("  上市別:", dict(Counter(m["market"] for m in master).most_common(8)))
    print("  目前狀態:", dict(Counter(m["status"] for m in master).most_common(8)))
    print(f"  with 下市日期: {sum(1 for m in master if m['delisted_on']):,}")

    equity = [m for m in master if m["exchange_listed_from"]]
    print(f"  ever exchange-listed (TSE上市日 or OTC上市日 present): {len(equity):,}")
    print(f"    of which now labelled UNPUB/PUB: "
          f"{sum(1 for m in equity if m['market'] in ('UNPUB', 'PUB')):,}"
          f"  <- 上市別 is rewritten on delisting")

    yrs, last_price = price_years()
    corpus_max = max(last_price.values())
    print(f"price corpus: {len(yrs):,} securities, max date {corpus_max}")

    # --- expected vs observed, per year -------------------------------------
    # Registered O-E status source: a full-year absence is only a contradiction
    # if nothing already known explains it. Doing the explaining HERE keeps the
    # verifier threshold-free — it then only has to check "unexplained > 0".
    status_by_id = _load_status_spans()

    report = {}
    by_id = {m["stock_id"]: m for m in equity}
    missing_ids_by_year = {}
    exits_observed = {}
    for a, b in zip(YEARS, YEARS[1:]):
        in_a = {s for s, ys in yrs.items() if a in ys}
        in_b = {s for s, ys in yrs.items() if b in ys}
        exits_observed[a] = len(in_a - in_b)
    print("\n  year  expected observed  missing  (%)   extra   of missing:"
          " delisted-later")
    for y in YEARS:
        y0, y1 = f"{y}-01-01", f"{y}-12-31"
        expected = {m["stock_id"] for m in equity
                    if m["exchange_listed_from"] <= y1
                    and (m["delisted_on"] is None or m["delisted_on"] >= y0)}
        observed = {s for s, ys in yrs.items() if y in ys}
        missing = expected - observed
        extra = observed - expected
        # The signature: the master says it was still listed after this year, so
        # it should have kept trading, yet the corpus has nothing for it.
        later = {s for s in missing
                 if by_id[s]["delisted_on"] is None or by_id[s]["delisted_on"] > y1}
        unexplained = {s for s in later
                       if not _covered_by_status(status_by_id.get(s, ()), y0, y1)}
        missing_ids_by_year[y] = sorted(missing)
        report[y] = {
            "expected_from_reference": len(expected),
            "observed_in_corpus": len(observed),
            "missing": len(missing),
            "missing_pct": round(100.0 * len(missing) / len(expected), 2) if expected else None,
            "in_corpus_not_in_reference": len(extra),
            "missing_though_listed_after_year_end": len(later),
            "unexplained_missing_though_listed": len(unexplained),
            "exits_observed": exits_observed.get(y),
            "exits_expected_from_reference": sum(
                1 for m in equity if m["delisted_on"] and m["delisted_on"][:4] == y),
        }
        print("  %s   %5d    %5d    %5d %6.2f   %4d      %4d  %4d   exits obs=%s exp=%d"
              % (y, len(expected), len(observed), len(missing),
                 report[y]["missing_pct"] or 0.0, len(extra), len(later),
                 len(unexplained), report[y]["exits_observed"],
                 report[y]["exits_expected_from_reference"]))

    # --- expected exits per year, from the master alone ----------------------
    exits = Counter()
    for m in equity:
        if m["delisted_on"]:
            exits[m["delisted_on"][:4]] += 1
    print("\nexpected delistings per year (master):")
    for y in YEARS:
        print("   %s  %3d" % (y, exits.get(y, 0)))

    # --- terminal-date clusters in the corpus --------------------------------
    term = Counter(d for s, d in last_price.items() if d < corpus_max)
    clusters = {d: n for d, n in term.items() if n >= 5}
    print("\nterminal-date clusters (>=5 securities ending the same day):")
    for d, n in sorted(clusters.items(), key=lambda kv: -kv[1])[:10]:
        print("   %s  %4d   master delistings that day: %d"
              % (d, n, sum(1 for m in equity if m["delisted_on"] == d)))

    payload = {
        "study": "D-1 price corpus vs independent security master",
        "read_only": True, "performance_computed": False,
        "diagnostic_only": True,
        "not_a_patch_list": True,
        "master": os.path.relpath(MASTER, REPO).replace("\\", "/"),
        "master_is_current_snapshot": True,
        "master_admissible_as_b0_runtime_source": False,
        "master_rows": len(master),
        "master_equity_with_listing_date": len(equity),
        "corpus_securities": len(yrs),
        "corpus_max_date": corpus_max,
        "per_year": report,
        "missing_ids_by_year": missing_ids_by_year,
        "expected_delistings_per_year_from_master": {y: exits.get(y, 0) for y in YEARS},
        "terminal_date_clusters": {
            d: {"corpus_terminations": n,
                "master_delistings_that_day": sum(1 for m in equity if m["delisted_on"] == d)}
            for d, n in sorted(clusters.items(), key=lambda kv: -kv[1])},
    }
    with open(OUT, "w", encoding="utf-8") as fh:
        json.dump(payload, fh, ensure_ascii=False, indent=1)
    a, b = _emit_csv(report, payload["terminal_date_clusters"])
    print("\nwrote", os.path.relpath(OUT, REPO))
    print("wrote", os.path.relpath(a, REPO))
    print("wrote", os.path.relpath(b, REPO))
    return 0


if __name__ == "__main__":
    sys.exit(main())
