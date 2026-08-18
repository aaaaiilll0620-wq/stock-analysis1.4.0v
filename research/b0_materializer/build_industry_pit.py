# -*- coding: utf-8 -*-
"""§2.3 · the canonical PIT TSE industry timeline, sealed for the materializer.

`industry_map.parquet` is a static current snapshot and 1,203 securities (49.4%)
have changed TSE industry at least once, so using it to compute historical
within-industry valuation is industry look-ahead for about half the universe.
§2.3 therefore requires a point-in-time step function, and this is where that
step function becomes a sealed input rather than a research by-product.

**This file does not invent the semantics.** They were established in
`research/p0_b09_value_reference/` and frozen by §2.3:

  * each 前N次 record is `(industry it BECAME, effective date)`; 前一次 is the
    most recent and normally agrees with the current TSE產業 column — verified
    there against 1316 上曜 / 1319 東陽 / 1229 聯華;
  * a security with no dated record at all falls back to `1900-01-01` + its
    current code, which is the one place a current value is used and is safe
    only because there is no change history to contradict;
  * the 92 securities whose current column disagrees with their latest dated
    record are `UNRESOLVED` from that record onward — NOT back-filled from the
    snapshot, and NOT assumed to have kept the old classification. UNRESOLVED
    means industry NA, which means Value NA, which means §4.1 complete-case
    drops them. Measured consequence, already in §2.3: a median of 41 securities
    per period, 2.303%.

What this file adds is a re-derivation and a CROSS-CHECK: the timeline is built
again here from the upstream workbook and then asserted row-for-row against the
existing frozen artefact. If the two ever disagree, this build aborts — because
that would mean the materializer had quietly acquired a second definition of
what a security's industry was, which is exactly the failure §2.3 exists to
prevent.

    python research/b0_materializer/build_industry_pit.py
"""
from __future__ import annotations

import hashlib
import json
import os
import sys

import openpyxl

HERE = os.path.dirname(os.path.abspath(__file__))
REPO = os.path.dirname(os.path.dirname(HERE))
sys.path.insert(0, REPO)

from core.b0_features import INDUSTRY_UNRESOLVED          # noqa: E402

IMPORTER_VERSION = "industry_pit_importer_v1"

IND_XLSX = os.path.join(REPO, "tej_exports", "DataExport0806", "產業類別",
                        "歷史產業類別.xlsx")
FROZEN_REFERENCE = os.path.join(
    REPO, "research", "p0_b09_value_reference", "pit_industry_timeline_v2.parquet")
OUT_PARQUET = os.path.join(REPO, "data", "b0", "industry_pit.parquet")
OUT_RECEIPT = os.path.join(HERE, "industry_pit_receipt.json")

CHANGE_PAIRS = (("前三次TSE產業變更", "前三次TSE產業變更日"),
                ("前二次TSE產業變更", "前二次TSE產業變更日"),
                ("前一次TSE產業變更", "前一次TSE產業變更日"))
NO_HISTORY_EFFECTIVE_FROM = "1900-01-01"


def _file_sha(path: str) -> str:
    h = hashlib.sha256()
    with open(path, "rb") as fh:
        for chunk in iter(lambda: fh.read(1 << 20), b""):
            h.update(chunk)
    return h.hexdigest()


def _code(v):
    if v is None:
        return None
    s = str(v).strip()
    return s.split()[0] if s and s != "." else None


def _date(v):
    if v is None:
        return None
    s = str(v).strip()
    if not s or s == ".":
        return None
    s = s.replace("/", "-")
    return s[:10] if len(s) >= 10 else None


def build_timeline():
    """(stock_id, effective_from, tse_ind_code, unresolved_from), ascending."""
    import pandas as pd

    wb = openpyxl.load_workbook(IND_XLSX, read_only=True, data_only=True)
    try:
        ws = wb[wb.sheetnames[0]]
        it = ws.iter_rows(values_only=True)
        idx = {c: i for i, c in enumerate(list(next(it)))}
        rows, unresolved = [], {}
        stats = {"securities": 0, "with_changes": 0, "no_history_fallback": 0,
                 "current_agrees_with_latest": 0, "unresolved_securities": 0}
        for r in it:
            if r[0] is None:
                continue
            sid = str(r[idx["代號"]]).strip()
            if not sid:
                continue
            stats["securities"] += 1
            recs = []
            for c_ind, c_date in CHANGE_PAIRS:
                code, d = _code(r[idx[c_ind]]), _date(r[idx[c_date]])
                if code and d:
                    recs.append((d, code))
            listed, first_ind = _date(r[idx["首次掛牌日期"]]), _code(
                r[idx["首次掛牌TSE產業"]])
            if listed and first_ind:
                recs.append((listed, first_ind))
            cur = _code(r[idx["TSE產業_代碼"]])
            if not recs:
                if cur:
                    recs.append((NO_HISTORY_EFFECTIVE_FROM, cur))
                    stats["no_history_fallback"] += 1
            recs = sorted(set(recs))
            if len(recs) > 1:
                stats["with_changes"] += 1
            if cur and recs:
                if recs[-1][1] == cur:
                    stats["current_agrees_with_latest"] += 1
                else:
                    # §2.3: the whole interval from that record onward is
                    # UNRESOLVED. No snapshot back-fill, no assumption that the
                    # old classification simply continued.
                    unresolved[sid] = recs[-1][0]
                    stats["unresolved_securities"] += 1
            for d, code in recs:
                rows.append({"stock_id": sid, "effective_from": d,
                             "tse_ind_code": code})
    finally:
        wb.close()
    tl = pd.DataFrame(rows).sort_values(
        ["stock_id", "effective_from"]).reset_index(drop=True)
    tl["unresolved_from"] = tl["stock_id"].map(unresolved)
    return tl, unresolved, stats


def industry_as_of(timeline, as_of: str) -> dict:
    """{stock_id: tse_ind_code or UNRESOLVED} at a session. The only resolver.

    Returning the sentinel rather than dropping the security is deliberate:
    `SecurityPitInputs.pit_industry` defaults to `UNRESOLVED`, and §3.2 turns
    that into Value = NA. A silently absent security and an UNRESOLVED one are
    the same outcome for Value but not the same fact, and the materializer
    should carry the fact.
    """
    sub = timeline[timeline["effective_from"] <= as_of]
    if sub.empty:
        return {}
    latest = sub.groupby("stock_id").last()
    out = {}
    for sid, row in latest.iterrows():
        unres = row["unresolved_from"]
        if isinstance(unres, str) and as_of >= unres:
            out[str(sid)] = INDUSTRY_UNRESOLVED
        else:
            out[str(sid)] = str(row["tse_ind_code"])
    return out


def assert_matches_frozen_reference(tl) -> dict:
    """A second construction that disagrees is a second definition. Abort."""
    import pandas as pd

    if not os.path.exists(FROZEN_REFERENCE):
        raise SystemExit("abort: frozen reference timeline missing at %s"
                         % os.path.relpath(FROZEN_REFERENCE, REPO))
    ref = pd.read_parquet(FROZEN_REFERENCE)
    cols = ["stock_id", "effective_from", "tse_ind_code"]
    a = tl[cols].astype(str).sort_values(cols).reset_index(drop=True)
    b = ref[cols].astype(str).sort_values(cols).reset_index(drop=True)
    if len(a) != len(b) or not a.equals(b):
        raise SystemExit(
            "abort: the re-derived PIT industry timeline differs from the frozen "
            "artefact (%d vs %d rows). §2.3's step function must have exactly one "
            "construction; investigate rather than pick one." % (len(a), len(b)))
    ref_unres = set(ref[ref["unresolved_from"].notna()]["stock_id"].astype(str))
    new_unres = set(tl[tl["unresolved_from"].notna()]["stock_id"].astype(str))
    if ref_unres != new_unres:
        raise SystemExit(
            "abort: UNRESOLVED sets differ (%d vs %d securities)"
            % (len(new_unres), len(ref_unres)))
    return {"reference": os.path.relpath(FROZEN_REFERENCE, REPO),
            "rows_identical": True, "unresolved_identical": True,
            "reference_sha256": _file_sha(FROZEN_REFERENCE)}


def main() -> None:
    import pandas as pd

    tl, unresolved, stats = build_timeline()
    cross = assert_matches_frozen_reference(tl)
    print("timeline rows=%d securities=%d unresolved=%d (matches frozen "
          "artefact row-for-row)" % (len(tl), tl["stock_id"].nunique(),
                                     len(unresolved)), flush=True)

    os.makedirs(os.path.dirname(OUT_PARQUET), exist_ok=True)
    tl.to_parquet(OUT_PARQUET, index=False)

    schema = json.dumps({c: str(tl[c].dtype) for c in tl.columns},
                        sort_keys=True).encode("utf-8")
    receipt = {
        "artefact": "data/b0/industry_pit.parquet",
        "builder": "research/b0_materializer/build_industry_pit.py",
        "clause": "§2.3 PIT TSE industry timeline",
        "importer_version": IMPORTER_VERSION,
        "upstream_sources": [{
            "file": os.path.relpath(IND_XLSX, REPO),
            "sha256": _file_sha(IND_XLSX)}],
        "content_sha256": _file_sha(OUT_PARQUET),
        "schema_sha256": hashlib.sha256(schema).hexdigest(),
        "bytes": os.path.getsize(OUT_PARQUET),
        "rows": int(len(tl)),
        "securities": int(tl["stock_id"].nunique()),
        "effective_from_min": str(tl["effective_from"].min()),
        "effective_from_max": str(tl["effective_from"].max()),
        "unresolved_securities": len(unresolved),
        "unresolved_semantics": (
            "industry = UNRESOLVED from the last dated record onward -> Value NA "
            "-> §4.1 complete-case; never back-filled from the current snapshot"),
        "no_history_fallback_securities": stats["no_history_fallback"],
        "current_snapshot_used_for_backfill": False,
        "cross_check": cross,
        "stats": stats,
        "performance_computed": False,
    }
    with open(OUT_RECEIPT, "w", encoding="utf-8", newline="\n") as fh:
        json.dump(receipt, fh, ensure_ascii=False, indent=1, sort_keys=True)
        fh.write("\n")
    print("unresolved securities: %d · no-history fallback: %d"
          % (len(unresolved), stats["no_history_fallback"]))
    print("wrote", os.path.relpath(OUT_PARQUET, REPO), "and",
          os.path.relpath(OUT_RECEIPT, REPO))


if __name__ == "__main__":
    main()
