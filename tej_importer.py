"""
TEJ 全市場歷史批次匯入器
================================================================================
用途:TEJ Pro 沒有 API,只能用本機用戶端手動批次匯出。這支腳本讀
      tej_exports/DataExport0806/ 底下的原始匯出檔 (2026-08-06 建立的唯一
      authoritative raw snapshot,取代已停用的 tej_exports/inbox*/),正規化
      欄位對齊 FinMind 慣例,合併去重後存成獨立的 Parquet 歷史庫。

      跟 core/data_cache.py 的 finmind_cache 刻意分開存放:兩者信任等級不同。
      TEJ 這份是「人工批次匯入、無 PIT 保證」(除下述已標記的例外),只供 §12
      全市場擴池規劃書 Phase 2 的粗篩規則回溯驗證用,不進生產環境的每日 PIT
      評分管線。

      舊 tej_exports/inbox*/ 已停用,程式不得再讀 —— 見
      docs/資料快照遷移_DataExport0806.md 的完整 manifest、涵蓋範圍證明、
      欄位落差與待刪除清單 (待 Codex 審查後才會真的刪 inbox*)。

DataExport0806 原始檔格式跟舊 inbox* 不同,注意:
  · 大部分類別是「一年一檔 .xlsx」或「整段一檔 .xlsx」,查詢精靈欄位順序不保證
    跨檔一致 (新版有些檔案代號/名稱合併成一欄),所以**一律用欄名對應,不用位置**。
  · margin_balance 改成「.zip 內包一個 UTF-16 + Tab 分隔的 .csv」(TEJ 大量匯出
    的新格式,跳過 Excel 106.8萬列上限);讀取見 _read_raw_table。
  · 部分類別 (institutional_flow / institutional_gross、fundamentals_quarterly /
    financial_statements、revenue_growth / monthly_revenue) 舊版是兩個各自獨立
    的匯出檔 (窄版/寬版分開匯出),新版只給一份最寬的匯出,兩個 dataset 都從
    同一份原始檔挑欄位產生 —— 不是兩份不同的原始資料。

支援的資料集 (--dataset 切換):

  price_valuation (預設,讀「個股股價、本益比2004-20260806」):
    未調整股價(日)+ 本益比/淨值比/殖利率,2004-01-02 起,涵蓋範圍等同舊 inbox/。

  institutional_flow (讀「法人回測2004-20260806」):
    外資/投信/自營買賣超(千股),2004-01-02 起 (舊 inbox_chip 同起點)。

  institutional_gross (讀「法人回測2004-20260806」,跟 institutional_flow 同一批原始檔):
    外資/投信買賣張數 + 持股率。舊 inbox_chip_gross 只有 2026-04-01~07-16 這段
    「種子」窗口;新原始檔把毛額欄位併進法人回測的寬版匯出,涵蓋回溯到
    2004-01-02,是實質擴大而不只是搬家。

  fundamentals_quarterly (讀「財報2004~202606」,7欄舊版子集,已被 financial_statements 取代):
    單季歸屬母公司淨利/EPS/營業利益。
    ⚠ ROE(A)稅後在新原始檔沒有對應欄位 (舊 7 欄匯出直接给 TEJ 算好的 ROE,新的
    寬版匯出沒有這欄) —— 本欄位這裡刻意留白 (NaN),不用 net_income/equity 自算
    替代,避免用跟 TEJ 官方口徑不同的近似值悄悄頂替。scripts/tej_universe_
    screen_validation.py 有讀這個欄位,需要真 ROE 的話,要嘛跟 TEJ 另外要
    ROE 欄,要嘛在消費端明確自算並標註口徑。

  financial_statements (讀「財報2004~202606」,IFRS 三大財報單季完整版):
    比舊版 16 欄豐富很多 (56 欄),新增「財報發布日」= 真實公告日 (舊版沒有,
    完全無 PIT 保證;這欄可以拿來做 PIT 對齊,是實質品質提升)。範圍從舊的
    2019+ 生產環境擴大到 2005-06 起 (舊 2005-2018 段要另外跑
    scripts/import_financials_2005_2018.py 補丁,新原始檔已經內建這段,
    該腳本已標記為 deprecated)。
    ⚠ 常續性稅後淨利 (recurring_net_income) 新原始檔沒有這欄,留白不補。

  revenue_growth (讀「月營收2004-202608」,4欄子集,已被 monthly_revenue 取代):
    單月營收成長率 (YoY,非合併)。範圍從舊版的 2004 起 (但只有成長率,無公告日)
    延續。

  monthly_revenue (讀「月營收2004-202608」,月營收完整版):
    比舊版多「單月營收與上月比%」「流通在外股數」「合併營收(Y/N)」;
    release_date=真實公告日 (PIT 對齊用,新原始檔仍有)。範圍從舊版的 2019+
    (完整版) 擴大到 2004 起 —— 這是實質品質提升,舊版 2004-2018 段只有
    revenue_growth 那種「只有成長率」的陽春版。
    ⚠ 「去年單月營收」「去年累計營收」新原始檔沒有這兩欄,留白不補 (目前沒有
    下游程式讀這兩欄,純粹欄位變薄,不影響現有計算)。

  industry_map (讀「產業類別/現在產業類別.xlsx」,靜態對照表,無日期):
    代號/名稱 + TSE產業_代碼/名稱 + TEJ產業_代碼/名稱 + TEJ子產業_代碼/名稱。
    欄名與舊 inbox_industry/Industry.xlsx 完全一致,原封不動搬家。

  margin_balance (讀「融資融券2004-20260806」,含 5 個 .zip + 1 個 .xlsx):
    融資餘額(張)、融券餘額(張),還額外多留融資買進/賣出/增減、融資使用率、
    券資比 (供研究,同舊版精神)。範圍從舊 inbox_margin 的 2026-04-01~07-16
    (種子窗口) 大幅擴大到 2004-01-02 起 —— 實質擴大而非搬家。

  tdcc_weekly (讀「集保大戶2019-20260806」,集保股權分散週頻):
    欄名與舊 inbox_tdcc 完全一致,2019 起 (無變化)。

  director_pledge (讀「集團分類+董監質押與持股比2019-202606/pledge.xlsx」):
    欄名與舊 inbox_pledge/pledge.xlsx 完全一致;逐列比對兩份檔案 100% 相同
    (同一份原始資料搬家,非重新匯出)。

用法:
  python tej_importer.py                                 # 匯入 price_valuation
  python tej_importer.py --dataset institutional_flow     # 匯入法人買賣超
  python tej_importer.py --dataset financial_statements   # 匯入三大財報完整版
  python tej_importer.py --cache-dir /tmp/tej_cache_test --dataset margin_balance
                                                            # 先寫測試目錄,不動生產 tej_cache

限制:欄位對應寫在 DATASETS[...]["rename"] (中文欄名 → 目標欄名),依賴的是欄名
      不是位置,查詢精靈勾選順序改變不影響 (但如果 TEJ 把欄位中文名字改了,
      對應就會失效 —— 那種情況下該欄會被跳過而不是報錯,執行後務必看
      log 有沒有預期外的「欄位缺失」)。
================================================================================
"""
import csv
import fnmatch
import hashlib
import io
import json
import os
import re
import shutil
import sys
import logging
import argparse
import zipfile
from pathlib import Path
from types import MappingProxyType

import openpyxl
import pandas as pd
import pyarrow as pa

project_root = os.path.dirname(os.path.abspath(__file__))
if project_root not in sys.path:
    sys.path.insert(0, project_root)

logging.basicConfig(level=logging.INFO, format="%(message)s")
logger = logging.getLogger(__name__)

TEJ_CACHE_DIR = Path(os.environ.get("TEJ_CACHE", str(Path.home() / "tej_cache")))
DATA_ROOT = Path(project_root) / "tej_exports" / "DataExport0806"
LEGACY_SUPPLEMENT_DIR = Path(project_root) / "tej_exports" / "legacy_supplement"
LEGACY_SUPPLEMENT_SCRIPT = Path(project_root) / "scripts" / "extract_legacy_supplement.py"
MANIFEST_CSV = Path(project_root) / "tej_exports" / "DataExport0806_manifest.csv"

ID_SPLIT_COLS = ("證券代碼", "股票代號")   # 合併格式「1101 台泥」用的欄名候選
SOURCE_COLUMN_ALIASES = {
    # 2026-08-10 補匯出的 Q2 財報使用「年/月」；值的語意與舊模板「年月」相同。
    # 只登記已人工核對的精確別名，不做模糊表頭猜測。
    "年月": ("年/月",),
}
_INVALID_ID_STRINGS = {"", "nan", "none", "nat", "null", "na"}


def _sha256_of(path: Path) -> str:
    h = hashlib.sha256()
    with open(path, "rb") as f:
        while True:
            chunk = f.read(8 * 1024 * 1024)
            if not chunk:
                break
            h.update(chunk)
    return h.hexdigest()


_manifest_cache: dict = None


def _load_manifest() -> dict:
    """讀 tej_exports/DataExport0806_manifest.csv (scripts/build_data_manifest.py 產生),
    回傳 {relpath: sha256}。整支程式只讀一次,cache 起來。"""
    global _manifest_cache
    if _manifest_cache is not None:
        return _manifest_cache
    if not MANIFEST_CSV.exists():
        raise FileNotFoundError(
            f"{MANIFEST_CSV} 不存在,先跑 python scripts/build_data_manifest.py 建立 manifest"
            f" (每個 dataset 讀取前都要先過 manifest preflight,見 §Round3 code review)")
    with open(MANIFEST_CSV, encoding="utf-8") as f:
        _manifest_cache = {row["relpath"]: row["sha256"] for row in csv.DictReader(f)}
    return _manifest_cache


def _manifest_expected_relpaths(spec: dict) -> dict:
    """套用跟 _source_files 一樣的選檔邏輯 (file_glob 或 *.xlsx/*.zip),但作用在 manifest
    的檔案清單上,不是磁碟——manifest preflight 要能在磁碟檔案被動過手腳時也抓得到。"""
    source_rel = spec["source_dir"].relative_to(DATA_ROOT).as_posix()
    prefix = source_rel + "/"
    manifest = _load_manifest()
    in_dir = {r: h for r, h in manifest.items() if r.startswith(prefix)}
    if "file_glob" in spec:
        pattern = spec["file_glob"]
        return {r: h for r, h in in_dir.items() if fnmatch.fnmatch(Path(r).name, pattern)}
    return {r: h for r, h in in_dir.items() if r.lower().endswith((".xlsx", ".zip"))}


def _manifest_preflight(files: list, spec: dict, dataset: str) -> None:
    """讀檔前先核對:這個 dataset 實際會讀的檔案集合,跟 manifest 記錄的子集合是否
    完全一致 (無缺、無多)、每個檔案的 SHA-256 是否吻合。任一項不符就 raise,不解析。
    manifest 是 tej_exports/DataExport0806_manifest.csv (scripts/build_data_manifest.py
    建的),代表 2026-08-06 snapshot 建立當下的檔案狀態——這裡要驗證的是「現在磁碟上的
    檔案」有沒有跟「那個時間點記錄的檔案」一致,不是驗證內容本身完整 (那是 §3 全量
    比對的責任)。"""
    expected = _manifest_expected_relpaths(spec)
    actual_relpaths = {f.relative_to(DATA_ROOT).as_posix() for f in files}
    expected_relpaths = set(expected)

    missing = expected_relpaths - actual_relpaths
    extra = actual_relpaths - expected_relpaths
    if missing or extra:
        raise ValueError(
            f"{dataset}:manifest preflight 失敗 —— manifest 記錄過但磁碟上找不到"
            f" {sorted(missing)};磁碟上有但 manifest 沒收錄過 {sorted(extra)}。"
            f" 檔案集合跟 tej_exports/DataExport0806_manifest.csv 對不上,先查是不是有人"
            f" 加了新檔/刪了檔/沒重跑 build_data_manifest.py,不要放行解析。")

    mismatched = []
    for f in files:
        rel = f.relative_to(DATA_ROOT).as_posix()
        if _sha256_of(f) != expected[rel]:
            mismatched.append(rel)
    if mismatched:
        raise ValueError(
            f"{dataset}:manifest preflight 失敗 —— {mismatched} 的 SHA-256 跟 manifest 記錄"
            f" 不符 (檔案內容被改過,或 manifest 過期),先重跑"
            f" scripts/build_data_manifest.py --verify 查原因,不要放行解析。")


def _check_valid_keys(path: Path, df: pd.DataFrame, check_date: bool) -> None:
    """stock_id/date 無效值 fail-closed:空白/NaN 的儲存格 str() 後會變成字面上的
    "nan"/"None" 字串,不是真的 null,原本的 dropna(subset=["stock_id"]) 完全抓不到、
    會靜默把這些列當有效資料留著。這裡先明確檢查、抓到就 raise,不讓它們被 dropna
    靜默吃掉。"""
    id_norm = df["stock_id"].astype(str).str.strip().str.lower()
    bad_id_mask = df["stock_id"].isna() | id_norm.isin(_INVALID_ID_STRINGS)
    n_bad_id = int(bad_id_mask.sum())
    n_bad_date = 0
    if check_date:
        n_bad_date = int(df["date"].isna().sum())
    if n_bad_id or n_bad_date:
        raise ValueError(
            f"{path.relative_to(DATA_ROOT)}:{n_bad_id} 列 stock_id 無效 (空白/NaN/nan字串),"
            f"{n_bad_date} 列 date 無法解析。不允許用 dropna 靜默丟掉這些列,先去原始檔"
            f"查是哪幾列 (可能是合計列/備註列混進資料裡)。")


def _read_xlsx_raw(path: Path):
    """Round 3 review 第 5 項:`row_numbers` 必須來自 XLSX 工作表本身的實體列
    identity,不能是 `range(2, 2+len(df))` 這種從 DataFrame 長度反推的公式
    (即使數值上常常剛好相等,那也只是巧合,不是保證)。這裡完全不用
    `pd.read_excel`/`pd.ExcelFile`,改用 `openpyxl` 直接逐列走訪工作表,
    `enumerate(ws.iter_rows(), start=1)` 給的就是 openpyxl 自己從工作表 XML
    算出來的實體列序號 (含完全空白的列——已實測 `EmptyCell` 一樣會被
    `iter_rows()` 迭代到,不會被跳過),`row_numbers` 直接來自這個列舉,
    不是任何長度公式的產物。

    不用 pandas 讀取,也就不需要 `dtype=str`/`keep_default_na` 這些參數
    ——儲存格值一律 `str(v)`(`None` 轉成空字串 `""`,跟 round 2 review 的
    `keep_default_na=False, na_values=[]` 語意一致:只有原生空白儲存格會變
    空字串,任何非空白字面 token 原樣保留,不經過 pandas 的 NA 字面值清單)。"""
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    try:
        sheet_name = wb.sheetnames[0]
        ws = wb[sheet_name]
        columns = None
        data_cols = None
        row_numbers = []
        for physical_row_idx, row in enumerate(ws.iter_rows(), start=1):
            values = ["" if c.value is None else str(c.value) for c in row]
            if physical_row_idx == 1:
                columns = values
                data_cols = {c: [] for c in columns}
                continue
            if len(values) < len(columns):
                values = values + [""] * (len(columns) - len(values))
            for col, v in zip(columns, values):
                data_cols[col].append(v)
            row_numbers.append(physical_row_idx)
        if columns is None:
            raise ValueError(f"{path.name}:工作表 {sheet_name!r} 是空的,連表頭都沒有")
        df = pd.DataFrame(data_cols, columns=columns)
    finally:
        wb.close()
    if len(row_numbers) != len(df):
        raise RuntimeError(
            f"{path.name}:row_numbers 長度 {len(row_numbers)} 跟解析出來的列數"
            f" {len(df)} 對不上 (內部一致性檢查失敗,不放行)")
    return df, sheet_name, row_numbers


def _read_zip_csv_raw(path: Path):
    """Round 3 review 第 5 項:ZIP 內 UTF-16 + Tab 分隔 csv 的實體列號,一律用
    Python 內建 `csv.reader` 自己逐筆邏輯紀錄解析,不再假手 `pd.read_csv`。

    `csv.reader` 支援 RFC 4180 的 quoted 多行欄位 (一個欄位裡合法內嵌實體
    換行字元)——這種情況下**一筆邏輯紀錄會跨過不只一個實體列**,單純逐筆
    紀錄 `+= 1` 算不出正確的實體列號,而且 TEJ 匯出格式本來就不預期有這種
    內嵌換行的儲存格。這裡明確偵測:任何欄位值含 `\\n`/`\\r`,直接 fail-closed
    (不是靜默算一個可能是錯的列號)。沒有這種欄位時,`csv.reader` 保證一筆
    邏輯紀錄對應一個實體列,`row_numbers` 用逐筆遞增產生,不是套用固定公式。

    完全不經過 pandas 的 NA 偵測機制 (不管 `keep_default_na` 這些參數)——
    `csv.reader` 對空欄位本來就回傳空字串,不是 NaN,任何非空白字面 token
    (含 `"N/A"`/`"NA"`/`"NULL"`/`"nan"`) 原樣保留,天然滿足 round 2 review
    第 4 項的要求,不需要額外設定。"""
    with zipfile.ZipFile(path) as z:
        csv_names = [n for n in z.namelist() if n.lower().endswith(".csv")]
        if not csv_names:
            raise ValueError(f"{path.name} 裡沒有 .csv")
        member = csv_names[0]
        with z.open(member) as f:
            text = f.read().decode("utf-16")

    reader = csv.reader(io.StringIO(text), delimiter="\t")
    all_rows = list(reader)
    if not all_rows:
        raise ValueError(f"{path.name}/{member}:內容是空的,連表頭都沒有")
    header = all_rows[0]
    ncols = len(header)

    data_rows = []
    row_numbers = []
    for physical_row_idx, fields in enumerate(all_rows[1:], start=2):
        for field in fields:
            if "\n" in field or "\r" in field:
                raise ValueError(
                    f"{path.name}/{member}:第 {physical_row_idx} 筆邏輯紀錄的某個欄位"
                    f"含有內嵌實體換行字元 (quoted 多行欄位)——這代表一筆邏輯紀錄跨了"
                    f"不只一個實體列,無法用逐筆遞增可靠算出實體列號,寧可 fail-closed"
                    f"也不冒充一個可能算錯的列號 (Round 3 review 第 5 項)。")
        if len(fields) < ncols:
            fields = fields + [""] * (ncols - len(fields))   # 實體空白列,補齊成全空字串
        elif len(fields) > ncols:
            raise ValueError(
                f"{path.name}/{member}:第 {physical_row_idx} 筆紀錄有 {len(fields)} 欄,"
                f"表頭只有 {ncols} 欄,格式異常,不放行")
        data_rows.append(fields)
        row_numbers.append(physical_row_idx)

    df = pd.DataFrame(data_rows, columns=header) if data_rows else pd.DataFrame(columns=header)
    if len(row_numbers) != len(df):
        raise RuntimeError(
            f"{path.name}/{member}:row_numbers 長度 {len(row_numbers)} 跟解析出來的列數"
            f" {len(df)} 對不上 (內部一致性檢查失敗,不放行)")
    return df, member, row_numbers


def _read_raw_table(path: Path, return_locators: bool = False):
    """讀一份 DataExport0806 原始檔,回傳未經欄位對應的 DataFrame。
    支援 .xlsx (TEJ 查詢精靈匯出,見 `_read_xlsx_raw`) 與 .zip (大量匯出用,
    內含一個 UTF-16 + Tab 分隔的 .csv,見 `_read_zip_csv_raw`;之所以用 .zip
    包 csv 是為了跳過 Excel 單檔 1,048,576 列上限)。

    `return_locators=True` 時額外回傳品質證據用的兩項 locator (§C.9):
      · `container_member`——.zip 是實際被讀取的 csv 成員檔名,.xlsx 是實際被
        讀取的工作表名稱。
      · `row_numbers`——跟回傳的 `df` 逐列對齊的「實體、已對齊表頭」列號
        (list[int],1-based,表頭算第 1 列)。**這是 `_read_xlsx_raw`/
        `_read_zip_csv_raw` 直接從來源檔案的實體結構產生的權威值** (Round 3
        review 第 5 項:不是 `index + 2` 這種從 DataFrame 長度反推的公式,
        也不再是套在另一個函式名字下的同一條公式——兩支 reader 都會在回傳前
        自行核對 `len(row_numbers) == len(df)`,對不上直接 raise)。"""
    if path.suffix.lower() == ".zip":
        df, container_member, row_numbers = _read_zip_csv_raw(path)
    else:
        df, container_member, row_numbers = _read_xlsx_raw(path)
    if not return_locators:
        return df
    return df, container_member, row_numbers


def _split_id_name(df: pd.DataFrame) -> pd.DataFrame:
    """欄位對齊:部分匯出把代號/名稱分兩欄 (代號,名稱),部分合併成一欄
    (如「1101 台泥」,新版查詢精靈常見)。統一拆成 stock_id / stock_name。"""
    df = df.copy()
    if "代號" in df.columns and "名稱" in df.columns:
        df["stock_id"] = df["代號"].astype(str).str.strip()
        df["stock_name"] = df["名稱"].astype(str).str.strip()
        return df
    id_col = next((c for c in ID_SPLIT_COLS if c in df.columns), None)
    if id_col is None:
        raise ValueError(f"找不到代號欄位 (代號/{'/'.join(ID_SPLIT_COLS)}),"
                          f"現有欄位:{list(df.columns)}")
    raw = df[id_col].astype(str).str.strip()
    parts = raw.str.split(n=1, expand=True)
    df["stock_id"] = parts[0].str.strip()
    df["stock_name"] = parts[1].str.strip() if 1 in parts.columns else ""
    return df


def _normalize_source_column_aliases(df: pd.DataFrame, source_name: str = "") -> pd.DataFrame:
    """將已核對的 TEJ 表頭別名正規化成 canonical 欄名。

    canonical 與 alias 同時存在時必須逐列一致，否則 fail-closed，
    不允許新匯出檔靜默覆蓋原欄位。
    """
    out = df.copy()
    for canonical, aliases in SOURCE_COLUMN_ALIASES.items():
        present = [alias for alias in aliases if alias in out.columns]
        if canonical in out.columns:
            for alias in present:
                left = out[canonical].astype(str).str.strip()
                right = out[alias].astype(str).str.strip()
                mismatch = ~(left.eq(right) | (out[canonical].isna() & out[alias].isna()))
                if mismatch.any():
                    raise ValueError(
                        f"{source_name or '<source>'}:canonical 欄位 {canonical!r} 與別名"
                        f" {alias!r} 有 {int(mismatch.sum())} 列不一致，禁止靜默覆蓋")
        elif present:
            if len(present) > 1:
                raise ValueError(f"{source_name or '<source>'}:{canonical!r} 同時出現多個別名 {present}")
            out[canonical] = out[present[0]]
    return out


def _parse_dates(series: pd.Series, date_format=None) -> pd.Series:
    """日期解析:同一類別的舊檔 (2004/12/31) 跟新檔 (純數字 20260806) 格式不同,
    純數字格式 pd.to_datetime 預設不會自動判讀 (回 NaT),需要二次嘗試。"""
    raw = series.astype(str).str.strip().str.replace(r"\.0$", "", regex=True)
    if date_format:
        formats = (date_format,) if isinstance(date_format, str) else tuple(date_format)
        if not formats:
            raise ValueError("date_format 序列不可為空")
        dt = pd.Series(pd.NaT, index=raw.index, dtype="datetime64[ns]")
        for fmt in formats:
            pending = dt.isna()
            if pending.any():
                dt.loc[pending] = pd.to_datetime(raw[pending], format=fmt, errors="coerce")
        return dt
    dt = pd.to_datetime(raw, errors="coerce")
    retry_mask = dt.isna() & raw.str.match(r"^\d{8}$")
    if retry_mask.any():
        dt.loc[retry_mask] = pd.to_datetime(raw[retry_mask], format="%Y%m%d", errors="coerce")
    return dt


DATASETS = {
    "price_valuation": {
        "source_dir": DATA_ROOT / "個股股價、本益比2004-20260806",
        "date_col": "年月日",
        "required_cols": ["年月日", "開盤價(元)", "最高價(元)", "最低價(元)",
                           "收盤價(元)", "成交量(千股)"],
        "min_rows": 4_000_000, "min_stocks": 1500, "expected_date_min": "2004-01-05",
        "rename": {
            "開盤價(元)": "open", "最高價(元)": "max", "最低價(元)": "min",
            "收盤價(元)": "close", "成交量(千股)": "_volume_thousand_shares",
            "本益比-TSE": "PER_TSE", "本益比-TEJ": "PER_TEJ",
            "股價淨值比-TSE": "PBR_TSE", "股價淨值比-TEJ": "PBR_TEJ",
            "股利殖利率-TSE": "dividend_yield_TSE", "股利殖利率-TEJ": "dividend_yield_TEJ",
        },
        "thousand_cols": {"_volume_thousand_shares": "Trading_Volume"},
        "numeric_cols": ["open", "max", "min", "close", "PER_TSE", "PER_TEJ",
                          "PBR_TSE", "PBR_TEJ", "dividend_yield_TSE", "dividend_yield_TEJ"],
    },
    "institutional_flow": {
        "source_dir": DATA_ROOT / "法人回測2004-20260806",
        "date_col": "年月日",
        "required_cols": ["年月日", "外資買賣超(千股)", "投信買賣超(千股)", "自營買賣超(千股)"],
        "min_rows": 4_000_000, "min_stocks": 1500, "expected_date_min": "2004-01-05",
        "rename": {
            "外資買賣超(千股)": "_foreign_net_thousand",
            "投信買賣超(千股)": "_trust_net_thousand",
            "自營買賣超(千股)": "_dealer_net_thousand",
        },
        "thousand_cols": {
            "_foreign_net_thousand": "foreign_net",
            "_trust_net_thousand": "trust_net",
            "_dealer_net_thousand": "dealer_net",
        },
        "numeric_cols": [],
    },
    # 財報寬版子集,對齊舊 fundamentals_quarterly 7 欄輸出。
    # ROE(A)稅後新原始檔沒有欄位,靠 legacy_supplement/roe_after_tax.parquet 補
    # (只有 2019-03~2026-03,舊來源本身就只有這段;見
    # scripts/extract_legacy_supplement.py 與 docs/資料快照遷移_DataExport0806.md §5)。
    "fundamentals_quarterly": {
        "source_dir": DATA_ROOT / "財報2004~202606",
        "date_col": "年月",
        "date_format": "%Y%m",
        "required_cols": ["證券代碼", "年月", "歸屬母公司淨利（損）", "每股盈餘", "營業利益"],
        "min_rows": 60_000, "min_stocks": 1000, "expected_date_min": "2005-12-31",
        "rename": {
            "歸屬母公司淨利（損）": "_net_income_thousand",
            "每股盈餘": "eps",
            "營業利益": "_operating_income_thousand",
        },
        "thousand_cols": {"_net_income_thousand": "net_income",
                           "_operating_income_thousand": "operating_income"},
        "numeric_cols": ["eps"],
        "supplement": LEGACY_SUPPLEMENT_DIR / "roe_after_tax.parquet",
    },
    "revenue_growth": {
        "source_dir": DATA_ROOT / "月營收2004-202608",
        "date_col": "年月",
        "date_format": "%Y%m",
        "required_cols": ["證券代碼", "年月", "單月營收成長率％"],
        "min_rows": 200_000, "min_stocks": 1500, "expected_date_min": "2004-01-31",
        "rename": {"單月營收成長率％": "revenue_yoy_pct"},
        "thousand_cols": {},
        "numeric_cols": ["revenue_yoy_pct"],
    },
    # recurring_net_income (常續性稅後淨利) 新原始檔沒有,靠 legacy_supplement 補 (2019-03+)。
    # revenue_last_year / cum_revenue_last_year 同理,見 monthly_revenue 的 supplement。
    "monthly_revenue": {
        "source_dir": DATA_ROOT / "月營收2004-202608",
        "date_col": "年月",
        "date_format": "%Y%m",
        "required_cols": ["證券代碼", "年月", "營收發布日", "單月營收成長率％", "單月營收(千元)"],
        "min_rows": 200_000, "min_stocks": 1500, "expected_date_min": "2004-01-31",
        "rename": {
            "營收發布日": "release_date",
            "單月營收成長率％": "revenue_yoy_pct",
            "單月營收(千元)": "_revenue_thousand",
            "累計營收(千元)": "_cum_revenue_thousand",
        },
        "extra_date_cols": {"release_date": "%Y%m%d"},
        "thousand_cols": {
            "_revenue_thousand": "revenue",
            "_cum_revenue_thousand": "cum_revenue",
        },
        "numeric_cols": ["revenue_yoy_pct"],
        "supplement": LEGACY_SUPPLEMENT_DIR / "revenue_last_year.parquet",
    },
    "financial_statements": {
        "source_dir": DATA_ROOT / "財報2004~202606",
        "date_col": "年月",
        # 舊快照用 202606；2026-08-10 補匯出 Q2 用 2026/06。只接受這兩種已核對格式。
        "date_format": ("%Y%m", "%Y/%m"),
        "required_cols": ["證券代碼", "年月", "季別", "財報發布日", "營業收入淨額",
                           "歸屬母公司淨利（損）", "每股盈餘", "資產總額", "負債總額",
                           "股東權益總額", "來自營運之現金流量"],
        "min_rows": 60_000, "min_stocks": 1000, "expected_date_min": "2005-12-31",
        "rename": {
            "季別": "quarter",
            "財報發布日": "release_date",
            "營業收入淨額": "_revenue_thousand",
            "營業毛利": "_gross_profit_thousand",
            "營業利益": "_operating_income_thousand",
            "歸屬母公司淨利（損）": "_net_income_thousand",
            "每股盈餘": "eps",
            "資產總額": "_total_assets_thousand",
            "負債總額": "_total_liab_thousand",
            "流動資產": "_current_assets_thousand",
            "流動負債": "_current_liab_thousand",
            "股東權益總額": "_equity_thousand",
            "來自營運之現金流量": "_ocf_thousand",
            "  購置不動產廠房設備（含預付）－CFI": "_capex_thousand",
        },
        "extra_date_cols": {"release_date": "%Y/%m/%d"},
        "thousand_cols": {
            "_revenue_thousand": "revenue",
            "_gross_profit_thousand": "gross_profit",
            "_operating_income_thousand": "operating_income",
            "_net_income_thousand": "net_income",
            "_total_assets_thousand": "total_assets",
            "_total_liab_thousand": "total_liabilities",
            "_current_assets_thousand": "current_assets",
            "_current_liab_thousand": "current_liabilities",
            "_equity_thousand": "equity",
            "_ocf_thousand": "operating_cash_flow",
            "_capex_thousand": "capex",
        },
        "numeric_cols": ["eps", "quarter"],
        # 常續性稅後淨利新原始檔沒有,靠 legacy_supplement 補 (只有 2019-03~2026-03)。
        "supplement": LEGACY_SUPPLEMENT_DIR / "recurring_net_income.parquet",
    },
    "institutional_gross": {
        "source_dir": DATA_ROOT / "法人回測2004-20260806",
        "date_col": "年月日",
        "required_cols": ["年月日", "外資買進張數", "外資賣出張數", "投信買進張數",
                           "投信賣出張數", "外資總投資股率%", "投信持股率%"],
        "min_rows": 4_000_000, "min_stocks": 1500, "expected_date_min": "2004-01-05",
        "rename": {
            "外資買進張數": "_foreign_buy_lots",
            "外資賣出張數": "_foreign_sell_lots",
            "投信買進張數": "_trust_buy_lots",
            "投信賣出張數": "_trust_sell_lots",
            "外資總投資股率%": "foreign_holding_pct",
            "投信持股率%": "trust_holding_pct",
        },
        "thousand_cols": {
            "_foreign_buy_lots": "foreign_buy",
            "_foreign_sell_lots": "foreign_sell",
            "_trust_buy_lots": "trust_buy",
            "_trust_sell_lots": "trust_sell",
        },
        "numeric_cols": ["foreign_holding_pct", "trust_holding_pct"],
    },
    "margin_balance": {
        "source_dir": DATA_ROOT / "融資融券2004-20260806",
        "date_col": "年月日",
        "required_cols": ["年月日", "融資餘額(張)", "融券餘額(張)"],
        "min_rows": 4_000_000, "min_stocks": 1500, "expected_date_min": "2004-01-05",
        "rename": {
            "融資餘額(張)": "margin_balance",
            "融資買進(張)": "margin_buy",
            "融資賣出(張)": "margin_sell",
            "融資增減(張)": "margin_change",
            "融資使用率": "margin_usage_rate",
            "融券餘額(張)": "short_balance",
            "券資比": "short_margin_ratio",
        },
        "thousand_cols": {},
        "numeric_cols": ["margin_balance", "margin_buy", "margin_sell", "margin_change",
                          "margin_usage_rate", "short_balance", "short_margin_ratio"],
    },
    "industry_map": {
        "source_dir": DATA_ROOT / "產業類別",
        "static": True,
        # 用「歷史」不用「現在」:「現在產業類別.xlsx」只有 1952 檔 (排除已下市/已變更者),
        # 「歷史產業類別.xlsx」前 12 欄跟「現在」一模一樣、多了異動歷程欄,但代號集合
        # 跟舊 inbox_industry/Industry.xlsx 的 2436 檔完全一致 (已逐檔比對)。
        # 用「現在」會讓 484 檔失去產業對照,是實質倒退。
        "file_glob": "歷史產業類別.xlsx",
        "required_cols": ["代號", "TSE產業_代碼", "TEJ產業_代碼", "TEJ子產業_代碼"],
        "min_stocks": 2000,
        "rename": {
            "TSE產業_代碼": "tse_ind_code", "TSE產業_名稱": "tse_ind_name",
            "TEJ產業_代碼": "tej_ind_code", "TEJ產業_名稱": "tej_ind_name",
            "TEJ子產業_代碼": "tej_subind_code", "TEJ子產業_名稱": "tej_subind_name",
        },
        "thousand_cols": {},
        "numeric_cols": [],
    },
    "tdcc_weekly": {
        "source_dir": DATA_ROOT / "集保大戶2019-20260806",
        "date_col": "年月日",
        "required_cols": ["年月日", "1000張以上  (比率)", "集保總人數", "集保總張數(千股)"],
        "min_rows": 300_000, "min_stocks": 1500, "expected_date_min": "2019-01-11",
        "rename": {
            "1000張以上  (比率)": "ratio_1000up",
            "1 張以下(比率)": "ratio_le1",
            "1 -5  張(比率)": "ratio_1to5",
            "5 -10 張(比率)": "ratio_5to10",
            "集保總人數": "holders",
            "集保總張數(千股)": "total_lots_thousand",
        },
        "thousand_cols": {},
        "numeric_cols": ["ratio_1000up", "ratio_le1", "ratio_1to5", "ratio_5to10",
                          "holders", "total_lots_thousand"],
    },
    "director_pledge": {
        "source_dir": DATA_ROOT / "集團分類+董監質押與持股比2019-202606",
        "date_col": "年月",
        "date_format": "%Y/%m",   # 這份「年月」是 "2026/06" 斜線格式,跟財報/月營收的純數字 "202606" 不同
        "required_cols": ["年月", "董監質押%", "董監持股%"],
        "min_rows": 80_000, "min_stocks": 1500, "expected_date_min": "2019-01-31",
        "rename": {
            "董監質押%": "pledge_pct",
            "董監持股%": "director_holding_pct",
            "集團名稱": "group_name",
        },
        "thousand_cols": {},
        "numeric_cols": ["pledge_pct", "director_holding_pct"],
    },
}


def _source_files(spec: dict) -> list[Path]:
    source_dir: Path = spec["source_dir"]
    if not source_dir.exists():
        raise FileNotFoundError(f"{source_dir} 不存在")
    if "file_glob" in spec:
        return sorted(source_dir.rglob(spec["file_glob"]))
    return sorted(list(source_dir.rglob("*.xlsx")) + list(source_dir.rglob("*.zip")))


def _check_required_cols(path: Path, raw: pd.DataFrame, spec: dict) -> None:
    """必要欄位 fail-closed:少了任何一個就整支腳本炸掉,不要默默跳過或產生縮水資料。
    這些是「這份檔案沒有它,這個 dataset 就沒有意義」的欄位 (見 §7 code review)。"""
    required = spec.get("required_cols") or []
    # `_split_id_name` 已將「代號+名稱」與「證券代碼=代號 名稱」兩種
    # TEJ 格式統一成 stock_id/stock_name。因此 required_cols 裡的「證券代碼」
    # 可由已通過 `_split_id_name` 強制檢查的 stock_id 滿足；其他欄位仍逐字符匹配。
    missing = [c for c in required
               if c not in raw.columns and not (c == "證券代碼" and "stock_id" in raw.columns)]
    if missing:
        raise ValueError(
            f"{path.relative_to(DATA_ROOT)}:缺少必要欄位 {missing} (現有欄位:{list(raw.columns)})。"
            f" 這通常代表 TEJ 換了欄位命名或這份檔案匯出時漏勾欄位 —— 不要放寬 required_cols "
            f"來讓它跑過去,先去確認原始檔本身是不是有問題。")


# =============================================================================
# Phase A1(DataExport0806 V2 隔離建置預註冊 §B/§C.1/§C.9):精確 schema 公式、
# 顯式型別凍結、in-memory 品質證據。只改 `_load_one` 這支確定性解析函式跟
# `load_source` 既有的跨檔彙總邏輯,不新增 builder/orchestrator/receipt/
# snapshot_id;呼叫端不傳 `dataset`/`evidence_sink` 時,行為/回傳型別
# 100% 跟舊版相容 (仍是一份 plain DataFrame)。
# =============================================================================

def _final_target_columns(spec: dict) -> list:
    """§B 精確公式:final_columns 的非 standard_keys、非 supplement 部分。
    `sorted((rename 全部值 − thousand_cols 中繼 key) ∪ thousand_cols 目標值)`——
    這個集合只跟 spec 本身有關,不因某個原始檔缺了某欄而變動 (§B 第 18 輪)。"""
    thousand_cols = spec.get("thousand_cols", {})
    rename_targets = set(spec["rename"].values())
    thousand_keys = set(thousand_cols.keys())
    thousand_targets = set(thousand_cols.values())
    return sorted((rename_targets - thousand_keys) | thousand_targets)


def _numeric_target_columns(spec: dict) -> set:
    """§C.1 顯式型別凍結:數值目標欄位 = numeric_cols ∪ thousand_cols 的目標值。
    其餘的最終欄位 (stock_id/stock_name/date/release_date/industry 代碼名稱/
    group_name) 一律是字串欄位——見 §B 表格逐 dataset 核對,11 個 dataset的
    rename 目標欄位沒有例外落在這兩個集合以外。"""
    return set(spec.get("numeric_cols", [])) | set(spec.get("thousand_cols", {}).values())


def _check_final_schema(df: pd.DataFrame, spec: dict, label: str) -> None:
    """§B/Phase A1 要求:拒絕多出、缺少或順序錯亂的最終欄位——欄位集合跟順序
    必須逐字等於 spec 凍結算出來的預期值,不因來源檔案的欄位差異而變動。"""
    if spec.get("static"):
        expected = ["stock_id", "stock_name"] + list(spec["rename"].values())
    else:
        expected = ["stock_id", "stock_name", "date"] + _final_target_columns(spec)
    actual = list(df.columns)
    if actual != expected:
        missing = [c for c in expected if c not in actual]
        extra = [c for c in actual if c not in expected]
        raise ValueError(
            f"{label}:最終欄位跟 §B 凍結的 schema 公式不符——缺少 {missing}、"
            f"多出 {extra} (實際順序 {actual},預期順序 {expected})。欄位集合與順序"
            f"必須逐字相符,不允許來源檔案的欄位差異影響最終 schema。")


def _to_string_column(series: pd.Series) -> pd.Series:
    """§C.1 顯式型別凍結:identifiers/names/dates/industry 代碼名稱/group_name
    一律正規化成 nullable 字串,不進 `pd.to_numeric` 或任何數值轉換路徑
    (industry code 尤其不能被誤轉成數字,前導零在數字型別下會被吃掉)。
    先用 `pd.isna` 逐格判斷再 `str().strip()`,避免 `.astype(str)` 把原生
    NaN 字面轉成錯誤的 "nan" 字串。

    Round 3 review 第 6 項修正:round 2 review 為了保留非空白字面 token (如
    `"N/A"`) 把讀檔階段的 pandas NA 偵測關掉之後,原生空白儲存格改用空字串
    `""` 表示,不再是原生 `NaN`——這支函式當時沒有跟著更新,`str("").strip()`
    == `""`,`pd.isna(v)` 對它是 `False`,於是空字串被當成「有內容的合法值」
    保留下來,不是 `pd.NA`。**canonical 字串欄位的空白定義是「原生缺值」+
    「trim 後變空字串」兩者的聯集**——這裡改成先 strip 再判斷是不是空字串,
    是空字串就正規化成 `None`(→`pd.NA`),不當成非空值保留。"""
    def _norm(v):
        if pd.isna(v):
            return None
        s = str(v).strip()
        return s if s else None
    return series.map(_norm).astype("string")


def _dedup_key_v1(dataset: str, source_relpath: str, source_container_member,
                   source_row_number: int, target_column: str) -> str:
    """§C.9 canonical `dedup_key_v1` 序列化 (第 18 輪凍結):固定欄位順序的
    JSON 陣列、ensure_ascii、緊湊分隔符號、UTF-8 編碼後 SHA-256,取代拼接
    字串再雜湊的分隔符號歧義寫法。"""
    canonical_array = [
        "dedup_key_v1", dataset, source_relpath, source_container_member,
        source_row_number, target_column,
    ]
    payload = json.dumps(canonical_array, ensure_ascii=True,
                          separators=(",", ":"), sort_keys=False).encode("utf-8")
    return hashlib.sha256(payload).hexdigest()


NULL_CAUSE_RETAINED_BLANK = "RETAINED_BLANK"
NULL_CAUSE_RETAINED_UNPARSEABLE = "RETAINED_UNPARSEABLE"
NULL_CAUSE_SOURCE_COLUMN_ABSENT = "SOURCE_COLUMN_ABSENT"
NULL_CAUSE_SUPPLEMENT_KEY_NOT_COVERED = "SUPPLEMENT_KEY_NOT_COVERED"
NULL_CAUSE_OTHER_UNEXPLAINED = "OTHER_UNEXPLAINED"
_FINAL_NULL_CAUSES = (NULL_CAUSE_RETAINED_BLANK, NULL_CAUSE_RETAINED_UNPARSEABLE,
                      NULL_CAUSE_SOURCE_COLUMN_ABSENT, NULL_CAUSE_SUPPLEMENT_KEY_NOT_COVERED,
                      NULL_CAUSE_OTHER_UNEXPLAINED)


def _classify_numeric_cells(*, dataset: str, path: Path, chinese_col: str, final_target: str,
                             source_series: pd.Series, unit_scale: float, container_member,
                             row_numbers, stock_id_series: pd.Series, date_series: pd.Series,
                             collect_records: bool):
    """§C.9 階段一:去重前的來源 cell 分類 (單一原始檔 × 單一目標欄位)。

    回傳 `(resulting, stage_one_counts, cell_records, provenance)`:
      · `resulting`——unit_scale 套用後、寫進候選欄位的 float64 Series
        (blank/unparseable 的位置是 NaN)。
      · `stage_one_counts`——source_row_count/column_present_row_count/
        column_absent_row_count/parsed_numeric_cell_count/blank_cell_count/
        unparseable_cell_count 六個分類,滿足
        `column_present_row_count = parsed + blank + unparseable` 跟
        `source_row_count = column_present_row_count + column_absent_row_count`
        兩個恆等式 (這裡固定 column_absent_row_count=0,因為只有來源欄位
        存在時才會呼叫這支函式)。
      · `cell_records`——只有 blank/unparseable 兩類的逐列 locator 證據
        (§C.9:sidecar 本身只收這兩類,避免膨脹);`collect_records=False`
        時固定回傳 `[]`,呼叫端沒要求品質證據時不做這筆額外工作。
      · `provenance`——跟 `source_series` 逐列對齊的 list,每一列是一個
        `(tag, ...)` tuple:`("PARSED", 最終數值)`/`(NULL_CAUSE_RETAINED_
        BLANK,)`/`(NULL_CAUSE_RETAINED_UNPARSEABLE, 原始 raw token)`。
        Round 3 review 第 1 項的核心:兩列同樣是 NaN,但一列是原生空白、
        一列是文字「.」,provenance tuple 不相等——這是判斷「安全完全重複」
        還是「衝突」的唯一依據,不能只比較 `resulting` 的最終值 (兩者的
        `resulting` 都是 NaN,單看最終值會誤判成安全重複)。

    `row_numbers`(Round 2 review 第 5 項,Round 3 review 第 5 項改為
    `_read_xlsx_raw`/`_read_zip_csv_raw` 從實體結構直接產生的權威值)是
    跟 `source_series` 逐列對齊的實體列號陣列;這支函式**不自己反推**列號,
    只在 `collect_records=True` 時原樣引用。
    """
    str_repr = source_series.map(lambda v: None if pd.isna(v) else str(v).strip())
    is_blank = pd.Series([s is None or s == "" for s in str_repr],
                          index=source_series.index, dtype=bool)
    parsed = pd.to_numeric(source_series, errors="coerce")
    is_unparseable = (~is_blank) & parsed.isna()
    is_parsed = (~is_blank) & (~is_unparseable)
    resulting = (parsed * unit_scale).astype("float64")

    n = len(source_series)
    stage_one = {
        "dataset": dataset, "source_column": chinese_col, "target_column": final_target,
        "source_row_count": int(n), "column_present_row_count": int(n),
        "column_absent_row_count": 0,
        "parsed_numeric_cell_count": int(is_parsed.sum()),
        "blank_cell_count": int(is_blank.sum()),
        "unparseable_cell_count": int(is_unparseable.sum()),
    }
    # 內部一致性防呆 (fail-closed):stage-one 分類的 blank+unparseable 加總,
    # 必須等於這欄實際寫出的 null 數——這個等式對不上代表分類邏輯本身有 bug,
    # 不能讓一份跟自己宣稱的 accounting 對不起來的資料溜過去。
    actual_null = int(resulting.isna().sum())
    expected_null = stage_one["blank_cell_count"] + stage_one["unparseable_cell_count"]
    if actual_null != expected_null:
        raise RuntimeError(
            f"{dataset}/{final_target}:最終 null 數 {actual_null} 跟階段一分類"
            f" blank+unparseable={expected_null} 對不上 (內部一致性檢查失敗)")

    is_blank_list = is_blank.tolist()
    is_unparseable_list = is_unparseable.tolist()
    str_repr_list = str_repr.tolist()
    resulting_list = resulting.tolist()
    provenance = []
    for pos in range(n):
        if is_blank_list[pos]:
            provenance.append((NULL_CAUSE_RETAINED_BLANK,))
        elif is_unparseable_list[pos]:
            provenance.append((NULL_CAUSE_RETAINED_UNPARSEABLE, str_repr_list[pos]))
        else:
            provenance.append(("PARSED", resulting_list[pos]))

    cell_records = []
    if collect_records:
        source_relpath = path.relative_to(DATA_ROOT).as_posix()
        source_sha256 = _sha256_of(path)
        stock_id_list = stock_id_series.tolist()
        date_list = date_series.tolist()
        for pos in range(n):
            tag = provenance[pos][0]
            if tag == "PARSED":
                continue
            blank = tag == NULL_CAUSE_RETAINED_BLANK
            row_number = row_numbers[pos]   # 權威值來自 reader,不在這裡反推
            cell_records.append({
                "dataset": dataset, "source_relpath": source_relpath,
                "source_file_sha256": source_sha256,
                "source_container_member": container_member,
                "source_row_number": row_number,
                "stock_id": stock_id_list[pos], "date": date_list[pos],
                "source_column": chinese_col, "target_column": final_target,
                "raw_token": None if blank else str_repr_list[pos],
                "is_blank": blank, "is_unparseable": not blank,
                "parser": "pd.to_numeric", "unit_scale_applied": float(unit_scale),
                "resulting_value": None,
                "dedup_key": _dedup_key_v1(dataset, source_relpath, container_member,
                                            row_number, final_target),
            })
    return resulting, stage_one, cell_records, provenance


def _coverage_status(is_present: bool, non_null_count: int) -> str:
    """§B 逐檔覆蓋矩陣的四選一狀態,這裡只用得到其中三種(第四種
    `NOT_APPLICABLE` 是 supplement-only 欄位,由 `load_source` 在組
    coverage_matrix 時直接標記,不經過這支函式)。"""
    if not is_present:
        return NULL_CAUSE_SOURCE_COLUMN_ABSENT
    return "PRESENT" if non_null_count > 0 else "PRESENT_ALL_NULL"


def _load_one(path: Path, spec: dict, *, dataset: str = "", collect_evidence: bool = False):
    """讀一份原始檔,套用共用管線 (§B):`_check_required_cols`/`_check_valid_keys`
    fail-closed、§B 精確 schema 公式 (缺席的目標欄位補一欄全 null,不讓欄位從
    輸出消失)、§C.1 顯式型別凍結 (數值 `.astype("float64")`、字串欄位顯式
    `"string"` dtype)。

    `collect_evidence=False`(預設,`load_source` 內部非 `return_evidence`
    呼叫方式)時只回傳 `df`,行為/回傳型別 100% 向後相容。`missing_source_
    columns`/`coverage_row` 兩項輕量 metadata **無論哪種模式都會計算**並寫進
    `df.attrs`——這兩項不需要碰 `DATA_ROOT`/檔案雜湊,呼叫端 (`load_source`)
    在每個檔案 concat 前立刻讀取,不依賴 concat 後的 `.attrs`,用來驅動
    「全部原始檔都缺同一欄」的 fail-closed 判定。

    `collect_evidence=True` 時額外回傳 `(df, file_evidence)`。`file_evidence`
    除了 §C.9 階段一的 `cell_records`/`stage_one_counts` 外,還有 Round 3
    review 新增的 `row_records`——**逐列**(不是逐 key 收斂後的字典)的完整
    provenance 清單,每一列含這個實體列的 locator (`source_relpath`/
    `source_file_sha256`/`source_container_member`/`source_row_number`/
    `stock_id`/`date`)、`values`(這一列每個最終欄位的值)、`provenance`
    (這一列每個目標欄位的 `(tag, ...)` 分類 tuple)。`load_source(...,
    return_evidence=True)` 會把多個檔案的 `row_records` 原樣串接、依 key
    分組後才做去重判定 (Round 3 review 第 2 項:conflict adjudication 之前
    不能先把每個檔案的列收斂進字典,那樣會靜默丟掉「兩個檔案 provenance
    不同但字典後蓋前」的證據)。這支函式本身仍然不寫 sidecar 檔案。"""
    if collect_evidence:
        raw, container_member, row_numbers = _read_raw_table(path, return_locators=True)
    else:
        raw, container_member, row_numbers = _read_raw_table(path), None, None
    raw = _normalize_source_column_aliases(raw, path.name)
    raw = _split_id_name(raw)
    _check_required_cols(path, raw, spec)

    rename = spec["rename"]
    thousand_cols = spec.get("thousand_cols", {})
    numeric_targets = _numeric_target_columns(spec)
    n_rows = len(raw)
    missing_source_columns = [c for c in rename if c not in raw.columns]
    if missing_source_columns:
        logger.info(f"  {path.name}:欄位缺失 (跳過對應,補一欄全 null,不讓欄位從輸出消失,"
                     f"§B 第 18 輪規則) {missing_source_columns}")

    df = pd.DataFrame(index=raw.index)
    df["stock_id"] = _to_string_column(raw["stock_id"])
    df["stock_name"] = _to_string_column(raw["stock_name"])

    source_relpath = path.relative_to(DATA_ROOT).as_posix() if collect_evidence else None
    source_sha256 = _sha256_of(path) if collect_evidence else None

    if spec.get("static"):
        coverage_row = {}
        provenance_by_target = {}
        targets = list(rename.values())
        for chinese_col, renamed_col in rename.items():
            present = chinese_col in raw.columns
            if present:
                df[renamed_col] = _to_string_column(raw[chinese_col])
                provenance_by_target[renamed_col] = [
                    (NULL_CAUSE_RETAINED_BLANK,) if pd.isna(v) else ("PARSED", v)
                    for v in df[renamed_col]
                ]
            else:
                df[renamed_col] = pd.array([pd.NA] * n_rows, dtype="string")
                provenance_by_target[renamed_col] = [(NULL_CAUSE_SOURCE_COLUMN_ABSENT,)] * n_rows
            coverage_row[renamed_col] = _coverage_status(present, int(df[renamed_col].notna().sum()))
        df = df[["stock_id", "stock_name"] + targets]
        _check_final_schema(df, spec, path.name)
        _check_valid_keys(path, df, check_date=False)
        df.attrs["missing_source_columns"] = missing_source_columns
        df.attrs["coverage_row"] = coverage_row
        if not collect_evidence:
            return df

        stock_id_list = df["stock_id"].tolist()
        stock_name_list = df["stock_name"].tolist()
        row_records = []
        for i in range(n_rows):
            row_records.append({
                "key": (stock_id_list[i],),
                "locator": {
                    "source_relpath": source_relpath, "source_file_sha256": source_sha256,
                    "source_container_member": container_member,
                    "source_row_number": row_numbers[i],
                    "stock_id": stock_id_list[i], "date": None,
                },
                "values": {"stock_id": stock_id_list[i], "stock_name": stock_name_list[i],
                           **{t: df[t].iloc[i] for t in targets}},
                "provenance": {t: provenance_by_target[t][i] for t in targets},
            })
        return df, {
            "source_relpath": source_relpath, "source_file_sha256": source_sha256,
            "source_container_member": container_member,
            "missing_source_columns": missing_source_columns, "coverage_row": coverage_row,
            "stage_one_counts": [], "cell_records": [], "row_records": row_records,
        }

    date_col = spec["date_col"]
    if date_col not in raw.columns:
        # date_col 已經在 required_cols 檢查過了 (每個 spec 的 required_cols 都包含
        # date_col),理論上走不到這裡;留著純防呆,一樣 fail-closed 不 return 空表。
        raise ValueError(f"{path.relative_to(DATA_ROOT)}:找不到日期欄 {date_col}")
    df["date"] = _to_string_column(
        _parse_dates(raw[date_col], spec.get("date_format")).dt.strftime("%Y-%m-%d"))

    extra_date_cols = spec.get("extra_date_cols") or {}
    stage_one_counts = []
    cell_records = []
    coverage_row = {}
    provenance_by_target = {}   # target_column -> [provenance tuple, ...] (這個檔案逐列對齊)

    for chinese_col, renamed_col in rename.items():
        final_target = thousand_cols.get(renamed_col, renamed_col)
        is_numeric = final_target in numeric_targets

        if chinese_col not in raw.columns:
            if is_numeric:
                df[final_target] = pd.Series([float("nan")] * n_rows, index=raw.index,
                                              dtype="float64")
            else:
                df[final_target] = pd.array([pd.NA] * n_rows, dtype="string")
            coverage_row[final_target] = NULL_CAUSE_SOURCE_COLUMN_ABSENT
            provenance_by_target[final_target] = [(NULL_CAUSE_SOURCE_COLUMN_ABSENT,)] * n_rows
            stage_one_counts.append({
                "dataset": dataset, "source_column": chinese_col, "target_column": final_target,
                "source_row_count": int(n_rows), "column_present_row_count": 0,
                "column_absent_row_count": int(n_rows),
                "parsed_numeric_cell_count": 0, "blank_cell_count": 0,
                "unparseable_cell_count": 0,
            })
            continue

        source_series = raw[chinese_col]
        if is_numeric:
            unit_scale = 1000.0 if renamed_col in thousand_cols else 1.0
            resulting, stage_one, records, provenance = _classify_numeric_cells(
                dataset=dataset, path=path, chinese_col=chinese_col, final_target=final_target,
                source_series=source_series, unit_scale=unit_scale,
                container_member=container_member, row_numbers=row_numbers,
                stock_id_series=df["stock_id"], date_series=df["date"],
                collect_records=collect_evidence,
            )
            df[final_target] = resulting
            stage_one_counts.append(stage_one)
            cell_records.extend(records)
        elif final_target in extra_date_cols:
            fmt = extra_date_cols[final_target]
            df[final_target] = _to_string_column(
                _parse_dates(source_series, fmt).dt.strftime("%Y-%m-%d"))
            provenance = [(NULL_CAUSE_RETAINED_BLANK,) if pd.isna(v) else ("PARSED", v)
                          for v in df[final_target]]
        else:
            df[final_target] = _to_string_column(source_series)
            provenance = [(NULL_CAUSE_RETAINED_BLANK,) if pd.isna(v) else ("PARSED", v)
                          for v in df[final_target]]

        coverage_row[final_target] = _coverage_status(True, int(df[final_target].notna().sum()))
        provenance_by_target[final_target] = provenance

    final_cols = ["stock_id", "stock_name", "date"] + _final_target_columns(spec)
    df = df[final_cols]
    _check_final_schema(df, spec, path.name)
    _check_valid_keys(path, df, check_date=True)

    df.attrs["missing_source_columns"] = missing_source_columns
    df.attrs["coverage_row"] = coverage_row
    if not collect_evidence:
        return df

    targets = _final_target_columns(spec)
    stock_id_list = df["stock_id"].tolist()
    stock_name_list = df["stock_name"].tolist()
    date_list = df["date"].tolist()
    row_records = []
    for i in range(n_rows):
        row_records.append({
            "key": (stock_id_list[i], date_list[i]),
            "locator": {
                "source_relpath": source_relpath, "source_file_sha256": source_sha256,
                "source_container_member": container_member,
                "source_row_number": row_numbers[i],
                "stock_id": stock_id_list[i], "date": date_list[i],
            },
            "values": {"stock_id": stock_id_list[i], "stock_name": stock_name_list[i],
                       "date": date_list[i], **{t: df[t].iloc[i] for t in targets}},
            "provenance": {t: provenance_by_target[t][i] for t in targets},
        })

    return df, {
        "source_relpath": source_relpath, "source_file_sha256": source_sha256,
        "source_container_member": container_member,
        "missing_source_columns": missing_source_columns, "coverage_row": coverage_row,
        "stage_one_counts": stage_one_counts, "cell_records": cell_records,
        "row_records": row_records,
    }


def _check_duplicate_key_conflicts(combined: pd.DataFrame, dataset: str, key_cols=None,
                                    evidence_sink: dict = None) -> None:
    """drop_duplicates(keep="last") 會默默丟掉重複 (stock_id, date) 的其中一列。
    這裡把重複鍵分兩種:
      · 「完全重複」(同 key、所有欄位值都相同)——單純同一份資料被匯出兩次,
        安全去重,但次數要記錄下來,不能無聲無息。
      · 「衝突重複」(同 key、至少一個欄位值不同)——代表原始檔之間互相打架,
        不管衝突筆數佔比多小都 raise,不設任何百分比容忍門檻 (Round 3 review:
        移除原本的 1% 容忍度,一個衝突就是一個需要人去查的問題)。

    `evidence_sink` 是 Phase A1 新增的**選用**參數 (dict):傳入時額外記錄
    `exact_duplicate_mapping`——每一個完全重複的保留 key,對應了幾筆被去重掉
    的來源列 (§C.9「去重映射」)。不傳的話行為跟舊版完全相同。"""
    key_cols = key_cols or ["stock_id", "date"]
    value_cols = [c for c in combined.columns if c not in key_cols + ["stock_name"]]
    dup_mask = combined.duplicated(subset=key_cols, keep=False)
    if not dup_mask.any():
        if evidence_sink is not None:
            evidence_sink["exact_duplicate_mapping"] = []
        return
    dup = combined[dup_mask]

    n_exact_dup_rows = 0
    conflicting_keys = []
    exact_duplicate_mapping = []
    for key, g in dup.groupby(key_cols):
        if g[value_cols].nunique(dropna=False).gt(1).any():
            conflicting_keys.append(key)
        else:
            n_exact_dup_rows += len(g) - 1     # 完全重複:len(g) 列只留 1 列,其餘算「去重掉的」
            exact_duplicate_mapping.append({"key": key, "n_source_rows": int(len(g))})

    if n_exact_dup_rows:
        logger.info(f"  {dataset}:{n_exact_dup_rows} 列是完全重複的 (stock_id, date) "
                    f"(所有欄位值都相同,同一份資料被匯出兩次),已去重")

    if evidence_sink is not None:
        evidence_sink["exact_duplicate_mapping"] = exact_duplicate_mapping

    if conflicting_keys:
        if evidence_sink is not None:
            evidence_sink["conflicting_keys"] = list(conflicting_keys)
        total_keys = combined[key_cols].drop_duplicates().shape[0]
        raise ValueError(
            f"{dataset}:{len(conflicting_keys)}/{total_keys} 個 (stock_id, date) 在不同"
            f"原始檔案間數值不一致 (例:{conflicting_keys[:5]})。不設容忍門檻,一個都不行——"
            f"這代表原始檔彼此口徑不同或有壞檔,先去查是哪些檔案衝突,不要用 keep='last' "
            f"默默選一個了事。")


# code-defined 的預期 schema (Round 5 review):receipt 的 schema 欄位是抽取腳本
# 自己寫的,receipt 被竄改或抽取腳本邏輯漏改時,「跟 receipt 一致」不能代表
# 「跟這支消費端程式真正需要的欄位一致」——要有一份獨立於 receipt 之外、寫死在
# 消費端程式碼裡的預期值。
SUPPLEMENT_SCHEMAS = {
    "roe_after_tax": ["stock_id", "date", "roe_after_tax"],
    "recurring_net_income": ["stock_id", "date", "recurring_net_income"],
    "revenue_last_year": ["stock_id", "date", "revenue_last_year", "cum_revenue_last_year"],
}

# receipt 必要欄位 (Round 6 review):這些欄位「缺席」跟「值不符」一樣要 raise。
# 原本每道比對都用 `if receipt.get(f) is not None and ...` 包起來,少一個欄位就
# 讓那道檢查靜默空轉——這正是本專案最貴的那種 bug:檢查看起來在,實際沒作用。
REQUIRED_RECEIPT_TOP_FIELDS = ("overall_status", "script_sha256", "outputs")
# 這幾個是拿去跟 _profile_supplement() 重算結果逐項比對的統計欄位。
REQUIRED_RECEIPT_PROFILE_FIELDS = ("row_count", "stock_count", "date_min", "date_max",
                                    "duplicate_key_row_count", "null_key_row_count", "dtypes")
# dedup 不能被重算比對 (它記錄的是投影前/合併時的去重過程,消費端手上只有最終
# parquet,重算不出「投影前的原始列有幾個重複鍵」這種資訊),所以是必要欄位但
# 走獨立的巢狀結構驗證 (見 _validate_dedup_metadata),不進 REQUIRED_RECEIPT_
# PROFILE_FIELDS 那組值比對迴圈。
REQUIRED_RECEIPT_OUTPUT_FIELDS = ("sha256", "schema", "dedup") + REQUIRED_RECEIPT_PROFILE_FIELDS

# dedup 巢狀結構驗證用 (Round 7 review):每份統計要有的欄位跟型別。
_DEDUP_STAT_FIELDS = {
    "stage": str,
    "checked_columns": list,
    "n_duplicate_key_rows": int,
    "n_exact_duplicate_rows_removed": int,
    "n_conflicting_keys": int,
}
_DEDUP_STAGES = frozenset({"raw_source", "projected"})
# 每個 supplement 名字底下,dedup.sources 應該恰好出現哪些來源條目 (Round 8
# review)。原本只驗證「每個出現的條目結構對不對」,沒有驗證「條目的集合本身
# 對不對」——來源被改名、多一個、或漏一個都不會被抓到。
_EXPECTED_DEDUP_SOURCE_NAMES = {
    "roe_after_tax": frozenset({"roe_after_tax"}),
    "recurring_net_income": frozenset({"recurring_net_income_2005_2018",
                                        "recurring_net_income_2019plus"}),
    "revenue_last_year": frozenset({"revenue_last_year"}),
}
_DEDUP_COUNT_FIELDS = ("n_duplicate_key_rows", "n_exact_duplicate_rows_removed", "n_conflicting_keys")
_OVERLAP_FIELDS = ("n_overlap_keys", "n_overlap_identical", "n_overlap_conflicting")


def _is_plain_int(v) -> bool:
    """bool 是 int 的子類別 (`isinstance(True, int) == True`),但 True/False 混進
    計數欄位明顯是型別錯誤,要排除掉才是真正的型別檢查。"""
    return isinstance(v, int) and not isinstance(v, bool)


def _validate_dedup_stat(stat, label: str) -> None:
    if not isinstance(stat, dict):
        raise ValueError(f"{label}:dedup 統計不是物件,而是 {type(stat).__name__}")
    missing = [f for f in _DEDUP_STAT_FIELDS if f not in stat]
    if missing:
        raise ValueError(f"{label}:dedup 統計缺少必要欄位 {missing}")
    if not isinstance(stat["stage"], str):
        raise ValueError(f"{label}:dedup 統計的 stage 型別應為 str,實際是"
                          f" {type(stat['stage']).__name__}")
    if not isinstance(stat["checked_columns"], list):
        raise ValueError(f"{label}:dedup 統計的 checked_columns 型別應為 list,實際是"
                          f" {type(stat['checked_columns']).__name__}")
    for field in _DEDUP_COUNT_FIELDS:
        if not _is_plain_int(stat[field]):
            raise ValueError(f"{label}:dedup 統計的 {field!r} 型別應為 int,實際是"
                              f" {type(stat[field]).__name__}")
        if stat[field] < 0:
            raise ValueError(f"{label}:dedup 統計的 {field!r}={stat[field]} 是負數,不合理")
    if stat["stage"] not in _DEDUP_STAGES:
        raise ValueError(f"{label}:dedup 統計的 stage={stat['stage']!r} 不是合法值"
                          f" {sorted(_DEDUP_STAGES)}")
    if stat["n_conflicting_keys"] != 0:
        raise ValueError(
            f"{label}:dedup 統計的 n_conflicting_keys={stat['n_conflicting_keys']} 不是 0"
            f"——正常的抽取邏輯一偵測到衝突就會 raise,不會把非零的衝突數寫進 receipt,"
            f"這裡出現非零代表 receipt 被竄改或抽取邏輯有漏洞,不能信任")


def _validate_dedup_metadata(name: str, dedup) -> None:
    """驗證 receipt 裡 `dedup` 欄位的巢狀結構 (Round 7 review,對應
    scripts/extract_legacy_supplement.py 的 `check_source_duplicates`/
    `_dedupe_or_raise`/`_combine_recurring_windows` 寫入的統計):

      · `sources` 底下每個來源條目要恰好是兩份統計 (raw_source + projected)。
      · 每份統計要有 stage/checked_columns/n_duplicate_key_rows/
        n_exact_duplicate_rows_removed/n_conflicting_keys,型別正確、計數非負、
        stage 是合法值、衝突數必須是 0 (非零代表抽取當下沒有 raise,矛盾)。
      · `recurring_net_income` 額外要有 `cross_window_overlap`,三個計數同樣
        型別正確、非負,且 n_overlap_conflicting 必須是 0、三個計數要滿足
        `n_overlap_keys == n_overlap_identical + n_overlap_conflicting` (Round 8
        review:原本沒驗證這個算術恆等式,三個數字各自「型別對、非負」不代表
        彼此加總得起來)。
      · `sources` 底下出現的來源名字集合,要恰好等於這個 supplement 名字凍結的
        預期集合 `_EXPECTED_DEDUP_SOURCE_NAMES[name]`(Round 8 review:原本只驗證
        「有出現的條目結構對不對」,來源被改名、多一個、少一個都不會被抓到)。

    任何缺欄、型別錯、負數、stage 不合法、非零衝突數、來源名字集合不符、或
    overlap 算術對不起來,都直接 raise——這些欄位本身就是「檢查有沒有真的發生
    過」的證據,證據本身結構壞掉不能放行。"""
    if not isinstance(dedup, dict):
        raise ValueError(f"{name}:receipt 的 dedup 欄位不是物件,而是 {type(dedup).__name__}")

    sources = dedup.get("sources")
    if not isinstance(sources, dict) or not sources:
        raise ValueError(f"{name}:receipt 的 dedup.sources 缺失或不是非空物件")

    expected_names = _EXPECTED_DEDUP_SOURCE_NAMES.get(name)
    if expected_names is None:
        raise ValueError(f"{name}:_EXPECTED_DEDUP_SOURCE_NAMES 沒有登記這個 supplement 名字的"
                          f"預期來源集合,先在 tej_importer.py 裡凍結它")
    actual_names = frozenset(sources.keys())
    if actual_names != expected_names:
        raise ValueError(
            f"{name}:dedup.sources 的來源名字集合 {sorted(actual_names)} 跟凍結的預期"
            f" {sorted(expected_names)} 不符 (缺少 {sorted(expected_names - actual_names)},"
            f"多出 {sorted(actual_names - expected_names)})——來源被改名/漏掉/多一個都不允許")

    for source_name, entries in sources.items():
        if not isinstance(entries, list) or len(entries) != 2:
            raise ValueError(f"{name}:dedup.sources[{source_name!r}] 應該恰好是兩份統計"
                              f" (raw_source + projected) 組成的陣列,實際是 {entries!r}")
        for stat in entries:
            _validate_dedup_stat(stat, f"{name}/{source_name}")
        stages = {stat["stage"] for stat in entries if isinstance(stat, dict) and "stage" in stat}
        if stages != _DEDUP_STAGES:
            raise ValueError(f"{name}:dedup.sources[{source_name!r}] 的 stage 標籤組合是"
                              f" {sorted(stages)},應該恰好是 {sorted(_DEDUP_STAGES)} 各一份"
                              f" (缺一份、標籤重複或寫錯都不行)")

    if name == "recurring_net_income":
        overlap = dedup.get("cross_window_overlap")
        if not isinstance(overlap, dict):
            raise ValueError(f"{name}:receipt 的 dedup.cross_window_overlap 缺失或不是物件"
                              f" (recurring_net_income 有兩個窗口,必須有重疊比對統計)")
        missing = [f for f in _OVERLAP_FIELDS if f not in overlap]
        if missing:
            raise ValueError(f"{name}:dedup.cross_window_overlap 缺少必要欄位 {missing}")
        for field in _OVERLAP_FIELDS:
            if not _is_plain_int(overlap[field]):
                raise ValueError(f"{name}:dedup.cross_window_overlap 的 {field!r} 型別應為"
                                  f" int,實際是 {type(overlap[field]).__name__}")
            if overlap[field] < 0:
                raise ValueError(f"{name}:dedup.cross_window_overlap 的 {field!r}="
                                  f"{overlap[field]} 是負數,不合理")
        expected_total = overlap["n_overlap_identical"] + overlap["n_overlap_conflicting"]
        if overlap["n_overlap_keys"] != expected_total:
            raise ValueError(
                f"{name}:dedup.cross_window_overlap 的算術對不起來——"
                f"n_overlap_keys={overlap['n_overlap_keys']} 應該等於"
                f" n_overlap_identical({overlap['n_overlap_identical']}) +"
                f" n_overlap_conflicting({overlap['n_overlap_conflicting']}) ="
                f" {expected_total},三個數字各自型別/非負都對不代表加總得起來")
        if overlap["n_overlap_conflicting"] != 0:
            raise ValueError(
                f"{name}:dedup.cross_window_overlap.n_overlap_conflicting="
                f"{overlap['n_overlap_conflicting']} 不是 0——正常的合併邏輯一偵測到重疊"
                f"衝突就會 raise,這裡出現非零代表 receipt 被竄改或合併邏輯有漏洞")


def _profile_supplement(df: pd.DataFrame, key_cols=("stock_id", "date")) -> dict:
    """跟 scripts/extract_legacy_supplement.py 的 `_profile` 同一套統計邏輯,在消費端
    獨立重算一次 (Round 5 review)。SHA-256 只驗證「檔案位元組沒被動過」,不驗證
    receipt 裡描述性的統計數字 (row_count/stock_count/date range/重複鍵/null 鍵)
    是不是真的跟檔案內容一致——receipt 可能被手動改過某個欄位而 sha256 沒變
    (例如直接編輯 json),或者抽取腳本本身的 `_profile` 邏輯跟這裡不同步。"""
    key_cols = list(key_cols)
    dup_mask = df.duplicated(subset=key_cols, keep=False)
    null_key_count = int(df[key_cols].isna().any(axis=1).sum())
    return {
        "schema": list(df.columns),
        "dtypes": {c: str(t) for c, t in df.dtypes.items()},
        "row_count": int(len(df)),
        "stock_count": int(df["stock_id"].nunique()),
        "date_min": str(df["date"].min()) if "date" in df.columns and len(df) else None,
        "date_max": str(df["date"].max()) if "date" in df.columns and len(df) else None,
        "duplicate_key_row_count": int(dup_mask.sum()),
        "null_key_row_count": null_key_count,
    }


def _verify_supplement(supplement_path: Path, dataset: str) -> pd.DataFrame:
    """消費 legacy_supplement 之前的完整驗證 (Round 4 review,Round 5 review 加強):
      · receipt.json 存在且 overall_status=PASS。
      · 現在這個 parquet 的 SHA-256 跟 receipt 記錄的一致 (檔案沒被改過)。
      · scripts/extract_legacy_supplement.py 現在的 SHA-256 跟 receipt 記錄的一致
        (抽取腳本沒有被改過卻沒重新產生 receipt——腳本改了邏輯,receipt 卻還是
        舊的,會讓人誤信一份跟現在程式碼對不上的資料)。**腳本本身不存在也要
        raise**(Round 5 review:原本只在腳本存在時才比對雜湊,腳本被誤刪/搬移時
        會直接跳過這道防線,不能因為「檔案不見了」就放行)。
      · schema 對照**程式碼凍結**的 `SUPPLEMENT_SCHEMAS`,不是只信 receipt 自己
        宣稱的 schema (Round 5 review)。
      · 重新計算 row_count/stock_count/date range/重複鍵/null 鍵並跟 receipt 記錄的
        比對,不只是信任 receipt 裡寫的數字 (Round 5 review)。
      · 上述 receipt 欄位**缺席也要 raise**(Round 6 review):原本每道比對都是
        `if receipt.get(f) is not None and ...`,receipt 少一個欄位就讓那道檢查
        靜默空轉,是 fail-open。現在先用 `REQUIRED_RECEIPT_*_FIELDS` 檢查欄位
        存在,才開始比值。
      · `dedup` 巢狀結構驗證 (Round 7 review):`dedup` 記錄抽取腳本的去重/重疊
        過程,消費端算不出來、不能重算比對,但結構要驗證——見
        `_validate_dedup_metadata`。
      · (stock_id, date) 唯一且非 null。
    任一項不符就 raise,不合併——寧可讓 dataset 匯入失敗,也不要用一份沒被驗證過
    的 supplement 悄悄污染輸出。"""
    receipt_path = supplement_path.parent / "receipt.json"
    if not supplement_path.exists():
        raise FileNotFoundError(
            f"{dataset}:{supplement_path} 不存在,先跑一次 scripts/extract_legacy_supplement.py"
            f" (見 docs/資料快照遷移_DataExport0806.md §5)")
    if not receipt_path.exists():
        raise FileNotFoundError(
            f"{dataset}:找不到 {receipt_path},supplement 沒有 receipt 不能信任,"
            f"先跑 scripts/extract_legacy_supplement.py 產生")

    with open(receipt_path, encoding="utf-8") as f:
        receipt = json.load(f)

    if receipt.get("overall_status") != "PASS":
        raise ValueError(f"{dataset}:{receipt_path} 的 overall_status="
                          f"{receipt.get('overall_status')!r},不是 'PASS',不能消費")

    name = supplement_path.stem
    output_receipt = (receipt.get("outputs") or {}).get(name)
    if output_receipt is None:
        raise ValueError(f"{dataset}:{receipt_path} 裡沒有 {name!r} 這個 output 的記錄")

    # Round 6 review:先確認 receipt 該有的欄位「存在」,再談值對不對。原本每道
    # 比對都寫成 `if receipt.get(f) is not None and ...`,receipt 少一個欄位就整道
    # 檢查被跳過 (fail-open)——只要把 row_count 那一行從 json 刪掉,重算比對就
    # 形同不存在。缺欄跟值不符一樣危險,都要 raise。
    missing_top = [f for f in REQUIRED_RECEIPT_TOP_FIELDS if f not in receipt]
    if missing_top:
        raise ValueError(
            f"{dataset}:{receipt_path} 缺少必要欄位 {missing_top}——receipt 不完整就不能"
            f"信任 (缺欄不等於通過,先重跑 scripts/extract_legacy_supplement.py 產生完整 receipt)")
    missing_fields = [f for f in REQUIRED_RECEIPT_OUTPUT_FIELDS if f not in output_receipt]
    if missing_fields:
        raise ValueError(
            f"{dataset}:{receipt_path} 的 outputs[{name!r}] 缺少必要欄位 {missing_fields}"
            f"——這些欄位是後面 schema/統計重算比對的依據,缺任何一個都會讓對應的檢查"
            f"變成空轉,不允許放行 (先重跑 scripts/extract_legacy_supplement.py)")

    actual_hash = _sha256_of(supplement_path)
    if actual_hash != output_receipt["sha256"]:
        raise ValueError(f"{dataset}:{supplement_path} 現在的 SHA-256 跟 receipt 記錄的不符"
                          f" (檔案被改過,或 receipt 過期),先重跑抽取腳本")

    if not LEGACY_SUPPLEMENT_SCRIPT.exists():
        raise FileNotFoundError(
            f"{dataset}:{LEGACY_SUPPLEMENT_SCRIPT} 不存在,無法比對抽取腳本 SHA-256 跟"
            f" {receipt_path} 記錄的是否一致,不能信任這份 supplement——腳本不見了不代表"
            f"可以跳過這道驗證,先確認腳本是不是被誤刪/搬移。")
    actual_script_hash = _sha256_of(LEGACY_SUPPLEMENT_SCRIPT)
    if actual_script_hash != receipt["script_sha256"]:
        raise ValueError(
            f"{dataset}:scripts/extract_legacy_supplement.py 現在的 SHA-256 跟"
            f" {receipt_path} 記錄的不符 (腳本被改過但沒有重新產生 receipt),"
            f"先重跑抽取腳本產生新 receipt")

    expected_schema = SUPPLEMENT_SCHEMAS.get(name)
    if expected_schema is None:
        raise ValueError(f"{dataset}:{name!r} 不在程式碼凍結的 SUPPLEMENT_SCHEMAS 清單裡,"
                          f"先在 tej_importer.py 裡登記它的預期 schema 再消費")

    supp = pd.read_parquet(supplement_path)
    if list(supp.columns) != expected_schema:
        raise ValueError(f"{dataset}:{supplement_path} 的 schema {list(supp.columns)} 跟"
                          f" 程式碼凍結的預期 {expected_schema} 不符 (不允許意外多出/少掉欄位)")
    if output_receipt["schema"] != expected_schema:
        raise ValueError(f"{dataset}:receipt 記錄的 schema {output_receipt['schema']} 跟程式碼"
                          f"凍結的預期 {expected_schema} 不符 (receipt 過期或被竄改,即使"
                          f" SHA-256 對得上這裡也要擋)")

    recomputed = _profile_supplement(supp)
    mismatches = [
        (field, output_receipt[field], recomputed[field])
        for field in REQUIRED_RECEIPT_PROFILE_FIELDS
        if output_receipt[field] != recomputed[field]
    ]
    if mismatches:
        raise ValueError(
            f"{dataset}:{supplement_path} 重新計算的統計跟 receipt 記錄的不一致"
            f" (receipt 可能過期或被竄改,即使 SHA-256 相符也不能信任):{mismatches}")

    # Round 7 review:dedup 記錄的是投影前/合併時的去重過程,消費端手上只有最終
    # parquet,重算不出「投影前原始列有幾個重複鍵」這種資訊,不能走上面的數字
    # 重算比對,但結構本身要驗證——缺欄/型別錯/負數/非零衝突數都代表這份紀錄
    # 不可信,不能只因為它「存在」就當作檢查真的發生過。
    _validate_dedup_metadata(name, output_receipt["dedup"])

    dup = supp.duplicated(subset=["stock_id", "date"], keep=False)
    if dup.any():
        raise ValueError(f"{dataset}:{supplement_path} 有 {int(dup.sum())} 列重複"
                          f" (stock_id, date),不能消費")
    null_key = supp[["stock_id", "date"]].isna().any(axis=1)
    if null_key.any():
        raise ValueError(f"{dataset}:{supplement_path} 有 {int(null_key.sum())} 列"
                          f" stock_id/date 是 null,不能消費")
    return supp


def _check_sanity_floor(combined: pd.DataFrame, dataset: str, spec: dict) -> None:
    """合併結果的最低限度檢查:列數/檔數/日期範圍明顯低於已知規模,代表漏檔或欄位
    對應壞掉導致資料被默默腰斬,禁止繼續寫出 (fail-closed,不要讓縮水結果溜過去)。
    門檻是先前完整跑過一次、跟生產環境逐列比對過後的實測值打對折,只抓「明顯腰斬」
    這種粗粒度問題。

    ⚠ 這是次要防線,不是完整性證明:通過這個檢查只代表「沒有腰斬式的明顯縮水」,
    不代表資料逐股逐列完整無缺——那要靠 §3 全量比對 (§Round3 review 已明確要求
    不能拿這種粗篩門檻宣稱完整性,manifest 也一樣不能,見
    docs/資料快照遷移_DataExport0806.md §1)。"""
    min_rows = spec.get("min_rows")
    if min_rows and len(combined) < min_rows:
        raise ValueError(f"{dataset}:合併後只有 {len(combined)} 列,低於預期下限 {min_rows}"
                          f" (可能漏檔或欄位對應壞掉),禁止寫出")
    min_stocks = spec.get("min_stocks")
    if min_stocks and combined["stock_id"].nunique() < min_stocks:
        raise ValueError(f"{dataset}:合併後只有 {combined['stock_id'].nunique()} 檔股票,"
                          f"低於預期下限 {min_stocks},禁止寫出")
    expected_date_min = spec.get("expected_date_min")
    if expected_date_min and "date" in combined.columns:
        actual_min = combined["date"].min()
        if actual_min > expected_date_min:
            raise ValueError(f"{dataset}:合併後最早日期是 {actual_min},比預期下限"
                              f" {expected_date_min} 還晚 (可能漏了早期的檔案),禁止寫出")


def _assert_supplement_no_column_overwrite(combined: pd.DataFrame, supp: pd.DataFrame,
                                            key_cols=("stock_id", "date")) -> None:
    """§C.8「只能新增原生欄位、不能覆寫」的明確執行期斷言 (Phase A1 新增,取代
    原本只有結構上不重疊、沒有程式碼真的檢查過的狀態)。supplement 的非 key 欄位
    如果已經存在於 combined,left join 會靜默把它變成 `_x`/`_y` 後綴或整欄覆蓋,
    兩種都不是我們要的語意,先 fail-closed。"""
    overlap_cols = (set(supp.columns) - set(key_cols)) & set(combined.columns)
    if overlap_cols:
        raise ValueError(
            f"supplement 欄位 {sorted(overlap_cols)} 已經存在於合併後的 combined——"
            f"supplement 只能新增原生欄位,不能覆寫既有欄位,先去查是不是 SUPPLEMENT_SCHEMAS"
            f" 的欄位名稱跟某個 DATASETS 目標欄位撞名了。")


def _profile_supplement_merge(combined: pd.DataFrame, supp: pd.DataFrame,
                               key_cols=("stock_id", "date")) -> dict:
    """§C.10 supplement provenance receipt 需要的 non-overlap/覆蓋率統計,在
    importer 範圍內先算好、回傳給未來 builder 寫 receipt 用 (Phase A1 只算數字,
    不寫檔案;Round 2 review 第 8 項擴充:merge 前後列數、原生/supplement 欄位
    清單、完整的 overlap/uncovered/supplement-only key 列表,不只是計數)。
    `post_merge_row_count` 由呼叫端 (`load_source`) 在 merge 完成後補上,這支
    函式本身只在 merge 前算,不知道 merge 後的結果。"""
    key_cols = list(key_cols)
    combined_keys = set(map(tuple, combined[key_cols].itertuples(index=False, name=None)))
    supp_keys = set(map(tuple, supp[key_cols].itertuples(index=False, name=None)))
    overlap_keys = sorted(combined_keys & supp_keys)
    uncovered_keys = sorted(combined_keys - supp_keys)
    supplement_only_keys = sorted(supp_keys - combined_keys)
    return {
        "pre_merge_row_count": int(len(combined)),
        "post_merge_row_count": None,
        "native_columns": list(combined.columns),
        "supplement_columns": [c for c in supp.columns if c not in key_cols],
        "supplement_row_count": int(len(supp)),
        "overlap_key_count": len(overlap_keys),
        "overlap_keys": overlap_keys,
        "rows_supplement_key_not_covered": len(uncovered_keys),
        "uncovered_keys": uncovered_keys,
        "supplement_only_key_count": len(supplement_only_keys),
        "supplement_only_keys": supplement_only_keys,
    }


def _arrow_type_metadata(df: pd.DataFrame) -> dict:
    """Round 3 review 第 6 項:evidence schema 要能區分「邏輯契約型別」
    (`logical_types`,例如 `"float64"`)、「pandas 實際型別」(`actual_dtypes`,
    例如 `"float64"`/`"string"`)、跟「未來寫進 Parquet 會得到的 Arrow 型別」
    三種——三者不是同一件事 (pandas 的 nullable `"string"` dtype 在 Arrow 裡
    通常對應 `string`,pandas `float64` 對應 Arrow `double`,不是每個 pandas
    dtype 名稱都剛好等於 Arrow 型別名稱)。這裡用 `pyarrow.Schema.from_pandas`
    純粹在記憶體裡推算「如果現在把這個 DataFrame 寫成 Parquet,Arrow 會用
    什麼型別」,不實際寫任何檔案。"""
    schema = pa.Schema.from_pandas(df, preserve_index=False)
    return {field.name: str(field.type) for field in schema}


def _check_merged_schema(df: pd.DataFrame, spec: dict, supplement_columns: list, label: str) -> None:
    """Round 2 review 第 3 項:supplement merge **之後**的最終 schema 驗證——
    `_check_final_schema` 只驗證 `_load_one` 單檔的輸出 (沒有 supplement 欄位),
    這支函式驗證 `load_source` merge 完成後的 combined:欄位集合/順序必須是
    `standard_keys + sorted(native targets) + approved supplement columns`
    (supplement 欄位维持 `SUPPLEMENT_SCHEMAS` 凍結的順序接在最後,不重新排序),
    而且每個 supplement 數值欄位的實際 dtype 必須是 `float64`(拒絕
    缺少/多出/順序錯亂/型別錯誤的 supplement 輸出)。"""
    expected = ["stock_id", "stock_name", "date"] + _final_target_columns(spec) + list(supplement_columns)
    actual = list(df.columns)
    if actual != expected:
        missing = [c for c in expected if c not in actual]
        extra = [c for c in actual if c not in expected]
        raise ValueError(
            f"{label}:merge 後最終欄位跟預期公式 (standard_keys + sorted(native targets) +"
            f" supplement columns) 不符——缺少 {missing}、多出 {extra} (實際順序 {actual},"
            f"預期順序 {expected})。")
    for col in supplement_columns:
        actual_dtype = str(df[col].dtype)
        if actual_dtype != "float64":
            raise ValueError(
                f"{label}:supplement 欄位 {col!r} 的實際 dtype 是 {actual_dtype!r},不是"
                f" 'float64'——supplement 數值欄位必須顯式轉型,不能沿用 parquet 裡原本"
                f"可能因為剛好沒有 NaN 而被推斷成 int64 的型別。")


def _deep_freeze_json_safe(obj):
    """遞迴把 dict/list/tuple/str/int/float/bool/None 組成的結構凍結成完全
    唯讀的形式:dict → `MappingProxyType`(值也遞迴凍結),list/tuple → tuple
    (元素也遞迴凍結),純量原樣回傳 (str/int/float/bool/None 本身就不可變)。
    任何不在這個型別集合裡的值直接 raise——`duplicate_evidence` 只允許
    JSON 相容的內容,不該塞 DataFrame/Series/自訂物件這類東西。

    這支函式在建構的當下就把輸入**完全重建**成新的容器,不共用呼叫端傳進來
    的任何 dict/list 物件——之後呼叫端對原始輸入做任何 mutate,都不會反映
    到這裡凍結出來的結果 (Round 3 review 收尾第 1 項:constructor-input
    mutation 不能影響已保存的證據)。"""
    if isinstance(obj, dict):
        return MappingProxyType({(k if isinstance(k, str) else str(k)): _deep_freeze_json_safe(v)
                                  for k, v in obj.items()})
    if isinstance(obj, (list, tuple)):
        return tuple(_deep_freeze_json_safe(v) for v in obj)
    if isinstance(obj, (str, int, float, bool)) or obj is None:
        return obj
    raise TypeError(
        f"duplicate_evidence 含有無法安全凍結/序列化的型別 {type(obj).__name__}: {obj!r}"
        f" (只允許 dict/list/tuple/str/int/float/bool/None)")


def _thaw_to_json_safe_copy(obj):
    """`_deep_freeze_json_safe` 的反向操作,把凍結結構轉回一份全新的、普通
    可變、JSON 相容的 dict/list/純量結構——**每次呼叫都重新遞迴建構獨立的
    容器**,不共用任何既有的 dict/list 物件。tuple 決定性地轉成 list (符合
    JSON 陣列語意)。呼叫端拿到回傳值後怎麼 mutate,都不會影響
    `DuplicateProvenanceConflictError` 內部保存的凍結快照,也不會影響下一次
    呼叫的回傳結果 (Round 3 review 收尾第 2 項)。"""
    if isinstance(obj, MappingProxyType):
        return {k: _thaw_to_json_safe_copy(v) for k, v in obj.items()}
    if isinstance(obj, tuple):
        return [_thaw_to_json_safe_copy(v) for v in obj]
    return obj


class DuplicateProvenanceConflictError(ValueError):
    """Round 3 review 第 4 項:provenance-aware 去重判定 (`_adjudicate_and_
    retain_with_provenance`,只有 `return_evidence=True` 會用到) 發現同一個
    (key, target_column) 底下,多筆來源列的 provenance 不一致時丟出——不是
    只比較最終值 (兩列都是 NaN,但一列是原生空白、一列是文字「.」,provenance
    不同,一樣算衝突)。

    繼承 `ValueError` 是刻意的:任何原本只 `except ValueError` (或完全不處理
    例外) 的既有呼叫端,行為不變。跟舊版差別在多帶了完整的衝突證據——舊版的
    衝突證據只存在區域變數裡,`raise` 之後呼叫端完全看不到。

    Round 3 review 收尾修正的 immutable/serializable 契約 (第一版
    `MappingProxyType(dict(duplicate_evidence))` 只做了淺層包裝,
    `json.dumps` 會因為 mappingproxy 本身跟巢狀 tuple 失敗,巢狀的
    list/dict 也還是可以被 mutate,兩個要求都沒有真的做到):

      · 建構子當下用 `_deep_freeze_json_safe` 把證據**完全重建**成內部私有
        (name-mangled) 的深層唯讀快照 (`MappingProxyType` + `tuple`
        遞迴組成),跟傳進來的原始物件完全脫鉤。
      · `duplicate_evidence` 這個 property **每次存取都重新呼叫
        `_thaw_to_json_safe_copy`**,回傳一份全新、普通可變、只由
        dict/list/str/int/float/bool/None 組成的深拷貝——`json.dumps(exc.
        duplicate_evidence, sort_keys=True)` 可以直接成功;拿到的結果
        怎麼 mutate,都不會影響下一次存取的結果,也不會影響例外物件內部
        保存的快照。
      · `to_dict()` 是 `duplicate_evidence` 的正式別名,語意完全相同,
        供偏好方法呼叫風格的呼叫端使用。"""

    def __init__(self, message: str, duplicate_evidence: dict):
        super().__init__(message)
        self.__frozen_evidence = _deep_freeze_json_safe(duplicate_evidence)

    @property
    def duplicate_evidence(self):
        return _thaw_to_json_safe_copy(self.__frozen_evidence)

    def to_dict(self):
        return self.duplicate_evidence


def _adjudicate_and_retain_with_provenance(*, dataset: str, spec: dict, file_row_records: list,
                                            native_targets: list, is_static: bool):
    """Round 3 review 第 1/2/3 項的核心,只有 `load_source(..., return_evidence=
    True)` 會呼叫。

    `file_row_records`:`[(file_index, row_record), ...]`。`file_index` 是
    `_source_files()` 排序後的檔案順序 (第 2 項凍結的 explicit source order
    第一層);`row_record` 是 `_load_one(..., collect_evidence=True)` 回傳的
    `row_records` 其中一筆 (`key`/`locator`/`values`/`provenance`)。**呼叫端
    必須依照檔案順序、檔案內原始列順序 append**——這支函式不重新排序輸入,
    只依賴呼叫端已經給的順序 + `locator["source_row_number"]` (第二層) 組出
    `(file_index, source_row_number)` 這個唯一決定「誰是最後一筆」的顯式
    tuple,不依賴 `pandas.sort_values` 的穩定排序這種間接、容易被之後的重構
    悄悄破壞的機制。

    回傳 `(retained_rows, duplicate_mapping_entries)`:
      · `retained_rows`——`{key: row_record}`,依 explicit source order 選出
        的最後一筆列,交給 `_assemble_combined_from_retained_rows` 組出最終
        DataFrame,也是 `_build_evidence_bundle` 算 `final_null_causes` 的
        依據。
      · `duplicate_mapping_entries`——只針對「這個 key 底下真的有多筆列」的
        (key, target_column) 組合各一個 dict:`key`/`target_column`/
        `retained_source_locator`/`contributing_source_locators` (frozen
        source order)/`sidecar_dedup_keys` (只有 blank/unparseable 那幾筆
        才有,`dedup_key_v1` 現算,不用另外儲存)/`source_row_count`/
        `removed_row_count`。

    Fail-closed (第 1 項):任一 (key, target_column) 的多筆列 provenance
    tuple 不完全一致就是衝突——**全部 key/target 都判定完才一次性 raise**
    `DuplicateProvenanceConflictError`(第 4 項),不會漏掉其他衝突,也不會
    在還沒判定完所有組合前就先回傳一個「部分正確」的 retained_rows。"""
    rows_by_key = {}
    for file_index, rr in file_row_records:
        key = rr["key"]
        source_order = (file_index, rr["locator"]["source_row_number"])
        rows_by_key.setdefault(key, []).append((source_order, rr))

    conflicts = []
    retained_rows = {}
    duplicate_mapping_entries = []

    for key, entries in rows_by_key.items():
        # 第 2 項:explicit source order 排序(不是 pandas sort_values 的穩定性),
        # 「最後一筆」就是這個排序下的最後一個元素。
        entries_sorted = sorted(entries, key=lambda e: e[0])
        retained_rows[key] = entries_sorted[-1][1]

        for target in native_targets:
            group = [(order, rr["provenance"][target], rr) for order, rr in entries_sorted]
            distinct = sorted({g[1] for g in group})
            n = len(group)
            if n > 1 and len(distinct) > 1:
                conflicts.append({
                    "key": key, "target_column": target,
                    "entries": [{"source_order": order, "locator": dict(rr["locator"]),
                                 "provenance": prov} for order, prov, rr in group],
                    "distinct_provenances": list(distinct),
                })
                continue
            if n > 1:
                contributing_locators = [dict(rr["locator"]) for _, _, rr in group]
                sidecar_dedup_keys = []
                for _, prov, rr in group:
                    if prov[0] in (NULL_CAUSE_RETAINED_BLANK, NULL_CAUSE_RETAINED_UNPARSEABLE):
                        loc = rr["locator"]
                        sidecar_dedup_keys.append(_dedup_key_v1(
                            dataset, loc["source_relpath"], loc["source_container_member"],
                            loc["source_row_number"], target))
                duplicate_mapping_entries.append({
                    "key": key, "target_column": target,
                    "retained_source_locator": dict(entries_sorted[-1][1]["locator"]),
                    "contributing_source_locators": contributing_locators,
                    "sidecar_dedup_keys": sidecar_dedup_keys,
                    "source_row_count": n, "removed_row_count": n - 1,
                })

    if conflicts:
        raise DuplicateProvenanceConflictError(
            f"{dataset}:{len(conflicts)} 個 (key, target_column) 組合的來源 provenance"
            f" 不一致——最終值可能剛好相同 (例如都是 NaN),但語意不同 (原生空白 vs"
            f" 特定無法解析文字 vs 欄位整檔缺席,彼此都不算同一件事),不能當作安全的"
            f"完全重複去重,fail-closed (Round 3 review 第 1 項)。",
            {"dataset": dataset, "conflicts": conflicts})

    return retained_rows, duplicate_mapping_entries


def _assemble_combined_from_retained_rows(spec: dict, retained_rows: dict,
                                           is_static: bool) -> pd.DataFrame:
    """把 `_adjudicate_and_retain_with_provenance` 選出的 `retained_rows` 組成
    最終 `combined` DataFrame。Round 3 review 第 2 項:「排序只能發生在保留列
    選出來之後」——這裡最後才對 key 排序,不會讓排序本身影響選中哪一列 (那件
    事已經在 `_adjudicate_and_retain_with_provenance` 的 explicit source order
    裡做完了)。逐欄顯式指定 dtype (跟 `_load_one`/`_check_final_schema` 同一套
    §C.1 契約),不依賴從 list-of-dict 建構 DataFrame 時的型別推斷。"""
    native_targets = list(spec["rename"].values()) if is_static else _final_target_columns(spec)
    numeric_targets = _numeric_target_columns(spec)
    sort_key = (lambda k: k[0]) if is_static else (lambda k: (k[0], k[1]))
    sorted_keys = sorted(retained_rows.keys(), key=sort_key)

    columns_data = {"stock_id": [], "stock_name": []}
    if not is_static:
        columns_data["date"] = []
    for t in native_targets:
        columns_data[t] = []
    for key in sorted_keys:
        values = retained_rows[key]["values"]
        columns_data["stock_id"].append(values["stock_id"])
        columns_data["stock_name"].append(values["stock_name"])
        if not is_static:
            columns_data["date"].append(values["date"])
        for t in native_targets:
            columns_data[t].append(values[t])

    combined = pd.DataFrame(columns_data)
    combined["stock_id"] = combined["stock_id"].astype("string")
    combined["stock_name"] = combined["stock_name"].astype("string")
    if not is_static:
        combined["date"] = combined["date"].astype("string")
    for t in native_targets:
        combined[t] = combined[t].astype("float64" if t in numeric_targets else "string")
    return combined


def _build_evidence_bundle(*, dataset: str, spec: dict, combined: pd.DataFrame,
                            file_evidences: list, retained_rows: dict,
                            duplicate_mapping_entries: list, supplement_columns: list,
                            merge_profile: dict, native_targets: list, is_static: bool) -> dict:
    """把逐檔證據 (`file_evidences`) 跟 provenance-aware 去重的結果
    (`retained_rows`/`duplicate_mapping_entries`,來自
    `_adjudicate_and_retain_with_provenance`) 彙總成 `load_source(dataset,
    return_evidence=True)` 回傳的凍結 evidence bundle。**不依賴 concat 後的
    `DataFrame.attrs`**——所有輸入都是 concat 前逐檔收集、明確傳進來的
    Python 物件。"""
    cell_records = []
    per_file_stage_one_counts = []
    coverage_matrix = []

    for fe in file_evidences:
        cell_records.extend(fe["cell_records"])
        per_file_stage_one_counts.append({
            "source_relpath": fe["source_relpath"], "source_file_sha256": fe["source_file_sha256"],
            "source_container_member": fe["source_container_member"], "counts": fe["stage_one_counts"],
        })
        for target, status in fe["coverage_row"].items():
            coverage_matrix.append({"source_relpath": fe["source_relpath"],
                                     "target_column": target, "status": status})
        for col in supplement_columns:
            coverage_matrix.append({"source_relpath": fe["source_relpath"],
                                     "target_column": col, "status": "NOT_APPLICABLE"})

    # duplicate_mapping (Round 3 review 第 3 項):`_adjudicate_and_retain_with_
    # provenance` 已經算出「一筆 entry 對應一個 (key, target_column)」的完整
    # target-level mapping,這裡原樣傳出去,不再做任何 key-level 的攤平。
    duplicate_mapping = duplicate_mapping_entries

    # final_null_causes (Round 2 review 第 6 項,Round 3 review 改成直接讀
    # `retained_rows` 的 provenance——這是 `_adjudicate_and_retain_with_
    # provenance` 已經用 explicit source order 選出的保留列,不需要再另外
    # 用字典覆蓋去推導,原因跟保留哪一列是同一份判定的結果)。
    final_null_causes = {}
    for target in native_targets:
        counts = {c: 0 for c in _FINAL_NULL_CAUSES}
        for row in retained_rows.values():
            tag = row["provenance"][target][0]
            if tag == "PARSED":
                continue
            if tag in counts:
                counts[tag] += 1
            else:
                counts[NULL_CAUSE_OTHER_UNEXPLAINED] += 1
        final_null_causes[target] = counts
        if counts[NULL_CAUSE_OTHER_UNEXPLAINED]:
            raise RuntimeError(
                f"{dataset}:目標欄位 {target!r} 有 {counts[NULL_CAUSE_OTHER_UNEXPLAINED]} 個"
                f" 最終 null 對不上任何已知原因 (stage-two accounting 有漏洞),fail-closed")

    supp_keys = None
    if merge_profile is not None:
        supp_keys = set(merge_profile.get("overlap_keys", [])) | set(
            merge_profile.get("supplement_only_keys", []))
    for col in supplement_columns:
        counts = {c: 0 for c in _FINAL_NULL_CAUSES}
        null_mask = combined[col].isna()
        for sid, dt in zip(combined.loc[null_mask, "stock_id"], combined.loc[null_mask, "date"]):
            if supp_keys is not None and (sid, dt) not in supp_keys:
                counts[NULL_CAUSE_SUPPLEMENT_KEY_NOT_COVERED] += 1
            else:
                counts[NULL_CAUSE_OTHER_UNEXPLAINED] += 1
        final_null_causes[col] = counts
        if counts[NULL_CAUSE_OTHER_UNEXPLAINED]:
            raise RuntimeError(
                f"{dataset}:supplement 欄位 {col!r} 有 {counts[NULL_CAUSE_OTHER_UNEXPLAINED]} 個"
                f" 最終 null 對不上任何已知原因 (SUPPLEMENT_KEY_NOT_COVERED 以外的 null,"
                f" stage-two accounting 有漏洞),fail-closed")

    numeric_targets = _numeric_target_columns(spec) | set(supplement_columns)
    schema_metadata = {
        "logical_types": {c: ("float64" if c in numeric_targets else "string")
                          for c in combined.columns},
        "actual_dtypes": {c: str(combined[c].dtype) for c in combined.columns},
        "arrow_types": _arrow_type_metadata(combined),
    }

    return {
        "cell_records": cell_records,
        "per_file_stage_one_counts": per_file_stage_one_counts,
        "coverage_matrix": coverage_matrix,
        "duplicate_mapping": duplicate_mapping,
        "final_null_causes": final_null_causes,
        "supplement_merge_profile": merge_profile,
        "schema": schema_metadata,
    }


def _load_source_fast_path(dataset: str, spec: dict, files: list, is_static: bool,
                            native_targets: list) -> pd.DataFrame:
    """`return_evidence=False` 的舊版行為,逐字保留(Round 2 之前就有,Round 3
    review 沒有要求修改這條路徑,只要求 `return_evidence=True` 的正式路徑
    provenance-aware——見 module 層級 `load_source` docstring 的說明)。
    值比較是 NaN-blind 的粗篩 (`_check_duplicate_key_conflicts`),`keep="last"`
    依賴 `sort_values` 的穩定排序,不追蹤 provenance。**不得被未來的正式
    builder 拿來當作唯一的去重依據**——正式建置必須呼叫
    `return_evidence=True` 才會走 `_adjudicate_and_retain_with_provenance`。"""
    frames = []
    per_file_coverage_rows = []
    for f in files:
        logger.info(f"讀取 {f.relative_to(DATA_ROOT)}")
        one = _load_one(f, spec)
        per_file_coverage_rows.append(one.attrs.get("coverage_row", {}))
        frames.append(one)

    _raise_if_target_absent_from_all_files(dataset, files, native_targets, per_file_coverage_rows)

    frames = [x for x in frames if not x.empty]
    if not frames:
        raise FileNotFoundError(f"{spec['source_dir']} 底下沒有能對應出必要欄位的檔案")
    combined = pd.concat(frames, ignore_index=True)

    if is_static:
        _check_duplicate_key_conflicts(combined, dataset, key_cols=["stock_id"])
        combined = (combined.drop_duplicates(subset=["stock_id"], keep="last")
                            .sort_values("stock_id").reset_index(drop=True))
        _check_sanity_floor(combined, dataset, spec)
        return combined

    _check_duplicate_key_conflicts(combined, dataset)
    combined = combined.sort_values(["stock_id", "date"])
    combined = combined.drop_duplicates(subset=["stock_id", "date"], keep="last")
    combined = combined.reset_index(drop=True)

    supplement_columns = []
    supplement_path = spec.get("supplement")
    if supplement_path is not None:
        supp = _verify_supplement(supplement_path, dataset)
        _assert_supplement_no_column_overwrite(combined, supp)
        supplement_columns = [c for c in supp.columns if c not in ("stock_id", "date")]
        merge_profile = _profile_supplement_merge(combined, supp)
        pre_merge_rows = len(combined)
        combined = combined.merge(supp, on=["stock_id", "date"], how="left")
        if len(combined) != pre_merge_rows:
            raise RuntimeError(
                f"{dataset}:merge 後列數從 {pre_merge_rows} 變成 {len(combined)} ——"
                f" supplement 造成了 row 膨脹 (fan-out),代表它裡面有重複鍵混進來。"
                f" _verify_supplement 已經檢查過 supplement 本身的重複鍵,這裡是合併後的"
                f" 第二層防線,兩層都要過。")
        for col in supplement_columns:
            combined[col] = combined[col].astype("float64")
        _check_duplicate_key_conflicts(combined, dataset)   # merge 後再查一次,防呆用
        merge_profile["post_merge_row_count"] = int(len(combined))
        combined.attrs["supplement_merge_profile"] = merge_profile

    _check_merged_schema(combined, spec, supplement_columns, dataset)
    _check_sanity_floor(combined, dataset, spec)
    return combined


def _raise_if_target_absent_from_all_files(dataset: str, files: list, native_targets: list,
                                            per_file_coverage_rows: list) -> None:
    """§B 第 17 輪 + Round 2 review 第 2 項:all-files-absent 的 fail-closed
    判定必須從 coverage 資料本身推導,不是另一套獨立維護的『缺欄集合』邏輯。"""
    if not per_file_coverage_rows:
        return
    for target in native_targets:
        statuses = [row.get(target) for row in per_file_coverage_rows]
        if statuses and all(s == NULL_CAUSE_SOURCE_COLUMN_ABSENT for s in statuses):
            raise ValueError(
                f"{dataset}:目標欄位 {target!r} 在這個 dataset 讀到的全部 {len(files)}"
                f" 個原始檔裡,coverage 狀態通通是 {NULL_CAUSE_SOURCE_COLUMN_ABSENT}——"
                f"不能被靜默接受成『全部都是 SOURCE_COLUMN_ABSENT 的 null』,先去確認"
                f" TEJ 是不是換了欄位命名,或這條 DATASETS 的 rename 映射該刪除"
                f" (§B 第 17 輪,Round 2 review 改成從 coverage 矩陣推導)。")


def _load_source_with_evidence(dataset: str, spec: dict, files: list, is_static: bool,
                                native_targets: list):
    """`return_evidence=True` 的 provenance-aware 正式路徑 (Round 3 review 第
    1/2/3/4 項)。跟 `_load_source_fast_path` 的關鍵差異:去重判定不看
    「最終值是不是一樣」,看「pre-dedup 的來源 provenance 是不是一樣」;
    retained row 的選擇用 explicit source order (檔案順序 + 實體列順序),
    不依賴 `pandas.sort_values` 的穩定排序;任何語意衝突在選出任何一筆保留
    列**之前**就先蒐集完、一次 raise `DuplicateProvenanceConflictError`。"""
    file_row_records = []
    file_evidences = []
    per_file_coverage_rows = []
    for file_index, f in enumerate(files):
        logger.info(f"讀取 {f.relative_to(DATA_ROOT)}")
        one, fe = _load_one(f, spec, dataset=dataset, collect_evidence=True)
        file_evidences.append(fe)
        per_file_coverage_rows.append(one.attrs.get("coverage_row", {}))
        for rr in fe["row_records"]:
            file_row_records.append((file_index, rr))

    _raise_if_target_absent_from_all_files(dataset, files, native_targets, per_file_coverage_rows)

    if not file_row_records:
        raise FileNotFoundError(f"{spec['source_dir']} 底下沒有能對應出必要欄位的檔案")

    retained_rows, duplicate_mapping_entries = _adjudicate_and_retain_with_provenance(
        dataset=dataset, spec=spec, file_row_records=file_row_records,
        native_targets=native_targets, is_static=is_static)
    combined = _assemble_combined_from_retained_rows(spec, retained_rows, is_static)

    if is_static:
        _check_final_schema(combined, spec, dataset)
        _check_sanity_floor(combined, dataset, spec)
        evidence_bundle = _build_evidence_bundle(
            dataset=dataset, spec=spec, combined=combined, file_evidences=file_evidences,
            retained_rows=retained_rows, duplicate_mapping_entries=duplicate_mapping_entries,
            supplement_columns=[], merge_profile=None, native_targets=native_targets, is_static=True)
        return combined, evidence_bundle

    supplement_columns = []
    merge_profile = None
    supplement_path = spec.get("supplement")
    if supplement_path is not None:
        supp = _verify_supplement(supplement_path, dataset)
        _assert_supplement_no_column_overwrite(combined, supp)
        supplement_columns = [c for c in supp.columns if c not in ("stock_id", "date")]
        merge_profile = _profile_supplement_merge(combined, supp)
        pre_merge_rows = len(combined)
        combined = combined.merge(supp, on=["stock_id", "date"], how="left")
        if len(combined) != pre_merge_rows:
            raise RuntimeError(
                f"{dataset}:merge 後列數從 {pre_merge_rows} 變成 {len(combined)} ——"
                f" supplement 造成了 row 膨脹 (fan-out),代表它裡面有重複鍵混進來。"
                f" _verify_supplement 已經檢查過 supplement 本身的重複鍵,這裡是合併後的"
                f" 第二層防線,兩層都要過。")
        for col in supplement_columns:
            combined[col] = combined[col].astype("float64")
        _check_duplicate_key_conflicts(combined, dataset)   # merge 後結構性防呆 (key 已保證唯一)
        merge_profile["post_merge_row_count"] = int(len(combined))
        combined.attrs["supplement_merge_profile"] = merge_profile

    _check_merged_schema(combined, spec, supplement_columns, dataset)
    _check_sanity_floor(combined, dataset, spec)

    evidence_bundle = _build_evidence_bundle(
        dataset=dataset, spec=spec, combined=combined, file_evidences=file_evidences,
        retained_rows=retained_rows, duplicate_mapping_entries=duplicate_mapping_entries,
        supplement_columns=supplement_columns, merge_profile=merge_profile,
        native_targets=native_targets, is_static=False)
    return combined, evidence_bundle


def load_source(dataset: str, *, return_evidence: bool = False):
    """讀某個 dataset 的全部原始檔、合併、去重、(若有) 合併 supplement。

    `return_evidence=False`(預設)時只回傳 `combined: DataFrame`,行為/回傳
    型別跟舊版完全相同,走 `_load_source_fast_path`(NaN-blind 的粗篩去重,
    值恰好相同就當作安全重複,不追蹤 provenance)。

    `return_evidence=True` 時走 `_load_source_with_evidence`,回傳
    `(combined, evidence_bundle)`。這條路徑的去重判定改成 provenance-aware
    (Round 3 review 第 1 項):兩個來源列的最終值恰好相同 (例如都是 NaN)
    不代表可以安全去重,還要 pre-dedup 的來源語意 (blank/unparseable 的
    原始 token/欄位整檔缺席) 也相同。任何語意衝突會丟出
    `DuplicateProvenanceConflictError`(`ValueError` 的子類別,向後相容既有
    只 `except ValueError` 的呼叫端),並帶著完整衝突證據的
    `.duplicate_evidence` 屬性。

    `evidence_bundle` 是凍結 schema 的 dict (見 `_build_evidence_bundle`
    docstring),鍵至少含 `cell_records`/`per_file_stage_one_counts`/
    `coverage_matrix`/`duplicate_mapping`/`final_null_causes`/
    `supplement_merge_profile`/`schema`。完全用 concat 前逐檔收集的 Python
    物件明確組出來,不依賴 `DataFrame.attrs` 撐過 `pd.concat`。

    **未來的正式 Phase B builder 必須使用 `return_evidence=True`**——快速
    路徑刻意保留 NaN-blind 的弱去重定義只是為了向後相容既有非正式呼叫端
    (Round 3 review 第 1 項最後一句的要求),不是給正式建置用的。"""
    spec = DATASETS[dataset]
    files = _source_files(spec)
    if not files:
        raise FileNotFoundError(f"{spec['source_dir']} 底下沒有找到任何 .xlsx/.zip 檔案")
    _manifest_preflight(files, spec, dataset)

    is_static = bool(spec.get("static"))
    native_targets = list(spec["rename"].values()) if is_static else _final_target_columns(spec)

    if not return_evidence:
        return _load_source_fast_path(dataset, spec, files, is_static, native_targets)
    return _load_source_with_evidence(dataset, spec, files, is_static, native_targets)


def _recover_or_clear_stale_backup(out_dir: Path, backup_dir: Path, label: str) -> None:
    """呼叫發布/commit 流程最開頭要先做的檢查 (Round 5 review)。

    正常情況下 backup_dir 只會在「階段二 commit 進行中」短暫存在,commit 成功後
    階段三就會清掉它。但如果上一次執行在 `out_dir.rename(backup_dir)` 之後、
    `staging_dir.rename(out_dir)` 完成之前被強制中斷 (例如程序被 kill,不是走
    except 分支的那種失敗——那種失敗函式內部已經會自動還原),下次呼叫時會看到
    「backup_dir 存在、out_dir 不存在」這個殘留狀態:backup_dir 此時是**唯一僅存**
    的舊資料,絕對不能盲目刪掉。

    規則:
      · backup_dir 不存在 → 沒有殘留,直接返回。
      · backup_dir 存在但 out_dir 不存在 → 唯一僅存的舊資料,先還原成 out_dir
        (rename 本身失敗就讓例外往外傳,不吞掉,backup_dir 保持原狀,不遺失資料;
        呼叫端可以之後再重試一次)。
      · 兩者都存在 → 這才是「真的清得掉」的殘留 backup (out_dir 已經是有效資料),
        安全地整個刪掉。
    """
    if not backup_dir.exists():
        return
    if not out_dir.exists():
        backup_dir.rename(out_dir)
        return
    shutil.rmtree(backup_dir)


def save_by_stock(df: pd.DataFrame, dataset: str, cache_dir: Path = TEJ_CACHE_DIR) -> int:
    """把整份 dataset 依 stock_id 拆檔寫入 cache_dir/<dataset>/。

    分三個明確階段,commit point 精確定義在「階段二」的 rename 成功那一刻
    (Round 4 review 修正——原本 Round 3 版本的 except 涵蓋了階段三的清理動作,
    導致清理失敗會讓函式 raise,即使資料其實已經正確發布了):

      階段一 (staging 寫入 + 驗證,commit point 之前):
        全部寫進 cache_dir/.<dataset>.staging/;驗證檔案數=股票數、
        每個檔案的 schema 跟輸入一致、(stock_id, date) 唯一且非 null、
        總列數=輸入列數。任何一項不符,staging 直接刪掉,out_dir 完全不動。

      階段二 (commit,atomic 目錄互換):
        `staging_dir.rename(out_dir)` 執行成功 = 已發布。這之前 (含 rename 本身
        失敗) 任何例外,out_dir 會被還原成呼叫前的內容。因為是整個目錄互換,
        這次資料裡已經不存在的舊股票 (例如下市股被排除) 不會變成殘留在 out_dir
        裡的 orphan parquet——舊版逐檔覆寫的寫法沒有清過這種 orphan。

      階段三 (commit 之後的收尾,不影響「有沒有發布成功」的判定):
        清理 backup_dir。這一步失敗只印警告,不 raise——commit 已經成功,
        資料已經生效,不能讓呼叫端誤以為整個操作失敗了。

    只在真的呼叫這支函式時才會動到 cache_dir;測試一律用 tmp_path,不動真正的
    生產 cache (見 tests/test_tej_data_migration.py)。

    函式最開頭會先跑 `_recover_or_clear_stale_backup`(Round 5 review):處理
    上一次執行在階段二 commit 中途被強制中斷、backup_dir 是唯一僅存舊資料的
    殘留情況,還原它而不是盲目刪除。
    """
    out_dir = cache_dir / dataset
    staging_dir = cache_dir / f".{dataset}.staging"
    backup_dir = cache_dir / f".{dataset}.rollback"
    cache_dir.mkdir(parents=True, exist_ok=True)

    # Round 5 review:先處理上次執行可能留下的殘留 backup_dir,見
    # _recover_or_clear_stale_backup 的docstring。這一步之後,backup_dir
    # 保證不存在 (要嘛本來就沒有,要嘛剛被還原成 out_dir,要嘛剛被安全清掉)。
    _recover_or_clear_stale_backup(out_dir, backup_dir, dataset)

    if staging_dir.exists():        # 上次失敗留下的殘留,清掉重來 (不動 out_dir)
        shutil.rmtree(staging_dir)
    staging_dir.mkdir(parents=True)

    # ---- 階段一:寫 staging + 驗證 schema/唯一鍵/列數,失敗的話 out_dir 不受影響 ----
    try:
        n = 0
        expected_cols = list(df.columns)
        for stock_id, g in df.groupby("stock_id"):
            p = staging_dir / f"{stock_id}.parquet"
            g.sort_values("date").reset_index(drop=True).to_parquet(p, index=False)
            n += 1

        expected_stocks = df["stock_id"].nunique()
        staged_files = list(staging_dir.glob("*.parquet"))
        if len(staged_files) != expected_stocks:
            raise RuntimeError(
                f"{dataset}:staging 寫出 {len(staged_files)} 檔,預期 {expected_stocks} 檔"
                f" (股票數對不上,不發布)")

        total_written = 0
        for p in staged_files:
            g = pd.read_parquet(p)
            if list(g.columns) != expected_cols:
                raise RuntimeError(f"{dataset}:{p.name} 的 schema {list(g.columns)} 跟"
                                    f" 輸入 {expected_cols} 不符,不發布")
            dup = g.duplicated(subset=["stock_id", "date"], keep=False)
            if dup.any():
                raise RuntimeError(f"{dataset}:{p.name} 有 {int(dup.sum())} 列重複"
                                    f" (stock_id, date),不發布")
            if g[["stock_id", "date"]].isna().any().any():
                raise RuntimeError(f"{dataset}:{p.name} 有 null 的 stock_id/date,不發布")
            total_written += len(g)
        if total_written != len(df):
            raise RuntimeError(
                f"{dataset}:staging 總列數 {total_written},跟輸入 {len(df)} 對不上"
                f" (不發布)")
    except Exception:
        if staging_dir.exists():
            shutil.rmtree(staging_dir)
        raise

    # ---- 階段二:commit (atomic 目錄互換)。commit point = staging_dir.rename(out_dir) ----
    # backup_dir 此刻保證不存在 (函式最開頭的 _recover_or_clear_stale_backup 已處理過,
    # 階段一過程中沒有其他程式碼會動到 backup_dir),不需要再盲目清一次。
    moved_old = False
    if out_dir.exists():
        out_dir.rename(backup_dir)
        moved_old = True
    try:
        staging_dir.rename(out_dir)
    except Exception:
        if moved_old and backup_dir.exists() and not out_dir.exists():
            backup_dir.rename(out_dir)     # commit 失敗,把舊版還原回去
        raise

    # ---- 階段三:commit 之後的收尾,失敗不代表整體失敗 ----
    if backup_dir.exists():
        try:
            shutil.rmtree(backup_dir)
        except Exception as e:
            logger.warning(f"{dataset}:發布成功,但清理舊備份 {backup_dir} 失敗:{e}"
                            f" (不影響 {out_dir} 內容,下次執行會自動清掉)")
    return n


def main():
    parser = argparse.ArgumentParser(description="TEJ 全市場歷史批次匯入 (讀 tej_exports/DataExport0806)")
    parser.add_argument("--dataset", choices=list(DATASETS), default="price_valuation")
    parser.add_argument("--cache-dir", default=str(TEJ_CACHE_DIR), help="輸出 Parquet 根目錄")
    args = parser.parse_args()

    df = load_source(args.dataset)
    if DATASETS[args.dataset].get("static"):
        out = Path(args.cache_dir) / f"{args.dataset}.parquet"
        out.parent.mkdir(parents=True, exist_ok=True)
        tmp = out.with_suffix(".parquet.tmp")
        df.to_parquet(tmp, index=False)
        os.replace(tmp, out)      # atomic:寫壞的話 out 保持原樣,不會有半份檔案
        logger.info(f"靜態對照表共 {len(df)} 檔,已寫入 {out}")
        return
    logger.info(f"合併後共 {len(df)} 列,{df['stock_id'].nunique()} 檔,"
                f"日期範圍 {df['date'].min()} ~ {df['date'].max()}")
    n = save_by_stock(df, args.dataset, Path(args.cache_dir))
    logger.info(f"已寫入 {n} 檔股票的 Parquet 至 {args.cache_dir}/{args.dataset}/")


if __name__ == "__main__":
    main()
