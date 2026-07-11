"""
回測執行器 — 支援自行輸入股票代號、參數網格搜尋、評級變動追蹤、路線一驗證
--------------------------------------------------------------------
用法:
  互動輸入:      python tests\run_backtest.py
  命令列指定:    python tests\run_backtest.py 2330 2454 3481
  路線一驗證:    python tests\run_backtest.py --validate      (省略代號=用分散化池)
  參數最佳化:    python tests\run_backtest.py --optimize
  評級變動追蹤:  python tests\run_backtest.py --track
  多週期穩健性:  python tests\run_backtest.py --cycle          (2022 空頭 vs 2023–2025 多頭排序力)
  分項因子歸因:  python tests\run_backtest.py --attribution    (五維度對排序力的邊際貢獻)
"""
import sys
import os
import argparse

project_root = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
if project_root not in sys.path:
    sys.path.insert(0, project_root)

import logging
from core.backtest import Backtester
from core.data_provider import DataProvider

logging.basicConfig(level=logging.WARNING)

# ============================================================================
# 分散化測試池 (~45 檔):刻意跨產業、跨市值、跨景氣循環、彼此低相關,
# 才是對系統鑑別力的公平檢驗 (避免全是高度連動的權值股)。
#   · 分位數多空 (前⅓−後⅓) 需要足夠檔數:12 檔 → 每邊僅 4 檔,價差噪音大;
#     ~45 檔 → 每邊約 15 檔,排序力量測才穩健。
#   · 刻意納入『金融股』(觸發 is_financial 現金流豁免) 與『中小型成長股』(觸發投信吸籌比),
#     讓 PIT 資料層修補真正被檢驗到。
#   · 全部選 2019 前已上市者,確保 2021–2022 空頭段也有完整歷史 (避免晚上市造成缺資料)。
# ============================================================================
DIVERSIFIED_POOL = [
    # 半導體 — 大型權值 / IC 設計成長
    "2330", "2454", "2303", "2379", "3034", "3443", "3661",  # 台積電 聯發科 聯電 瑞昱 聯詠 創意 世芯-KY
    # 半導體 — 記憶體 / 二線 (景氣循環、主力洗籌碼強)
    "2408", "2344", "2337",                                  # 南亞科 華邦電 旺宏
    # 電子代工 / 組裝
    "2317", "2382", "2357", "4938", "3231", "2356",          # 鴻海 廣達 華碩 和碩 緯創 英業達
    # 光學 / 散熱 / 電源 / 被動元件
    "3008", "2308", "3017", "3324", "2327",                  # 大立光 台達電 奇鋐 雙鴻 國巨
    # 面板 (高波動循環)
    "2409", "3481",                                          # 友達 群創
    # 金融 (觸發 is_financial:銀行/保險現金流豁免)
    "2881", "2882", "2891", "2886", "2884", "2892",          # 富邦金 國泰金 中信金 兆豐金 玉山金 第一金
    # 塑化 / 鋼鐵 / 水泥 (景氣循環,財報落後景氣)
    "1301", "1303", "2002", "1101",                          # 台塑 南亞 中鋼 台泥
    # 航運 (高波動循環)
    "2603", "2609", "2615",                                  # 長榮 陽明 萬海
    # 電信 / 防禦
    "2412", "3045",                                          # 中華電 台灣大
    # 內需 / 食品 / 零售
    "1216", "2912",                                          # 統一 統一超
    # 紡織 / 製鞋 (外銷景氣)
    "1476", "9910",                                          # 儒鴻 豐泰
    # 機械 / 自動化 (中小型成長)
    "2049", "1590",                                          # 上銀 亞德客
    # 生技 (低相關、獨立題材)
    "6446",                                                  # 藥華藥
]

DEFAULT_SYMBOLS = DIVERSIFIED_POOL


def parse_symbols(raw: str):
    if not raw:
        return []
    for sep in ("，", ",", "、", ";", "；"):
        raw = raw.replace(sep, " ")
    out = []
    for tok in raw.split():
        tok = tok.strip().upper()
        if tok and tok not in out:
            out.append(tok)
    return out


def fetch_names(symbols):
    try:
        DataProvider._ensure_login()
        info = DataProvider._api.get_data(dataset="TaiwanStockInfo")
        if info is None or info.empty:
            return {}
        return {str(r.get("stock_id", "")).strip(): str(r.get("stock_name", "")).strip()
                for _, r in info.iterrows()
                if str(r.get("stock_id", "")).strip() in symbols and r.get("stock_name")}
    except Exception:
        return {}


MODE_ALIAS = {"c": "conservative", "b": "balanced", "a": "aggressive",
              "保守": "conservative", "平衡": "balanced", "積極": "aggressive"}


def normalize_mode(raw, default="balanced"):
    if not raw:
        return default
    r = str(raw).strip().lower()
    if r in ("conservative", "balanced", "aggressive"):
        return r
    return MODE_ALIAS.get(r, default)


def build_backtester(symbols, mode, use_cache=True, refresh=False):
    """撈名稱 + 建立 Backtester + 載入歷史資料 (較耗時,故抽出共用)。
    use_cache=True → 走本機 Parquet 快取 (需先跑 build_cache.py);refresh=True → 快取再補增量。"""
    if not symbols:
        symbols = list(DEFAULT_SYMBOLS)
        print(f"未指定代號 → 使用分散化測試池 ({len(symbols)} 檔,跨產業低相關)")
    print(f"\n將回測 {len(symbols)} 檔:{' '.join(symbols)}  (模式:{mode})")
    # 快取模式省下『撈名稱』那次 API (0 API 目標);報表以代號顯示。
    names = {} if use_cache else (print("撈取股票名稱...") or fetch_names(symbols))
    bt = Backtester(symbols=symbols, names=names, mode=mode)
    if use_cache:
        from core.backtest import cached_fetch_history
        tag = "本機快取 + 補增量" if refresh else "本機快取 (0 API)"
        print(f"從{tag}載入歷史資料集...")
        bt.load(fetcher=lambda s: cached_fetch_history(s, refresh=refresh))
    else:
        print("正在撈取歷史資料集 (每檔約 8 個資料集,檔數多時請耐心等候)...")
        bt.load()
    return bt


def run_action(bt, action, *, start="2023-01-01", end="2025-12-31", rebalance="M",
               holding=20, exit_mode="dynamic_stop", benchmark="0050",
               ratings=("強勢買進", "強烈推薦"), top_n=5, weighting="score", track=False):
    """依 action 執行對應分析。action ∈ backtest / validate / neutral / ranked / optimize / cycle / attribution。"""
    if action == "optimize":
        bt.optimize(start=start, end=end, rebalance=rebalance, holding_days=holding)
    elif action == "validate":
        bt.validate_signal(start=start, end=end, rebalance=rebalance, holding_days=holding)
    elif action == "neutral":
        bt.market_neutral_curve(start=start, end=end, rebalance=rebalance, benchmark=benchmark)
    elif action == "ranked":
        bt.ranked_equity_curve(start=start, end=end, rebalance=rebalance,
                               top_n=top_n, weighting=weighting, benchmark=benchmark)
    elif action == "cycle":
        bt.cycle_robustness(rebalance=rebalance, holding_days=holding, benchmark=benchmark)
    elif action == "attribution":
        bt.factor_attribution(start=start, end=end, rebalance=rebalance, holding_days=holding)
    else:  # backtest (標準:分桶報酬 + 權益曲線)
        print(f"開始 point-in-time 滾動回測 (出場模式:{exit_mode})...\n")
        records = bt.run(start=start, end=end, rebalance=rebalance,
                         holding_days=holding, exit_mode=exit_mode)
        bt.summarize(records)
        if track:
            bt.track_rating_changes(records)
        print(f"權益曲線買進評級:{', '.join(ratings)}")
        bt.equity_curve(start=start, end=end, rebalance=rebalance,
                        strategy_ratings=ratings, benchmark=benchmark)
        return records          # 回傳供匯出 Excel
    return None


def _open_file(path):
    """用系統預設程式開啟檔案 (跨平台);失敗僅提示,不中斷。"""
    try:
        if sys.platform == "win32":
            os.startfile(os.path.abspath(path))          # type: ignore[attr-defined]
        elif sys.platform == "darwin":
            import subprocess; subprocess.run(["open", path], check=False)
        else:
            import subprocess; subprocess.run(["xdg-open", path], check=False)
        print(f"📂 已自動開啟:{path}")
    except Exception as e:
        print(f"(檔案已存檔,但自動開啟失敗:{e})")


def export_backtest_excel(records, mode="balanced", path=None, auto_open=True):
    """
    把回測結果匯出成『美化排版』的 Excel:
      ① 分桶績效 (各評級平均報酬/勝率/樣本數 + 多空價差)
      ② 個股排行 (依平均報酬排序,含各評級次數與多數評級)
      ③ 完整明細 (逐評級日×逐檔的原始記錄)
    排版:標題列、深藍表頭白字、凍結窗格、隔列淺底、報酬紅綠標色、評級色塊、繁中字體。
    """
    import pandas as pd
    from datetime import datetime
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        from openpyxl.utils import get_column_letter
        from openpyxl.chart import BarChart, Reference
        from openpyxl.chart.shapes import GraphicalProperties
        from openpyxl.drawing.fill import PatternFillProperties, ColorChoice
    except ImportError:
        print("⚠️ 需要 openpyxl 才能匯出 Excel (pip install openpyxl)。")
        return None
    if records is None or len(records) == 0:
        print("⚠️ 無回測記錄可匯出。")
        return None

    RSUPER, RSTRONG, RWATCH, RAVOID = "強勢買進", "強烈推薦", "觀望追蹤", "謹慎避開"
    ORDER = [RSUPER, RSTRONG, RWATCH, RAVOID]
    FNT = "Microsoft JhengHei"        # Windows 繁中字體;跨平台自動 fallback

    title_font = Font(name=FNT, size=16, bold=True, color="1F3864")
    sub_font = Font(name=FNT, size=10, color="595959")
    head_font = Font(name=FNT, size=12, bold=True, color="FFFFFF")
    body_font = Font(name=FNT, size=11)
    bold_body = Font(name=FNT, size=11, bold=True)
    pos_font = Font(name=FNT, size=11, bold=True, color="1B7A34")   # 綠:正報酬
    neg_font = Font(name=FNT, size=11, bold=True, color="C00000")   # 紅:負報酬
    head_fill = PatternFill("solid", fgColor="1F3864")
    band_fill = PatternFill("solid", fgColor="F2F6FC")
    thin = Side(style="thin", color="D9D9D9")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    center = Alignment(horizontal="center", vertical="center")
    left = Alignment(horizontal="left", vertical="center")
    rating_fill = {RSUPER: "C6EFCE", RSTRONG: "E2EFDA", RWATCH: "FFF2CC", RAVOID: "F2F2F2"}
    rating_col = {RSUPER: "1B7A34", RSTRONG: "375623", RWATCH: "9C6500", RAVOID: "808080"}

    df = records.copy()
    df["forward_return"] = pd.to_numeric(df.get("forward_return"), errors="coerce")

    def _sheet_header(ws, title, subtitle, ncol):
        ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=ncol)
        c = ws.cell(row=1, column=1, value=title); c.font = title_font; c.alignment = left
        ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=ncol)
        c = ws.cell(row=2, column=1, value=subtitle); c.font = sub_font; c.alignment = left
        ws.row_dimensions[1].height = 24

    def _write_table(ws, headers, rows, start_row, widths, ret_cols=(), rating_col_idx=None,
                     num_cols=()):
        # 表頭
        for j, h in enumerate(headers, 1):
            c = ws.cell(row=start_row, column=j, value=h)
            c.font = head_font; c.fill = head_fill; c.alignment = center; c.border = border
        ws.row_dimensions[start_row].height = 22
        # 內容
        for i, row in enumerate(rows):
            r = start_row + 1 + i
            banded = (i % 2 == 1)
            for j, val in enumerate(row, 1):
                c = ws.cell(row=r, column=j, value=val)
                c.font = body_font; c.border = border
                c.alignment = left if j == 1 else center
                if banded:
                    c.fill = band_fill
                if j in num_cols and isinstance(val, (int, float)):
                    c.number_format = '0.0'
                if j in ret_cols and isinstance(val, (int, float)):
                    c.number_format = '+0.00"%";-0.00"%"'
                    c.font = pos_font if val > 0 else (neg_font if val < 0 else bold_body)
                if rating_col_idx and j == rating_col_idx and val in rating_fill:
                    c.fill = PatternFill("solid", fgColor=rating_fill[val])
                    c.font = Font(name=FNT, size=11, bold=True, color=rating_col[val])
                    c.alignment = center
        for j, w in enumerate(widths, 1):
            ws.column_dimensions[get_column_letter(j)].width = w
        ws.freeze_panes = ws.cell(row=start_row + 1, column=1)

    ts = datetime.now().strftime("%Y-%m-%d %H:%M")
    period = f"{df['as_of'].min()} ~ {df['as_of'].max()}" if "as_of" in df.columns else ""
    wb = Workbook()

    # ---- ① 分桶績效 ----
    ws1 = wb.active; ws1.title = "① 分桶績效"
    _sheet_header(ws1, "回測分桶績效 — 各評級後續報酬",
                  f"模式:{mode}　｜　區間:{period}　｜　產生:{ts}", 6)
    rows = []
    present = [r for r in ORDER if r in set(df["rating"])]
    for r in present:
        g = df[df["rating"] == r]["forward_return"].dropna()
        if len(g) == 0:
            continue
        win = float((g > 0).mean() * 100)
        rows.append([r, int(len(g)), round(float(g.mean()), 2), round(float(g.median()), 2),
                     round(win, 1), round(float(g.std()), 2)])
    _write_table(ws1, ["評級", "樣本數", "平均報酬%", "中位數%", "勝率%", "標準差"], rows, 4,
                 [16, 10, 13, 12, 10, 10], ret_cols=(3, 4), rating_col_idx=1, num_cols=(5, 6))
    # 小長條圖:各評級平均報酬% (放在表格右側 H4)
    if rows:
        n = len(rows)
        chart = BarChart()
        chart.type = "col"
        chart.title = "各評級平均報酬%"
        chart.y_axis.title = "平均報酬%"
        chart.x_axis.title = "評級"
        chart.legend = None
        chart.height, chart.width = 7.5, 13
        data = Reference(ws1, min_col=3, min_row=4, max_row=4 + n)      # 含表頭當序列名
        cats = Reference(ws1, min_col=1, min_row=5, max_row=4 + n)
        chart.add_data(data, titles_from_data=True)
        chart.set_categories(cats)
        try:
            chart.series[0].graphicalProperties = GraphicalProperties(solidFill="4472C4")
        except Exception:
            pass
        ws1.add_chart(chart, "H4")
    # 多空價差
    means = {r: df[df["rating"] == r]["forward_return"].mean() for r in present}
    base_r = next((r for r in (RSUPER, RSTRONG, RWATCH) if r in means), None)
    foot = 4 + 1 + len(rows) + 1
    if base_r and RAVOID in means:
        sp = float(means[base_r] - means[RAVOID])
        c = ws1.cell(row=foot, column=1,
                     value=f"多空價差 ({base_r} − {RAVOID}):{sp:+.2f}%  (正且大代表評級有鑑別力)")
        c.font = Font(name=FNT, size=11, bold=True, color=("1B7A34" if sp > 0 else "C00000"))
        ws1.merge_cells(start_row=foot, start_column=1, end_row=foot, end_column=6)

    # ---- ② 個股排行 ----
    ws2 = wb.create_sheet("② 個股排行")
    _sheet_header(ws2, "個股排行 — 依回測期平均報酬排序",
                  f"模式:{mode}　｜　各欄為該股被評為各評級的次數", 8)
    name_map = (df.drop_duplicates("symbol").set_index("symbol")["name"].to_dict()
                if "name" in df.columns else {})
    stock_rows = []
    for sym, g in df.groupby("symbol"):
        cnt = g["rating"].value_counts().to_dict()
        avg = float(g["forward_return"].mean())
        dom = g["rating"].mode().iloc[0] if not g["rating"].mode().empty else ""
        stock_rows.append([f"{sym} {name_map.get(sym, '')}".strip(),
                           int(cnt.get(RSUPER, 0)), int(cnt.get(RSTRONG, 0)),
                           int(cnt.get(RWATCH, 0)), int(cnt.get(RAVOID, 0)),
                           round(avg, 2), dom])
    stock_rows.sort(key=lambda x: x[5], reverse=True)
    _write_table(ws2, ["個股", "強勢", "強推", "觀望", "避開", "平均報酬%", "多數評級"],
                 stock_rows, 4, [22, 8, 8, 8, 8, 13, 12],
                 ret_cols=(6,), rating_col_idx=7)

    # ---- ③ 完整明細 ----
    ws3 = wb.create_sheet("③ 完整明細")
    cols = [c for c in ["as_of", "symbol", "name", "rating", "total_score", "whale_score",
                        "valuation_status", "forward_return", "exit_reason", "bars_held",
                        "exit_date"] if c in df.columns]
    hdr_map = {"as_of": "評級日", "symbol": "代號", "name": "名稱", "rating": "評級",
               "total_score": "綜合分", "whale_score": "籌碼分", "valuation_status": "估值狀態",
               "forward_return": "後續報酬%", "exit_reason": "出場原因", "bars_held": "持有K數",
               "exit_date": "出場日"}
    _sheet_header(ws3, "完整回測明細 (Point-in-Time)", f"共 {len(df)} 筆　｜　模式:{mode}", len(cols))
    det = df.sort_values(["as_of", "symbol"]) if "as_of" in df.columns else df
    rows3, ret_idx = [], cols.index("forward_return") + 1 if "forward_return" in cols else None
    for _, r in det.iterrows():
        rows3.append([round(float(r[c]), 2) if c in ("total_score", "whale_score", "forward_return")
                      and pd.notna(r[c]) else r[c] for c in cols])
    _write_table(ws3, [hdr_map[c] for c in cols], rows3, 4,
                 [12, 9, 12, 11, 9, 9, 16, 12, 15, 9, 12],
                 ret_cols=(ret_idx,) if ret_idx else (),
                 rating_col_idx=(cols.index("rating") + 1) if "rating" in cols else None)

    if path is None:
        out_dir = os.path.join(project_root, "outputs", "excel")
        os.makedirs(out_dir, exist_ok=True)
        path = os.path.join(out_dir, f"回測排行_{mode}_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx")
    try:
        wb.save(path)
        print(f"✅ 已匯出美化 Excel:{path}")
        if auto_open:
            _open_file(path)
        return path
    except Exception as e:
        print(f"⚠️ Excel 匯出失敗:{e}")
        return None


# 互動選單的功能對照
ACTIONS = {
    "1": ("backtest", "標準回測 (分桶報酬 + 權益曲線)"),
    "2": ("validate", "樣本外驗證 (絕對報酬 vs 市場中性)"),
    "3": ("neutral", "市場中性權益曲線 (買前⅓、放空後⅓)"),
    "4": ("ranked", "排序配置 (每期買綜合分前 N 名)"),
    "5": ("optimize", "參數網格搜尋 (train/test)"),
    "6": ("cycle", "多市場週期穩健性 (2022 空頭 vs 2023–2025 多頭排序力)"),
    "7": ("attribution", "分項因子歸因 (五維度對排序力的邊際貢獻)"),
}


def _ask(prompt, default=""):
    try:
        s = input(prompt).strip()
    except (EOFError, KeyboardInterrupt):
        return None
    return s if s else default


def interactive_menu(defaults):
    """按下『執行』(無命令列參數) 時進入:可輸入代號、切換模式、選擇分析功能,並可連續操作。"""
    print("\n" + "=" * 64)
    print("📊 台股 PIT 回測系統 — 互動模式")
    print("   · 輸入代號 (2330 2454 …)  → 回測這幾檔 (空白/逗號皆可)")
    print("   · 直接 Enter              → 使用分散化測試池")
    print("   · 輸入 m                   → 切換策略模式 (保守/平衡/積極)")
    print("   · 輸入 q                   → 離開")
    print("=" * 64)

    mode = normalize_mode(defaults.get("mode"), "balanced")
    print(f"目前策略模式:{mode}")
    bt = None
    loaded_key = None                      # (symbols, mode);相同就重用已載入的資料,不重撈

    while True:
        raw = _ask(f"\n[{mode}] 請輸入代號 (Enter=測試池 / m=換模式 / q=離開) ▶ ")
        if raw is None or raw.lower() in ("q", "quit", "exit"):
            print("已離開。")
            break
        if raw.lower() in ("m", "mode", "模式"):
            sel = _ask("  模式:1) conservative 保守  2) balanced 平衡  3) aggressive 積極 (名稱或 c/b/a) ▶ ")
            mode = normalize_mode({"1": "conservative", "2": "balanced", "3": "aggressive"}.get(sel, sel), mode)
            print(f"  ✅ 已切換為:{mode}")
            loaded_key = None              # 換模式→重建
            continue

        symbols = parse_symbols(raw) if raw else list(DEFAULT_SYMBOLS)

        # 選功能
        print("  要跑哪個分析?")
        for k, (_, desc) in ACTIONS.items():
            print(f"    {k}) {desc}")
        sel = _ask("  請選擇 (1-7,預設 1) ▶ ", "1")
        action = ACTIONS.get(sel, ACTIONS["1"])[0]

        top_n, weighting, ratings = defaults["top_n"], defaults["weighting"], defaults["ratings"]
        if action == "ranked":
            top_n = int(_ask("    買前幾名? (預設 5) ▶ ", "5") or 5)
            weighting = "equal" if _ask("    配重 1)分數加權 2)等權 (預設 1) ▶ ", "1") == "2" else "score"

        # 相同 (代號+模式) 就重用已載入資料,避免重撈
        key = (tuple(symbols), mode)
        if bt is None or key != loaded_key:
            bt = build_backtester(symbols, mode,
                                  use_cache=defaults.get("use_cache", False),
                                  refresh=defaults.get("refresh", False))
            loaded_key = key

        try:
            recs = run_action(bt, action, start=defaults["start"], end=defaults["end"],
                              rebalance=defaults["rebalance"], holding=defaults["holding"],
                              exit_mode=defaults["exit_mode"], benchmark=defaults["benchmark"],
                              ratings=ratings, top_n=top_n, weighting=weighting, track=defaults["track"])
            # 跑完標準回測 → 詢問是否匯出美化 Excel 排行
            if action == "backtest" and recs is not None and len(recs) > 0:
                if (_ask("  要把排行匯出成美化 Excel 嗎? (y/n,預設 y) ▶ ", "y") or "y").lower() in ("y", "yes"):
                    export_backtest_excel(recs, mode=mode)
        except Exception as e:
            print(f"⚠️ 執行時發生錯誤:{e}")
        print("\n(可繼續輸入下一組代號 / m 換模式 / 選其他分析 / q 離開)")


def main():
    parser = argparse.ArgumentParser(description="Point-in-time 回測執行器")
    parser.add_argument("symbols", nargs="*", help="股票代號;省略則用分散化測試池")
    parser.add_argument("-m", "--mode", default="balanced", help="conservative/balanced/aggressive")
    parser.add_argument("--start", default="2023-01-01")
    parser.add_argument("--end", default="2025-12-31")
    parser.add_argument("--rebalance", default="M", help="M(月) 或 W(週)")
    parser.add_argument("--holding", type=int, default=20)
    parser.add_argument("--exit-mode", default="dynamic_stop",
                        choices=["dynamic_stop", "horizon"],
                        help="出場方式:dynamic_stop(預設,動態籌碼支撐停損+移動停利) / horizon(固定持有)")
    parser.add_argument("--only-strong", action="store_true",
                        help="權益曲線只買最高級(強勢買進)")
    parser.add_argument("--with-watch", action="store_true",
                        help="權益曲線納入觀望追蹤(較分散、較接近大盤)")
    parser.add_argument("--benchmark", default="0050")
    parser.add_argument("--optimize", action="store_true", help="參數網格搜尋 + train/test 驗證")
    parser.add_argument("--track", action="store_true", help="追蹤評級變動")
    parser.add_argument("--validate", action="store_true", help="路線一驗證:只買強推 vs 買進(強推+觀望)")
    parser.add_argument("--neutral", action="store_true", help="市場中性權益曲線:每期買前⅓、放空後⅓")
    parser.add_argument("--ranked", action="store_true",
                        help="排序配置 long-only:每期買綜合分前 N 名 (利用已驗證的排序能力)")
    parser.add_argument("--cycle", action="store_true",
                        help="多市場週期穩健性:跨 2022 空頭 / 2023–2025 多頭段比較排序力")
    parser.add_argument("--attribution", action="store_true",
                        help="分項因子歸因:拆解五維度對市場中性排序力的邊際貢獻")
    parser.add_argument("--top-n", type=int, default=5, help="--ranked 時買進的檔數 (預設 5)")
    parser.add_argument("--weighting", default="score", choices=["score", "equal"],
                        help="--ranked 的配重:score(分數加權,預設) / equal(等權)")
    parser.add_argument("--cache", action="store_true",
                        help="(預設已啟用) 走本機 Parquet 快取,回測 0 次 API;保留供相容")
    parser.add_argument("--cache-refresh", action="store_true",
                        help="走本機快取但先補抓增量到最新 (少量 API)")
    parser.add_argument("--live", action="store_true",
                        help="強制即時抓 FinMind 全歷史 (會大量 API);預設走本機快取 0 API")
    parser.add_argument("--no-regime", action="store_true",
                        help="關閉市場 regime 動態權重 (A/B 對照:預設開啟,空頭段自動降動能加重基本面)")
    parser.add_argument("--interactive", action="store_true", help="強制互動輸入代號")
    parser.add_argument("--export", default=None, choices=["xlsx"],
                        help="標準回測後把排行匯出成美化 Excel")
    args, _ = parser.parse_known_args()

    # 買進評級 (標準回測 / 互動預設)
    if args.only_strong:
        ratings = ("強勢買進",)
    elif args.with_watch:
        ratings = ("強勢買進", "強烈推薦", "觀望追蹤")
    else:
        ratings = ("強勢買進", "強烈推薦")          # 預設:集中買買進級,不再closet-index大盤

    # 預設走本機快取 (0 API);--live 才即時抓。--cache-refresh 走快取並補增量。
    use_cache = (not args.live)
    defaults = dict(mode=args.mode, start=args.start, end=args.end, rebalance=args.rebalance,
                    holding=args.holding, exit_mode=args.exit_mode, benchmark=args.benchmark,
                    ratings=ratings, top_n=args.top_n, weighting=args.weighting, track=args.track,
                    use_cache=use_cache, refresh=args.cache_refresh)

    # 直接按『執行』(無任何命令列參數) 或 --interactive → 進入互動選單
    if args.interactive or len(sys.argv) <= 1:
        interactive_menu(defaults)
        return

    # 有命令列參數 → 一次性執行 (維持原本 CLI 行為)
    symbols = parse_symbols(" ".join(args.symbols)) if args.symbols else []
    bt = build_backtester(symbols, args.mode,
                          use_cache=use_cache, refresh=args.cache_refresh)
    bt.use_regime = not args.no_regime      # 市場 regime 動態權重 (預設開;--no-regime 關閉做對照)
    action = ("optimize" if args.optimize else "validate" if args.validate
              else "neutral" if args.neutral else "ranked" if args.ranked
              else "cycle" if args.cycle else "attribution" if args.attribution
              else "backtest")
    recs = run_action(bt, action, start=args.start, end=args.end, rebalance=args.rebalance,
                      holding=args.holding, exit_mode=args.exit_mode, benchmark=args.benchmark,
                      ratings=ratings, top_n=args.top_n, weighting=args.weighting, track=args.track)
    if args.export and action == "backtest" and recs is not None:
        export_backtest_excel(recs, mode=args.mode)


if __name__ == "__main__":
    main()