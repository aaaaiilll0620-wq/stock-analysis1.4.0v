# -*- coding: utf-8 -*-
"""TEJ DataExport0806 遷移的 fail-closed 機制稽核測試 (Round 3 code review)。

規格:docs/資料快照遷移_DataExport0806.md。

只用本檔 synthetic fixture (寫在 `tmp_path` 底下) ——**不讀取**任何真實的
`tej_exports/DataExport0806`、`tej_exports/inbox*`、`~/tej_cache`。測到的模組層級
常數 (`tej_importer.DATA_ROOT`/`MANIFEST_CSV`) 一律用 `monkeypatch` 換成 tmp_path
底下的路徑,不動真的專案資料。

涵蓋:必要欄位缺失、stock_id/date 無效值、重複鍵衝突 (含「低於舊 1% 門檻也要
raise」)、manifest preflight (缺檔/多檔/hash 不符)、save_by_stock 的 staging
rollback 與 no-orphan、legacy supplement 的重複/null 鍵防呆、
_full_population_diff.py 的 outer-join/欄位缺失/NaN 不對稱/精確誤差偵測。
"""
from __future__ import annotations

import csv
import hashlib
import importlib
import json
import shutil
import sys
import zipfile
from pathlib import Path

import numpy as np
import openpyxl
import pandas as pd
import pytest

REPO_ROOT = Path(__file__).resolve().parent.parent
if str(REPO_ROOT) not in sys.path:
    sys.path.insert(0, str(REPO_ROOT))
if str(REPO_ROOT / "scripts") not in sys.path:
    sys.path.insert(0, str(REPO_ROOT / "scripts"))

tej_importer = pytest.importorskip("tej_importer")
extract_legacy_supplement = pytest.importorskip("extract_legacy_supplement")
diffmod = pytest.importorskip("_full_population_diff")


def test_normalize_source_column_aliases_accepts_verified_year_month_variant():
    raw = pd.DataFrame({"代號": ["2330"], "名稱": ["台積電"], "年/月": ["2026/06"]})
    got = tej_importer._normalize_source_column_aliases(raw, "q2.xlsx")
    assert got["年月"].tolist() == ["2026/06"]


def test_normalize_source_column_aliases_rejects_conflicting_canonical_value():
    raw = pd.DataFrame({"年月": ["202603"], "年/月": ["2026/06"]})
    with pytest.raises(ValueError, match="禁止靜默覆蓋"):
        tej_importer._normalize_source_column_aliases(raw, "conflict.xlsx")


def test_required_security_code_accepts_normalized_stock_id(tmp_path, monkeypatch):
    monkeypatch.setattr(tej_importer, "DATA_ROOT", tmp_path)
    raw = pd.DataFrame({"stock_id": ["2330"], "年月": ["202606"]})
    tej_importer._check_required_cols(
        tmp_path / "q2.xlsx", raw, {"required_cols": ["證券代碼", "年月"]})


def test_parse_dates_accepts_only_explicit_financial_month_formats():
    raw = pd.Series(["202603", "2026/06", "合計"])
    got = tej_importer._parse_dates(raw, ("%Y%m", "%Y/%m"))
    assert got.iloc[0] == pd.Timestamp("2026-03-01")
    assert got.iloc[1] == pd.Timestamp("2026-06-01")
    assert pd.isna(got.iloc[2])


def _sha256_of(path: Path) -> str:
    return hashlib.sha256(path.read_bytes()).hexdigest()


@pytest.fixture
def patched_root(tmp_path, monkeypatch):
    """把 tej_importer 的 DATA_ROOT/MANIFEST_CSV 換成 tmp_path 底下的路徑,
    並清掉 manifest cache (模組層級全域變數,測試之間會互相汙染)。"""
    data_root = tmp_path / "DataExport0806"
    data_root.mkdir()
    manifest_csv = tmp_path / "manifest.csv"
    monkeypatch.setattr(tej_importer, "DATA_ROOT", data_root)
    monkeypatch.setattr(tej_importer, "MANIFEST_CSV", manifest_csv)
    monkeypatch.setattr(tej_importer, "_manifest_cache", None)
    return data_root, manifest_csv


def _write_manifest(manifest_csv: Path, entries: dict):
    """entries: {relpath: sha256}"""
    with open(manifest_csv, "w", newline="", encoding="utf-8") as f:
        w = csv.writer(f)
        w.writerow(["relpath", "size_bytes", "sha256", "mtime_utc"])
        for rel, sha in entries.items():
            w.writerow([rel, 0, sha, "2026-01-01T00:00:00+00:00"])


# =============================================================================
# 1. required_cols fail-closed
# =============================================================================

def test_required_cols_missing_raises(patched_root):
    data_root, _ = patched_root
    raw = pd.DataFrame({"代號": ["1101"], "名稱": ["台泥"], "年月日": ["2026/01/02"]})
    spec = {"required_cols": ["年月日", "開盤價(元)"]}
    p = data_root / "dummy.xlsx"
    with pytest.raises(ValueError, match="缺少必要欄位"):
        tej_importer._check_required_cols(p, raw, spec)


def test_required_cols_present_passes(patched_root):
    data_root, _ = patched_root
    raw = pd.DataFrame({"代號": ["1101"], "名稱": ["台泥"], "年月日": ["2026/01/02"],
                         "開盤價(元)": [10.0]})
    spec = {"required_cols": ["年月日", "開盤價(元)"]}
    p = data_root / "dummy.xlsx"
    tej_importer._check_required_cols(p, raw, spec)   # 不 raise 就算過


# =============================================================================
# 2. stock_id / date 無效值 fail-closed (空白/NaN/"nan"字串,不是靜默 dropna)
# =============================================================================

@pytest.mark.parametrize("bad_id", ["", "nan", "None", "NaN", "  ", np.nan])
def test_invalid_stock_id_raises(patched_root, bad_id):
    data_root, _ = patched_root
    df = pd.DataFrame({"stock_id": ["1101", bad_id], "stock_name": ["台泥", "X"],
                        "date": ["2026-01-02", "2026-01-03"]})
    p = data_root / "dummy.xlsx"
    with pytest.raises(ValueError, match="stock_id 無效"):
        tej_importer._check_valid_keys(p, df, check_date=True)


def test_invalid_date_raises(patched_root):
    data_root, _ = patched_root
    df = pd.DataFrame({"stock_id": ["1101", "1102"], "stock_name": ["台泥", "亞泥"],
                        "date": ["2026-01-02", np.nan]})
    p = data_root / "dummy.xlsx"
    with pytest.raises(ValueError, match="date 無法解析"):
        tej_importer._check_valid_keys(p, df, check_date=True)


def test_valid_keys_pass(patched_root):
    data_root, _ = patched_root
    df = pd.DataFrame({"stock_id": ["1101", "1102"], "stock_name": ["台泥", "亞泥"],
                        "date": ["2026-01-02", "2026-01-03"]})
    p = data_root / "dummy.xlsx"
    tej_importer._check_valid_keys(p, df, check_date=True)   # 不 raise 就算過


# =============================================================================
# 3. 重複鍵衝突 —— 移除 1% 容忍門檻,任何衝突都要 raise
# =============================================================================

def test_exact_duplicate_rows_do_not_raise():
    df = pd.DataFrame({
        "stock_id": ["1101", "1101"], "stock_name": ["台泥", "台泥"],
        "date": ["2026-01-02", "2026-01-02"], "close": [10.0, 10.0],
    })
    tej_importer._check_duplicate_key_conflicts(df, "dummy")   # 完全重複,不 raise


def test_single_conflicting_duplicate_raises_even_below_old_1pct_threshold():
    """9999 個乾淨的 key + 1 個衝突 key,衝突率 = 1/10000 = 0.01%,遠低於舊版的 1%
    門檻。Round 3 review 要求移除門檻,這裡驗證即使佔比極小也一樣 raise。"""
    n_clean = 9999
    clean = pd.DataFrame({
        "stock_id": [f"{1000+i}" for i in range(n_clean)],
        "stock_name": ["X"] * n_clean,
        "date": ["2026-01-02"] * n_clean,
        "close": [10.0] * n_clean,
    })
    conflict = pd.DataFrame({
        "stock_id": ["9999", "9999"], "stock_name": ["Y", "Y"],
        "date": ["2026-01-02", "2026-01-02"], "close": [10.0, 20.0],   # 同 key 不同值
    })
    combined = pd.concat([clean, conflict], ignore_index=True)
    with pytest.raises(ValueError, match="數值不一致"):
        tej_importer._check_duplicate_key_conflicts(combined, "dummy")


def test_static_dataset_duplicate_conflict_uses_stock_id_only_key():
    df = pd.DataFrame({
        "stock_id": ["1101", "1101"], "stock_name": ["台泥", "台泥"],
        "tse_ind_code": ["M1100", "M9999"],   # 同代號、不同產業代碼 → 衝突
    })
    with pytest.raises(ValueError, match="數值不一致"):
        tej_importer._check_duplicate_key_conflicts(df, "industry_map", key_cols=["stock_id"])


# =============================================================================
# 4. manifest preflight —— 缺檔/多檔/hash 不符都要在解析前 raise
# =============================================================================

def _make_source_dir(data_root: Path, name: str, files: dict) -> Path:
    d = data_root / name
    d.mkdir()
    for fname, content in files.items():
        (d / fname).write_bytes(content)
    return d


def test_manifest_preflight_passes_when_everything_matches(patched_root):
    data_root, manifest_csv = patched_root
    src_dir = _make_source_dir(data_root, "cat", {"a.xlsx": b"hello", "b.xlsx": b"world"})
    files = sorted(src_dir.glob("*.xlsx"))
    entries = {f.relative_to(data_root).as_posix(): _sha256_of(f) for f in files}
    _write_manifest(manifest_csv, entries)
    spec = {"source_dir": src_dir}
    tej_importer._manifest_preflight(files, spec, "dummy")   # 不 raise 就算過


def test_manifest_preflight_raises_on_missing_file(patched_root):
    data_root, manifest_csv = patched_root
    src_dir = _make_source_dir(data_root, "cat", {"a.xlsx": b"hello"})
    files = sorted(src_dir.glob("*.xlsx"))
    entries = {f.relative_to(data_root).as_posix(): _sha256_of(f) for f in files}
    entries["cat/b.xlsx"] = "deadbeef" * 8   # manifest 記錄了一個磁碟上沒有的檔案
    _write_manifest(manifest_csv, entries)
    spec = {"source_dir": src_dir}
    with pytest.raises(ValueError, match="manifest preflight 失敗"):
        tej_importer._manifest_preflight(files, spec, "dummy")


def test_manifest_preflight_raises_on_extra_file(patched_root):
    data_root, manifest_csv = patched_root
    src_dir = _make_source_dir(data_root, "cat", {"a.xlsx": b"hello", "b.xlsx": b"world"})
    files = sorted(src_dir.glob("*.xlsx"))
    a = src_dir / "a.xlsx"
    entries = {a.relative_to(data_root).as_posix(): _sha256_of(a)}   # 只記錄 a,漏了 b
    _write_manifest(manifest_csv, entries)
    spec = {"source_dir": src_dir}
    with pytest.raises(ValueError, match="manifest preflight 失敗"):
        tej_importer._manifest_preflight(files, spec, "dummy")


def test_manifest_preflight_raises_on_hash_mismatch(patched_root):
    data_root, manifest_csv = patched_root
    src_dir = _make_source_dir(data_root, "cat", {"a.xlsx": b"hello"})
    files = sorted(src_dir.glob("*.xlsx"))
    entries = {files[0].relative_to(data_root).as_posix(): "0" * 64}   # 假雜湊,一定不符
    _write_manifest(manifest_csv, entries)
    spec = {"source_dir": src_dir}
    with pytest.raises(ValueError, match="SHA-256"):
        tej_importer._manifest_preflight(files, spec, "dummy")


# =============================================================================
# 5. save_by_stock —— atomic staged publish:no-orphan + rollback
# =============================================================================

def _stock_df(stock_ids, date="2026-01-02"):
    return pd.DataFrame({
        "stock_id": stock_ids, "stock_name": ["X"] * len(stock_ids),
        "date": [date] * len(stock_ids), "close": list(range(len(stock_ids))),
    })


def test_save_by_stock_no_orphan_after_shrinking_universe(tmp_path):
    cache_dir = tmp_path / "cache"
    df1 = _stock_df(["1101", "1102", "1103"])
    tej_importer.save_by_stock(df1, "price_valuation", cache_dir)
    assert (cache_dir / "price_valuation" / "1103.parquet").exists()

    df2 = _stock_df(["1101", "1102"])   # 1103 這次資料裡消失了 (例如下市)
    tej_importer.save_by_stock(df2, "price_valuation", cache_dir)

    out_dir = cache_dir / "price_valuation"
    assert not (out_dir / "1103.parquet").exists(), "舊股票的 parquet 沒有被清掉,留下 orphan"
    assert (out_dir / "1101.parquet").exists()
    assert (out_dir / "1102.parquet").exists()
    assert len(list(out_dir.glob("*.parquet"))) == 2
    # staging/rollback 暫存目錄不應該留下來
    assert not (cache_dir / ".price_valuation.staging").exists()
    assert not (cache_dir / ".price_valuation.rollback").exists()


def test_save_by_stock_rollback_on_failure_preserves_old_data(tmp_path, monkeypatch):
    cache_dir = tmp_path / "cache"
    df1 = _stock_df(["1101", "1102"])
    tej_importer.save_by_stock(df1, "price_valuation", cache_dir)
    out_dir = cache_dir / "price_valuation"
    original_1101 = pd.read_parquet(out_dir / "1101.parquet")

    # 模擬寫到一半失敗:groupby 之後、驗證階段強制炸掉
    real_nunique = pd.Series.nunique

    def _boom(self, *a, **kw):
        raise RuntimeError("simulated failure during staging validation")

    monkeypatch.setattr(pd.Series, "nunique", _boom)
    df2 = _stock_df(["1101", "1102", "1104"])
    with pytest.raises(RuntimeError, match="simulated failure"):
        tej_importer.save_by_stock(df2, "price_valuation", cache_dir)
    monkeypatch.setattr(pd.Series, "nunique", real_nunique)

    # out_dir 必須維持失敗前的原樣,不是空的、不是新資料、也不是半寫的殘檔
    assert (out_dir / "1101.parquet").exists()
    assert (out_dir / "1102.parquet").exists()
    assert not (out_dir / "1104.parquet").exists()
    restored = pd.read_parquet(out_dir / "1101.parquet")
    pd.testing.assert_frame_equal(original_1101, restored)
    # 暫存目錄清乾淨,沒有殘留
    assert not (cache_dir / ".price_valuation.staging").exists()
    assert not (cache_dir / ".price_valuation.rollback").exists()


# =============================================================================
# 6. legacy supplement:發布前的重複/null 鍵防呆 (_profile)
# =============================================================================

def test_supplement_profile_detects_duplicate_key():
    df = pd.DataFrame({"stock_id": ["1101", "1101"], "date": ["2026-01-02", "2026-01-02"],
                        "roe_after_tax": [1.0, 2.0]})
    profile = extract_legacy_supplement._profile(df, key_cols=["stock_id", "date"])
    assert profile["duplicate_key_row_count"] == 2


def test_supplement_profile_detects_null_key():
    df = pd.DataFrame({"stock_id": ["1101", None], "date": ["2026-01-02", "2026-01-03"],
                        "roe_after_tax": [1.0, 2.0]})
    profile = extract_legacy_supplement._profile(df, key_cols=["stock_id", "date"])
    assert profile["null_key_row_count"] == 1


def test_supplement_profile_clean_data_passes():
    df = pd.DataFrame({"stock_id": ["1101", "1102"], "date": ["2026-01-02", "2026-01-02"],
                        "roe_after_tax": [1.0, 2.0]})
    profile = extract_legacy_supplement._profile(df, key_cols=["stock_id", "date"])
    assert profile["duplicate_key_row_count"] == 0
    assert profile["null_key_row_count"] == 0


# =============================================================================
# 7. _full_population_diff.py —— outer-join/欄位缺失/NaN 不對稱/精確誤差/
#    重複鍵優先於 merge/status taxonomy/n_both_null (Round 4 review 更新)
# =============================================================================

PV_SPEC = {"cols": ["close"], "key_cols": ["stock_id", "date"]}
IM_SPEC = {"cols": ["tse_ind_code"], "key_cols": ["stock_id"], "static": True}


def _write_dataset(root: Path, dataset: str, frames: dict):
    """frames: {stock_id: DataFrame}"""
    d = root / dataset
    d.mkdir(parents=True)
    for sid, df in frames.items():
        df.to_parquet(d / f"{sid}.parquet", index=False)


def _write_static_dataset(root: Path, dataset: str, df: pd.DataFrame):
    root.mkdir(parents=True, exist_ok=True)
    df.to_parquet(root / f"{dataset}.parquet", index=False)


def test_diff_detects_missing_old_key_via_outer_join(tmp_path):
    old_root, new_root = tmp_path / "old", tmp_path / "new"
    old_df = pd.DataFrame({"stock_id": ["1101"] * 2, "date": ["2026-01-02", "2026-01-03"],
                            "close": [10.0, 11.0]})
    new_df = pd.DataFrame({"stock_id": ["1101"], "date": ["2026-01-02"],   # 少了 01-03 這列
                            "close": [10.0]})
    _write_dataset(old_root, "price_valuation", {"1101": old_df})
    _write_dataset(new_root, "price_valuation", {"1101": new_df})

    result = diffmod.diff_dataset("price_valuation", PV_SPEC, old_root, new_root)
    assert result["missing_keys_count"] == 1, "inner join 會把這列漏掉的日期吃掉,outer join 不應該"
    assert result["structural_status"] == "FAIL"
    assert result["overall_status"] == "FAIL"


def test_diff_detects_missing_column(tmp_path):
    old_root, new_root = tmp_path / "old", tmp_path / "new"
    old_df = pd.DataFrame({"stock_id": ["1101"], "date": ["2026-01-02"], "close": [10.0]})
    new_df = pd.DataFrame({"stock_id": ["1101"], "date": ["2026-01-02"]})   # 沒有 close 欄
    _write_dataset(old_root, "price_valuation", {"1101": old_df})
    _write_dataset(new_root, "price_valuation", {"1101": new_df})

    result = diffmod.diff_dataset("price_valuation", PV_SPEC, old_root, new_root)
    assert "close" in result["missing_columns_new"]
    assert result["structural_status"] == "FAIL"
    assert result["overall_status"] == "FAIL"


def test_diff_detects_nan_asymmetry_as_mismatch(tmp_path):
    old_root, new_root = tmp_path / "old", tmp_path / "new"
    old_df = pd.DataFrame({"stock_id": ["1101"], "date": ["2026-01-02"], "close": [10.0]})
    new_df = pd.DataFrame({"stock_id": ["1101"], "date": ["2026-01-02"], "close": [np.nan]})
    _write_dataset(old_root, "price_valuation", {"1101": old_df})
    _write_dataset(new_root, "price_valuation", {"1101": new_df})

    result = diffmod.diff_dataset("price_valuation", PV_SPEC, old_root, new_root)
    assert result["columns"]["close"]["n_null_mismatch"] == 1, \
        "一邊有值一邊 NaN 一定要算 mismatch (舊版 abs(x-NaN)>1 是 False,會漏掉)"
    assert result["value_status"] == "DIFF_UNRESOLVED"
    assert result["overall_status"] == "REVIEW_REQUIRED"


def test_diff_detects_small_diff_below_old_arbitrary_threshold(tmp_path):
    """diff = 0.5,舊版「絕對誤差 > 1」的門檻會把它當作相等而漏掉。新版報 exact
    diff,只要不是完全相等就算進 n_value_mismatch,不再套用沒有凍結依據的門檻。"""
    old_root, new_root = tmp_path / "old", tmp_path / "new"
    old_df = pd.DataFrame({"stock_id": ["1101"], "date": ["2026-01-02"], "close": [10.0]})
    new_df = pd.DataFrame({"stock_id": ["1101"], "date": ["2026-01-02"], "close": [10.5]})
    _write_dataset(old_root, "price_valuation", {"1101": old_df})
    _write_dataset(new_root, "price_valuation", {"1101": new_df})

    result = diffmod.diff_dataset("price_valuation", PV_SPEC, old_root, new_root)
    assert result["columns"]["close"]["n_value_mismatch"] == 1
    assert result["columns"]["close"]["n_exact_equal"] == 0
    assert result["value_status"] == "DIFF_UNRESOLVED"
    assert result["overall_status"] == "REVIEW_REQUIRED", \
        "有數值差異不能是 PASS,一律要 REVIEW_REQUIRED,只有完全相等才是 EXACT_PASS"


def test_diff_exact_match_reports_exact_pass(tmp_path):
    old_root, new_root = tmp_path / "old", tmp_path / "new"
    df = pd.DataFrame({"stock_id": ["1101"], "date": ["2026-01-02"], "close": [10.0]})
    _write_dataset(old_root, "price_valuation", {"1101": df})
    _write_dataset(new_root, "price_valuation", {"1101": df.copy()})

    result = diffmod.diff_dataset("price_valuation", PV_SPEC, old_root, new_root)
    assert result["structural_status"] == "PASS"
    assert result["value_status"] == "EXACT_PASS"
    assert result["overall_status"] == "EXACT_PASS"
    assert result["missing_keys_count"] == 0
    assert result["columns"]["close"]["n_exact_equal"] == 1
    assert result["columns"]["close"]["n_value_mismatch"] == 0


def test_diff_n_both_null_counted_separately(tmp_path):
    old_root, new_root = tmp_path / "old", tmp_path / "new"
    df = pd.DataFrame({"stock_id": ["1101"], "date": ["2026-01-02"], "close": [np.nan]})
    _write_dataset(old_root, "price_valuation", {"1101": df})
    _write_dataset(new_root, "price_valuation", {"1101": df.copy()})

    result = diffmod.diff_dataset("price_valuation", PV_SPEC, old_root, new_root)
    cr = result["columns"]["close"]
    assert cr["n_both_null"] == 1
    assert cr["n_null_mismatch"] == 0
    assert cr["n_exact_equal"] == 0          # 兩邊都 null,不算「兩邊都有值且相等」
    assert cr["n_value_mismatch"] == 0
    assert result["value_status"] == "EXACT_PASS", "兩邊都 null 語意上是相等,不該觸發 DIFF_UNRESOLVED"


def test_diff_detects_dtype_drift_not_masked_by_stringification(tmp_path):
    """old 是數值 dtype、new 被讀成字串 dtype,但字串化後剛好長得一樣
    ("10.0" == "10.0")。Round 5 review 要求這種 dtype 不相容不能被字串比較掩蓋
    掉,必須直接算 mismatch 並標記 dtype_status=INCOMPATIBLE。"""
    old_root, new_root = tmp_path / "old", tmp_path / "new"
    old_df = pd.DataFrame({"stock_id": ["1101"], "date": ["2026-01-02"], "close": [10.0]})
    new_df = pd.DataFrame({"stock_id": ["1101"], "date": ["2026-01-02"], "close": ["10.0"]})
    _write_dataset(old_root, "price_valuation", {"1101": old_df})
    _write_dataset(new_root, "price_valuation", {"1101": new_df})

    result = diffmod.diff_dataset("price_valuation", PV_SPEC, old_root, new_root)
    cr = result["columns"]["close"]
    assert cr["dtype_status"] == "INCOMPATIBLE"
    assert cr["n_value_mismatch"] == 1
    assert cr["n_exact_equal"] == 0
    assert result["value_status"] == "DIFF_UNRESOLVED"
    assert result["overall_status"] == "REVIEW_REQUIRED"


def test_diff_same_dtype_reports_dtype_status_ok(tmp_path):
    old_root, new_root = tmp_path / "old", tmp_path / "new"
    df = pd.DataFrame({"stock_id": ["1101"], "date": ["2026-01-02"], "close": [10.0]})
    _write_dataset(old_root, "price_valuation", {"1101": df})
    _write_dataset(new_root, "price_valuation", {"1101": df.copy()})

    result = diffmod.diff_dataset("price_valuation", PV_SPEC, old_root, new_root)
    assert result["columns"]["close"]["dtype_status"] == "OK"


def test_diff_duplicate_keys_block_merge_not_cross_join(tmp_path):
    """old 側同一個 (stock_id, date) 出現兩次、值不同——如果沒先擋下來就直接
    merge,pandas 會 cross join 出 2 列,把統計數字弄假。這裡驗證:一偵測到
    重複鍵就整個 dataset 判 structural FAIL,而且完全不做數值比對
    (value_status=SKIPPED_DUE_TO_STRUCTURAL_FAIL),不是硬著頭皮 merge。"""
    old_root, new_root = tmp_path / "old", tmp_path / "new"
    old_df = pd.DataFrame({"stock_id": ["1101", "1101"], "date": ["2026-01-02", "2026-01-02"],
                            "close": [10.0, 99.0]})   # 同 key 兩列、值不同
    new_df = pd.DataFrame({"stock_id": ["1101"], "date": ["2026-01-02"], "close": [10.0]})
    _write_dataset(old_root, "price_valuation", {"1101": old_df})
    _write_dataset(new_root, "price_valuation", {"1101": new_df})

    result = diffmod.diff_dataset("price_valuation", PV_SPEC, old_root, new_root)
    assert result["old_duplicate_key_count"] == 1
    assert "old_duplicate_keys" in result["structural_fail_reasons"]
    assert result["structural_status"] == "FAIL"
    assert result["value_status"] == "SKIPPED_DUE_TO_STRUCTURAL_FAIL"
    assert result["columns"] == {}, "重複鍵時完全不該跑數值比對"


def test_diff_new_duplicate_keys_also_block_merge(tmp_path):
    old_root, new_root = tmp_path / "old", tmp_path / "new"
    old_df = pd.DataFrame({"stock_id": ["1101"], "date": ["2026-01-02"], "close": [10.0]})
    new_df = pd.DataFrame({"stock_id": ["1101", "1101"], "date": ["2026-01-02", "2026-01-02"],
                            "close": [10.0, 20.0]})
    _write_dataset(old_root, "price_valuation", {"1101": old_df})
    _write_dataset(new_root, "price_valuation", {"1101": new_df})

    result = diffmod.diff_dataset("price_valuation", PV_SPEC, old_root, new_root)
    assert result["new_duplicate_key_count"] == 1
    assert "new_duplicate_keys" in result["structural_fail_reasons"]
    assert result["structural_status"] == "FAIL"


def test_diff_industry_map_static_dataset(tmp_path):
    """industry_map 是靜態單檔、key=stock_id (無 date),Round 4 review 要求它
    納入跟其他 10 個 dataset 同一套框架,不是 main() 裡另外簡陋比對。"""
    old_root, new_root = tmp_path / "old", tmp_path / "new"
    old_df = pd.DataFrame({"stock_id": ["1101", "1102"], "tse_ind_code": ["M1100", "M1100"]})
    new_df = pd.DataFrame({"stock_id": ["1101", "1102"], "tse_ind_code": ["M1100", "M9999"]})
    _write_static_dataset(old_root / "industry_map_parent", "industry_map", old_df)
    _write_static_dataset(new_root / "industry_map_parent", "industry_map", new_df)

    result = diffmod.diff_dataset("industry_map", IM_SPEC,
                                   old_root / "industry_map_parent", new_root / "industry_map_parent")
    assert result["structural_status"] == "PASS"
    assert result["value_status"] == "DIFF_UNRESOLVED"
    assert result["columns"]["tse_ind_code"]["n_value_mismatch"] == 1


def test_diff_receipt_is_exclusive_create_collision_safe(tmp_path, monkeypatch):
    """receipt 用 open(...,"x") 排他建立,撞名要直接炸掉,不能靜默覆寫掉舊 receipt。"""
    monkeypatch.setattr(diffmod, "RECEIPT_DIR", tmp_path)
    receipt = {"dummy": True}
    p1 = diffmod._write_receipt(receipt)
    assert p1.exists()
    # 模擬撞名:直接用同一個路徑再開一次 "x" 模式
    with pytest.raises(FileExistsError):
        with open(p1, "x", encoding="utf-8") as f:
            f.write("{}")


def test_diff_receipt_includes_input_file_hashes(tmp_path):
    old_root, new_root = tmp_path / "old", tmp_path / "new"
    df = pd.DataFrame({"stock_id": ["1101"], "date": ["2026-01-02"], "close": [10.0]})
    _write_dataset(old_root, "price_valuation", {"1101": df})
    _write_dataset(new_root, "price_valuation", {"1101": df.copy()})

    result = diffmod.diff_dataset("price_valuation", PV_SPEC, old_root, new_root)
    assert len(result["input_files"]["old"]) == 1
    assert len(result["input_files"]["new"]) == 1
    assert result["input_files"]["old"][0]["sha256"]
    assert result["input_files"]["new"][0]["sha256"]


# =============================================================================
# 8. tej_importer._verify_supplement —— 消費 legacy supplement 前的完整驗證
# =============================================================================

def _valid_dedup_stat(stage: str) -> dict:
    return {
        "stage": stage,
        "checked_columns": ["stock_id", "date"],
        "n_duplicate_key_rows": 0,
        "n_exact_duplicate_rows_removed": 0,
        "n_conflicting_keys": 0,
    }


def _valid_dedup(name: str) -> dict:
    """組出一份會通過 `_validate_dedup_metadata` 的 dedup 結構 (Round 7)。
    recurring_net_income 有兩個窗口來源 + cross_window_overlap,其餘 supplement
    只有一個同名來源。"""
    if name == "recurring_net_income":
        pair = [_valid_dedup_stat("raw_source"), _valid_dedup_stat("projected")]
        return {
            "sources": {
                "recurring_net_income_2005_2018": pair,
                "recurring_net_income_2019plus": [_valid_dedup_stat("raw_source"),
                                                    _valid_dedup_stat("projected")],
            },
            "cross_window_overlap": {"n_overlap_keys": 0, "n_overlap_identical": 0,
                                      "n_overlap_conflicting": 0},
        }
    return {"sources": {name: [_valid_dedup_stat("raw_source"), _valid_dedup_stat("projected")]}}


def _write_supplement(supp_dir: Path, name: str, df: pd.DataFrame, receipt_overrides: dict = None):
    """建 receipt 時用 tej_importer._profile_supplement 算完整統計 (Round 5:消費端
    要重新計算並比對這些欄位,receipt 裡必須真的有,不然測試沒有測到這道檢查),
    並附上一份會通過結構驗證的 `dedup` (Round 7)。"""
    supp_dir.mkdir(parents=True, exist_ok=True)
    p = supp_dir / f"{name}.parquet"
    df.to_parquet(p, index=False)
    profile = tej_importer._profile_supplement(df)
    receipt = {
        "script_sha256": _sha256_of(REPO_ROOT / "scripts" / "extract_legacy_supplement.py"),
        "overall_status": "PASS",
        "outputs": {name: {"sha256": _sha256_of(p), **profile, "dedup": _valid_dedup(name)}},
    }
    if receipt_overrides:
        receipt.update(receipt_overrides)
    with open(supp_dir / "receipt.json", "w", encoding="utf-8") as f:
        json.dump(receipt, f)
    return p


def test_verify_supplement_passes_when_everything_matches(tmp_path):
    df = pd.DataFrame({"stock_id": ["1101"], "date": ["2019-03-01"], "roe_after_tax": [1.0]})
    p = _write_supplement(tmp_path / "supp", "roe_after_tax", df)
    out = tej_importer._verify_supplement(p, "fundamentals_quarterly")
    pd.testing.assert_frame_equal(out.reset_index(drop=True), df)


def test_verify_supplement_missing_receipt_raises(tmp_path):
    supp_dir = tmp_path / "supp"
    supp_dir.mkdir()
    df = pd.DataFrame({"stock_id": ["1101"], "date": ["2019-03-01"], "roe_after_tax": [1.0]})
    df.to_parquet(supp_dir / "roe_after_tax.parquet", index=False)
    with pytest.raises(FileNotFoundError, match="receipt"):
        tej_importer._verify_supplement(supp_dir / "roe_after_tax.parquet", "fundamentals_quarterly")


def test_verify_supplement_non_pass_receipt_raises(tmp_path):
    df = pd.DataFrame({"stock_id": ["1101"], "date": ["2019-03-01"], "roe_after_tax": [1.0]})
    p = _write_supplement(tmp_path / "supp", "roe_after_tax", df, {"overall_status": "FAIL"})
    with pytest.raises(ValueError, match="PASS"):
        tej_importer._verify_supplement(p, "fundamentals_quarterly")


def test_verify_supplement_tampered_parquet_raises(tmp_path):
    df = pd.DataFrame({"stock_id": ["1101"], "date": ["2019-03-01"], "roe_after_tax": [1.0]})
    p = _write_supplement(tmp_path / "supp", "roe_after_tax", df)
    # receipt 寫完之後,parquet 內容被動過手腳 (雜湊對不上了)
    df2 = pd.DataFrame({"stock_id": ["1101"], "date": ["2019-03-01"], "roe_after_tax": [999.0]})
    df2.to_parquet(p, index=False)
    with pytest.raises(ValueError, match="SHA-256"):
        tej_importer._verify_supplement(p, "fundamentals_quarterly")


def test_verify_supplement_tampered_script_raises(tmp_path, monkeypatch):
    df = pd.DataFrame({"stock_id": ["1101"], "date": ["2019-03-01"], "roe_after_tax": [1.0]})
    p = _write_supplement(tmp_path / "supp", "roe_after_tax", df,
                           {"script_sha256": "0" * 64})   # 假雜湊,一定跟現在的腳本對不上
    with pytest.raises(ValueError, match="extract_legacy_supplement.py"):
        tej_importer._verify_supplement(p, "fundamentals_quarterly")


def test_verify_supplement_schema_mismatch_raises(tmp_path):
    df = pd.DataFrame({"stock_id": ["1101"], "date": ["2019-03-01"], "roe_after_tax": [1.0]})
    p = _write_supplement(tmp_path / "supp", "roe_after_tax", df)
    # receipt 記錄的 schema 跟實際 parquet 的 schema 對不上 (例如腳本改過欄位順序沒同步)
    receipt_path = p.parent / "receipt.json"
    with open(receipt_path, encoding="utf-8") as f:
        receipt = json.load(f)
    receipt["outputs"]["roe_after_tax"]["schema"] = ["stock_id", "date", "other_col"]
    with open(receipt_path, "w", encoding="utf-8") as f:
        json.dump(receipt, f)
    with pytest.raises(ValueError, match="schema"):
        tej_importer._verify_supplement(p, "fundamentals_quarterly")


def test_verify_supplement_duplicate_key_raises(tmp_path):
    df = pd.DataFrame({"stock_id": ["1101", "1101"], "date": ["2019-03-01", "2019-03-01"],
                        "roe_after_tax": [1.0, 2.0]})
    p = _write_supplement(tmp_path / "supp", "roe_after_tax", df)
    with pytest.raises(ValueError, match="重複"):
        tej_importer._verify_supplement(p, "fundamentals_quarterly")


def test_verify_supplement_null_key_raises(tmp_path):
    df = pd.DataFrame({"stock_id": ["1101", None], "date": ["2019-03-01", "2019-04-01"],
                        "roe_after_tax": [1.0, 2.0]})
    p = _write_supplement(tmp_path / "supp", "roe_after_tax", df)
    with pytest.raises(ValueError, match="null"):
        tej_importer._verify_supplement(p, "fundamentals_quarterly")


def test_verify_supplement_missing_script_raises(tmp_path, monkeypatch):
    """Round 5 review:原本只在腳本存在時才比對雜湊,腳本被誤刪/搬移時完全跳過
    這道驗證。現在腳本不存在也要 raise,不能因為「檔案不見了」就放行。"""
    df = pd.DataFrame({"stock_id": ["1101"], "date": ["2019-03-01"], "roe_after_tax": [1.0]})
    p = _write_supplement(tmp_path / "supp", "roe_after_tax", df)
    monkeypatch.setattr(tej_importer, "LEGACY_SUPPLEMENT_SCRIPT", tmp_path / "nonexistent_script.py")
    with pytest.raises(FileNotFoundError, match="抽取腳本"):
        tej_importer._verify_supplement(p, "fundamentals_quarterly")


def test_verify_supplement_schema_must_match_code_defined_not_just_self_consistent_receipt(tmp_path):
    """receipt 的 schema 欄位是抽取腳本自己寫的,可能跟實際 parquet 自洽 (沒被
    竄改) 但仍然跟這支消費端程式碼凍結的 SUPPLEMENT_SCHEMAS 不符 (例如多了一個
    非預期欄位)。receipt 自洽不能取代程式碼凍結的 schema 檢查。"""
    df = pd.DataFrame({"stock_id": ["1101"], "date": ["2019-03-01"], "roe_after_tax": [1.0],
                        "unexpected_extra_col": [999]})
    p = _write_supplement(tmp_path / "supp", "roe_after_tax", df)   # receipt schema 跟這份 df 自洽
    with pytest.raises(ValueError, match="程式碼凍結"):
        tej_importer._verify_supplement(p, "fundamentals_quarterly")


def test_verify_supplement_recomputed_row_count_mismatch_raises(tmp_path):
    """receipt 裡的 row_count 被竄改成跟實際內容不符的假數字 (SHA-256 沒動,因為
    parquet 本身沒被改)——消費端必須重新計算一次比對,不能只信 receipt 寫的數字。"""
    df = pd.DataFrame({"stock_id": ["1101", "1102"], "date": ["2019-03-01", "2019-04-01"],
                        "roe_after_tax": [1.0, 2.0]})
    p = _write_supplement(tmp_path / "supp", "roe_after_tax", df)
    receipt_path = p.parent / "receipt.json"
    with open(receipt_path, encoding="utf-8") as f:
        receipt = json.load(f)
    receipt["outputs"]["roe_after_tax"]["row_count"] = 999
    with open(receipt_path, "w", encoding="utf-8") as f:
        json.dump(receipt, f)
    with pytest.raises(ValueError, match="重新計算"):
        tej_importer._verify_supplement(p, "fundamentals_quarterly")


def test_profile_implementations_agree_across_producer_and_consumer(tmp_path):
    """receipt 由 extract_legacy_supplement._profile 產生,消費端用
    tej_importer._profile_supplement 重算後逐欄比對——這是兩份各自獨立的實作,
    只要有一邊改了欄位或算法沒同步,正常的 supplement 就會被誤判成竄改。這個
    測試把兩者的耦合鎖住 (含 parquet round-trip:receipt 是寫檔前算的,驗證是
    讀檔後算的,dtype 必須撐得過來回一趟)。"""
    df = pd.DataFrame({"stock_id": ["1101", "1102", "1102"],
                        "date": ["2019-03-01", "2019-03-01", "2019-06-01"],
                        "roe_after_tax": [1.5, np.nan, 2.0]})
    producer = extract_legacy_supplement._profile(df, key_cols=["stock_id", "date"])
    p = tmp_path / "roe_after_tax.parquet"
    df.to_parquet(p, index=False)
    consumer = tej_importer._profile_supplement(pd.read_parquet(p))
    for field in tej_importer.REQUIRED_RECEIPT_PROFILE_FIELDS:
        assert producer[field] == consumer[field], f"{field} 兩邊算出來不一致"
    assert producer["schema"] == consumer["schema"]


@pytest.mark.parametrize("field", list(tej_importer.REQUIRED_RECEIPT_OUTPUT_FIELDS))
def test_verify_supplement_missing_receipt_output_field_raises(tmp_path, field):
    """Round 6 review:receipt 少一個欄位不能讓對應的檢查靜默空轉。原本每道比對
    都是 `if receipt.get(f) is not None and ...`,把 row_count 那行從 json 刪掉,
    重算比對就形同不存在——缺欄跟值不符一樣危險,兩者都要 raise。"""
    df = pd.DataFrame({"stock_id": ["1101"], "date": ["2019-03-01"], "roe_after_tax": [1.0]})
    p = _write_supplement(tmp_path / "supp", "roe_after_tax", df)
    receipt_path = p.parent / "receipt.json"
    with open(receipt_path, encoding="utf-8") as f:
        receipt = json.load(f)
    del receipt["outputs"]["roe_after_tax"][field]
    with open(receipt_path, "w", encoding="utf-8") as f:
        json.dump(receipt, f)
    with pytest.raises(ValueError, match="缺少必要欄位"):
        tej_importer._verify_supplement(p, "fundamentals_quarterly")


@pytest.mark.parametrize("field", list(tej_importer.REQUIRED_RECEIPT_TOP_FIELDS))
def test_verify_supplement_missing_receipt_top_field_raises(tmp_path, field):
    df = pd.DataFrame({"stock_id": ["1101"], "date": ["2019-03-01"], "roe_after_tax": [1.0]})
    p = _write_supplement(tmp_path / "supp", "roe_after_tax", df)
    receipt_path = p.parent / "receipt.json"
    with open(receipt_path, encoding="utf-8") as f:
        receipt = json.load(f)
    del receipt[field]
    with open(receipt_path, "w", encoding="utf-8") as f:
        json.dump(receipt, f)
    # overall_status 缺席會先被既有的 PASS 檢查擋下 (訊息不同),其餘走缺欄檢查;
    # 兩種都必須 raise,不能放行。
    with pytest.raises(ValueError):
        tej_importer._verify_supplement(p, "fundamentals_quarterly")


def test_verify_supplement_null_receipt_field_still_compared(tmp_path):
    """欄位存在但值是 null:不能因為「是 None」就跳過比對 (原本的 is not None
    守衛正是這樣被繞過的),要跟重算結果比並 raise。"""
    df = pd.DataFrame({"stock_id": ["1101"], "date": ["2019-03-01"], "roe_after_tax": [1.0]})
    p = _write_supplement(tmp_path / "supp", "roe_after_tax", df)
    receipt_path = p.parent / "receipt.json"
    with open(receipt_path, encoding="utf-8") as f:
        receipt = json.load(f)
    receipt["outputs"]["roe_after_tax"]["row_count"] = None
    with open(receipt_path, "w", encoding="utf-8") as f:
        json.dump(receipt, f)
    with pytest.raises(ValueError, match="重新計算"):
        tej_importer._verify_supplement(p, "fundamentals_quarterly")


def test_verify_supplement_recomputed_stock_count_mismatch_raises(tmp_path):
    df = pd.DataFrame({"stock_id": ["1101", "1102"], "date": ["2019-03-01", "2019-04-01"],
                        "roe_after_tax": [1.0, 2.0]})
    p = _write_supplement(tmp_path / "supp", "roe_after_tax", df)
    receipt_path = p.parent / "receipt.json"
    with open(receipt_path, encoding="utf-8") as f:
        receipt = json.load(f)
    receipt["outputs"]["roe_after_tax"]["stock_count"] = 999
    with open(receipt_path, "w", encoding="utf-8") as f:
        json.dump(receipt, f)
    with pytest.raises(ValueError, match="重新計算"):
        tej_importer._verify_supplement(p, "fundamentals_quarterly")


# =============================================================================
# 8b. dedup 巢狀結構驗證 (Round 7 review)
# =============================================================================

def _rewrite_dedup(p: Path, name: str, mutate) -> None:
    """讀回 receipt,對 outputs[name]["dedup"] 套用 `mutate` 後寫回。"""
    receipt_path = p.parent / "receipt.json"
    with open(receipt_path, encoding="utf-8") as f:
        receipt = json.load(f)
    mutate(receipt["outputs"][name]["dedup"])
    with open(receipt_path, "w", encoding="utf-8") as f:
        json.dump(receipt, f)


def test_verify_supplement_dedup_missing_sources_raises(tmp_path):
    df = pd.DataFrame({"stock_id": ["1101"], "date": ["2019-03-01"], "roe_after_tax": [1.0]})
    p = _write_supplement(tmp_path / "supp", "roe_after_tax", df)
    _rewrite_dedup(p, "roe_after_tax", lambda d: d.pop("sources"))
    with pytest.raises(ValueError, match="dedup.sources"):
        tej_importer._verify_supplement(p, "fundamentals_quarterly")


def test_verify_supplement_dedup_source_entry_not_pair_raises(tmp_path):
    """只有一份統計 (缺 projected 或缺 raw_source) 必須被抓到。"""
    df = pd.DataFrame({"stock_id": ["1101"], "date": ["2019-03-01"], "roe_after_tax": [1.0]})
    p = _write_supplement(tmp_path / "supp", "roe_after_tax", df)

    def _drop_one(d):
        d["sources"]["roe_after_tax"] = d["sources"]["roe_after_tax"][:1]
    _rewrite_dedup(p, "roe_after_tax", _drop_one)
    with pytest.raises(ValueError, match="恰好是兩份統計"):
        tej_importer._verify_supplement(p, "fundamentals_quarterly")


def test_verify_supplement_dedup_stat_missing_field_raises(tmp_path):
    df = pd.DataFrame({"stock_id": ["1101"], "date": ["2019-03-01"], "roe_after_tax": [1.0]})
    p = _write_supplement(tmp_path / "supp", "roe_after_tax", df)

    def _drop_field(d):
        del d["sources"]["roe_after_tax"][0]["n_conflicting_keys"]
    _rewrite_dedup(p, "roe_after_tax", _drop_field)
    with pytest.raises(ValueError, match="缺少必要欄位"):
        tej_importer._verify_supplement(p, "fundamentals_quarterly")


def test_verify_supplement_dedup_stat_wrong_type_raises(tmp_path):
    df = pd.DataFrame({"stock_id": ["1101"], "date": ["2019-03-01"], "roe_after_tax": [1.0]})
    p = _write_supplement(tmp_path / "supp", "roe_after_tax", df)

    def _wrong_type(d):
        d["sources"]["roe_after_tax"][0]["n_conflicting_keys"] = "0"    # 字串,不是 int
    _rewrite_dedup(p, "roe_after_tax", _wrong_type)
    with pytest.raises(ValueError, match="型別應為"):
        tej_importer._verify_supplement(p, "fundamentals_quarterly")


def test_verify_supplement_dedup_stat_bool_rejected_as_int(tmp_path):
    """bool 是 int 的子類別,但混進計數欄位是型別錯誤,不能被 isinstance(x, int) 誤判通過。"""
    df = pd.DataFrame({"stock_id": ["1101"], "date": ["2019-03-01"], "roe_after_tax": [1.0]})
    p = _write_supplement(tmp_path / "supp", "roe_after_tax", df)

    def _bool_count(d):
        d["sources"]["roe_after_tax"][0]["n_duplicate_key_rows"] = False
    _rewrite_dedup(p, "roe_after_tax", _bool_count)
    with pytest.raises(ValueError, match="型別應為"):
        tej_importer._verify_supplement(p, "fundamentals_quarterly")


def test_verify_supplement_dedup_negative_count_raises(tmp_path):
    df = pd.DataFrame({"stock_id": ["1101"], "date": ["2019-03-01"], "roe_after_tax": [1.0]})
    p = _write_supplement(tmp_path / "supp", "roe_after_tax", df)

    def _negative(d):
        d["sources"]["roe_after_tax"][0]["n_duplicate_key_rows"] = -1
    _rewrite_dedup(p, "roe_after_tax", _negative)
    with pytest.raises(ValueError, match="負數"):
        tej_importer._verify_supplement(p, "fundamentals_quarterly")


def test_verify_supplement_dedup_duplicate_stage_label_raises(tmp_path):
    """兩份統計都標 raw_source (漏了 projected)——stage 標籤組合不對要 raise。"""
    df = pd.DataFrame({"stock_id": ["1101"], "date": ["2019-03-01"], "roe_after_tax": [1.0]})
    p = _write_supplement(tmp_path / "supp", "roe_after_tax", df)

    def _dup_stage(d):
        d["sources"]["roe_after_tax"][1]["stage"] = "raw_source"
    _rewrite_dedup(p, "roe_after_tax", _dup_stage)
    with pytest.raises(ValueError, match="stage 標籤組合"):
        tej_importer._verify_supplement(p, "fundamentals_quarterly")


def test_verify_supplement_dedup_nonzero_conflicting_keys_raises(tmp_path):
    """正常的抽取邏輯偵測到衝突時已經 raise 了,receipt 裡不該出現非零衝突數——
    出現就代表 receipt 被竄改或抽取邏輯有漏洞,一律 raise。"""
    df = pd.DataFrame({"stock_id": ["1101"], "date": ["2019-03-01"], "roe_after_tax": [1.0]})
    p = _write_supplement(tmp_path / "supp", "roe_after_tax", df)

    def _conflict(d):
        d["sources"]["roe_after_tax"][0]["n_conflicting_keys"] = 1
    _rewrite_dedup(p, "roe_after_tax", _conflict)
    with pytest.raises(ValueError, match="n_conflicting_keys"):
        tej_importer._verify_supplement(p, "fundamentals_quarterly")


def test_verify_supplement_dedup_recurring_net_income_missing_overlap_raises(tmp_path):
    df = pd.DataFrame({"stock_id": ["1101"], "date": ["2019-03-01"], "recurring_net_income": [1.0]})
    p = _write_supplement(tmp_path / "supp", "recurring_net_income", df)
    _rewrite_dedup(p, "recurring_net_income", lambda d: d.pop("cross_window_overlap"))
    with pytest.raises(ValueError, match="cross_window_overlap"):
        tej_importer._verify_supplement(p, "fundamentals_quarterly")


def test_verify_supplement_dedup_recurring_net_income_nonzero_overlap_conflict_raises(tmp_path):
    df = pd.DataFrame({"stock_id": ["1101"], "date": ["2019-03-01"], "recurring_net_income": [1.0]})
    p = _write_supplement(tmp_path / "supp", "recurring_net_income", df)

    def _overlap_conflict(d):
        d["cross_window_overlap"]["n_overlap_conflicting"] = 2
    _rewrite_dedup(p, "recurring_net_income", _overlap_conflict)
    with pytest.raises(ValueError, match="n_overlap_conflicting"):
        tej_importer._verify_supplement(p, "fundamentals_quarterly")


def test_verify_supplement_dedup_valid_structure_passes(tmp_path):
    """健全性測試:_valid_dedup 產生的結構本身要能通過,免得上面所有「要 raise」
    的測試其實是被別的原因擋下來的假陽性。"""
    df = pd.DataFrame({"stock_id": ["1101"], "date": ["2019-03-01"], "recurring_net_income": [1.0]})
    p = _write_supplement(tmp_path / "supp", "recurring_net_income", df)
    out = tej_importer._verify_supplement(p, "financial_statements")
    pd.testing.assert_frame_equal(out.reset_index(drop=True), df)


# ---- Round 8 A:凍結的來源名字集合 + overlap 算術恆等式 ----

def test_verify_supplement_dedup_missing_source_name_raises(tmp_path):
    """recurring_net_income 應該恰好有兩個來源條目,漏掉一個 (改名/被刪) 要 raise。"""
    df = pd.DataFrame({"stock_id": ["1101"], "date": ["2019-03-01"], "recurring_net_income": [1.0]})
    p = _write_supplement(tmp_path / "supp", "recurring_net_income", df)

    def _drop_source(d):
        del d["sources"]["recurring_net_income_2019plus"]
    _rewrite_dedup(p, "recurring_net_income", _drop_source)
    with pytest.raises(ValueError, match="來源名字集合"):
        tej_importer._verify_supplement(p, "financial_statements")


def test_verify_supplement_dedup_extra_source_name_raises(tmp_path):
    df = pd.DataFrame({"stock_id": ["1101"], "date": ["2019-03-01"], "roe_after_tax": [1.0]})
    p = _write_supplement(tmp_path / "supp", "roe_after_tax", df)

    def _add_source(d):
        d["sources"]["unexpected_extra_source"] = [_valid_dedup_stat("raw_source"),
                                                     _valid_dedup_stat("projected")]
    _rewrite_dedup(p, "roe_after_tax", _add_source)
    with pytest.raises(ValueError, match="來源名字集合"):
        tej_importer._verify_supplement(p, "fundamentals_quarterly")


def test_verify_supplement_dedup_renamed_source_raises(tmp_path):
    """來源名字被改掉 (例如打錯字/腳本改了命名沒同步)——集合對不上,一樣要 raise。"""
    df = pd.DataFrame({"stock_id": ["1101"], "date": ["2019-03-01"], "revenue_last_year": [1.0],
                        "cum_revenue_last_year": [2.0]})
    p = _write_supplement(tmp_path / "supp", "revenue_last_year", df)

    def _rename_source(d):
        d["sources"]["revenue_last_yr"] = d["sources"].pop("revenue_last_year")
    _rewrite_dedup(p, "revenue_last_year", _rename_source)
    with pytest.raises(ValueError, match="來源名字集合"):
        tej_importer._verify_supplement(p, "monthly_revenue")


def test_verify_supplement_dedup_overlap_arithmetic_mismatch_raises(tmp_path):
    """n_overlap_keys 應該等於 n_overlap_identical + n_overlap_conflicting——三個
    數字各自型別對、非負,不代表加總對得起來 (例如漏算了幾個重疊 key)。"""
    df = pd.DataFrame({"stock_id": ["1101"], "date": ["2019-03-01"], "recurring_net_income": [1.0]})
    p = _write_supplement(tmp_path / "supp", "recurring_net_income", df)

    def _bad_arithmetic(d):
        d["cross_window_overlap"]["n_overlap_keys"] = 5   # identical=0, conflicting=0, 但 keys=5
    _rewrite_dedup(p, "recurring_net_income", _bad_arithmetic)
    with pytest.raises(ValueError, match="算術對不起來"):
        tej_importer._verify_supplement(p, "financial_statements")


def test_load_source_raises_on_supplement_row_multiplication(patched_root, monkeypatch):
    """supplement 如果 (理論上不該發生,前面幾道檢查也會擋,這裡測的是最後一道
    防線) 帶有重複鍵,merge 後 combined 的列數會膨脹 (pandas left join 對重複鍵
    做 fan-out)——load_source 要抓到這個並 raise,不能讓多出來的列悄悄流進輸出。
    直接 monkeypatch _verify_supplement 回傳一份帶重複鍵的 supplement,模擬
    「前面的防線萬一失守」的情境,驗證合併後的列數檢查有獨立擋住。"""
    data_root, manifest_csv = patched_root
    src_dir = _make_source_dir(data_root, "cat", {"a.xlsx": b"dummy"})
    files = sorted(src_dir.glob("*.xlsx"))
    entries = {f.relative_to(data_root).as_posix(): _sha256_of(f) for f in files}
    _write_manifest(manifest_csv, entries)

    combined = pd.DataFrame({"stock_id": ["1101"], "stock_name": ["X"], "date": ["2026-01-02"],
                              "close": [10.0]})
    bad_supp = pd.DataFrame({"stock_id": ["1101", "1101"], "date": ["2026-01-02", "2026-01-02"],
                              "extra_col": [1.0, 2.0]})   # 重複鍵,left join 會 fan-out 成 2 列

    fake_supp_path = data_root / "fake_supplement.parquet"
    fake_supp_path.touch()
    spec = {
        "source_dir": src_dir, "date_col": "年月日",
        "required_cols": [], "rename": {}, "thousand_cols": {}, "numeric_cols": [],
        "supplement": fake_supp_path,
    }
    monkeypatch.setitem(tej_importer.DATASETS, "_test_fanout_dataset", spec)
    monkeypatch.setattr(tej_importer, "_load_one", lambda f, s: combined.copy())
    monkeypatch.setattr(tej_importer, "_verify_supplement", lambda path, ds: bad_supp)

    with pytest.raises(RuntimeError, match="fan-out|膨脹"):
        tej_importer.load_source("_test_fanout_dataset")


# =============================================================================
# 9. extract_legacy_supplement —— check_source_keys / enforce_output_spec /
#    publish_staging (Round 4 review)
# =============================================================================

def test_check_source_keys_raises_on_invalid_stock_id():
    df = pd.DataFrame({"stock_id": ["1101", "nan"], "date": ["2019-03-01", "2019-04-01"]})
    with pytest.raises(ValueError, match="stock_id 無效"):
        extract_legacy_supplement.check_source_keys(df, "dummy_source")


def test_check_source_keys_raises_on_invalid_date():
    df = pd.DataFrame({"stock_id": ["1101", "1102"], "date": ["2019-03-01", np.nan]})
    with pytest.raises(ValueError, match="date 無法解析"):
        extract_legacy_supplement.check_source_keys(df, "dummy_source")


def test_check_source_keys_passes_clean_data():
    df = pd.DataFrame({"stock_id": ["1101", "1102"], "date": ["2019-03-01", "2019-04-01"]})
    extract_legacy_supplement.check_source_keys(df, "dummy_source")   # 不 raise 就算過


def test_dedupe_or_raise_removes_exact_duplicates_and_reports_stats():
    df = pd.DataFrame({"stock_id": ["1101", "1101", "1102"],
                        "date": ["2019-03-01", "2019-03-01", "2019-03-01"],
                        "roe_after_tax": [1.5, 1.5, 2.0]})
    out, stats = extract_legacy_supplement._dedupe_or_raise(df, ["stock_id", "date"], "roe_after_tax")
    assert len(out) == 2
    assert stats["n_exact_duplicate_rows_removed"] == 1
    assert stats["n_duplicate_key_rows"] == 2
    assert stats["n_conflicting_keys"] == 0
    assert stats["stage"] == "projected"


def test_dedupe_or_raise_raises_on_conflicting_values():
    df = pd.DataFrame({"stock_id": ["1101", "1101"], "date": ["2019-03-01", "2019-03-01"],
                        "roe_after_tax": [1.5, 9.9]})
    with pytest.raises(ValueError, match="數值不一致"):
        extract_legacy_supplement._dedupe_or_raise(df, ["stock_id", "date"], "roe_after_tax")


def test_dedupe_or_raise_treats_null_vs_value_as_conflict():
    """一邊有值一邊 NaN 不能被當成「完全重複」安全去重,nunique(dropna=False) 要
    把這種不對稱也算進「不只一種值」。"""
    df = pd.DataFrame({"stock_id": ["1101", "1101"], "date": ["2019-03-01", "2019-03-01"],
                        "roe_after_tax": [1.5, np.nan]})
    with pytest.raises(ValueError, match="數值不一致"):
        extract_legacy_supplement._dedupe_or_raise(df, ["stock_id", "date"], "roe_after_tax")


def test_dedupe_or_raise_passthrough_when_no_duplicates():
    df = pd.DataFrame({"stock_id": ["1101", "1102"], "date": ["2019-03-01", "2019-03-01"],
                        "roe_after_tax": [1.5, 2.0]})
    out, stats = extract_legacy_supplement._dedupe_or_raise(df, ["stock_id", "date"], "roe_after_tax")
    pd.testing.assert_frame_equal(out.reset_index(drop=True), df.reset_index(drop=True))
    assert stats["n_duplicate_key_rows"] == 0
    assert stats["n_exact_duplicate_rows_removed"] == 0


# ---- Round 6:投影前的原始來源層衝突檢查 ----

def test_check_source_duplicates_catches_conflict_in_non_extracted_column():
    """核心的 Round 6 缺口:同一個 (stock_id, date) 在來源檔出現兩次,**目標欄位
    (roe_after_tax) 剛好相同**,但來源檔其他欄位 (net_income) 互相矛盾。投影成
    supplement 三欄之後,這兩列會長得完全一樣、被當成「安全的完全重複」去掉,
    來源檔自相矛盾這件事就永遠沒有人看見。原始來源層必須先擋下來。"""
    raw = pd.DataFrame({
        "stock_id": ["1101", "1101"], "date": ["2019-03-01", "2019-03-01"],
        "net_income": [100.0, 999.0],      # 來源檔內部矛盾,但不是我們要抽的欄位
        "roe_after_tax": [1.5, 1.5],       # 目標欄位剛好一致
    })
    # 投影後看起來是無害的完全重複 —— 舊邏輯就是在這裡放行的
    projected = raw[["stock_id", "date", "roe_after_tax"]]
    out, proj_stats = extract_legacy_supplement._dedupe_or_raise(
        projected, ["stock_id", "date"], "roe_after_tax")
    assert len(out) == 1 and proj_stats["n_conflicting_keys"] == 0

    # 原始來源層要抓到它
    with pytest.raises(ValueError, match="數值不一致"):
        extract_legacy_supplement.check_source_duplicates(raw, ["stock_id", "date"], "roe_after_tax")


def test_check_source_duplicates_reports_stage_and_checked_columns():
    raw = pd.DataFrame({"stock_id": ["1101"], "date": ["2019-03-01"],
                        "net_income": [100.0], "roe_after_tax": [1.5]})
    stats = extract_legacy_supplement.check_source_duplicates(raw, ["stock_id", "date"], "roe_after_tax")
    assert stats["stage"] == "raw_source"
    assert "net_income" in stats["checked_columns"], "原始來源層要把非目標欄位也納入比對"
    assert stats["n_conflicting_keys"] == 0


def test_check_source_duplicates_allows_fully_identical_rows():
    raw = pd.DataFrame({"stock_id": ["1101", "1101"], "date": ["2019-03-01", "2019-03-01"],
                        "net_income": [100.0, 100.0], "roe_after_tax": [1.5, 1.5]})
    stats = extract_legacy_supplement.check_source_duplicates(raw, ["stock_id", "date"], "roe_after_tax")
    assert stats["n_exact_duplicate_rows_removed"] == 1
    assert stats["n_conflicting_keys"] == 0


def test_combine_recurring_windows_dedupes_identical_overlap_and_reports_stats():
    old_window = pd.DataFrame({"stock_id": ["1101"], "date": ["2019-03-31"],
                                "recurring_net_income": [1000.0]})
    new_window = pd.DataFrame({"stock_id": ["1101"], "date": ["2019-03-31"],
                                "recurring_net_income": [1000.0]})   # 同 key 同值,重疊但一致
    out, stats = extract_legacy_supplement._combine_recurring_windows(old_window, new_window)
    assert len(out) == 1
    assert out.iloc[0]["recurring_net_income"] == 1000.0
    assert stats["n_overlap_keys"] == 1
    assert stats["n_overlap_identical"] == 1
    assert stats["n_overlap_conflicting"] == 0


def test_combine_recurring_windows_raises_on_conflicting_overlap():
    old_window = pd.DataFrame({"stock_id": ["1101"], "date": ["2019-03-31"],
                                "recurring_net_income": [1000.0]})
    new_window = pd.DataFrame({"stock_id": ["1101"], "date": ["2019-03-31"],
                                "recurring_net_income": [2000.0]})   # 同 key 不同值
    with pytest.raises(ValueError, match="重疊"):
        extract_legacy_supplement._combine_recurring_windows(old_window, new_window)


def test_combine_recurring_windows_no_overlap_keeps_both():
    old_window = pd.DataFrame({"stock_id": ["1101"], "date": ["2005-12-31"],
                                "recurring_net_income": [500.0]})
    new_window = pd.DataFrame({"stock_id": ["1101"], "date": ["2019-03-31"],
                                "recurring_net_income": [1000.0]})
    out, stats = extract_legacy_supplement._combine_recurring_windows(old_window, new_window)
    assert len(out) == 2
    assert stats["n_overlap_keys"] == 0


def test_enforce_output_spec_rejects_wrong_schema(monkeypatch):
    df = pd.DataFrame({"stock_id": ["1101"], "date": ["2019-03-01"], "roe_after_tax": [1.0],
                        "unexpected_extra_col": [1]})
    profile = extract_legacy_supplement._profile(df, ["stock_id", "date"])
    with pytest.raises(ValueError, match="schema"):
        extract_legacy_supplement.enforce_output_spec(df, "roe_after_tax", profile)


def test_enforce_output_spec_rejects_below_min_rows():
    df = pd.DataFrame({"stock_id": ["1101"], "date": ["2019-03-01"], "roe_after_tax": [1.0]})
    profile = extract_legacy_supplement._profile(df, ["stock_id", "date"])
    with pytest.raises(ValueError, match="下限"):
        extract_legacy_supplement.enforce_output_spec(df, "roe_after_tax", profile)


def test_enforce_output_spec_rejects_empty_dataframe():
    df = pd.DataFrame(columns=["stock_id", "date", "roe_after_tax"])
    profile = extract_legacy_supplement._profile(df, ["stock_id", "date"])
    with pytest.raises(ValueError, match="下限"):
        extract_legacy_supplement.enforce_output_spec(df, "roe_after_tax", profile)


def test_enforce_output_spec_rejects_duplicate_and_null_keys():
    dup_df = pd.DataFrame({"stock_id": ["1101", "1101"], "date": ["2019-03-01", "2019-03-01"],
                            "roe_after_tax": [1.0, 2.0]})
    profile = extract_legacy_supplement._profile(dup_df, ["stock_id", "date"])
    with pytest.raises(ValueError, match="重複"):
        extract_legacy_supplement.enforce_output_spec(dup_df, "roe_after_tax", profile)


def _make_dir_with_file(root: Path, rel_name: str, content: bytes = b"x"):
    root.mkdir(parents=True, exist_ok=True)
    (root / rel_name).write_bytes(content)


def test_supplement_recover_or_clear_stale_backup_restores_when_output_missing(tmp_path):
    out = tmp_path / "out"
    backup = tmp_path / "backup"
    _make_dir_with_file(backup, "only-copy.parquet", b"only-copy-data")
    extract_legacy_supplement._recover_or_clear_stale_backup(out, backup, "dummy")
    assert (out / "only-copy.parquet").read_bytes() == b"only-copy-data"
    assert not backup.exists()


def test_supplement_recover_or_clear_stale_backup_cleans_when_both_exist(tmp_path):
    out = tmp_path / "out"
    backup = tmp_path / "backup"
    _make_dir_with_file(out, "active.parquet", b"active-data")
    _make_dir_with_file(backup, "stale.parquet", b"stale-data")
    extract_legacy_supplement._recover_or_clear_stale_backup(out, backup, "dummy")
    assert (out / "active.parquet").read_bytes() == b"active-data"
    assert not backup.exists()


def test_supplement_recover_or_clear_stale_backup_failed_restore_then_recovers_on_retry(
        tmp_path, monkeypatch):
    out = tmp_path / "out"
    backup = tmp_path / "backup"
    _make_dir_with_file(backup, "only-copy.parquet", b"only-copy-data")

    real_rename = Path.rename

    def _boom_rename(self, target):
        if self == backup:
            raise RuntimeError("simulated restore failure")
        return real_rename(self, target)

    monkeypatch.setattr(Path, "rename", _boom_rename)
    with pytest.raises(RuntimeError, match="simulated restore failure"):
        extract_legacy_supplement._recover_or_clear_stale_backup(out, backup, "dummy")
    monkeypatch.setattr(Path, "rename", real_rename)

    assert backup.exists() and not out.exists()

    extract_legacy_supplement._recover_or_clear_stale_backup(out, backup, "dummy")
    assert (out / "only-copy.parquet").read_bytes() == b"only-copy-data"
    assert not backup.exists()


def test_publish_staging_recovers_stale_backup_before_committing_new_staging(tmp_path):
    """publish_staging 呼叫端整合測試:上一輪執行留下「只有 backup、沒有 out」的
    殘留,這一輪帶著新的 staging 資料呼叫,應該先安全還原舊資料再完成新資料的
    commit,不遺失任何一份資料。"""
    staging = tmp_path / "staging"
    out = tmp_path / "out"
    backup = tmp_path / "backup"
    _make_dir_with_file(staging, "new.parquet", b"new-data")
    _make_dir_with_file(backup, "old.parquet", b"old-data")   # 殘留:只有 backup

    extract_legacy_supplement.publish_staging(staging, out, backup)

    assert (out / "new.parquet").read_bytes() == b"new-data"
    assert not staging.exists()
    assert not backup.exists()


def test_publish_staging_success_swaps_directories(tmp_path):
    staging = tmp_path / "staging"
    out = tmp_path / "out"
    backup = tmp_path / "backup"
    _make_dir_with_file(staging, "new.parquet", b"new-data")

    extract_legacy_supplement.publish_staging(staging, out, backup)

    assert (out / "new.parquet").read_bytes() == b"new-data"
    assert not staging.exists()
    assert not backup.exists()


def test_publish_staging_restores_old_output_when_commit_rename_fails(tmp_path, monkeypatch):
    """commit point (staging→out 的 rename) 本身失敗——舊 out_dir 必須被還原,
    不能讓 out_dir 憑空消失。"""
    staging = tmp_path / "staging"
    out = tmp_path / "out"
    backup = tmp_path / "backup"
    _make_dir_with_file(staging, "new.parquet", b"new-data")
    _make_dir_with_file(out, "old.parquet", b"old-data")

    real_rename = Path.rename

    def _boom_rename(self, target):
        if self == staging:
            raise RuntimeError("simulated rename failure during commit")
        return real_rename(self, target)

    monkeypatch.setattr(Path, "rename", _boom_rename)
    with pytest.raises(RuntimeError, match="simulated rename failure"):
        extract_legacy_supplement.publish_staging(staging, out, backup)
    monkeypatch.setattr(Path, "rename", real_rename)

    assert (out / "old.parquet").read_bytes() == b"old-data", "commit 失敗,舊資料要被還原"
    assert not backup.exists()


def test_publish_staging_cleanup_failure_does_not_raise(tmp_path, monkeypatch):
    """commit point 之後的 backup 清理失敗,不該讓整個發布被回報成失敗——
    這時候新資料其實已經生效了。"""
    staging = tmp_path / "staging"
    out = tmp_path / "out"
    backup = tmp_path / "backup"
    _make_dir_with_file(staging, "new.parquet", b"new-data")
    _make_dir_with_file(out, "old.parquet", b"old-data")

    real_rmtree = shutil.rmtree

    def _boom_rmtree(path, *a, **kw):
        if Path(path) == backup:
            raise RuntimeError("simulated cleanup failure")
        return real_rmtree(path, *a, **kw)

    monkeypatch.setattr(shutil, "rmtree", _boom_rmtree)
    try:
        extract_legacy_supplement.publish_staging(staging, out, backup)   # 不該 raise
    finally:
        monkeypatch.setattr(shutil, "rmtree", real_rmtree)

    assert (out / "new.parquet").read_bytes() == b"new-data", "commit 已經成功,新資料要生效"


# =============================================================================
# 10. save_by_stock 階段三 (commit 之後的收尾) 失敗不該讓整個函式回報失敗
# =============================================================================

def test_save_by_stock_backup_cleanup_failure_does_not_raise(tmp_path, monkeypatch):
    cache_dir = tmp_path / "cache"
    df1 = _stock_df(["1101", "1102"])
    tej_importer.save_by_stock(df1, "price_valuation", cache_dir)

    real_rmtree = shutil.rmtree
    backup_dir = cache_dir / ".price_valuation.rollback"

    def _boom_rmtree(path, *a, **kw):
        if Path(path) == backup_dir:
            raise RuntimeError("simulated backup cleanup failure")
        return real_rmtree(path, *a, **kw)

    monkeypatch.setattr(shutil, "rmtree", _boom_rmtree)
    try:
        df2 = _stock_df(["1101", "1103"])
        n = tej_importer.save_by_stock(df2, "price_valuation", cache_dir)   # 不該 raise
    finally:
        monkeypatch.setattr(shutil, "rmtree", real_rmtree)

    assert n == 2
    out_dir = cache_dir / "price_valuation"
    assert (out_dir / "1101.parquet").exists()
    assert (out_dir / "1103.parquet").exists(), "commit 已經成功,新資料要生效 (即使清理殘留失敗)"


def test_recover_or_clear_stale_backup_noop_when_backup_absent(tmp_path):
    out = tmp_path / "out"
    backup = tmp_path / "backup"
    tej_importer._recover_or_clear_stale_backup(out, backup, "dummy")   # 不 raise 就算過
    assert not out.exists()
    assert not backup.exists()


def test_recover_or_clear_stale_backup_cleans_when_both_exist(tmp_path):
    out = tmp_path / "out"
    backup = tmp_path / "backup"
    _make_dir_with_file(out, "active.parquet", b"active-data")
    _make_dir_with_file(backup, "stale.parquet", b"stale-data")
    tej_importer._recover_or_clear_stale_backup(out, backup, "dummy")
    assert (out / "active.parquet").read_bytes() == b"active-data"
    assert not backup.exists()


def test_recover_or_clear_stale_backup_restores_when_output_missing(tmp_path):
    out = tmp_path / "out"
    backup = tmp_path / "backup"
    _make_dir_with_file(backup, "only-copy.parquet", b"only-copy-data")
    tej_importer._recover_or_clear_stale_backup(out, backup, "dummy")
    assert (out / "only-copy.parquet").read_bytes() == b"only-copy-data", \
        "backup 是唯一僅存的舊資料,必須被還原成 out_dir,不能盲目刪掉"
    assert not backup.exists()


def test_recover_or_clear_stale_backup_failed_restore_then_recovers_on_retry(tmp_path, monkeypatch):
    """第一次呼叫時還原 (rename) 本身失敗——資料不能遺失,backup 要保持原狀。
    第二次 (不 monkeypatch) 呼叫要能安全完成還原。"""
    out = tmp_path / "out"
    backup = tmp_path / "backup"
    _make_dir_with_file(backup, "only-copy.parquet", b"only-copy-data")

    real_rename = Path.rename

    def _boom_rename(self, target):
        if self == backup:
            raise RuntimeError("simulated restore failure")
        return real_rename(self, target)

    monkeypatch.setattr(Path, "rename", _boom_rename)
    with pytest.raises(RuntimeError, match="simulated restore failure"):
        tej_importer._recover_or_clear_stale_backup(out, backup, "dummy")
    monkeypatch.setattr(Path, "rename", real_rename)

    # 還原失敗:backup 保持原狀,沒有資料遺失,out 依然不存在
    assert backup.exists()
    assert not out.exists()
    assert (backup / "only-copy.parquet").read_bytes() == b"only-copy-data"

    # 新的一次呼叫 (沒有 monkeypatch) 要能安全完成還原
    tej_importer._recover_or_clear_stale_backup(out, backup, "dummy")
    assert (out / "only-copy.parquet").read_bytes() == b"only-copy-data"
    assert not backup.exists()


def test_save_by_stock_recovers_stale_backup_and_completes(tmp_path):
    """save_by_stock 呼叫端整合測試:上次執行留下「只有 backup、沒有 out_dir」的
    殘留 (模擬強制中斷),下一次呼叫要能自動還原並正常完成發布,不遺失舊資料
    也不會把 backup 誤判成可以直接刪除的殘留。"""
    cache_dir = tmp_path / "cache"
    df1 = _stock_df(["1101", "1102"])
    tej_importer.save_by_stock(df1, "price_valuation", cache_dir)
    out_dir = cache_dir / "price_valuation"
    backup_dir = cache_dir / ".price_valuation.rollback"

    # 模擬「commit 中途被強制中斷」的殘留狀態:out_dir 被搬成 backup_dir,
    # 新資料沒有寫進去 (staging_dir 也不存在,不是這支函式的暫存殘留)。
    out_dir.rename(backup_dir)
    assert not out_dir.exists()
    assert backup_dir.exists()

    df2 = _stock_df(["1101", "1102", "1103"])
    n = tej_importer.save_by_stock(df2, "price_valuation", cache_dir)

    assert n == 3
    assert (out_dir / "1103.parquet").exists()
    assert not backup_dir.exists()


def test_save_by_stock_rejects_schema_mismatch_in_staged_file(tmp_path, monkeypatch):
    """驗證階段要真的檢查 schema,不是只看檔案數/總列數對不對。"""
    cache_dir = tmp_path / "cache"
    df = _stock_df(["1101", "1102"])

    real_to_parquet = pd.DataFrame.to_parquet
    call_count = {"n": 0}

    def _corrupting_to_parquet(self, path, *a, **kw):
        call_count["n"] += 1
        if call_count["n"] == 1:
            corrupted = self.drop(columns=["close"])   # 第一個寫出的檔案缺一欄
            return real_to_parquet(corrupted, path, *a, **kw)
        return real_to_parquet(self, path, *a, **kw)

    monkeypatch.setattr(pd.DataFrame, "to_parquet", _corrupting_to_parquet)
    try:
        with pytest.raises(RuntimeError, match="schema"):
            tej_importer.save_by_stock(df, "price_valuation", cache_dir)
    finally:
        monkeypatch.setattr(pd.DataFrame, "to_parquet", real_to_parquet)

    assert not (cache_dir / "price_valuation").exists(), "驗證失敗,不該發布任何東西"
    assert not (cache_dir / ".price_valuation.staging").exists()


# =============================================================================
# 11. Phase A1(預註冊 §B/§C.1/§C.9):精確 schema 公式、顯式型別凍結、
#     in-memory 品質證據。全部用本檔 synthetic fixture,不讀真實 DataExport0806。
# =============================================================================

def _write_xlsx(path: Path, df: pd.DataFrame) -> Path:
    path.parent.mkdir(parents=True, exist_ok=True)
    df.to_excel(path, index=False)
    return path


def _numeric_spec(source_dir: Path, **overrides) -> dict:
    spec = {
        "source_dir": source_dir,
        "date_col": "年月日",
        "required_cols": ["代號", "名稱", "年月日", "收盤價(元)"],
        "rename": {
            "收盤價(元)": "close",
            "成交量(千股)": "_volume_thousand_shares",
            "本益比": "PER",
        },
        "thousand_cols": {"_volume_thousand_shares": "Trading_Volume"},
        "numeric_cols": ["close", "PER"],
    }
    spec.update(overrides)
    return spec


def _static_spec(source_dir: Path, **overrides) -> dict:
    spec = {
        "source_dir": source_dir,
        "static": True,
        "required_cols": ["代號", "名稱", "產業代碼"],
        "rename": {"產業代碼": "ind_code", "產業名稱": "ind_name"},
        "thousand_cols": {},
        "numeric_cols": [],
    }
    spec.update(overrides)
    return spec


# ---- 11.1 精確 11-dataset schema 公式(直接對照真實 DATASETS,不用 synthetic) ----

_EXPECTED_FINAL_TARGETS = {
    "price_valuation": ["PBR_TEJ", "PBR_TSE", "PER_TEJ", "PER_TSE", "Trading_Volume",
                         "close", "dividend_yield_TEJ", "dividend_yield_TSE", "max", "min", "open"],
    "institutional_flow": ["dealer_net", "foreign_net", "trust_net"],
    "fundamentals_quarterly": ["eps", "net_income", "operating_income"],
    "revenue_growth": ["revenue_yoy_pct"],
    "monthly_revenue": ["cum_revenue", "release_date", "revenue", "revenue_yoy_pct"],
    "financial_statements": ["capex", "current_assets", "current_liabilities", "eps", "equity",
                              "gross_profit", "net_income", "operating_cash_flow",
                              "operating_income", "quarter", "release_date", "revenue",
                              "total_assets", "total_liabilities"],
    "institutional_gross": ["foreign_buy", "foreign_holding_pct", "foreign_sell", "trust_buy",
                             "trust_holding_pct", "trust_sell"],
    "margin_balance": ["margin_balance", "margin_buy", "margin_change", "margin_sell",
                        "margin_usage_rate", "short_balance", "short_margin_ratio"],
    "tdcc_weekly": ["holders", "ratio_1000up", "ratio_1to5", "ratio_5to10", "ratio_le1",
                     "total_lots_thousand"],
    "director_pledge": ["director_holding_pct", "group_name", "pledge_pct"],
}


@pytest.mark.parametrize("dataset", list(_EXPECTED_FINAL_TARGETS))
def test_final_target_columns_matches_frozen_B_table(dataset):
    """§B 凍結表格(第 18 輪)逐 dataset 核對:11 個非 static dataset 裡的 10 個,
    `_final_target_columns` 算出來的集合跟順序,必須逐字等於預註冊文件表格。"""
    spec = tej_importer.DATASETS[dataset]
    assert tej_importer._final_target_columns(spec) == sorted(_EXPECTED_FINAL_TARGETS[dataset])


def test_industry_map_static_column_order_matches_frozen_table():
    spec = tej_importer.DATASETS["industry_map"]
    expected = ["stock_id", "stock_name", "tse_ind_code", "tse_ind_name",
                "tej_ind_code", "tej_ind_name", "tej_subind_code", "tej_subind_name"]
    actual = ["stock_id", "stock_name"] + list(spec["rename"].values())
    assert actual == expected


def test_final_target_columns_excludes_thousand_intermediate_keys():
    """§B 第 18 輪修正的核心點:公式要扣掉 thousand_cols 的中繼 key,不能只是
    『rename 的全部值』。"""
    spec = tej_importer.DATASETS["price_valuation"]
    targets = tej_importer._final_target_columns(spec)
    assert "_volume_thousand_shares" not in targets
    assert "Trading_Volume" in targets


# ---- 11.2 顯式型別凍結:數值 float64 / 字串 "string" dtype,含前導零保留 ----

def test_load_one_numeric_columns_are_explicit_float64(tmp_path):
    spec = _numeric_spec(tmp_path / "src")
    raw = pd.DataFrame({"代號": ["1101"], "名稱": ["台泥"], "年月日": ["2026/01/02"],
                         "收盤價(元)": [10], "成交量(千股)": [5], "本益比": [12.3]})
    p = _write_xlsx(tmp_path / "src" / "a.xlsx", raw)
    df = tej_importer._load_one(p, spec)
    assert df["close"].dtype == "float64"
    assert df["PER"].dtype == "float64"
    assert df["Trading_Volume"].dtype == "float64"
    assert df["Trading_Volume"].iloc[0] == 5000.0   # 千股轉換:5 * 1000


def test_load_one_string_columns_are_nullable_string_dtype(tmp_path):
    spec = _numeric_spec(tmp_path / "src")
    raw = pd.DataFrame({"代號": ["1101"], "名稱": ["台泥"], "年月日": ["2026/01/02"],
                         "收盤價(元)": [10.0]})
    p = _write_xlsx(tmp_path / "src" / "a.xlsx", raw)
    df = tej_importer._load_one(p, spec)
    assert df["stock_id"].dtype == "string"
    assert df["stock_name"].dtype == "string"
    assert df["date"].dtype == "string"


def test_load_one_preserves_leading_zero_in_industry_code(tmp_path):
    """§C.1:industry code 不能被誤轉成數字,前導零要保留(來源本身是文字儲存格,
    不是被 Excel 存成數字格式)。"""
    spec = _static_spec(tmp_path / "src")
    raw = pd.DataFrame({"代號": ["1101", "2330"], "名稱": ["台泥", "台積電"],
                         "產業代碼": ["0050", "0099"], "產業名稱": ["水泥工業", "半導體業"]})
    p = _write_xlsx(tmp_path / "src" / "industry.xlsx", raw)
    df = tej_importer._load_one(p, spec)
    assert df["ind_code"].tolist() == ["0050", "0099"]
    assert df["ind_code"].dtype == "string"


# ---- 11.3 source-column-absent materialization vs all-files-absent fail-closed ----

def test_load_one_materializes_missing_target_column_as_null_not_dropped(tmp_path):
    """單一檔案缺一個 rename 來源欄位:目標欄位必須還在最終 schema 裡,值全部是
    null,不能整欄從輸出消失(§B 第 17 輪核心規則)。"""
    spec = _numeric_spec(tmp_path / "src")
    raw = pd.DataFrame({"代號": ["1101"], "名稱": ["台泥"], "年月日": ["2026/01/02"],
                         "收盤價(元)": [10.0]})   # 沒有「本益比」「成交量(千股)」欄
    p = _write_xlsx(tmp_path / "src" / "a.xlsx", raw)
    df = tej_importer._load_one(p, spec)
    assert list(df.columns) == ["stock_id", "stock_name", "date"] + tej_importer._final_target_columns(spec)
    assert df["PER"].isna().all()
    assert df["Trading_Volume"].isna().all()
    assert df["PER"].dtype == "float64"


def test_load_source_raises_when_target_absent_from_all_files(patched_root):
    """全部原始檔都缺同一個 rename 來源欄位——不能被靜默接受成『全部都是
    SOURCE_COLUMN_ABSENT 的 null』,load_source 要 fail-closed。"""
    data_root, manifest_csv = patched_root
    src_dir = data_root / "cat"
    spec = _numeric_spec(src_dir)
    raw1 = pd.DataFrame({"代號": ["1101"], "名稱": ["台泥"], "年月日": ["2026/01/02"],
                          "收盤價(元)": [10.0]})
    raw2 = pd.DataFrame({"代號": ["2330"], "名稱": ["台積電"], "年月日": ["2026/01/02"],
                          "收盤價(元)": [500.0]})
    p1 = _write_xlsx(src_dir / "a.xlsx", raw1)
    p2 = _write_xlsx(src_dir / "b.xlsx", raw2)
    entries = {f.relative_to(data_root).as_posix(): _sha256_of(f) for f in [p1, p2]}
    _write_manifest(manifest_csv, entries)

    import unittest.mock as mock
    with mock.patch.dict(tej_importer.DATASETS, {"_test_absent_everywhere": spec}):
        with pytest.raises(ValueError, match="SOURCE_COLUMN_ABSENT"):
            tej_importer.load_source("_test_absent_everywhere")


def test_load_source_does_not_raise_when_target_absent_from_only_some_files(patched_root):
    data_root, manifest_csv = patched_root
    src_dir = data_root / "cat"
    spec = _numeric_spec(src_dir)
    # 「成交量(千股)」兩個檔都有 (不是全部缺席);「本益比」只有 b.xlsx 缺,
    # 不是全部原始檔都缺——只有這種情況才不該 fail-closed。
    raw1 = pd.DataFrame({"代號": ["1101"], "名稱": ["台泥"], "年月日": ["2026/01/02"],
                          "收盤價(元)": [10.0], "成交量(千股)": [100.0], "本益比": [12.3]})
    raw2 = pd.DataFrame({"代號": ["2330"], "名稱": ["台積電"], "年月日": ["2026/01/03"],
                          "收盤價(元)": [500.0], "成交量(千股)": [200.0]})   # 這個檔缺「本益比」
    p1 = _write_xlsx(src_dir / "a.xlsx", raw1)
    p2 = _write_xlsx(src_dir / "b.xlsx", raw2)
    entries = {f.relative_to(data_root).as_posix(): _sha256_of(f) for f in [p1, p2]}
    _write_manifest(manifest_csv, entries)

    import unittest.mock as mock
    with mock.patch.dict(tej_importer.DATASETS, {"_test_absent_partial": spec}):
        combined = tej_importer.load_source("_test_absent_partial")
    by_stock = combined.set_index("stock_id")
    assert pd.isna(by_stock.loc["2330", "PER"]), "缺席的那一格應該是 NaN"
    assert by_stock.loc["1101", "PER"] == 12.3


# ---- 11.4 §C.9 階段一品質證據:blank / 文字「.」/ 其他無法解析 token ----

def test_load_one_evidence_classifies_blank_vs_unparseable_vs_parsed(tmp_path, monkeypatch):
    monkeypatch.setattr(tej_importer, "DATA_ROOT", tmp_path)   # source_relpath locator 用
    spec = _numeric_spec(tmp_path / "src")
    raw = pd.DataFrame({
        "代號": ["1101", "1102", "1103", "1104"],
        "名稱": ["A", "B", "C", "D"],
        "年月日": ["2026/01/02"] * 4,
        # Round 2 review 第 4 項修正後,"N/A" 這種 pandas 預設 na_values 字面
        # token 現在會正確保留成原始文字 (不再被讀檔本身吃成 NaN),可以直接用。
        "收盤價(元)": [10.5, np.nan, ".", "N/A"],
    })
    p = _write_xlsx(tmp_path / "src" / "a.xlsx", raw)
    df, fe = tej_importer._load_one(p, spec, dataset="price_valuation", collect_evidence=True)
    evidence = fe["cell_records"]

    by_stock = {r["stock_id"]: r for r in evidence if r["target_column"] == "close"}
    assert set(by_stock) == {"1102", "1103", "1104"}, "只有 blank/unparseable 這兩類進 sidecar"

    assert by_stock["1102"]["is_blank"] is True
    assert by_stock["1102"]["is_unparseable"] is False
    assert by_stock["1102"]["raw_token"] is None

    assert by_stock["1103"]["is_blank"] is False
    assert by_stock["1103"]["is_unparseable"] is True
    assert by_stock["1103"]["raw_token"] == "."

    assert by_stock["1104"]["is_blank"] is False
    assert by_stock["1104"]["is_unparseable"] is True
    assert by_stock["1104"]["raw_token"] == "N/A", \
        "字面 token 'N/A' 不能在讀檔階段就被 pandas 預設 na_values 吃成 NaN"

    assert df.loc[df["stock_id"] == "1101", "close"].iloc[0] == 10.5
    assert df["close"].isna().sum() == 3


def test_load_one_evidence_blank_whitespace_only_string(tmp_path, monkeypatch):
    """純空白字串(不是原生 NaN)也要判定成 is_blank,raw_token 一樣是 null。"""
    monkeypatch.setattr(tej_importer, "DATA_ROOT", tmp_path)   # source_relpath locator 用
    spec = _numeric_spec(tmp_path / "src")
    raw = pd.DataFrame({"代號": ["1101"], "名稱": ["A"], "年月日": ["2026/01/02"],
                         "收盤價(元)": ["   "]})
    p = _write_xlsx(tmp_path / "src" / "a.xlsx", raw)
    df, fe = tej_importer._load_one(p, spec, dataset="price_valuation", collect_evidence=True)
    rec = next(r for r in fe["cell_records"] if r["target_column"] == "close")
    assert rec["is_blank"] is True
    assert rec["raw_token"] is None


def test_load_one_evidence_disabled_by_default_no_extra_work(tmp_path):
    """`collect_evidence=False`(預設,`load_source` 目前的呼叫方式)不應該產生
    任何證據副作用,只回傳跟舊版一樣的單一 DataFrame(不是 tuple)。"""
    spec = _numeric_spec(tmp_path / "src")
    raw = pd.DataFrame({"代號": ["1101"], "名稱": ["A"], "年月日": ["2026/01/02"],
                         "收盤價(元)": ["."]})
    p = _write_xlsx(tmp_path / "src" / "a.xlsx", raw)
    df = tej_importer._load_one(p, spec)   # 舊版呼叫方式:兩個 positional 參數
    assert isinstance(df, pd.DataFrame)
    assert df["close"].isna().all()


def test_dedup_key_v1_is_deterministic_and_field_sensitive():
    k1 = tej_importer._dedup_key_v1("price_valuation", "a/b.xlsx", "Sheet1", 5, "close")
    k2 = tej_importer._dedup_key_v1("price_valuation", "a/b.xlsx", "Sheet1", 5, "close")
    assert k1 == k2
    assert len(k1) == 64
    k3 = tej_importer._dedup_key_v1("price_valuation", "a/b.xlsx", "Sheet1", 6, "close")
    assert k1 != k3
    k4 = tej_importer._dedup_key_v1("price_valuation", "a/b.xlsx", None, 5, "close")
    assert k1 != k4, "source_container_member 是 None 時要用 JSON null,不是跟字串 'Sheet1' 混淆"


# ---- 11.5 stage-specific counts / 最終 null 原因對帳 ----

def test_load_one_stage_one_counts_reconcile_with_final_null_count(tmp_path, monkeypatch):
    monkeypatch.setattr(tej_importer, "DATA_ROOT", tmp_path)
    spec = _numeric_spec(tmp_path / "src")
    raw = pd.DataFrame({
        "代號": ["1101", "1102", "1103", "1104", "1105"],
        "名稱": ["A", "B", "C", "D", "E"],
        "年月日": ["2026/01/02"] * 5,
        "收盤價(元)": [10.5, 11.0, np.nan, ".", "x"],
    })
    p = _write_xlsx(tmp_path / "src" / "a.xlsx", raw)
    df, fe = tej_importer._load_one(p, spec, dataset="price_valuation", collect_evidence=True)
    counts = next(c for c in fe["stage_one_counts"] if c["target_column"] == "close")
    assert counts["source_row_count"] == 5
    assert counts["column_present_row_count"] == 5
    assert counts["column_absent_row_count"] == 0
    assert counts["parsed_numeric_cell_count"] == 2
    assert counts["blank_cell_count"] == 1
    assert counts["unparseable_cell_count"] == 2
    assert counts["parsed_numeric_cell_count"] + counts["blank_cell_count"] + \
        counts["unparseable_cell_count"] == counts["column_present_row_count"]
    assert int(df["close"].isna().sum()) == counts["blank_cell_count"] + counts["unparseable_cell_count"]


def test_load_one_stage_one_counts_for_absent_column(tmp_path, monkeypatch):
    monkeypatch.setattr(tej_importer, "DATA_ROOT", tmp_path)
    spec = _numeric_spec(tmp_path / "src")
    raw = pd.DataFrame({"代號": ["1101"], "名稱": ["A"], "年月日": ["2026/01/02"],
                         "收盤價(元)": [10.0]})   # 沒有「本益比」欄
    p = _write_xlsx(tmp_path / "src" / "a.xlsx", raw)
    df, fe = tej_importer._load_one(p, spec, dataset="price_valuation", collect_evidence=True)
    counts = next(c for c in fe["stage_one_counts"] if c["target_column"] == "PER")
    assert counts["column_absent_row_count"] == 1
    assert counts["column_present_row_count"] == 0
    assert counts["source_row_count"] == counts["column_present_row_count"] + counts["column_absent_row_count"]


# ---- 11.6 完全重複去重映射 / 衝突重複防呆(既有 fail-closed 機制的證據延伸) ----

def test_check_duplicate_key_conflicts_evidence_records_exact_duplicate_mapping():
    df = pd.DataFrame({
        "stock_id": ["1101", "1101", "1102"], "stock_name": ["台泥", "台泥", "亞泥"],
        "date": ["2026-01-02", "2026-01-02", "2026-01-02"], "close": [10.0, 10.0, 20.0],
    })
    evidence = {}
    tej_importer._check_duplicate_key_conflicts(df, "dummy", evidence_sink=evidence)
    assert evidence["exact_duplicate_mapping"] == [{"key": ("1101", "2026-01-02"), "n_source_rows": 2}]


def test_check_duplicate_key_conflicts_evidence_records_conflicting_keys_before_raise():
    df = pd.DataFrame({
        "stock_id": ["1101", "1101"], "stock_name": ["台泥", "台泥"],
        "date": ["2026-01-02", "2026-01-02"], "close": [10.0, 99.0],
    })
    evidence = {}
    with pytest.raises(ValueError, match="數值不一致"):
        tej_importer._check_duplicate_key_conflicts(df, "dummy", evidence_sink=evidence)
    assert evidence["conflicting_keys"] == [("1101", "2026-01-02")]


# ---- 11.7 supplement no-overwrite / non-overlap accounting ----

def test_assert_supplement_no_column_overwrite_raises_on_collision():
    combined = pd.DataFrame({"stock_id": ["1101"], "date": ["2019-03-01"], "eps": [1.0]})
    supp = pd.DataFrame({"stock_id": ["1101"], "date": ["2019-03-01"], "eps": [2.0]})
    with pytest.raises(ValueError, match="不能覆寫"):
        tej_importer._assert_supplement_no_column_overwrite(combined, supp)


def test_assert_supplement_no_column_overwrite_passes_when_disjoint():
    combined = pd.DataFrame({"stock_id": ["1101"], "date": ["2019-03-01"], "eps": [1.0]})
    supp = pd.DataFrame({"stock_id": ["1101"], "date": ["2019-03-01"], "roe_after_tax": [2.0]})
    tej_importer._assert_supplement_no_column_overwrite(combined, supp)   # 不 raise 就算過


def test_profile_supplement_merge_counts_overlap_and_uncovered_keys():
    combined = pd.DataFrame({"stock_id": ["1101", "1102", "1103"],
                              "date": ["2019-03-01", "2019-03-01", "2019-03-01"]})
    supp = pd.DataFrame({"stock_id": ["1101", "1102", "1199"],
                          "date": ["2019-03-01", "2019-03-01", "2019-03-01"],
                          "roe_after_tax": [1.0, 2.0, 3.0]})
    profile = tej_importer._profile_supplement_merge(combined, supp)
    assert profile["pre_merge_row_count"] == 3
    assert profile["supplement_row_count"] == 3
    assert profile["overlap_key_count"] == 2
    assert profile["rows_supplement_key_not_covered"] == 1   # 1103 沒被 supplement 覆蓋到
    assert profile["supplement_only_key_count"] == 1         # 1199 只在 supplement 裡


def test_load_source_records_supplement_merge_profile_in_attrs(patched_root):
    data_root, manifest_csv = patched_root
    src_dir = data_root / "cat"
    spec = _numeric_spec(src_dir, required_cols=["代號", "名稱", "年月日"],
                          rename={}, thousand_cols={}, numeric_cols=[])
    raw = pd.DataFrame({"代號": ["1101"], "名稱": ["台泥"], "年月日": ["2019/03/01"]})
    p = _write_xlsx(src_dir / "a.xlsx", raw)
    entries = {p.relative_to(data_root).as_posix(): _sha256_of(p)}
    _write_manifest(manifest_csv, entries)

    supp_df = pd.DataFrame({"stock_id": ["1101"], "date": ["2019-03-01"], "roe_after_tax": [1.5]})
    supp_p = _write_supplement(data_root / "supp", "roe_after_tax", supp_df)
    spec["supplement"] = supp_p

    import unittest.mock as mock
    with mock.patch.dict(tej_importer.DATASETS, {"_test_supplement_profile": spec}):
        combined = tej_importer.load_source("_test_supplement_profile")
    assert combined.attrs["supplement_merge_profile"]["overlap_key_count"] == 1
    assert combined.attrs["supplement_merge_profile"]["rows_supplement_key_not_covered"] == 0


# ---- 11.8 backward compatibility / fail-closed schema ----

def test_load_one_default_call_is_backward_compatible_with_save_by_stock(tmp_path):
    """既有 DataFrame-only 呼叫端(save_by_stock/main)不需要知道 evidence API 存在,
    _load_one(path, spec) 兩個 positional 參數的舊呼叫方式要繼續可用。"""
    spec = _numeric_spec(tmp_path / "src")
    raw = pd.DataFrame({"代號": ["1101"], "名稱": ["台泥"], "年月日": ["2026/01/02"],
                         "收盤價(元)": [10.0], "成交量(千股)": [5.0], "本益比": [12.0]})
    p = _write_xlsx(tmp_path / "src" / "a.xlsx", raw)
    df = tej_importer._load_one(p, spec)
    cache_dir = tmp_path / "cache"
    n = tej_importer.save_by_stock(df, "_bwcompat_test", cache_dir)
    assert n == 1
    assert (cache_dir / "_bwcompat_test" / "1101.parquet").exists()


def test_check_final_schema_rejects_extra_column():
    spec = _numeric_spec(Path("unused"))
    good_cols = ["stock_id", "stock_name", "date"] + tej_importer._final_target_columns(spec)
    df = pd.DataFrame({c: [] for c in good_cols + ["unexpected_extra"]})
    with pytest.raises(ValueError, match="多出"):
        tej_importer._check_final_schema(df, spec, "label")


def test_check_final_schema_rejects_missing_column():
    spec = _numeric_spec(Path("unused"))
    good_cols = ["stock_id", "stock_name", "date"] + tej_importer._final_target_columns(spec)
    df = pd.DataFrame({c: [] for c in good_cols[:-1]})
    with pytest.raises(ValueError, match="缺少"):
        tej_importer._check_final_schema(df, spec, "label")


def test_check_final_schema_rejects_misordered_columns():
    spec = _numeric_spec(Path("unused"))
    good_cols = ["stock_id", "stock_name", "date"] + tej_importer._final_target_columns(spec)
    misordered = good_cols[:-2] + [good_cols[-1], good_cols[-2]]
    df = pd.DataFrame({c: [] for c in misordered})
    with pytest.raises(ValueError, match="順序|不符"):
        tej_importer._check_final_schema(df, spec, "label")


def test_check_final_schema_passes_for_static_insertion_order():
    spec = _static_spec(Path("unused"))
    df = pd.DataFrame({c: [] for c in ["stock_id", "stock_name", "ind_code", "ind_name"]})
    tej_importer._check_final_schema(df, spec, "label")   # 不 raise 就算過


# =============================================================================
# 12. Phase A1 Round 2 review 修復(Codex 拒收 A1,要求補完 evidence 契約)。
#     全部用本檔 synthetic fixture,不讀真實 DataExport0806。
# =============================================================================

def _write_zip_csv(path: Path, df: pd.DataFrame, member_name: str = "data.csv") -> Path:
    """組一份跟 `_read_raw_table` 的 .zip 分支相容的來源檔:UTF-16 + Tab 分隔的
    .csv 包在 .zip 裡。用純文字組 csv 內容 (不用 `df.to_csv`),這樣才能塞進
    刻意的實體空白列 (df.to_csv 不會產生空白列)。"""
    path.parent.mkdir(parents=True, exist_ok=True)
    header = "\t".join(df.columns)
    lines = [header]
    for _, row in df.iterrows():
        lines.append("\t".join("" if pd.isna(v) else str(v) for v in row))
    csv_text = "\n".join(lines) + "\n"
    with zipfile.ZipFile(path, "w") as z:
        z.writestr(member_name, csv_text.encode("utf-16"))
    return path


def _write_zip_csv_raw_text(path: Path, text: str, member_name: str = "data.csv") -> Path:
    """跟上面類似,但直接給完整 csv 文字內容 (含實體空白列),供列號測試用。"""
    path.parent.mkdir(parents=True, exist_ok=True)
    with zipfile.ZipFile(path, "w") as z:
        z.writestr(member_name, text.encode("utf-16"))
    return path


# ---- 12.1 Round 2 review 第 1 項:load_source(..., return_evidence=True) ----

def test_load_source_return_evidence_false_returns_plain_dataframe_unchanged(patched_root):
    data_root, manifest_csv = patched_root
    src_dir = data_root / "cat"
    spec = _numeric_spec(src_dir)
    raw = pd.DataFrame({"代號": ["1101"], "名稱": ["台泥"], "年月日": ["2026/01/02"],
                         "收盤價(元)": [10.0], "成交量(千股)": [5.0], "本益比": [12.0]})
    p = _write_xlsx(src_dir / "a.xlsx", raw)
    entries = {p.relative_to(data_root).as_posix(): _sha256_of(p)}
    _write_manifest(manifest_csv, entries)

    import unittest.mock as mock
    with mock.patch.dict(tej_importer.DATASETS, {"_test_re_false": spec}):
        result = tej_importer.load_source("_test_re_false")
    assert isinstance(result, pd.DataFrame), "不傳 return_evidence 必須維持舊版回傳型別(單一 DataFrame)"


def test_load_source_return_evidence_true_returns_frozen_bundle_keys(patched_root):
    data_root, manifest_csv = patched_root
    src_dir = data_root / "cat"
    spec = _numeric_spec(src_dir)
    raw = pd.DataFrame({"代號": ["1101"], "名稱": ["台泥"], "年月日": ["2026/01/02"],
                         "收盤價(元)": [10.0], "成交量(千股)": [5.0], "本益比": [12.0]})
    p = _write_xlsx(src_dir / "a.xlsx", raw)
    entries = {p.relative_to(data_root).as_posix(): _sha256_of(p)}
    _write_manifest(manifest_csv, entries)

    import unittest.mock as mock
    with mock.patch.dict(tej_importer.DATASETS, {"_test_re_true": spec}):
        result = tej_importer.load_source("_test_re_true", return_evidence=True)
    assert isinstance(result, tuple) and len(result) == 2
    combined, evidence = result
    assert isinstance(combined, pd.DataFrame)
    expected_keys = {"cell_records", "per_file_stage_one_counts", "coverage_matrix",
                      "duplicate_mapping", "final_null_causes", "supplement_merge_profile", "schema"}
    assert expected_keys <= set(evidence.keys()), \
        f"缺少凍結的 bundle key:{expected_keys - set(evidence.keys())}"
    assert "logical_types" in evidence["schema"] and "actual_dtypes" in evidence["schema"]


# ---- 12.2 Round 2 review 第 2 項:完整逐檔逐欄位 coverage matrix ----

def test_load_source_coverage_matrix_has_entry_per_file_per_native_column(patched_root):
    data_root, manifest_csv = patched_root
    src_dir = data_root / "cat"
    spec = _numeric_spec(src_dir)
    raw1 = pd.DataFrame({"代號": ["1101"], "名稱": ["台泥"], "年月日": ["2026/01/02"],
                          "收盤價(元)": [10.0], "成交量(千股)": [5.0], "本益比": [12.0]})
    raw2 = pd.DataFrame({"代號": ["2330"], "名稱": ["台積電"], "年月日": ["2026/01/02"],
                          "收盤價(元)": [500.0], "成交量(千股)": [8.0]})   # 這檔缺「本益比」
    p1 = _write_xlsx(src_dir / "a.xlsx", raw1)
    p2 = _write_xlsx(src_dir / "b.xlsx", raw2)
    entries = {f.relative_to(data_root).as_posix(): _sha256_of(f) for f in [p1, p2]}
    _write_manifest(manifest_csv, entries)

    import unittest.mock as mock
    with mock.patch.dict(tej_importer.DATASETS, {"_test_cov_matrix": spec}):
        _, evidence = tej_importer.load_source("_test_cov_matrix", return_evidence=True)
    matrix = evidence["coverage_matrix"]
    native_targets = tej_importer._final_target_columns(spec)
    by_file = {}
    for row in matrix:
        by_file.setdefault(row["source_relpath"], {})[row["target_column"]] = row["status"]
    assert set(by_file) == {"cat/a.xlsx", "cat/b.xlsx"}
    for target in native_targets:
        assert target in by_file["cat/a.xlsx"], f"每個 (file, target) 組合都要有一筆矩陣紀錄:{target}"
        assert target in by_file["cat/b.xlsx"]
    assert by_file["cat/a.xlsx"]["PER"] == "PRESENT"
    assert by_file["cat/b.xlsx"]["PER"] == "SOURCE_COLUMN_ABSENT"


def test_load_source_coverage_matrix_present_all_null_status(patched_root):
    """欄位存在,但這個檔案裡全部是空白/無法解析——PRESENT_ALL_NULL,不是 PRESENT。"""
    data_root, manifest_csv = patched_root
    src_dir = data_root / "cat"
    spec = _numeric_spec(src_dir)
    raw = pd.DataFrame({"代號": ["1101", "1102"], "名稱": ["台泥", "亞泥"],
                         "年月日": ["2026/01/02", "2026/01/02"],
                         "收盤價(元)": [10.0, 11.0], "本益比": [".", np.nan],
                         "成交量(千股)": [100.0, 200.0]})
    p = _write_xlsx(src_dir / "a.xlsx", raw)
    entries = {p.relative_to(data_root).as_posix(): _sha256_of(p)}
    _write_manifest(manifest_csv, entries)

    import unittest.mock as mock
    with mock.patch.dict(tej_importer.DATASETS, {"_test_cov_allnull": spec}):
        _, evidence = tej_importer.load_source("_test_cov_allnull", return_evidence=True)
    per_status = {row["target_column"]: row["status"] for row in evidence["coverage_matrix"]
                  if row["source_relpath"] == "cat/a.xlsx"}
    assert per_status["PER"] == "PRESENT_ALL_NULL"
    assert per_status["close"] == "PRESENT"


def test_load_source_coverage_matrix_marks_supplement_columns_not_applicable(patched_root):
    data_root, manifest_csv = patched_root
    src_dir = data_root / "cat"
    spec = dict(tej_importer.DATASETS["fundamentals_quarterly"])
    spec["source_dir"] = src_dir
    spec["min_rows"] = None
    spec["min_stocks"] = None
    spec["expected_date_min"] = None
    raw = pd.DataFrame({"證券代碼": ["1101 台泥"], "年月": ["202603"],
                         "歸屬母公司淨利（損）": [1000.0], "每股盈餘": [1.5], "營業利益": [900.0]})
    p = _write_xlsx(src_dir / "a.xlsx", raw)
    entries = {p.relative_to(data_root).as_posix(): _sha256_of(p)}
    _write_manifest(manifest_csv, entries)
    supp_df = pd.DataFrame({"stock_id": ["1101"], "date": ["2026-03-01"], "roe_after_tax": [8.5]})
    spec["supplement"] = _write_supplement(data_root / "supp", "roe_after_tax", supp_df)

    import unittest.mock as mock
    with mock.patch.dict(tej_importer.DATASETS, {"_test_cov_supp": spec}):
        _, evidence = tej_importer.load_source("_test_cov_supp", return_evidence=True)
    supp_rows = [r for r in evidence["coverage_matrix"] if r["target_column"] == "roe_after_tax"]
    assert supp_rows and all(r["status"] == "NOT_APPLICABLE" for r in supp_rows)


def test_load_source_all_files_absent_error_names_source_column_absent_status(patched_root):
    """Round 2 review 第 2 項最後一句:all-files-absent 判定要能從 coverage 矩陣
    推導——這裡驗證錯誤訊息直接引用矩陣狀態字面值,不是另一套獨立措辭。"""
    data_root, manifest_csv = patched_root
    src_dir = data_root / "cat"
    spec = _numeric_spec(src_dir)
    raw = pd.DataFrame({"代號": ["1101"], "名稱": ["台泥"], "年月日": ["2026/01/02"],
                         "收盤價(元)": [10.0]})   # 「成交量(千股)」「本益比」都缺
    p = _write_xlsx(src_dir / "a.xlsx", raw)
    entries = {p.relative_to(data_root).as_posix(): _sha256_of(p)}
    _write_manifest(manifest_csv, entries)

    import unittest.mock as mock
    with mock.patch.dict(tej_importer.DATASETS, {"_test_absent_matrix": spec}):
        with pytest.raises(ValueError, match="SOURCE_COLUMN_ABSENT"):
            tej_importer.load_source("_test_absent_matrix", return_evidence=True)


# ---- 12.3 Round 2 review 第 3 項:merge 後最終 schema + supplement 顯式 float64 ----

def test_check_merged_schema_rejects_wrong_dtype_supplement_column():
    spec = _numeric_spec(Path("unused"))
    good_cols = ["stock_id", "stock_name", "date"] + tej_importer._final_target_columns(spec)
    df = pd.DataFrame({c: [] for c in good_cols + ["roe_after_tax"]})
    df["roe_after_tax"] = df["roe_after_tax"].astype("int64")   # 錯誤型別
    with pytest.raises(ValueError, match="float64"):
        tej_importer._check_merged_schema(df, spec, ["roe_after_tax"], "label")


def test_check_merged_schema_rejects_missing_supplement_column():
    spec = _numeric_spec(Path("unused"))
    good_cols = ["stock_id", "stock_name", "date"] + tej_importer._final_target_columns(spec)
    df = pd.DataFrame({c: [] for c in good_cols})   # 少了預期的 supplement 欄位
    with pytest.raises(ValueError, match="缺少"):
        tej_importer._check_merged_schema(df, spec, ["roe_after_tax"], "label")


def test_check_merged_schema_passes_with_correct_float64_supplement_column():
    spec = _numeric_spec(Path("unused"))
    good_cols = ["stock_id", "stock_name", "date"] + tej_importer._final_target_columns(spec)
    df = pd.DataFrame({c: pd.array([], dtype="float64") if c not in ("stock_id", "stock_name", "date")
                        else pd.array([], dtype="string") for c in good_cols})
    df["roe_after_tax"] = pd.array([], dtype="float64")
    tej_importer._check_merged_schema(df, spec, ["roe_after_tax"], "label")   # 不 raise 就算過


def _build_supplemented_spec(real_dataset: str, src_dir: Path, supplement_path: Path) -> dict:
    spec = dict(tej_importer.DATASETS[real_dataset])
    spec["source_dir"] = src_dir
    spec["supplement"] = supplement_path
    spec["min_rows"] = None
    spec["min_stocks"] = None
    spec["expected_date_min"] = None
    return spec


def test_load_source_fundamentals_quarterly_final_order_with_supplement(patched_root):
    data_root, manifest_csv = patched_root
    src_dir = data_root / "cat"
    raw = pd.DataFrame({"證券代碼": ["1101 台泥"], "年月": ["202603"],
                         "歸屬母公司淨利（損）": [1000.0], "每股盈餘": [1.5], "營業利益": [900.0]})
    p = _write_xlsx(src_dir / "a.xlsx", raw)
    entries = {p.relative_to(data_root).as_posix(): _sha256_of(p)}
    _write_manifest(manifest_csv, entries)
    supp_df = pd.DataFrame({"stock_id": ["1101"], "date": ["2026-03-01"], "roe_after_tax": [8.5]})
    supp_p = _write_supplement(data_root / "supp", "roe_after_tax", supp_df)
    spec = _build_supplemented_spec("fundamentals_quarterly", src_dir, supp_p)

    import unittest.mock as mock
    with mock.patch.dict(tej_importer.DATASETS, {"_test_fq_order": spec}):
        combined = tej_importer.load_source("_test_fq_order")
    expected = (["stock_id", "stock_name", "date"] + tej_importer._final_target_columns(spec)
                + ["roe_after_tax"])
    assert list(combined.columns) == expected
    assert combined["roe_after_tax"].dtype == "float64"
    assert combined.loc[0, "roe_after_tax"] == 8.5


def test_load_source_monthly_revenue_final_order_with_supplement(patched_root):
    data_root, manifest_csv = patched_root
    src_dir = data_root / "cat"
    raw = pd.DataFrame({"證券代碼": ["1101 台泥"], "年月": ["202603"],
                         "營收發布日": ["20260310"], "單月營收成長率％": [3.2],
                         "單月營收(千元)": [500.0], "累計營收(千元)": [1500.0]})
    p = _write_xlsx(src_dir / "a.xlsx", raw)
    entries = {p.relative_to(data_root).as_posix(): _sha256_of(p)}
    _write_manifest(manifest_csv, entries)
    supp_df = pd.DataFrame({"stock_id": ["1101"], "date": ["2026-03-01"],
                             "revenue_last_year": [480.0], "cum_revenue_last_year": [1400.0]})
    supp_p = _write_supplement(data_root / "supp", "revenue_last_year", supp_df)
    spec = _build_supplemented_spec("monthly_revenue", src_dir, supp_p)

    import unittest.mock as mock
    with mock.patch.dict(tej_importer.DATASETS, {"_test_mr_order": spec}):
        combined = tej_importer.load_source("_test_mr_order")
    expected = (["stock_id", "stock_name", "date"] + tej_importer._final_target_columns(spec)
                + ["revenue_last_year", "cum_revenue_last_year"])
    assert list(combined.columns) == expected
    assert combined["revenue_last_year"].dtype == "float64"
    assert combined["cum_revenue_last_year"].dtype == "float64"


def test_load_source_financial_statements_final_order_with_supplement(patched_root):
    data_root, manifest_csv = patched_root
    src_dir = data_root / "cat"
    raw = pd.DataFrame({
        "證券代碼": ["1101 台泥"], "年月": ["202603"], "季別": [1],
        "財報發布日": ["2026/03/15"], "營業收入淨額": [10000.0], "營業毛利": [2000.0],
        "營業利益": [1500.0], "歸屬母公司淨利（損）": [1000.0], "每股盈餘": [1.5],
        "資產總額": [50000.0], "負債總額": [20000.0], "流動資產": [15000.0],
        "流動負債": [8000.0], "股東權益總額": [30000.0], "來自營運之現金流量": [1200.0],
        "  購置不動產廠房設備（含預付）－CFI": [300.0],
    })
    p = _write_xlsx(src_dir / "a.xlsx", raw)
    entries = {p.relative_to(data_root).as_posix(): _sha256_of(p)}
    _write_manifest(manifest_csv, entries)
    supp_df = pd.DataFrame({"stock_id": ["1101"], "date": ["2026-03-01"],
                             "recurring_net_income": [950.0]})
    supp_p = _write_supplement(data_root / "supp", "recurring_net_income", supp_df)
    spec = _build_supplemented_spec("financial_statements", src_dir, supp_p)

    import unittest.mock as mock
    with mock.patch.dict(tej_importer.DATASETS, {"_test_fs_order": spec}):
        combined = tej_importer.load_source("_test_fs_order")
    expected = (["stock_id", "stock_name", "date"] + tej_importer._final_target_columns(spec)
                + ["recurring_net_income"])
    assert list(combined.columns) == expected
    assert combined["recurring_net_income"].dtype == "float64"
    assert combined["quarter"].dtype == "float64"


# ---- 12.4 Round 2 review 第 4 項:讀檔階段保留每一個非空白字面 token ----

@pytest.mark.parametrize("token", ["N/A", "NA", "NULL", "nan", "another_bad_token"])
def test_read_raw_table_xlsx_preserves_nonblank_na_like_literals(tmp_path, token):
    raw = pd.DataFrame({"a": ["10.5", token]})
    p = tmp_path / "a.xlsx"
    raw.to_excel(p, index=False)
    df = tej_importer._read_raw_table(p)
    assert df["a"].iloc[1] == token, f"字面 token {token!r} 不能在讀檔階段被吃成 NaN"
    assert not pd.isna(df["a"].iloc[1])


@pytest.mark.parametrize("token", ["N/A", "NA", "NULL", "nan", "another_bad_token"])
def test_read_raw_table_zip_preserves_nonblank_na_like_literals(tmp_path, token):
    p = tmp_path / "a.zip"
    text = f"a\n10.5\n{token}\n"
    _write_zip_csv_raw_text(p, text)
    df = tej_importer._read_raw_table(p)
    assert df["a"].iloc[1] == token, f"字面 token {token!r} 不能在讀檔階段被吃成 NaN"
    assert not pd.isna(df["a"].iloc[1])


def test_read_raw_table_xlsx_true_blank_and_whitespace_and_dot_distinguishable(tmp_path):
    raw = pd.DataFrame({"a": ["10.5", None, "   ", "."]})
    p = tmp_path / "a.xlsx"
    raw.to_excel(p, index=False)
    df = tej_importer._read_raw_table(p)
    assert str(df["a"].iloc[1]).strip() == ""    # 真正空白
    assert str(df["a"].iloc[2]).strip() == ""    # 純空白字串,strip 後也是空
    assert df["a"].iloc[3] == "."                # 文字「.」原樣保留,不是空白


def test_read_raw_table_zip_true_blank_and_dot_distinguishable(tmp_path):
    p = tmp_path / "a.zip"
    text = "a\n10.5\n\n.\n"   # 中間一列是實體空白列
    _write_zip_csv_raw_text(p, text)
    df = tej_importer._read_raw_table(p)
    assert len(df) == 3, "skip_blank_lines=False,實體空白列要保留成一列"
    assert str(df["a"].iloc[1]).strip() == ""
    assert df["a"].iloc[2] == "."


def test_load_one_end_to_end_preserves_all_seven_required_tokens(tmp_path, monkeypatch):
    """整合測試:透過 `_load_one` 走完整條管線,驗證第 4 項要求的 7 種輸入
    (真空白、純空白字串、「.」、N/A、NA、NULL、nan) 加上另一個任意無效 token,
    全部被正確分類成 blank 或 unparseable,沒有任何一個在半路被 pandas 吃掉。"""
    monkeypatch.setattr(tej_importer, "DATA_ROOT", tmp_path)
    spec = _numeric_spec(tmp_path / "src")
    ids = ["1101", "1102", "1103", "1104", "1105", "1106", "1107", "1108"]
    values = [10.5, np.nan, "   ", ".", "N/A", "NA", "NULL", "nan"]
    raw = pd.DataFrame({"代號": ids, "名稱": ["X"] * 8, "年月日": ["2026/01/02"] * 8,
                         "收盤價(元)": values})
    p = _write_xlsx(tmp_path / "src" / "a.xlsx", raw)
    df, fe = tej_importer._load_one(p, spec, dataset="price_valuation", collect_evidence=True)
    recorded = {r["stock_id"] for r in fe["cell_records"] if r["target_column"] == "close"}
    assert recorded == set(ids[1:]), "除了成功解析的 10.5 外,其餘 7 筆都要進 sidecar"
    tokens_by_id = {r["stock_id"]: r for r in fe["cell_records"] if r["target_column"] == "close"}
    for sid, blank in [("1102", True), ("1103", True)]:
        assert tokens_by_id[sid]["is_blank"] is blank
    for sid, tok in [("1104", "."), ("1105", "N/A"), ("1106", "NA"),
                      ("1107", "NULL"), ("1108", "nan")]:
        assert tokens_by_id[sid]["is_unparseable"] is True
        assert tokens_by_id[sid]["raw_token"] == tok


# ---- 12.5 Round 2 review 第 5 項:讀檔器產生真正的實體列號,不靠 index+2 反推 ----

def test_read_raw_table_xlsx_row_numbers_align_with_physical_rows(tmp_path):
    raw = pd.DataFrame({"a": ["10.5", "11.0", "12.0"]})
    p = tmp_path / "a.xlsx"
    raw.to_excel(p, index=False)
    df, container_member, row_numbers = tej_importer._read_raw_table(p, return_locators=True)
    assert row_numbers == [2, 3, 4]
    assert container_member == "Sheet1"


def test_read_raw_table_zip_row_numbers_align_with_physical_rows_after_blank_line(tmp_path):
    """關鍵回歸測試:實體第 3 列是空白列。若列號還是用 `index + 2` 反推 (在
    `skip_blank_lines=True` 的舊行為下),第 4 實體列會被錯誤標成第 3 列。
    這裡驗證 reader 產生的 row_numbers 正確對齊到實體列號 (2,3,4,5),不是
    因為空白列被跳過而錯位的 (2,3,4)。"""
    p = tmp_path / "a.zip"
    text = "a\n10.5\n\n12.0\n13.0\n"   # 實體列:1=header,2=10.5,3=blank,4=12.0,5=13.0
    _write_zip_csv_raw_text(p, text)
    df, container_member, row_numbers = tej_importer._read_raw_table(p, return_locators=True)
    assert len(df) == 4
    assert row_numbers == [2, 3, 4, 5]
    assert df["a"].iloc[3] == "13.0"


def test_load_one_cell_record_row_number_correct_after_blank_physical_line_in_zip(tmp_path, monkeypatch):
    """端到端驗證:zip 來源檔案中間有一列實體空白列,`_load_one` 產生的
    cell_records 裡,空白列之後那一列的 `source_row_number` 要對到正確的實體
    列號,不是被空白列跳過所汙染的錯誤值 (Round 2 review 第 5 項核心情境)。"""
    monkeypatch.setattr(tej_importer, "DATA_ROOT", tmp_path)
    spec = _numeric_spec(tmp_path / "src")
    # 實體列:1=header,2=1101 資料(收盤價 10.5),3=1102 資料整列空白(視為空白列),
    # 4=1103 資料 unparseable「.」
    text = ("代號\t名稱\t年月日\t收盤價(元)\n"
            "1101\tA\t2026/01/02\t10.5\n"
            "1102\tB\t2026/01/02\t\n"
            "1103\tC\t2026/01/02\t.\n")
    p = _write_zip_csv_raw_text(tmp_path / "src" / "a.zip", text)
    df, fe = tej_importer._load_one(p, spec, dataset="price_valuation", collect_evidence=True)
    by_stock = {r["stock_id"]: r for r in fe["cell_records"] if r["target_column"] == "close"}
    assert by_stock["1102"]["source_row_number"] == 3
    assert by_stock["1103"]["source_row_number"] == 4, \
        "第 3 實體列是空白列但沒有被跳過,第 4 實體列的列號不能因此錯位成 3"


# ---- 12.6 Round 2 review 第 6 項:stage-two 最終 null 原因,跨 concat/去重/merge 對帳 ----

def test_load_source_final_null_causes_reconcile_across_files(patched_root):
    """單一檔案裡「本益比」整欄缺席,會直接觸發 all-files-absent fail-closed
    (這是第 2 項要驗證的行為),所以這裡刻意用兩個檔案:file A 有「本益比」,
    file B 沒有——不是全部都缺席,不會 fail-closed,但 file B 那些列的 PER
    最終 null 原因應該是 SOURCE_COLUMN_ABSENT。"""
    data_root, manifest_csv = patched_root
    src_dir = data_root / "cat"
    spec = _numeric_spec(src_dir)
    raw1 = pd.DataFrame({
        "代號": ["1101", "1102"], "名稱": ["A", "B"], "年月日": ["2026/01/02"] * 2,
        "收盤價(元)": [10.5, "."], "成交量(千股)": [1.0, 2.0], "本益比": [12.0, 13.0],
    })
    raw2 = pd.DataFrame({
        "代號": ["1103"], "名稱": ["C"], "年月日": ["2026/01/02"],
        "收盤價(元)": [np.nan], "成交量(千股)": [3.0],
    })   # 這個檔缺「本益比」整欄
    p1 = _write_xlsx(src_dir / "a.xlsx", raw1)
    p2 = _write_xlsx(src_dir / "b.xlsx", raw2)
    entries = {f.relative_to(data_root).as_posix(): _sha256_of(f) for f in [p1, p2]}
    _write_manifest(manifest_csv, entries)

    import unittest.mock as mock
    with mock.patch.dict(tej_importer.DATASETS, {"_test_nullcause_xf": spec}):
        combined, evidence = tej_importer.load_source("_test_nullcause_xf", return_evidence=True)
    close_causes = evidence["final_null_causes"]["close"]
    assert close_causes["RETAINED_UNPARSEABLE"] == 1   # 1102 的「.」
    assert close_causes["RETAINED_BLANK"] == 1          # 1103 的 NaN
    assert close_causes["OTHER_UNEXPLAINED"] == 0
    assert sum(close_causes.values()) == int(combined["close"].isna().sum())

    per_causes = evidence["final_null_causes"]["PER"]
    assert per_causes["SOURCE_COLUMN_ABSENT"] == 1   # 只有 1103 (來自缺欄的 file B)
    assert per_causes["OTHER_UNEXPLAINED"] == 0
    assert sum(per_causes.values()) == int(combined["PER"].isna().sum())


def test_load_source_final_null_causes_survive_cross_file_exact_duplicate(patched_root):
    """同一個 (stock_id, date) 在兩個檔案裡完全重複(值也相同,含空白狀態
    本身),安全去重成一列——這一個保留下來的 null 原因,只能算一次,不能因為
    去重前有兩筆證據就重複計數兩次。"""
    data_root, manifest_csv = patched_root
    src_dir = data_root / "cat"
    spec = _numeric_spec(src_dir)
    raw1 = pd.DataFrame({"代號": ["1101"], "名稱": ["台泥"], "年月日": ["2026/01/02"],
                          "收盤價(元)": ["."], "成交量(千股)": [1.0], "本益比": [12.0]})
    raw2 = pd.DataFrame({"代號": ["1101"], "名稱": ["台泥"], "年月日": ["2026/01/02"],
                          "收盤價(元)": ["."], "成交量(千股)": [1.0], "本益比": [12.0]})   # 完全重複
    p1 = _write_xlsx(src_dir / "a.xlsx", raw1)
    p2 = _write_xlsx(src_dir / "b.xlsx", raw2)
    entries = {f.relative_to(data_root).as_posix(): _sha256_of(f) for f in [p1, p2]}
    _write_manifest(manifest_csv, entries)

    import unittest.mock as mock
    with mock.patch.dict(tej_importer.DATASETS, {"_test_nullcause_exactdup": spec}):
        combined, evidence = tej_importer.load_source("_test_nullcause_exactdup", return_evidence=True)
    assert len(combined) == 1
    close_causes = evidence["final_null_causes"]["close"]
    assert close_causes["RETAINED_UNPARSEABLE"] == 1, "去重後只剩一列,原因只能算一次"
    assert sum(close_causes.values()) == 1 == int(combined["close"].isna().sum())


def test_load_source_final_null_causes_supplement_key_not_covered(patched_root):
    data_root, manifest_csv = patched_root
    src_dir = data_root / "cat"
    raw = pd.DataFrame({"證券代碼": ["1101 台泥", "2330 台積電"], "年月": ["202603", "202603"],
                         "歸屬母公司淨利（損）": [1000.0, 2000.0], "每股盈餘": [1.5, 3.0],
                         "營業利益": [900.0, 1800.0]})
    p = _write_xlsx(src_dir / "a.xlsx", raw)
    entries = {p.relative_to(data_root).as_posix(): _sha256_of(p)}
    _write_manifest(manifest_csv, entries)
    # supplement 只覆蓋 1101,2330 沒被覆蓋到
    supp_df = pd.DataFrame({"stock_id": ["1101"], "date": ["2026-03-01"], "roe_after_tax": [8.5]})
    supp_p = _write_supplement(data_root / "supp", "roe_after_tax", supp_df)
    spec = _build_supplemented_spec("fundamentals_quarterly", src_dir, supp_p)

    import unittest.mock as mock
    with mock.patch.dict(tej_importer.DATASETS, {"_test_nullcause_supp": spec}):
        combined, evidence = tej_importer.load_source("_test_nullcause_supp", return_evidence=True)
    roe_causes = evidence["final_null_causes"]["roe_after_tax"]
    assert roe_causes["SUPPLEMENT_KEY_NOT_COVERED"] == 1
    assert roe_causes["OTHER_UNEXPLAINED"] == 0
    assert sum(roe_causes.values()) == int(combined["roe_after_tax"].isna().sum())


# ---- 12.7 Round 2 review 第 7 項:完整去重證據映射 ----

def test_load_source_duplicate_mapping_target_level_locators_and_sidecar_keys(patched_root):
    """Round 3 review 第 3 項:兩個檔案對同一個 (stock_id, date) 匯出完全相同
    的一列 (provenance 也一致,安全去重),涵蓋一個 unparseable 目標
    (close,有 sidecar row) 跟一個 parsed 目標 (PER,沒有 sidecar row 但一樣
    要有 source locator)。"""
    data_root, manifest_csv = patched_root
    src_dir = data_root / "cat"
    spec = _numeric_spec(src_dir)
    raw1 = pd.DataFrame({"代號": ["1101"], "名稱": ["台泥"], "年月日": ["2026/01/02"],
                          "收盤價(元)": ["."], "成交量(千股)": [1.0], "本益比": [12.0]})
    raw2 = pd.DataFrame({"代號": ["1101"], "名稱": ["台泥"], "年月日": ["2026/01/02"],
                          "收盤價(元)": ["."], "成交量(千股)": [1.0], "本益比": [12.0]})   # 同 key、provenance 也一致
    p1 = _write_xlsx(src_dir / "a.xlsx", raw1)
    p2 = _write_xlsx(src_dir / "b.xlsx", raw2)
    entries = {f.relative_to(data_root).as_posix(): _sha256_of(f) for f in [p1, p2]}
    _write_manifest(manifest_csv, entries)

    import unittest.mock as mock
    with mock.patch.dict(tej_importer.DATASETS, {"_test_dupmap": spec}):
        combined, evidence = tej_importer.load_source("_test_dupmap", return_evidence=True)
    assert len(combined) == 1
    mapping = evidence["duplicate_mapping"]
    assert isinstance(mapping, list), "第 3 項:duplicate_mapping 是 target-level 的扁平 list"

    close_entries = [e for e in mapping if e["target_column"] == "close"]
    assert len(close_entries) == 1
    entry = close_entries[0]
    assert entry["key"] == ("1101", "2026-01-02")
    assert entry["source_row_count"] == 2
    assert entry["removed_row_count"] == 1
    assert len(entry["contributing_source_locators"]) == 2
    assert entry["retained_source_locator"]["source_relpath"] == "cat/b.xlsx", \
        "explicit source order:檔案順序在後的 b.xlsx 才是保留列"
    assert len(entry["sidecar_dedup_keys"]) == 2, \
        "兩個原始檔各自的 blank/unparseable 證據都要保留,不能因為去重就少一筆"
    assert len(set(entry["sidecar_dedup_keys"])) == 2, "兩筆證據來自不同檔案,dedup_key 應該不同"

    per_entries = [e for e in mapping if e["target_column"] == "PER"]
    assert len(per_entries) == 1
    assert per_entries[0]["sidecar_dedup_keys"] == [], "parsed cell 沒有 sidecar row"
    assert len(per_entries[0]["contributing_source_locators"]) == 2, \
        "parsed cell 也要有 source locator (第 3 項要求)"


def _trigger_close_conflict(dataset_key, data_root, manifest_csv, spec):
    """建 a/b 兩個檔案,close 欄同 key 不同 parsed 值,回傳觸發的例外實例。"""
    src_dir = spec["source_dir"]
    raw1 = pd.DataFrame({"代號": ["1101"], "名稱": ["台泥"], "年月日": ["2026/01/02"],
                          "收盤價(元)": [10.0], "成交量(千股)": [1.0], "本益比": [12.0]})
    raw2 = pd.DataFrame({"代號": ["1101"], "名稱": ["台泥"], "年月日": ["2026/01/02"],
                          "收盤價(元)": [99.0], "成交量(千股)": [1.0], "本益比": [12.0]})   # close 衝突
    p1 = _write_xlsx(src_dir / "a.xlsx", raw1)
    p2 = _write_xlsx(src_dir / "b.xlsx", raw2)
    entries = {f.relative_to(data_root).as_posix(): _sha256_of(f) for f in [p1, p2]}
    _write_manifest(manifest_csv, entries)

    import unittest.mock as mock
    with mock.patch.dict(tej_importer.DATASETS, {dataset_key: spec}):
        with pytest.raises(tej_importer.DuplicateProvenanceConflictError) as excinfo:
            tej_importer.load_source(dataset_key, return_evidence=True)
    return excinfo.value


def test_load_source_duplicate_provenance_conflict_preserves_evidence_after_raise(patched_root):
    """Round 3 review 第 4 項:只檢查『有 raise』不夠,要能從例外物件本身拿到
    完整衝突證據 (raise 之前蒐集到的東西,不能因為例外往外傳就遺失)。內容用
    exported JSON 相容形式核對 (key/provenance 是 list,不是 tuple——見
    收尾修正的 immutable/serializable 契約)。"""
    data_root, manifest_csv = patched_root
    spec = _numeric_spec(data_root / "cat")
    exc = _trigger_close_conflict("_test_dupmap_conflict", data_root, manifest_csv, spec)

    assert isinstance(exc, ValueError), "必須是 ValueError 子類別,既有 except ValueError 呼叫端要相容"
    evidence = exc.duplicate_evidence
    conflicts = evidence["conflicts"]
    close_conflicts = [c for c in conflicts if c["target_column"] == "close"]
    assert len(close_conflicts) == 1
    conflict = close_conflicts[0]
    assert conflict["key"] == ["1101", "2026-01-02"]
    assert len(conflict["entries"]) == 2
    provenances = {tuple(e["provenance"]) for e in conflict["entries"]}
    assert provenances == {("PARSED", 10.0), ("PARSED", 99.0)}
    assert len(conflict["distinct_provenances"]) == 2


def test_duplicate_evidence_is_directly_json_dumpable(patched_root):
    data_root, manifest_csv = patched_root
    spec = _numeric_spec(data_root / "cat")
    exc = _trigger_close_conflict("_test_dupmap_json", data_root, manifest_csv, spec)
    dumped = json.dumps(exc.duplicate_evidence, sort_keys=True)
    reloaded = json.loads(dumped)
    assert reloaded["dataset"] == "_test_dupmap_json"


def test_duplicate_evidence_to_dict_matches_property():
    exc = tej_importer.DuplicateProvenanceConflictError(
        "boom", {"dataset": "d", "conflicts": [{"key": ("1101", "2026-01-02"),
                                                  "target_column": "close",
                                                  "entries": [{"source_order": (0, 2),
                                                               "locator": {"stock_id": "1101"},
                                                               "provenance": ("PARSED", 1.0)}],
                                                  "distinct_provenances": [("PARSED", 1.0)]}]})
    assert exc.to_dict() == exc.duplicate_evidence


def test_duplicate_evidence_constructor_input_mutation_does_not_affect_stored_evidence():
    """建構之後修改傳進去的原始 dict/list,不能反映到例外物件已保存的證據。"""
    conflicts = [{"key": ("1101", "2026-01-02"), "target_column": "close",
                  "entries": [{"source_order": (0, 2), "locator": {"stock_id": "1101"},
                               "provenance": ("PARSED", 1.0)}],
                  "distinct_provenances": [("PARSED", 1.0)]}]
    duplicate_evidence = {"dataset": "d", "conflicts": conflicts}
    exc = tej_importer.DuplicateProvenanceConflictError("boom", duplicate_evidence)

    # 建構之後才 mutate 原始輸入物件。
    conflicts.append({"key": ("9999", "2026-01-02"), "target_column": "PER",
                       "entries": [], "distinct_provenances": []})
    duplicate_evidence["dataset"] = "TAMPERED"
    conflicts[0]["target_column"] = "TAMPERED"

    stored = exc.duplicate_evidence
    assert stored["dataset"] == "d"
    assert len(stored["conflicts"]) == 1
    assert stored["conflicts"][0]["target_column"] == "close"


def test_duplicate_evidence_returned_snapshot_mutation_does_not_leak_to_next_read():
    """對 `duplicate_evidence` 回傳值的 mutate,不能影響下一次存取的結果。"""
    exc = tej_importer.DuplicateProvenanceConflictError(
        "boom", {"dataset": "d", "conflicts": [{"key": ("1101", "2026-01-02"),
                                                  "target_column": "close", "entries": [],
                                                  "distinct_provenances": []}]})
    first = exc.duplicate_evidence
    first["conflicts"].append({"key": ["tampered"], "target_column": "x",
                                "entries": [], "distinct_provenances": []})
    first["dataset"] = "TAMPERED"

    second = exc.duplicate_evidence
    assert second["dataset"] == "d"
    assert len(second["conflicts"]) == 1
    assert second != first

    third = exc.duplicate_evidence
    assert third == second, "重複存取要相等 (值相等,不需要是同一個物件)"
    assert third is not second, "但每次都要是全新的容器,不能共用同一個 list/dict"


def test_duplicate_evidence_rejects_non_json_safe_value():
    with pytest.raises(TypeError):
        tej_importer.DuplicateProvenanceConflictError(
            "boom", {"dataset": "d", "bad": pd.DataFrame({"a": [1]})})


def test_provenance_conflict_still_fail_closed_for_blank_vs_dot_after_evidence_fix(patched_root):
    """收尾修正只動了例外的證據封裝,既有的 blank-vs-dot fail-closed 行為不能
    跟著變弱。"""
    data_root, manifest_csv = patched_root
    spec = _numeric_spec(data_root / "cat")
    raw1 = pd.DataFrame({"代號": ["1101"], "名稱": ["台泥"], "年月日": ["2026/01/02"],
                          "收盤價(元)": [np.nan], "成交量(千股)": [1.0], "本益比": [12.0]})
    raw2 = pd.DataFrame({"代號": ["1101"], "名稱": ["台泥"], "年月日": ["2026/01/02"],
                          "收盤價(元)": ["."], "成交量(千股)": [1.0], "本益比": [12.0]})
    p1 = _write_xlsx(data_root / "cat" / "a.xlsx", raw1)
    p2 = _write_xlsx(data_root / "cat" / "b.xlsx", raw2)
    entries = {f.relative_to(data_root).as_posix(): _sha256_of(f) for f in [p1, p2]}
    _write_manifest(manifest_csv, entries)
    import unittest.mock as mock
    with mock.patch.dict(tej_importer.DATASETS, {"_test_blank_vs_dot_after_fix": spec}):
        with pytest.raises(tej_importer.DuplicateProvenanceConflictError) as excinfo:
            tej_importer.load_source("_test_blank_vs_dot_after_fix", return_evidence=True)
    conflicts = excinfo.value.duplicate_evidence["conflicts"]
    close_conflict = next(c for c in conflicts if c["target_column"] == "close")
    tags = {e["provenance"][0] for e in close_conflict["entries"]}
    assert tags == {"RETAINED_BLANK", "RETAINED_UNPARSEABLE"}


def test_provenance_conflict_still_fail_closed_for_parsed_values_after_evidence_fix(patched_root):
    data_root, manifest_csv = patched_root
    spec = _numeric_spec(data_root / "cat")
    exc = _trigger_close_conflict("_test_parsed_conflict_after_fix", data_root, manifest_csv, spec)
    conflicts = exc.duplicate_evidence["conflicts"]
    close_conflict = next(c for c in conflicts if c["target_column"] == "close")
    assert {tuple(p) for p in close_conflict["distinct_provenances"]} == {("PARSED", 10.0), ("PARSED", 99.0)}


# ---- 12.8 Round 2 review 第 8 項:完整 supplement provenance ----

def test_profile_supplement_merge_returns_full_provenance_detail():
    combined = pd.DataFrame({"stock_id": ["1101", "1102", "1103"],
                              "stock_name": ["A", "B", "C"],
                              "date": ["2019-03-01", "2019-03-01", "2019-03-01"],
                              "eps": [1.0, 2.0, 3.0]})
    supp = pd.DataFrame({"stock_id": ["1101", "1102", "1199"],
                          "date": ["2019-03-01", "2019-03-01", "2019-03-01"],
                          "roe_after_tax": [1.0, 2.0, 3.0]})
    profile = tej_importer._profile_supplement_merge(combined, supp)
    assert profile["pre_merge_row_count"] == 3
    assert profile["post_merge_row_count"] is None, "merge 後列數由 load_source 補上,這支函式不知道"
    assert profile["native_columns"] == ["stock_id", "stock_name", "date", "eps"]
    assert profile["supplement_columns"] == ["roe_after_tax"]
    assert profile["overlap_key_count"] == 2
    assert set(profile["overlap_keys"]) == {("1101", "2019-03-01"), ("1102", "2019-03-01")}
    assert profile["rows_supplement_key_not_covered"] == 1
    assert profile["uncovered_keys"] == [("1103", "2019-03-01")]
    assert profile["supplement_only_key_count"] == 1
    assert profile["supplement_only_keys"] == [("1199", "2019-03-01")]


def test_load_source_supplement_merge_profile_fills_post_merge_row_count(patched_root):
    data_root, manifest_csv = patched_root
    src_dir = data_root / "cat"
    raw = pd.DataFrame({"證券代碼": ["1101 台泥"], "年月": ["202603"],
                         "歸屬母公司淨利（損）": [1000.0], "每股盈餘": [1.5], "營業利益": [900.0]})
    p = _write_xlsx(src_dir / "a.xlsx", raw)
    entries = {p.relative_to(data_root).as_posix(): _sha256_of(p)}
    _write_manifest(manifest_csv, entries)
    supp_df = pd.DataFrame({"stock_id": ["1101"], "date": ["2026-03-01"], "roe_after_tax": [8.5]})
    supp_p = _write_supplement(data_root / "supp", "roe_after_tax", supp_df)
    spec = _build_supplemented_spec("fundamentals_quarterly", src_dir, supp_p)

    import unittest.mock as mock
    with mock.patch.dict(tej_importer.DATASETS, {"_test_postmerge_count": spec}):
        combined = tej_importer.load_source("_test_postmerge_count")
    profile = combined.attrs["supplement_merge_profile"]
    assert profile["post_merge_row_count"] == len(combined) == 1


# ---- 12.9 static dataset (industry_map) 走 return_evidence=True 的獨立路徑 ----

def test_load_source_return_evidence_true_works_for_static_dataset(patched_root):
    data_root, manifest_csv = patched_root
    src_dir = data_root / "cat"
    spec = _static_spec(src_dir)
    raw = pd.DataFrame({"代號": ["1101", "2330"], "名稱": ["台泥", "台積電"],
                         "產業代碼": ["0050", "0099"], "產業名稱": ["水泥工業", "半導體業"]})
    p = _write_xlsx(src_dir / "industry.xlsx", raw)
    entries = {p.relative_to(data_root).as_posix(): _sha256_of(p)}
    _write_manifest(manifest_csv, entries)

    import unittest.mock as mock
    with mock.patch.dict(tej_importer.DATASETS, {"_test_static_evidence": spec}):
        combined, evidence = tej_importer.load_source("_test_static_evidence", return_evidence=True)
    assert list(combined["stock_id"]) == ["1101", "2330"]
    assert evidence["supplement_merge_profile"] is None
    assert evidence["duplicate_mapping"] == []
    assert evidence["cell_records"] == [], "industry_map 沒有數值目標欄位,不會有 §C.9 cell 證據"
    coverage_targets = {row["target_column"] for row in evidence["coverage_matrix"]}
    assert coverage_targets == {"ind_code", "ind_name"}
    for target in ("ind_code", "ind_name"):
        assert evidence["final_null_causes"][target] == {
            "RETAINED_BLANK": 0, "RETAINED_UNPARSEABLE": 0, "SOURCE_COLUMN_ABSENT": 0,
            "SUPPLEMENT_KEY_NOT_COVERED": 0, "OTHER_UNEXPLAINED": 0,
        }
    assert evidence["schema"]["logical_types"]["ind_code"] == "string"


# =============================================================================
# 13. Phase A1 Round 3 review 修復(Codex 拒收 round 2 修復,要求 provenance
#     正確的去重)。全部用本檔 synthetic fixture,不讀真實 DataExport0806。
# =============================================================================

def _load_evidence(spec_overrides_dataset_key, data_root, manifest_csv, spec, files_dict):
    """files_dict: {filename: DataFrame},依檔名排序寫入(等於 _source_files 的
    排序方式,亦即 explicit source order 的第一層)。"""
    src_dir = spec["source_dir"]
    written = []
    for fname, df in files_dict.items():
        written.append(_write_xlsx(src_dir / fname, df))
    entries = {f.relative_to(data_root).as_posix(): _sha256_of(f) for f in written}
    _write_manifest(manifest_csv, entries)
    import unittest.mock as mock
    with mock.patch.dict(tej_importer.DATASETS, {spec_overrides_dataset_key: spec}):
        return tej_importer.load_source(spec_overrides_dataset_key, return_evidence=True)


# ---- 13.1 Round 3 review 第 1 項:duplicate 語意要看 provenance,不是只看最終值 ----

def test_provenance_conflict_blank_vs_dot(patched_root):
    """兩個檔案同一個 key:一個是真空白、一個是文字「.」——最終值都是 NaN,
    但語意不同,必須 fail-closed。Codex 直接舉的重現案例。"""
    data_root, manifest_csv = patched_root
    spec = _numeric_spec(data_root / "cat")
    raw1 = pd.DataFrame({"代號": ["1101"], "名稱": ["台泥"], "年月日": ["2026/01/02"],
                          "收盤價(元)": [np.nan], "成交量(千股)": [1.0], "本益比": [12.0]})
    raw2 = pd.DataFrame({"代號": ["1101"], "名稱": ["台泥"], "年月日": ["2026/01/02"],
                          "收盤價(元)": ["."], "成交量(千股)": [1.0], "本益比": [12.0]})
    import unittest.mock as mock
    with mock.patch.dict(tej_importer.DATASETS, {"_test_blank_vs_dot": spec}):
        p1 = _write_xlsx(data_root / "cat" / "a.xlsx", raw1)
        p2 = _write_xlsx(data_root / "cat" / "b.xlsx", raw2)
        entries = {f.relative_to(data_root).as_posix(): _sha256_of(f) for f in [p1, p2]}
        _write_manifest(manifest_csv, entries)
        with pytest.raises(tej_importer.DuplicateProvenanceConflictError) as excinfo:
            tej_importer.load_source("_test_blank_vs_dot", return_evidence=True)
    conflicts = excinfo.value.duplicate_evidence["conflicts"]
    close_conflict = next(c for c in conflicts if c["target_column"] == "close")
    tags = {e["provenance"][0] for e in close_conflict["entries"]}
    assert tags == {"RETAINED_BLANK", "RETAINED_UNPARSEABLE"}


def test_provenance_conflict_dot_vs_n_slash_a(patched_root):
    """兩個不同的無法解析 token (「.」跟「N/A」)——即使兩個都屬於
    RETAINED_UNPARSEABLE 這個 tag,raw token 不同,provenance tuple 不相等,
    一樣是衝突。"""
    data_root, manifest_csv = patched_root
    spec = _numeric_spec(data_root / "cat")
    raw1 = pd.DataFrame({"代號": ["1101"], "名稱": ["台泥"], "年月日": ["2026/01/02"],
                          "收盤價(元)": ["."], "成交量(千股)": [1.0], "本益比": [12.0]})
    raw2 = pd.DataFrame({"代號": ["1101"], "名稱": ["台泥"], "年月日": ["2026/01/02"],
                          "收盤價(元)": ["N/A"], "成交量(千股)": [1.0], "本益比": [12.0]})
    import unittest.mock as mock
    with mock.patch.dict(tej_importer.DATASETS, {"_test_dot_vs_na": spec}):
        p1 = _write_xlsx(data_root / "cat" / "a.xlsx", raw1)
        p2 = _write_xlsx(data_root / "cat" / "b.xlsx", raw2)
        entries = {f.relative_to(data_root).as_posix(): _sha256_of(f) for f in [p1, p2]}
        _write_manifest(manifest_csv, entries)
        with pytest.raises(tej_importer.DuplicateProvenanceConflictError) as excinfo:
            tej_importer.load_source("_test_dot_vs_na", return_evidence=True)
    conflicts = excinfo.value.duplicate_evidence["conflicts"]
    close_conflict = next(c for c in conflicts if c["target_column"] == "close")
    tokens = {e["provenance"][1] for e in close_conflict["entries"]}
    assert tokens == {".", "N/A"}


def test_provenance_conflict_absent_column_vs_blank(patched_root):
    """一個檔案整欄缺席 (SOURCE_COLUMN_ABSENT),另一個檔案該欄存在但這一列是
    空白 (RETAINED_BLANK)——最終值都是 NaN,語意不同,必須 fail-closed。"""
    data_root, manifest_csv = patched_root
    spec = _numeric_spec(data_root / "cat")
    raw1 = pd.DataFrame({"代號": ["1101"], "名稱": ["台泥"], "年月日": ["2026/01/02"],
                          "收盤價(元)": [10.0], "成交量(千股)": [1.0]})   # 這個檔缺「本益比」整欄
    raw2 = pd.DataFrame({"代號": ["1101"], "名稱": ["台泥"], "年月日": ["2026/01/02"],
                          "收盤價(元)": [10.0], "成交量(千股)": [2.0], "本益比": [np.nan]})   # 這個檔有欄但空白
    import unittest.mock as mock
    with mock.patch.dict(tej_importer.DATASETS, {"_test_absent_vs_blank": spec}):
        p1 = _write_xlsx(data_root / "cat" / "a.xlsx", raw1)
        p2 = _write_xlsx(data_root / "cat" / "b.xlsx", raw2)
        entries = {f.relative_to(data_root).as_posix(): _sha256_of(f) for f in [p1, p2]}
        _write_manifest(manifest_csv, entries)
        with pytest.raises(tej_importer.DuplicateProvenanceConflictError) as excinfo:
            tej_importer.load_source("_test_absent_vs_blank", return_evidence=True)
    conflicts = excinfo.value.duplicate_evidence["conflicts"]
    per_conflict = next(c for c in conflicts if c["target_column"] == "PER")
    tags = {e["provenance"][0] for e in per_conflict["entries"]}
    assert tags == {"SOURCE_COLUMN_ABSENT", "RETAINED_BLANK"}


def test_provenance_same_token_exact_duplicate_is_safe(patched_root):
    """兩個檔案同一個 key、同一個目標欄位,都是同一個無法解析 token「.」——
    provenance tuple 完全相同,是安全的完全重複,不 raise。"""
    data_root, manifest_csv = patched_root
    spec = _numeric_spec(data_root / "cat")
    raw1 = pd.DataFrame({"代號": ["1101"], "名稱": ["台泥"], "年月日": ["2026/01/02"],
                          "收盤價(元)": ["."], "成交量(千股)": [1.0], "本益比": [12.0]})
    raw2 = pd.DataFrame({"代號": ["1101"], "名稱": ["台泥"], "年月日": ["2026/01/02"],
                          "收盤價(元)": ["."], "成交量(千股)": [1.0], "本益比": [12.0]})
    combined, evidence = _load_evidence("_test_same_token_safe", data_root, manifest_csv, spec,
                                         {"a.xlsx": raw1, "b.xlsx": raw2})
    assert len(combined) == 1
    assert pd.isna(combined.loc[0, "close"])


# ---- 13.2 Round 3 review 第 2 項:retained row 選擇要用 explicit source order ----

def test_retained_row_selection_uses_explicit_file_order_not_sort_stability(patched_root):
    """兩個檔案對同一個 key 的每個目標欄位都回報完全相同的值/provenance
    (安全重複,不衝突——不同 parsed 數值本身就是合法的衝突,不能拿來測「哪一
    筆被選為保留列」)。沒辦法從『哪個值被保留』看出保留了誰的列,但可以從
    duplicate_mapping 的 `retained_source_locator` 看出來:必須是檔名排序在
    後的 b.xlsx,不是任何巧合。"""
    data_root, manifest_csv = patched_root
    spec = _numeric_spec(data_root / "cat")
    raw_a = pd.DataFrame({"代號": ["1101"], "名稱": ["台泥"], "年月日": ["2026/01/02"],
                          "收盤價(元)": ["."], "成交量(千股)": [1.0], "本益比": [12.0]})
    raw_b = pd.DataFrame({"代號": ["1101"], "名稱": ["台泥"], "年月日": ["2026/01/02"],
                          "收盤價(元)": ["."], "成交量(千股)": [1.0], "本益比": [12.0]})
    combined, evidence = _load_evidence("_test_explicit_order", data_root, manifest_csv, spec,
                                         {"a.xlsx": raw_a, "b.xlsx": raw_b})
    close_entry = next(e for e in evidence["duplicate_mapping"] if e["target_column"] == "close")
    assert close_entry["retained_source_locator"]["source_relpath"] == "cat/b.xlsx", \
        "檔名排序在後的 b.xlsx 才是保留列 (explicit source order)"


def test_retained_row_selection_stable_when_file_creation_order_reversed(patched_root):
    """Round 3 review 第 2 項要求的「file order 故意反過來」測試:即使把
    「實際建立檔案」的先後順序反過來 (先建立 b.xlsx 再建立 a.xlsx),
    `_source_files()` 仍然用檔名字典序排序,`retained_source_locator` 的結果
    應該完全一樣——排序依據是凍結的檔名排序,不是任何建立時間這種不穩定的
    side effect。"""
    data_root, manifest_csv = patched_root
    spec = _numeric_spec(data_root / "cat")
    raw_a = pd.DataFrame({"代號": ["1101"], "名稱": ["台泥"], "年月日": ["2026/01/02"],
                          "收盤價(元)": ["."], "成交量(千股)": [1.0], "本益比": [12.0]})
    raw_b = pd.DataFrame({"代號": ["1101"], "名稱": ["台泥"], "年月日": ["2026/01/02"],
                          "收盤價(元)": ["."], "成交量(千股)": [1.0], "本益比": [12.0]})
    import unittest.mock as mock
    with mock.patch.dict(tej_importer.DATASETS, {"_test_explicit_order_rev": spec}):
        pb = _write_xlsx(data_root / "cat" / "b.xlsx", raw_b)   # b 先建立
        pa = _write_xlsx(data_root / "cat" / "a.xlsx", raw_a)   # a 後建立
        entries = {f.relative_to(data_root).as_posix(): _sha256_of(f) for f in [pb, pa]}
        _write_manifest(manifest_csv, entries)
        combined, evidence = tej_importer.load_source("_test_explicit_order_rev", return_evidence=True)
    close_entry = next(e for e in evidence["duplicate_mapping"] if e["target_column"] == "close")
    assert close_entry["retained_source_locator"]["source_relpath"] == "cat/b.xlsx", \
        "檔名排序 (a<b) 決定 source order,不是實際建立檔案的先後順序"


def test_adjudicate_and_retain_uses_file_index_not_input_list_position(tmp_path):
    """直接測 `_adjudicate_and_retain_with_provenance`:輸入的 `file_row_records`
    即使沒有事先排序,只要 `(file_index, source_row_number)` 正確標註,一樣要
    選出正確的保留列 (第 2 項:不依賴輸入串列本身的順序或任何 dict 插入順序)。
    兩筆列的 values/provenance 完全相同 (安全重複),差異只在 locator 的
    `source_relpath`,藉此在不觸發衝突的情況下驗證『保留了誰的列』。"""
    spec = _numeric_spec(tmp_path / "src")

    def _mk(stock_id, date, source_relpath, row_number):
        return {
            "key": (stock_id, date),
            "locator": {"source_relpath": source_relpath, "source_file_sha256": "x",
                        "source_container_member": "Sheet1", "source_row_number": row_number,
                        "stock_id": stock_id, "date": date},
            "values": {"stock_id": stock_id, "stock_name": "X", "date": date,
                       "close": 10.0, "Trading_Volume": 1000.0, "PER": 12.0},
            "provenance": {"close": ("PARSED", 10.0), "Trading_Volume": ("PARSED", 1000.0),
                           "PER": ("PARSED", 12.0)},
        }

    # file_index=1 (較後的檔案) 先出現在輸入 list 裡,file_index=0 後出現——
    # 刻意打亂輸入順序,驗證函式看的是 file_index 不是 list 位置。
    file_row_records = [
        (1, _mk("1101", "2026-01-02", "b.xlsx", 2)),
        (0, _mk("1101", "2026-01-02", "a.xlsx", 2)),
    ]
    native_targets = tej_importer._final_target_columns(spec)
    retained, mapping = tej_importer._adjudicate_and_retain_with_provenance(
        dataset="price_valuation", spec=spec, file_row_records=file_row_records,
        native_targets=native_targets, is_static=False)
    assert retained[("1101", "2026-01-02")]["locator"]["source_relpath"] == "b.xlsx", \
        "file_index=1 比較大,即使它在輸入 list 裡先出現,還是要被選為保留列"


# ---- 13.3 Round 3 review 第 5 項:讀檔器要用真正的實體結構,不是公式 ----

def test_read_zip_csv_fails_closed_on_embedded_newline_in_quoted_field(tmp_path):
    p = tmp_path / "a.zip"
    # RFC 4180 quoted 欄位內嵌實體換行:一筆邏輯紀錄跨了兩個實體列。
    text = 'a\tb\n1101\t"多行\n內容"\n'
    _write_zip_csv_raw_text(p, text)
    with pytest.raises(ValueError, match="內嵌實體換行"):
        tej_importer._read_raw_table(p)


def test_read_xlsx_row_numbers_come_from_genuine_worksheet_row_identity(tmp_path):
    """用 openpyxl 直接組一份工作表,故意在資料區中間留一個完全沒有寫入任何
    儲存格的實體列 (不透過 pandas to_excel,連 openpyxl 都不主動幫忙補列)——
    工作表的 `max_row` 天然涵蓋這個空隙,`iter_rows()` 一樣會迭代到它,
    `row_numbers` 必須反映這個實體結構,不是「假設沒有空隙」的公式。"""
    p = tmp_path / "a.xlsx"
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.append(["a", "b"])          # 實體第 1 列:表頭
    ws.append(["1101", "10.5"])    # 實體第 2 列
    ws.cell(row=4, column=1, value="1103")   # 故意跳過第 3 列,直接寫第 4 列
    ws.cell(row=4, column=2, value="12.0")
    wb.save(p)

    df, sheet, row_numbers = tej_importer._read_raw_table(p, return_locators=True)
    assert sheet == "Sheet"
    assert len(df) == 3   # 實體第 2、3(全空白)、4 列都要被讀到
    assert row_numbers == [2, 3, 4]
    assert str(df["a"].iloc[1]).strip() == "", "實體第 3 列完全沒寫入,要是空白列,不是被跳過"
    assert df["a"].iloc[2] == "1103"


def test_read_zip_csv_row_numbers_reconcile_with_parsed_row_count(tmp_path):
    p = tmp_path / "a.zip"
    text = "a\tb\n1101\t10.5\n1102\t11.0\n1103\t12.0\n"
    _write_zip_csv_raw_text(p, text)
    df, member, row_numbers = tej_importer._read_raw_table(p, return_locators=True)
    assert len(row_numbers) == len(df) == 3
    assert row_numbers == [2, 3, 4]


# ---- 13.4 Round 3 review 第 6 項:canonical 字串空白正規化 + Arrow schema metadata ----

def test_to_string_column_normalizes_native_blank_and_stripped_empty_to_na():
    s = pd.Series(["  ", None, "台泥", ""], dtype="object")
    out = tej_importer._to_string_column(s)
    assert out.dtype == "string"
    assert pd.isna(out.iloc[0]), "純空白字串 trim 後是空的,要正規化成 pd.NA"
    assert pd.isna(out.iloc[1])
    assert out.iloc[2] == "台泥"
    assert pd.isna(out.iloc[3]), "空字串本身也要正規化成 pd.NA,不能被當成非空值保留"


def test_load_one_blank_group_name_normalizes_to_na_not_empty_string(tmp_path, monkeypatch):
    """director_pledge 的 group_name 是 §C.1 點名的字串欄位;來源儲存格真正
    空白時,最終欄位要是 `pd.NA`,不是空字串 `""`。"""
    monkeypatch.setattr(tej_importer, "DATA_ROOT", tmp_path)
    spec = dict(tej_importer.DATASETS["director_pledge"])
    spec["source_dir"] = tmp_path / "src"
    raw = pd.DataFrame({"證券代碼": ["1101 台泥"], "年月": ["2026/03"],
                         "董監質押%": [5.0], "董監持股%": [10.0], "集團名稱": [None]})
    p = _write_xlsx(tmp_path / "src" / "a.xlsx", raw)
    df, fe = tej_importer._load_one(p, spec, dataset="director_pledge", collect_evidence=True)
    assert pd.isna(df.loc[0, "group_name"])
    assert fe["coverage_row"]["group_name"] == "PRESENT_ALL_NULL"
    assert fe["row_records"][0]["provenance"]["group_name"] == ("RETAINED_BLANK",)


def test_load_one_blank_industry_name_normalizes_to_na(tmp_path):
    spec = _static_spec(tmp_path / "src")
    raw = pd.DataFrame({"代號": ["1101"], "名稱": ["台泥"], "產業代碼": ["0050"],
                         "產業名稱": [None]})
    p = _write_xlsx(tmp_path / "src" / "industry.xlsx", raw)
    df = tej_importer._load_one(p, spec)
    assert pd.isna(df.loc[0, "ind_name"])


def test_arrow_type_metadata_distinguishes_numeric_string_and_date_columns():
    df = pd.DataFrame({
        "stock_id": pd.array(["1101"], dtype="string"),
        "date": pd.array(["2026-01-02"], dtype="string"),
        "close": pd.array([10.5], dtype="float64"),
    })
    arrow_types = tej_importer._arrow_type_metadata(df)
    assert set(arrow_types) == {"stock_id", "date", "close"}
    assert arrow_types["close"] == "double"
    assert "string" in arrow_types["stock_id"].lower()
    assert "string" in arrow_types["date"].lower()


def test_evidence_bundle_schema_has_three_distinct_type_dimensions(patched_root):
    data_root, manifest_csv = patched_root
    spec = _numeric_spec(data_root / "cat")
    raw = pd.DataFrame({"代號": ["1101"], "名稱": ["台泥"], "年月日": ["2026/01/02"],
                         "收盤價(元)": [10.0], "成交量(千股)": [1.0], "本益比": [12.0]})
    combined, evidence = _load_evidence("_test_arrow_schema", data_root, manifest_csv, spec,
                                         {"a.xlsx": raw})
    schema = evidence["schema"]
    assert set(schema) == {"logical_types", "actual_dtypes", "arrow_types"}
    assert schema["logical_types"]["close"] == "float64"
    assert schema["actual_dtypes"]["close"] == "float64"
    assert schema["arrow_types"]["close"] == "double"
    assert schema["logical_types"]["stock_id"] == "string"
    assert "string" in schema["arrow_types"]["stock_id"].lower()
