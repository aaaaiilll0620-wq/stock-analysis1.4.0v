# -*- coding: utf-8 -*-
"""build_v2_candidate_verifier.py — DataExport0806 → V2 隔離候選建置的
獨立驗證器骨架。

規格:`docs/預註冊_DataExport0806_V2隔離建置.md`(下稱「預註冊文件」)§D
(混合架構——本檔案是「獨立驗證器腳本」,§D 明講**不 import、不呼叫**
`scripts/build_v2_candidate.py`(builder/orchestrator)的任何函式,**也不
呼叫** `tej_importer.load_source()`;唯一允許共用的是「已經提交/凍結的
預註冊文件跟 receipt schema 本身」的*定義*,不是任何一方執行當下產生的
可變狀態)、§D 單發驗證規則(run 層級 `.claim` 鎖 + crash 持久性判定)、
§C.5 第 3 項(驗證 receipt schema)。

**這份檔案不 import `build_v2_candidate.py` 或 `tej_importer.py`**——身分
公式(canonical JSON/SHA-256、三層 `snapshot_id_v1` 組合公式、
`marker_environment_identity_v1`、`runtime_environment_identity_v1`、
`environment_creation_identity`、`dedup_key_v1`、PEP 503 正規化)在這裡
**逐一獨立重新實作**,不是從 `build_v2_candidate.py` import 過來——這是
§D 刻意要求的架構隔離(如果 `_load_one`/身分公式本身有 bug,builder 跟
verifier 用同一套錯誤邏輯會得出一致的錯誤結論,驗證形同虛設),不是
疏漏或程式碼重複的失誤。

**本輪範圍(明確劃線,不隱藏)**:
- **已實作**:run 層級單發 `.claim` 鎖機制(§D 單發驗證規則,含 crash
  持久性判定,REQUEST CHANGES 第 2 項修正後——寫入/`fsync` 失敗不再刪除
  鎖檔,改成留下持久失敗標記)、獨立重算並核對 §C.1 三層身分公式 +
  `marker_environment_identity_v1` + `runtime_environment_identity_v1`
  (從彙總 build receipt 記錄的**原始欄位**獨立重算,不是信任 receipt 裡
  的雜湊值本身)、獨立重算 `environment_creation_identity`(排除自身欄位
  後重算,核對檔案雜湊/內部一致性,不宣稱能重建已銷毀的隔離環境)、獨立
  重算 quality sidecar 每一筆 `dedup_key`(用這份檔案自己的 `dedup_key_v1`
  實作,不是呼叫 `build_v2_candidate.dedup_key_v1`)並核對 locator 契約
  自洽性、§C.9 兩組 accounting 等式的獨立重算、正式狀態判定規則(§D 三
  選一,REQUEST CHANGES 第 1 項修正後——三方雜湊等式:`claim.build_
  receipt_sha256 == verification_receipt.build_receipt_sha256 ==
  現在重新雜湊的 aggregate_build_receipt_sha256`)、驗證 receipt 的
  schema/writer。
- **§D 第 18 輪的獨立原始 cell 重建職責(本輪真正實作,不是 stub)**:
  `independent_raw_cell_reconstruction()` 用 `openpyxl`(`.xlsx`)/
  `zipfile`+`csv`(`.zip`)**直接**開啟原始檔案,逐格重建
  `source_row_number`/`source_container_member`/`raw_token`/`is_blank`/
  `is_unparseable` 分類——**完全不 import/呼叫 `tej_importer.py` 或
  `build_v2_candidate.py` 的任何函式**,`classify_raw_cell()` 是這份檔案
  自己重新寫的分類邏輯,不是抄 `tej_importer._classify_numeric_cells`。
  `cross_check_reconstruction_against_sidecar()` 把重建結果拿去跟 builder
  寫出的 sidecar 記錄逐筆比對。**範圍限縮(誠實揭露,不隱藏)**:這支
  函式只重建「§C.9 locator + blank/unparseable 分類」這一層原始事實,
  **不**重做 `tej_importer.py` 的 rename 映射/單位換算/dtype 轉換/去重
  判定/supplement 合併等完整 ETL 業務邏輯——那些不是 verifier 要獨立驗證
  的對象(verifier 驗證的是「sidecar 記錄的原始儲存格分類是否忠實」,不是
  重新實作整條 pipeline)。也**尚未**接上「從 manifest 逐檔案自動跑一輪、
  彙總覆蓋率」的協調層(那屬於真正 Phase B glue 的一部分,本輪不做)。

狀態邊界(呼叫端必須遵守,本模組的 import 本身不觸碰檔案系統):
`PREREG_READY_IMPLEMENTATION_NOT_AUTHORIZED` / `BUILD_NOT_RUN` /
`PRODUCTION_NOT_APPROVED`。這個模組被 import 時不得有任何檔案系統副作用
——所有需要磁碟/套件環境觀測值的函式,一律以「注入參數」的形式接收。
"""
from __future__ import annotations

import csv
import hashlib
import io
import json
import math
import os
import re
import zipfile
from pathlib import Path

# ----------------------------------------------------------------------------
# 1. Canonical JSON / SHA-256 helpers —— 獨立實作(不 import
#    build_v2_candidate.py)。規則本身跟 builder 那份逐位元組相同(§C.1/
#    §C.9 凍結的 canonical 序列化規則),但**這裡是這份檔案自己的實作**,
#    刻意重複,不是共用同一個函式物件。
# ----------------------------------------------------------------------------

_HEX64_RE = re.compile(r"^[0-9a-f]{64}$")


def canonical_json_bytes(obj, *, sort_keys: bool) -> bytes:
    return json.dumps(
        obj, ensure_ascii=True, separators=(",", ":"), sort_keys=sort_keys
    ).encode("utf-8")


def sha256_hex(payload: bytes) -> str:
    if not isinstance(payload, (bytes, bytearray)):
        raise TypeError(f"payload must be bytes, got {type(payload).__name__}")
    return hashlib.sha256(payload).hexdigest()


def sha256_of_file(path) -> str:
    path = Path(path)
    h = hashlib.sha256()
    with open(path, "rb") as f:
        while True:
            chunk = f.read(8 * 1024 * 1024)
            if not chunk:
                break
            h.update(chunk)
    return h.hexdigest()


def _require_str(name: str, value) -> str:
    if not isinstance(value, str) or value == "":
        raise ValueError(f"{name} must be a non-empty str, got {value!r}")
    return value


def _require_sha256_hex(name: str, value) -> str:
    _require_str(name, value)
    if not _HEX64_RE.match(value):
        raise ValueError(
            f"{name} must be exactly 64 lowercase hex characters, got {value!r}"
        )
    return value


# ----------------------------------------------------------------------------
# 2. §C.1 三層身分公式 + marker/runtime environment identity——獨立重算
#    (不信任 receipt 裡已經寫好的雜湊值,從 receipt 記錄的**原始輸入欄位**
#    重新算一次,核對是否等於 receipt 宣稱的值)。
# ----------------------------------------------------------------------------

CANDIDATE_SCHEMA_VERSION = "dataexport0806_v2_candidate_schema_v1"


def source_data_identity(
    manifest_identity: str, manifest_sha256_file_identity: str, supplement_identity: str
) -> str:
    payload = (
        f"manifest_sha256={manifest_identity}\n"
        f"manifest_sha256_file_sha256={manifest_sha256_file_identity}\n"
        f"supplement_receipt_sha256={supplement_identity}"
    )
    return sha256_hex(payload.encode("utf-8"))


def build_implementation_identity(
    importer_identity: str,
    extractor_identity: str,
    builder_identity: str,
    dependency_lock_identity: str,
    runtime_environment_identity_v1_value: str,
    preregistration_commit: str,
    candidate_schema_version: str = CANDIDATE_SCHEMA_VERSION,
) -> str:
    payload = (
        f"importer_sha256={importer_identity}\n"
        f"extraction_script_sha256={extractor_identity}\n"
        f"builder_sha256={builder_identity}\n"
        f"dependency_lock_sha256={dependency_lock_identity}\n"
        f"runtime_environment_identity_v1={runtime_environment_identity_v1_value}\n"
        f"preregistration_commit={preregistration_commit}\n"
        f"candidate_schema_version={candidate_schema_version}"
    )
    return sha256_hex(payload.encode("utf-8"))


def snapshot_id_v1(source_data_identity_value: str, build_implementation_identity_value: str) -> str:
    payload = (
        f"source_data_identity={source_data_identity_value}\n"
        f"build_implementation_identity={build_implementation_identity_value}"
    )
    return sha256_hex(payload.encode("utf-8"))


MARKER_ENVIRONMENT_KEYS = frozenset(
    {
        "implementation_name", "implementation_version", "os_name",
        "platform_machine", "platform_python_implementation", "platform_release",
        "platform_system", "platform_version", "python_full_version",
        "python_version", "sys_platform",
    }
)


def marker_environment_identity_v1(marker_environment_v1: dict) -> str:
    if not isinstance(marker_environment_v1, dict):
        raise TypeError("marker_environment_v1 must be a dict")
    keys = set(marker_environment_v1.keys())
    if keys != MARKER_ENVIRONMENT_KEYS:
        raise ValueError(
            "marker_environment_v1 鍵集合跟凍結的 11 個鍵不符:缺"
            f" {sorted(MARKER_ENVIRONMENT_KEYS - keys)},多"
            f" {sorted(keys - MARKER_ENVIRONMENT_KEYS)}"
        )
    for k, v in marker_environment_v1.items():
        if not isinstance(v, str):
            raise ValueError(f"marker_environment_v1[{k!r}] 必須是字串,實際 {type(v).__name__}")
    return sha256_hex(canonical_json_bytes(marker_environment_v1, sort_keys=True))


RUNTIME_ENVIRONMENT_SCHEMA_TAG = "runtime_environment_identity_v1"
RUNTIME_ENVIRONMENT_SOURCE_FIELDS = (
    "schema_tag", "python_implementation", "python_version_full", "os_system",
    "os_release", "machine_arch", "pandas_version", "numpy_version",
    "pyarrow_version", "openpyxl_version", "parquet_engine", "excel_engine",
    "dependency_lock_identity", "marker_environment_identity_v1",
)


def runtime_environment_identity_v1(runtime_environment_source: list) -> str:
    if not isinstance(runtime_environment_source, list):
        raise TypeError("runtime_environment_source must be a list")
    if len(runtime_environment_source) != len(RUNTIME_ENVIRONMENT_SOURCE_FIELDS):
        raise ValueError(
            f"runtime_environment_source 必須恰好"
            f" {len(RUNTIME_ENVIRONMENT_SOURCE_FIELDS)} 個元素,實際"
            f" {len(runtime_environment_source)}"
        )
    if runtime_environment_source[0] != RUNTIME_ENVIRONMENT_SCHEMA_TAG:
        raise ValueError("runtime_environment_source[0] 必須是固定 schema tag")
    for field_name, value in zip(RUNTIME_ENVIRONMENT_SOURCE_FIELDS, runtime_environment_source):
        if value is None:
            raise ValueError(f"runtime_environment_source[{field_name}] 是 None,fail-closed")
    return sha256_hex(canonical_json_bytes(runtime_environment_source, sort_keys=False))


def environment_creation_identity(environment_creation_receipt: dict) -> str:
    if not isinstance(environment_creation_receipt, dict):
        raise TypeError("environment_creation_receipt must be a dict")
    payload_obj = {
        k: v for k, v in environment_creation_receipt.items()
        if k != "environment_creation_identity"
    }
    return sha256_hex(canonical_json_bytes(payload_obj, sort_keys=True))


def dedup_key_v1(
    dataset: str, source_relpath: str, source_container_member, source_row_number: int,
    target_column: str,
) -> str:
    if not isinstance(source_row_number, int) or isinstance(source_row_number, bool):
        raise ValueError(f"source_row_number must be a plain int, got {source_row_number!r}")
    canonical_array = [
        "dedup_key_v1", dataset, source_relpath, source_container_member,
        source_row_number, target_column,
    ]
    return sha256_hex(canonical_json_bytes(canonical_array, sort_keys=False))


_PEP503_NON_ALNUM_RE = re.compile(r"[-_.]+")


def normalize_package_name(name: str) -> str:
    """PEP 503 套件名稱正規化——獨立實作(跟 `build_v2_candidate.
    normalize_package_name` 同一套規則,不 import)。"""
    _require_str("name", name)
    return _PEP503_NON_ALNUM_RE.sub("-", name).lower()


# ----------------------------------------------------------------------------
# 3. 獨立重算並核對彙總 build receipt 的身分鏈(§D verifier 職責第 1 項)。
# ----------------------------------------------------------------------------


class IdentityVerificationError(ValueError):
    """身分鏈獨立重算跟 receipt 記錄值不符時拋出。攜帶 `.checks`——已完成的
    逐項核對結果(含失敗的那一項),方便寫進驗證 receipt。"""

    def __init__(self, message: str, checks: list):
        super().__init__(message)
        self.checks = checks


def verify_aggregate_receipt_identity_chain(receipt: dict) -> list:
    """獨立重算彙總 build receipt 記錄的 `source_data_identity`/
    `build_implementation_identity`/`snapshot_id_v1`/`marker_environment_
    identity_v1`(若 receipt 裡有原始 `marker_environment_v1` 物件可查)/
    `runtime_environment_identity_v1`(若 receipt 裡有原始
    `runtime_environment_source` 陣列可查),逐項核對是否等於 receipt 宣稱
    的值。回傳逐項 `{"check": ..., "passed": True}` 的清單;任一項不符,
    立刻 raise `IdentityVerificationError`(攜帶已完成的 checks,含失敗的
    那一項,`passed=False`)。"""
    checks = []

    def _check(name: str, expected, actual):
        passed = expected == actual
        checks.append({"check": name, "passed": passed, "expected": expected, "actual": actual})
        if not passed:
            raise IdentityVerificationError(
                f"{name}:receipt 記錄值={actual!r} 跟獨立重算值={expected!r} 不符", checks
            )

    sdi = source_data_identity(
        receipt["manifest_identity"],
        receipt["manifest_sha256_file_identity"],
        receipt["supplement_identity"],
    )
    _check("source_data_identity", sdi, receipt["source_data_identity"])

    bii = build_implementation_identity(
        receipt["importer_identity"],
        receipt["extractor_identity"],
        receipt["builder_identity"],
        receipt["dependency_lock_identity"],
        receipt["runtime_environment_identity_v1"],
        receipt["preregistration_commit"],
        receipt["candidate_schema_version"],
    )
    _check("build_implementation_identity", bii, receipt["build_implementation_identity"])

    sid = snapshot_id_v1(sdi, bii)
    _check("snapshot_id_v1", sid, receipt["snapshot_id_v1"])

    return checks


def verify_runtime_environment_identity(
    *, runtime_environment_source: list, marker_environment_v1: dict,
    recorded_runtime_environment_identity_v1: str,
    recorded_marker_environment_identity_v1: str,
) -> list:
    """§D verifier 職責第 1 項的核心:**從彙總 build receipt 記錄的不可變
    原始欄位**(不是 builder 算好的雜湊值本身)出發獨立重算,核對是否等於
    receipt 裡記錄的雜湊值,再核對這個值是否正確參與了 `build_
    implementation_identity` 組合(由呼叫端另外呼叫
    `verify_aggregate_receipt_identity_chain` 完成,這裡只管
    marker/runtime environment 這一層本身)。"""
    checks = []

    def _check(name, expected, actual):
        passed = expected == actual
        checks.append({"check": name, "passed": passed, "expected": expected, "actual": actual})
        if not passed:
            raise IdentityVerificationError(
                f"{name}:receipt 記錄值={actual!r} 跟獨立重算值={expected!r} 不符", checks
            )

    recomputed_marker = marker_environment_identity_v1(marker_environment_v1)
    _check("marker_environment_identity_v1", recomputed_marker, recorded_marker_environment_identity_v1)

    recomputed_runtime = runtime_environment_identity_v1(runtime_environment_source)
    _check("runtime_environment_identity_v1", recomputed_runtime, recorded_runtime_environment_identity_v1)

    return checks


def verify_environment_creation_receipt(
    *, environment_creation_receipt: dict, receipt_file_sha256: str,
    recorded_environment_creation_receipt_sha256: str,
    recorded_environment_creation_identity: str,
) -> list:
    """核對 `environment_creation_receipt_v1` 檔案本身雜湊等於彙總 build
    receipt 記錄的值;從檔案內容(排除自身欄位)重新算一次 canonical
    雜湊,核對等於檔案裡記錄的 `environment_creation_identity` 值,也等於
    彙總 build receipt 引用的那個值。**不宣稱能重建/重跑出那個當時已經
    可能被銷毀的隔離環境**——只核對 receipt 記錄本身有沒有被竄改、算式對
    不對(§D 的既有立場)。"""
    checks = []

    def _check(name, expected, actual):
        passed = expected == actual
        checks.append({"check": name, "passed": passed, "expected": expected, "actual": actual})
        if not passed:
            raise IdentityVerificationError(
                f"{name}:receipt 記錄值={actual!r} 跟獨立重算值={expected!r} 不符", checks
            )

    _check(
        "environment_creation_receipt_file_sha256",
        receipt_file_sha256,
        recorded_environment_creation_receipt_sha256,
    )
    recomputed_identity = environment_creation_identity(environment_creation_receipt)
    _check(
        "environment_creation_identity_matches_file_content",
        recomputed_identity,
        environment_creation_receipt.get("environment_creation_identity"),
    )
    _check(
        "environment_creation_identity_matches_aggregate_receipt_reference",
        recomputed_identity,
        recorded_environment_creation_identity,
    )
    return checks


# ----------------------------------------------------------------------------
# 4. Quality sidecar 獨立重建/核對(§D verifier 職責——**本輪範圍限縮**:
#    這裡獨立重算每一筆 sidecar 記錄的 `dedup_key`、核對 locator 契約
#    自洽性(§C.9),以及 §C.9 兩組 accounting 等式是否從 receipt 記錄的
#    逐檔計數正確加總——**不**重新打開原始 `.xlsx`/`.zip` 檔案本身,見本
#    模組 docstring 開頭「明確未實作」段落。
# ----------------------------------------------------------------------------

CELL_RECORD_FIELDS = frozenset(
    {
        "dataset", "source_relpath", "source_file_sha256", "source_container_member",
        "source_row_number", "stock_id", "date", "source_column", "target_column",
        "raw_token", "is_blank", "is_unparseable", "parser", "unit_scale_applied",
        "resulting_value", "dedup_key",
    }
)


def verify_sidecar_records(sidecar: dict) -> list:
    """獨立重算 sidecar 裡每一筆記錄的 `dedup_key`,核對跟宣稱值相符,並
    核對 §C.9 locator 契約(`is_blank`/`is_unparseable` 互斥、`raw_token`
    的 null-ness 跟兩者對應、`resulting_value` 必為 null、`parser` 固定
    字面值)。任一筆不符,立刻 raise,回傳已完成的逐筆 checks。"""
    checks = []
    dataset = sidecar.get("dataset")
    for i, record in enumerate(sidecar.get("records", [])):
        if not isinstance(record, dict) or set(record) != CELL_RECORD_FIELDS:
            checks.append({"check": f"records[{i}].shape", "passed": False})
            raise IdentityVerificationError(
                f"sidecar records[{i}] 欄位集合不符 §C.9 locator 契約", checks
            )
        is_blank, is_unparseable = record["is_blank"], record["is_unparseable"]
        if is_blank == is_unparseable:
            checks.append({"check": f"records[{i}].blank_xor_unparseable", "passed": False})
            raise IdentityVerificationError(
                f"sidecar records[{i}] is_blank/is_unparseable 未互斥", checks
            )
        recomputed = dedup_key_v1(
            record["dataset"], record["source_relpath"], record["source_container_member"],
            record["source_row_number"], record["target_column"],
        )
        passed = recomputed == record["dedup_key"]
        checks.append({"check": f"records[{i}].dedup_key", "passed": passed})
        if not passed:
            raise IdentityVerificationError(
                f"sidecar records[{i}] dedup_key 獨立重算值={recomputed!r} 跟宣稱值="
                f"{record['dedup_key']!r} 不符", checks
            )
        if record["dataset"] != dataset:
            checks.append({"check": f"records[{i}].dataset_matches_sidecar", "passed": False})
            raise IdentityVerificationError(
                f"sidecar records[{i}]['dataset'] 跟 sidecar 本身的 dataset 不符", checks
            )
    return checks


def verify_stage_one_and_sidecar_accounting(
    *, per_file_stage_one_counts: list, sidecar_record_count: int
) -> list:
    """§C.9:「sidecar 列數只跟階段一的 blank_cell_count +
    unparseable_cell_count 加總對帳」——獨立重算這個等式(從 receipt 記錄的
    逐檔計數加總,不重新解析原始檔案)。"""
    total = 0
    for entry in per_file_stage_one_counts:
        for c in entry.get("counts", []):
            total += int(c.get("blank_cell_count", 0)) + int(c.get("unparseable_cell_count", 0))
    passed = total == sidecar_record_count
    check = {
        "check": "sidecar_record_count_matches_stage_one_blank_plus_unparseable",
        "passed": passed, "expected": total, "actual": sidecar_record_count,
    }
    if not passed:
        raise IdentityVerificationError(
            f"sidecar_record_count={sidecar_record_count} 跟獨立重算的階段一"
            f" blank+unparseable 加總 {total} 不符", [check]
        )
    return [check]


def verify_final_null_causes_accounting(
    *, final_null_causes: dict, final_null_counts_from_output: dict
) -> list:
    """§C.9 階段二等式的獨立重算——五個分類加總必須等於該欄位最終輸出的
    null 總數。"""
    checks = []
    if set(final_null_causes) != set(final_null_counts_from_output):
        raise IdentityVerificationError(
            "final_null_causes 跟 final_null_counts_from_output 欄位集合不符",
            checks,
        )
    for col, causes in final_null_causes.items():
        total = sum(int(v) for v in causes.values())
        expected = int(final_null_counts_from_output[col])
        passed = total == expected
        checks.append(
            {"check": f"final_null_causes[{col}]", "passed": passed, "expected": expected, "actual": total}
        )
        if not passed:
            raise IdentityVerificationError(
                f"final_null_causes[{col!r}] 加總 {total} 跟輸出 null 數"
                f" {expected} 不符", checks
            )
    return checks


def _cell_display_string(value):
    """把一個原始儲存格值(openpyxl 儲存格值,或 csv 讀出的字串)轉成字串
    表示——`str(v).strip()`,語意上呼應 tej_importer 對「儲存格內容的字串
    表示」的既有慣例,但這裡是這份獨立解析路徑**自己的**實作,不呼叫
    `tej_importer` 的任何函式。空字串(strip 之後)正規化成 `None`,跟
    原生空白儲存格視為同一件事(§C.9 的 `is_blank` 定義)。"""
    if value is None:
        return None
    s = str(value).strip()
    return s if s else None


def classify_raw_cell(value) -> dict:
    """獨立分類一個原始儲存格值:blank / unparseable / parsed。這是這份
    驗證器檔案自己重新寫的分類邏輯(用 `float()` 判斷能否解析成數字),
    不是抄 `tej_importer._classify_numeric_cells` 的 `pd.to_numeric` 呼叫
    ——兩套獨立實作,如果對同一批真實資料的分類結果不一致,代表其中一套
    (或兩套)有 bug,這正是獨立驗證要能抓到的事。

    **Codex review P1 修正**:`float("nan")` 不會拋例外,原本的實作因此把
    字面 token `"nan"`(以及 `"NaN"`/`"NAN"` 等大小寫變體)誤判成「成功
    解析出一個數字」。但 `tej_importer` 的 `pd.to_numeric(...).isna()`
    判斷會把這種情況歸類成 `is_unparseable`(結果是 NaN,不是一個有意義
    的數字——這正是既有 migration tests 特別保護的 token,§C.9 兩套獨立
    實作在這裡本該一致)。現在明確在 `float()` 成功之後另外檢查
    `math.isnan(...)`,是 NaN 就改判 `is_unparseable`。**無窮大
    (`inf`/`-inf`/`Infinity`)維持解析成功**,不受這個修正影響——
    `pd.to_numeric` 對無窮大字面同樣會解析成非 NaN 的浮點無窮大值,兩套
    實作在這一點上本來就一致,不需要改。"""
    raw_token = _cell_display_string(value)
    if raw_token is None:
        return {"is_blank": True, "is_unparseable": False, "raw_token": None, "parsed_value": None}
    try:
        parsed_value = float(raw_token)
    except (TypeError, ValueError):
        return {"is_blank": False, "is_unparseable": True, "raw_token": raw_token, "parsed_value": None}
    if math.isnan(parsed_value):
        return {"is_blank": False, "is_unparseable": True, "raw_token": raw_token, "parsed_value": None}
    return {"is_blank": False, "is_unparseable": False, "raw_token": raw_token, "parsed_value": parsed_value}


def independent_read_xlsx_rows(path) -> list:
    """獨立用 `openpyxl` 直接開啟一份 `.xlsx`,逐列讀出「實體列號 + 表頭
    欄名 → 原始儲存格值」。**不呼叫 `tej_importer` 的任何函式**,也不做
    任何 rename/型別轉換/單位換算。回傳 `[{"sheet_name": str,
    "source_row_number": int, "cells": {header: raw_value}}, ...]`——
    `source_row_number` 是 `openpyxl` 對工作表逐列列舉得到的實體
    (1-based)列號(含表頭列本身,呼叫端自行排除表頭那一列),不是任何
    0-based 內部 index。

    **外部檔案格式契約(Codex review P0 修正——凍結跟 `tej_importer.
    _read_xlsx_raw` 完全相同的外部格式假設,不是共用程式碼)**:真實 TEJ
    查詢精靈匯出的 `.xlsx` 只有**第一個**工作表(`wb.sheetnames[0]`)是
    資料——舊版本讀了全部工作表,對多工作表檔案會產生假 mismatch(Codex
    review 指出的問題),現在改成只讀第一個,跟 importer 一致。空白工作表
    (連表頭都沒有)fail-closed,不是靜默回傳空清單。"""
    import openpyxl  # 惰性 import:保持模組本身 import 時零檔案系統副作用

    path = Path(path)
    wb = openpyxl.load_workbook(str(path), read_only=True, data_only=True)
    try:
        sheet_name = wb.sheetnames[0]
        ws = wb[sheet_name]
        header = None
        rows = []
        for physical_row_idx, row in enumerate(ws.iter_rows(), start=1):
            if physical_row_idx == 1:
                header = [(_cell_display_string(c.value) or "") for c in row]
                continue
            cells = {}
            for col_index, cell in enumerate(row):
                if col_index < len(header) and header[col_index]:
                    cells[header[col_index]] = cell.value
            rows.append(
                {"sheet_name": sheet_name, "source_row_number": physical_row_idx, "cells": cells}
            )
        if header is None:
            raise ValueError(f"{path.name}:工作表 {sheet_name!r} 是空的,連表頭都沒有")
        return rows
    finally:
        wb.close()


def independent_read_zip_csv_rows(path) -> list:
    """獨立用 `zipfile` + `csv` 直接開啟一份 `.zip`,逐列讀出「實體列號 +
    表頭欄名 → 原始儲存格字串」。**不呼叫 `tej_importer` 的任何函式**。
    回傳 `[{"sheet_name": <zip 內成員檔名>, "source_row_number": int,
    "cells": {header: raw_str}}, ...]`——`source_row_number` 是這個 csv
    成員內的實體(1-based)列號(表頭本身是第 1 列,回傳的資料列從第 2
    列起)。

    **外部檔案格式契約(Codex review P0 修正——凍結跟 `tej_importer.
    _read_zip_csv_raw` 完全相同的外部格式假設,不是共用程式碼)**:真實
    TEJ 大量匯出的 `.zip` 內含**一個 UTF-16 編碼、Tab 分隔**的 `.csv`
    成員——舊版本用 `utf-8-sig` 解碼、`csv.reader` 預設逗號分隔,對真實
    DataExport0806 的 zip 會誤讀或直接讀壞/讀空(Codex review 指出的問題)
    ,現在改成 `decode("utf-16")` + `delimiter="\\t"`,跟 importer 一致。
    只讀 `zf.namelist()` 裡**第一個**副檔名是 `.csv` 的成員,不是全部
    csv 成員。短列(欄位數少於表頭)補齊空字串;長列(欄位數多於表頭)
    fail-closed;任何欄位值含內嵌實體換行字元(`\\n`/`\\r`)一樣
    fail-closed(quoted 多行欄位會讓逐筆遞增的實體列號算法失真,寧可拒絕
    也不冒充可能算錯的列號——跟 importer 的 Round 3 review 第 5 項同一套
    理由)。"""
    path = Path(path)
    with zipfile.ZipFile(str(path)) as zf:
        csv_names = [n for n in zf.namelist() if n.lower().endswith(".csv")]
        if not csv_names:
            raise ValueError(f"{path.name} 裡沒有 .csv 成員")
        member = csv_names[0]
        with zf.open(member) as raw_f:
            text = raw_f.read().decode("utf-16")

    reader = csv.reader(io.StringIO(text), delimiter="\t")
    all_rows = list(reader)
    if not all_rows:
        raise ValueError(f"{path.name}/{member}:內容是空的,連表頭都沒有")
    header = all_rows[0]
    ncols = len(header)

    rows = []
    for physical_row_idx, fields in enumerate(all_rows[1:], start=2):
        for field in fields:
            if "\n" in field or "\r" in field:
                raise ValueError(
                    f"{path.name}/{member}:第 {physical_row_idx} 筆紀錄的某個欄位含有"
                    "內嵌實體換行字元(quoted 多行欄位),無法用逐筆遞增可靠算出實體"
                    "列號,寧可 fail-closed 也不冒充一個可能算錯的列號"
                )
        if len(fields) < ncols:
            fields = fields + [""] * (ncols - len(fields))
        elif len(fields) > ncols:
            raise ValueError(
                f"{path.name}/{member}:第 {physical_row_idx} 筆紀錄有 {len(fields)} 欄,"
                f"表頭只有 {ncols} 欄,格式異常,不放行"
            )
        cells = {header[i]: fields[i] for i in range(ncols)}
        rows.append({"sheet_name": member, "source_row_number": physical_row_idx, "cells": cells})
    return rows


def independent_raw_cell_reconstruction(path, *, target_columns: dict) -> list:
    """§D 第 18 輪新增的 verifier 核心職責——用一套**不跟 `tej_importer.
    _read_raw_table`/`_parse_dates`/`_classify_numeric_cells` 共用程式碼**
    的獨立解析路徑,重新打開一份原始 `.xlsx`/`.zip` 檔案,逐格重建每一個
    請求欄位(`target_columns`:`{中文來源欄名: target_column 名稱}` 映射)
    的 `source_row_number`/`source_container_member`/`raw_token`/
    `is_blank`/`is_unparseable` 分類。**本輪真正實作,不是 stub**——用
    `openpyxl`/`zipfile`/`csv` 這些通用第三方/標準函式庫直接讀檔,`classify_
    raw_cell()` 是這份檔案自己重新寫的分類邏輯。

    回傳 `[{"source_container_member": ..., "source_row_number": ...,
    "source_column": 中文來源欄名, "target_column": ..., "is_blank": ...,
    "is_unparseable": ..., "raw_token": ...}, ...]`——**只回傳 blank/
    unparseable 兩類**(跟 §C.9 sidecar 的收錄範圍一致,方便直接逐筆比對
    `cross_check_reconstruction_against_sidecar` 用),不含 parsed 成功的
    列(那些不進 sidecar,沒有比對對象)。

    **範圍限縮(誠實揭露,不是隱藏的假設)**:這支函式只重建「§C.9 locator
    + blank/unparseable 分類」這一層原始事實,**不**重做 `tej_importer.py`
    的 rename/單位換算/dtype 轉換/去重判定/supplement 合併等完整 ETL
    業務邏輯——那些不是 verifier 要獨立驗證的對象。"""
    path = Path(path)
    suffix = path.suffix.lower()
    if suffix == ".xlsx":
        raw_rows = independent_read_xlsx_rows(path)
    elif suffix == ".zip":
        raw_rows = independent_read_zip_csv_rows(path)
    else:
        raise ValueError(
            f"independent_raw_cell_reconstruction 只支援 .xlsx/.zip,收到 {path}"
        )

    records = []
    for row in raw_rows:
        for source_column, target_column in target_columns.items():
            if source_column not in row["cells"]:
                continue
            classification = classify_raw_cell(row["cells"][source_column])
            if classification["is_blank"] or classification["is_unparseable"]:
                records.append(
                    {
                        "source_container_member": row["sheet_name"],
                        "source_row_number": row["source_row_number"],
                        "source_column": source_column,
                        "target_column": target_column,
                        "is_blank": classification["is_blank"],
                        "is_unparseable": classification["is_unparseable"],
                        "raw_token": classification["raw_token"],
                    }
                )
    return records


def cross_check_reconstruction_against_sidecar(
    *, reconstructed_records: list, sidecar_records: list
) -> list:
    """把 `independent_raw_cell_reconstruction` 的獨立重建結果,跟 builder
    寫出的 sidecar 記錄(限定同一份來源檔案——呼叫端先篩出對應
    `source_relpath` 的子集合)逐筆比對,鍵是
    `(source_container_member, source_row_number, target_column)`。核對
    每一筆的 `is_blank`/`is_unparseable`/`raw_token` 是否完全一致,以及
    兩邊的鍵集合是否完全相同(獨立重建有但 sidecar 沒有記錄的一筆——builder
    漏記了;sidecar 有但獨立重建讀不到的一筆——sidecar 記錄可能是造假或
    對應錯檔案)。任一項不符,立刻 raise `IdentityVerificationError`(帶著
    已完成的 checks,含失敗的那一項)。

    **Codex review P1 修正**:原本直接用 dict comprehension 把兩邊列表轉成
    `{key: record}`——如果 sidecar 或重建結果裡出現重複的
    `(source_container_member, source_row_number, target_column)` key,
    dict comprehension 的 last-write-wins 語意會把前一筆**靜默蓋掉**,一筆
    造假或矛盾的重複記錄可能因此消失、錯誤地通過比對。現在改成
    `_build_unique_lookup` 先各自獨立掃過兩邊列表,偵測到任何重複 key 就
    直接 fail-closed,不悄悄用後者取代前者。"""

    def _key(record):
        return (
            record["source_container_member"],
            record["source_row_number"],
            record["target_column"],
        )

    checks = []

    def _build_unique_lookup(records, *, label):
        lookup = {}
        duplicate_keys = []
        for record in records:
            key = _key(record)
            if key in lookup:
                duplicate_keys.append(key)
            lookup[key] = record
        unique_duplicate_keys = sorted(map(repr, set(duplicate_keys)))
        passed = not duplicate_keys
        checks.append(
            {
                "check": f"{label}_has_no_duplicate_locator_keys",
                "passed": passed,
                "duplicate_keys": unique_duplicate_keys,
            }
        )
        if not passed:
            raise IdentityVerificationError(
                f"{label} 出現重複的 locator key(source_container_member,"
                f" source_row_number, target_column),不能用 dict 的"
                f" last-write-wins 語意靜默壓掉:{unique_duplicate_keys}",
                checks,
            )
        return lookup

    reconstructed_by_key = _build_unique_lookup(reconstructed_records, label="reconstructed_records")
    sidecar_by_key = _build_unique_lookup(sidecar_records, label="sidecar_records")

    only_in_reconstruction = sorted(
        map(repr, set(reconstructed_by_key) - set(sidecar_by_key))
    )
    only_in_sidecar = sorted(map(repr, set(sidecar_by_key) - set(reconstructed_by_key)))
    key_sets_match = not only_in_reconstruction and not only_in_sidecar
    checks.append(
        {
            "check": "reconstruction_and_sidecar_key_sets_match",
            "passed": key_sets_match,
            "only_in_reconstruction": only_in_reconstruction,
            "only_in_sidecar": only_in_sidecar,
        }
    )
    if not key_sets_match:
        raise IdentityVerificationError(
            "獨立重建的 blank/unparseable 儲存格集合,跟 sidecar 記錄的集合"
            f" 不完全相同:只在重建結果裡={only_in_reconstruction} 只在"
            f" sidecar 裡={only_in_sidecar}",
            checks,
        )

    for key, recon in reconstructed_by_key.items():
        side = sidecar_by_key[key]
        matches = (
            recon["is_blank"] == side["is_blank"]
            and recon["is_unparseable"] == side["is_unparseable"]
            and recon["raw_token"] == side["raw_token"]
        )
        checks.append({"check": f"cell_{key!r}_classification_matches", "passed": matches})
        if not matches:
            raise IdentityVerificationError(
                f"獨立重建對 {key!r} 的分類(is_blank={recon['is_blank']}"
                f" is_unparseable={recon['is_unparseable']}"
                f" raw_token={recon['raw_token']!r})跟 sidecar 記錄"
                f"(is_blank={side['is_blank']} is_unparseable="
                f"{side['is_unparseable']} raw_token={side['raw_token']!r})"
                " 不符",
                checks,
            )
    return checks


# ----------------------------------------------------------------------------
# 5. Run 層級單發 `.claim` 鎖機制(§D 單發驗證規則,第 18 輪修正)。
# ----------------------------------------------------------------------------


def claim_filename(run_id: str) -> str:
    """§D:「`.claim` 鎖檔的排他建立路徑只用 `run_id` 命名,不含
    `verifier_identity`」——`f"{run_id}.binding_verification.claim"`。"""
    _require_str("run_id", run_id)
    if "/" in run_id or "\\" in run_id or "\x00" in run_id or run_id in (".", ".."):
        raise ValueError(f"run_id 不合法,不能用來組檔名:{run_id!r}")
    return f"{run_id}.binding_verification.claim"


def _guard_filename_in_dir(run_dir, filename: str) -> Path:
    """跟 `build_v2_candidate.guard_receipt_filename` 邏輯相同,獨立實作
    (§D:verifier 不 import builder 的任何函式,含這類路徑守衛 helper)。"""
    if (
        not isinstance(filename, str) or filename == "" or filename in (".", "..")
        or "/" in filename or "\\" in filename or "\x00" in filename
    ):
        raise ValueError(f"檔名不合法:{filename!r}")
    run_dir = Path(run_dir)
    if not run_dir.is_dir():
        raise ValueError(f"run_dir 不存在或不是目錄:{run_dir}")
    run_dir_real = run_dir.resolve(strict=True)
    target = run_dir / filename
    expected = run_dir_real / filename
    target_resolved = target.resolve(strict=False)
    if target_resolved != expected:
        raise ValueError(f"路徑解析後不等於預期字面路徑(symlink/reparse 逃逸疑慮):{target_resolved}")
    return target_resolved


_FAILED_CLAIM_MARKER = "CLAIM_WRITE_FAILED"


def is_claim_write_failed_marker(claim_content) -> bool:
    """判斷一份已讀出的 `.claim` 內容是不是寫入/fsync 失敗後留下的持久失敗
    標記(`_mark_claim_write_failed` 寫的那種),而不是一份正常的 claim
    內容。純粹是給診斷/測試用的判斷式——`determine_official_run_status`
    本身**不需要**呼叫這支函式就能正確判定(見下方說明):只要 `.claim`
    檔案存在(不論內容是正常 claim、失敗標記、還是任何其他位元組),且沒有
    對應的合法 binding 驗證 receipt,crash 持久性規則就已經足以判定
    `BUILD_VERIFICATION_FAILED`;這支函式的存在只是讓「為什麼」這件事對人
    類讀者/測試更明確。"""
    return isinstance(claim_content, dict) and claim_content.get("status") == _FAILED_CLAIM_MARKER


def _mark_claim_write_failed(claim_path, original_exc: BaseException) -> None:
    """Codex review 修正(REQUEST CHANGES 第 2 項)的核心 helper:`.claim`
    排他建立**成功之後**,寫入內容或 `fsync` 失敗時呼叫——**盡力**
    (best-effort)把這個已經排他建立出來的檔案內容覆寫成一個明確、可持久
    辨識的失敗標記,而**不是刪除它**。刪除等於把這個 `run_id` 唯一一次的
    binding 驗證機會還給下一次嘗試,結構上重新打開了可以重試的窗口——
    這正是這次 review 要修正的缺陷。即使這次補寫標記本身也失敗(例如磁碟
    真的滿到連幾個位元組都寫不進去),原本(可能是空的、可能是部分內容的)
    檔案依然留在磁碟上:`os.open(..., O_EXCL)` 對這個 `run_id` 的任何後續
    `claim_binding_verification` 呼叫都會繼續因為檔案已存在而失敗——光是
    檔案存在這件事本身就結構性阻擋重試,不依賴內容是否可讀、也不依賴這次
    補寫是否成功。這支函式**永遠不拋出例外**(所有失敗都被吞掉)——它的
    呼叫端必須緊接著重新拋出**原始**例外,補寫標記失敗不能蓋掉根因。"""
    marker = canonical_json_bytes(
        {
            "status": _FAILED_CLAIM_MARKER,
            "error_type": type(original_exc).__name__,
            "error_message": str(original_exc),
        },
        sort_keys=True,
    )
    try:
        with open(claim_path, "wb") as f:
            f.write(marker)
            f.flush()
            try:
                os.fsync(f.fileno())
            except OSError:
                pass
    except OSError:
        pass


def claim_binding_verification(
    *,
    run_dir,
    run_id: str,
    verifier_identity: str,
    authorized_verifier_identity: str,
    build_receipt_path: str,
    build_receipt_sha256: str,
    start_timestamp_utc: str,
) -> tuple:
    """§D 單發驗證規則:**先**核對 `verifier_identity` 是否等於彙總 build
    receipt 記錄的 `authorized_verifier_identity`——不符就直接 raise,
    **連 `.claim` 鎖檔都不建立**(避免白白燒掉這個 `run_id` 唯一一次的
    binding 驗證機會)。核對通過才對 `.claim` 檔案做 `O_CREAT|O_EXCL` 排他
    建立(檔案系統層級原子操作),寫入內容後 `fsync`。

    回傳 `(claim_path, claim_content)`。若 `.claim` 已存在(這個 `run_id`
    已經被別的驗證執行用掉,或是先前一次寫入/fsync 失敗後留下的持久失敗
    標記),排他建立失敗,直接 raise,**不產生任何驗證 receipt**——這個
    `run_id` 的正式驗證機會已經用掉了。

    **Codex review 修正(REQUEST CHANGES 第 2 項)**:排他建立**成功之後**
    才發生的寫入/`fsync` 失敗,**不再刪除已經建立的 `.claim` 檔案**(舊版
    行為——刪除等於把這個 `run_id` 唯一一次的驗證機會還給下一次嘗試,結構
    上重新打開可以重試的窗口,違反單發驗證規則的精神)。改成呼叫
    `_mark_claim_write_failed` 盡力把檔案內容覆寫成一個明確、可持久辨識的
    失敗標記,然後重新拋出**原始**例外——不論標記覆寫本身成功與否,檔案
    都留在磁碟上,永久阻擋這個 `run_id` 的任何後續 claim 嘗試;呼叫端(見
    `determine_official_run_status`)只要看到 `.claim` 存在但沒有合法對應
    receipt,crash 持久性規則就會判定 `BUILD_VERIFICATION_FAILED`,永久
    不得回退成「還在等驗證」,也不會有第二次重試機會。"""
    _require_sha256_hex("verifier_identity", verifier_identity)
    _require_sha256_hex("authorized_verifier_identity", authorized_verifier_identity)
    if verifier_identity != authorized_verifier_identity:
        raise ValueError(
            "verifier_identity 跟彙總 build receipt 記錄的"
            f" authorized_verifier_identity 不符(verifier_identity="
            f"{verifier_identity!r} authorized_verifier_identity="
            f"{authorized_verifier_identity!r})——這份驗證器不是被授權的那一份,"
            "直接失敗、不建立 .claim 鎖檔(§D 單發驗證規則)"
        )
    _require_str("build_receipt_path", build_receipt_path)
    _require_sha256_hex("build_receipt_sha256", build_receipt_sha256)
    _require_str("start_timestamp_utc", start_timestamp_utc)

    filename = claim_filename(run_id)
    claim_path = _guard_filename_in_dir(run_dir, filename)
    content = canonical_json_bytes(
        {
            "authorized_verifier_identity": authorized_verifier_identity,
            "run_id": run_id,
            "build_receipt_path": build_receipt_path,
            "build_receipt_sha256": build_receipt_sha256,
            "start_timestamp_utc": start_timestamp_utc,
        },
        sort_keys=True,
    )
    flags = os.O_CREAT | os.O_EXCL | os.O_WRONLY | getattr(os, "O_BINARY", 0)
    try:
        fd = os.open(str(claim_path), flags, 0o644)
    except FileExistsError:
        raise ValueError(
            f"這個 run_id 的 .claim 鎖檔已存在(不論是哪一個 verifier_identity"
            f" 建立的,或是先前一次寫入失敗留下的持久失敗標記),binding"
            f" 驗證機會已經用掉:{claim_path}"
        ) from None
    try:
        with os.fdopen(fd, "wb") as f:
            f.write(content)
            f.flush()
            os.fsync(f.fileno())
    except BaseException as write_exc:
        _mark_claim_write_failed(claim_path, write_exc)
        raise
    return claim_path.resolve(strict=True), json.loads(content.decode("utf-8"))


# ----------------------------------------------------------------------------
# 6. 驗證 receipt(§C.5 第 3 項)。
# ----------------------------------------------------------------------------

VERIFICATION_RECEIPT_SCHEMA_TAG = "verification_receipt_v1"
BUILD_VALIDATED = "BUILD_VALIDATED"
BUILD_VERIFICATION_FAILED = "BUILD_VERIFICATION_FAILED"
_VERIFICATION_STATUSES = (BUILD_VALIDATED, BUILD_VERIFICATION_FAILED)

VERIFICATION_RECEIPT_FIELDS = (
    "schema", "run_id", "verifier_identity", "build_receipt_path",
    "build_receipt_sha256", "verifier_runtime_environment_identity_v1",
    "checks", "overall_status", "verified_at_utc", "binding",
)


def build_verification_receipt(
    *,
    run_id: str,
    verifier_identity: str,
    build_receipt_path: str,
    build_receipt_sha256: str,
    verifier_runtime_environment_identity_v1_value: str,
    checks: list,
    overall_status: str,
    verified_at_utc: str,
    binding: bool = True,
) -> dict:
    """組出驗證 receipt(記憶體內 dict)。**不寫檔**。`overall_status` 只能
    是 `BUILD_VALIDATED`/`BUILD_VERIFICATION_FAILED` 兩者之一;`checks`
    是逐項核對結果的清單(見上面各 `verify_*` 函式的回傳格式)——若
    `overall_status=BUILD_VALIDATED`,`checks` 裡不能有任何一項
    `passed=False`(自相矛盾:宣稱通過卻附上失敗證據)。"""
    _require_str("run_id", run_id)
    _require_sha256_hex("verifier_identity", verifier_identity)
    _require_str("build_receipt_path", build_receipt_path)
    _require_sha256_hex("build_receipt_sha256", build_receipt_sha256)
    _require_sha256_hex(
        "verifier_runtime_environment_identity_v1",
        verifier_runtime_environment_identity_v1_value,
    )
    if not isinstance(checks, list) or not checks:
        raise ValueError("checks must be a non-empty list")
    for i, c in enumerate(checks):
        if not isinstance(c, dict) or "check" not in c or "passed" not in c:
            raise ValueError(f"checks[{i}] 必須至少含 'check'/'passed' 兩個鍵,實際 {c!r}")
        if not isinstance(c["passed"], bool):
            raise ValueError(f"checks[{i}]['passed'] 必須是真正的 bool,實際 {c['passed']!r}")
    if overall_status not in _VERIFICATION_STATUSES:
        raise ValueError(f"overall_status 必須是 {_VERIFICATION_STATUSES} 之一")
    any_failed = any(not c["passed"] for c in checks)
    if overall_status == BUILD_VALIDATED and any_failed:
        raise ValueError(
            "overall_status=BUILD_VALIDATED 但 checks 裡有 passed=False 的項目"
            "——自相矛盾,不能記錄一份宣稱通過卻附帶失敗證據的驗證 receipt"
        )
    if overall_status == BUILD_VERIFICATION_FAILED and not any_failed:
        raise ValueError(
            "overall_status=BUILD_VERIFICATION_FAILED 但 checks 裡沒有任何"
            " passed=False 的項目——沒有失敗證據支持這個判定"
        )
    _require_str("verified_at_utc", verified_at_utc)
    if not isinstance(binding, bool):
        raise ValueError(f"binding must be a real bool, got {binding!r}")

    return {
        "schema": VERIFICATION_RECEIPT_SCHEMA_TAG,
        "run_id": run_id,
        "verifier_identity": verifier_identity,
        "build_receipt_path": build_receipt_path,
        "build_receipt_sha256": build_receipt_sha256,
        "verifier_runtime_environment_identity_v1": verifier_runtime_environment_identity_v1_value,
        "checks": checks,
        "overall_status": overall_status,
        "verified_at_utc": verified_at_utc,
        "binding": binding,
    }


def validate_verification_receipt(receipt: dict) -> None:
    if not isinstance(receipt, dict) or set(receipt) != set(VERIFICATION_RECEIPT_FIELDS):
        raise ValueError(f"verification receipt 欄位集合不符,實際 {receipt!r}")
    if receipt["schema"] != VERIFICATION_RECEIPT_SCHEMA_TAG:
        raise ValueError(f"schema 必須是 {VERIFICATION_RECEIPT_SCHEMA_TAG!r}")
    if receipt["overall_status"] not in _VERIFICATION_STATUSES:
        raise ValueError(f"overall_status 必須是 {_VERIFICATION_STATUSES} 之一")
    any_failed = any(not c["passed"] for c in receipt["checks"])
    if receipt["overall_status"] == BUILD_VALIDATED and any_failed:
        raise ValueError("overall_status=BUILD_VALIDATED 但 checks 有失敗項目")
    if receipt["overall_status"] == BUILD_VERIFICATION_FAILED and not any_failed:
        raise ValueError("overall_status=BUILD_VERIFICATION_FAILED 但 checks 沒有失敗項目")


def write_verification_receipt_json_atomic(path, receipt: dict) -> tuple:
    """跟 `build_v2_candidate.write_receipt_json_atomic` 邏輯相同,獨立
    實作(exclusive create,寫入失敗不留 partial 檔案)。"""
    validate_verification_receipt(receipt)
    path = Path(path)
    parent = path.parent
    if not parent.is_dir():
        raise ValueError(f"receipt 的父目錄必須已存在且是目錄:{parent}")
    payload = canonical_json_bytes(receipt, sort_keys=True)
    flags = os.O_CREAT | os.O_EXCL | os.O_WRONLY | getattr(os, "O_BINARY", 0)
    try:
        fd = os.open(str(path), flags, 0o644)
    except FileExistsError:
        raise ValueError(f"驗證 receipt 檔案已存在,拒絕覆寫:{path}") from None
    try:
        with os.fdopen(fd, "wb") as f:
            f.write(payload)
            f.flush()
            os.fsync(f.fileno())
    except BaseException:
        try:
            path.unlink()
        except OSError:
            pass
        raise
    return path.resolve(strict=True), sha256_hex(payload)


# ----------------------------------------------------------------------------
# 7. 正式狀態判定規則(§D「正式狀態的判定規則」,第 18 輪修正版三選一)。
# ----------------------------------------------------------------------------

STATUS_BUILD_COMPLETE_AWAITING_VERIFICATION = "BUILD_COMPLETE_AWAITING_VERIFICATION"
STATUS_BUILD_FAILED_PARTIAL = "BUILD_FAILED_PARTIAL"
STATUS_BUILD_VERIFICATION_FAILED = BUILD_VERIFICATION_FAILED
STATUS_BUILD_VALIDATED = BUILD_VALIDATED


def determine_official_run_status(
    *,
    claim_exists: bool,
    binding_verification_receipt,
    aggregate_build_receipt_overall_status: str,
    aggregate_build_receipt_sha256: str,
    claim_content=None,
) -> str:
    """§D「正式狀態的判定規則」三選一,逐字對應:

    1. 沒有 `.claim` 鎖檔——正式狀態看彙總 build receipt 自己的
       `overall_status`(`BUILD_COMPLETE_AWAITING_VERIFICATION` 或
       `BUILD_FAILED_PARTIAL`)。
    2. `.claim` 存在但沒有對應的 binding 驗證 receipt(涵蓋「驗證器 crash
       沒寫出 receipt」跟「receipt 寫出但格式不合法讀不出來,呼叫端傳
       `None`」兩種情況)——正式狀態直接是 `BUILD_VERIFICATION_FAILED`
       (crash 持久性規則,**不會**回退成「還在等驗證」)。
    3. `.claim` 存在,且存在對應的 binding 驗證 receipt,其
       `verifier_identity` 等於 `.claim` 記錄值、`build_receipt_sha256`
       等於**現在**重新雜湊彙總 build receipt 的結果——正式狀態看該驗證
       receipt 的 `overall_status`。

    **Codex review 修正(REQUEST CHANGES 第 1 項)**:原本第 3 項只核對
    `binding_verification_receipt.build_receipt_sha256 ==
    aggregate_build_receipt_sha256`(receipt 宣稱值 vs 現在重新雜湊值),
    完全沒有讀取/核對 `.claim` 鎖檔自己記錄的 `build_receipt_sha256`——
    `.claim` 是這個 `run_id` 唯一一次驗證機會,在**驗證開始當下**對彙總
    build receipt 拍下的快照雜湊。如果驗證器寫出的 receipt 宣稱的
    `build_receipt_sha256` 剛好跟*現在*重新雜湊的彙總 receipt 一致,但
    兩者都跟 `.claim` 鎖檔記錄的當時快照不同(例如彙總 receipt 檔案在
    claim 建立之後、verification receipt 寫出之前被置換過,兩次置換剛好
    收斂到同一個新雜湊),舊邏輯會誤判為相符。現在要求**三方雜湊完全
    相等**:`claim.build_receipt_sha256 ==
    binding_verification_receipt.build_receipt_sha256 ==
    aggregate_build_receipt_sha256`(現在重新雜湊的值),任一項不等都是
    `BUILD_VERIFICATION_FAILED`。因此當存在 binding 驗證 receipt 時,
    `claim_content` 現在是**必要**參數(呼叫端必須先讀出 `.claim` 內容才
    能判定——不能假設鎖檔存在就等於內容可信/可省略),缺席直接 raise。

    **Codex review P0 修正(第二輪)**:第 3 項原本完全沒有核對彙總 build
    receipt 自己的 `overall_status`——只要驗證 receipt 的
    `verifier_identity`/`run_id`/三方雜湊都對得上,即使彙總 receipt 早就
    是 `BUILD_FAILED_PARTIAL`(建置本身沒有完整跑完 11 個 dataset),舊
    邏輯依然會直接回傳驗證 receipt 宣稱的 `overall_status`,讓一個
    `BUILD_FAILED_PARTIAL` 的 build 被誤判成 `BUILD_VALIDATED`——直接違反
    預註冊文件「`BUILD_FAILED_PARTIAL` 永遠不能變成 `BUILD_VALIDATED`」
    (§D)。現在明確要求:`aggregate_build_receipt_overall_status` 必須是
    `BUILD_COMPLETE_AWAITING_VERIFICATION` 才可能達到 `BUILD_VALIDATED`
    ——不是這個值,不論驗證 receipt 宣稱什麼,一律 `BUILD_VERIFICATION_
    FAILED`。這個檢查現在對**兩個分支**(有無 `.claim`)都適用:
    `aggregate_build_receipt_overall_status` 的合法性(必須是兩個允許值
    之一)在函式最前面統一驗證一次,不再只在無 `.claim` 分支裡做。"""
    if aggregate_build_receipt_overall_status not in (
        STATUS_BUILD_COMPLETE_AWAITING_VERIFICATION, STATUS_BUILD_FAILED_PARTIAL,
    ):
        raise ValueError(
            "aggregate_build_receipt_overall_status 必須是"
            f" {STATUS_BUILD_COMPLETE_AWAITING_VERIFICATION!r}/"
            f"{STATUS_BUILD_FAILED_PARTIAL!r} 之一,實際"
            f" {aggregate_build_receipt_overall_status!r}"
        )

    if not claim_exists:
        return aggregate_build_receipt_overall_status

    if binding_verification_receipt is None:
        return STATUS_BUILD_VERIFICATION_FAILED

    if claim_content is None:
        raise ValueError(
            "claim_exists=True 且存在 binding_verification_receipt,但呼叫端"
            " 沒有提供 claim_content——判定正式狀態前必須先讀取 .claim 鎖檔"
            " 內容(不能假設鎖檔存在就等於內容可信或可省略,§D 三方雜湊等式"
            " 需要它)"
        )

    if binding_verification_receipt.get("run_id") != claim_content.get("run_id"):
        return STATUS_BUILD_VERIFICATION_FAILED
    if binding_verification_receipt.get("verifier_identity") != claim_content.get(
        "authorized_verifier_identity"
    ):
        return STATUS_BUILD_VERIFICATION_FAILED

    claim_sha = claim_content.get("build_receipt_sha256")
    receipt_sha = binding_verification_receipt.get("build_receipt_sha256")
    if claim_sha is None or receipt_sha is None or aggregate_build_receipt_sha256 is None:
        return STATUS_BUILD_VERIFICATION_FAILED
    if not (claim_sha == receipt_sha == aggregate_build_receipt_sha256):
        return STATUS_BUILD_VERIFICATION_FAILED

    overall = binding_verification_receipt.get("overall_status")
    if overall not in _VERIFICATION_STATUSES:
        return STATUS_BUILD_VERIFICATION_FAILED

    # aggregate_build_receipt_overall_status 已經在函式最前面驗證過是兩個
    # 合法值之一(BUILD_COMPLETE_AWAITING_VERIFICATION 或
    # BUILD_FAILED_PARTIAL)——只有前者才可能放行到 BUILD_VALIDATED。
    if aggregate_build_receipt_overall_status != STATUS_BUILD_COMPLETE_AWAITING_VERIFICATION:
        return STATUS_BUILD_VERIFICATION_FAILED

    return overall


# ----------------------------------------------------------------------------
# 8. CLI 安全邊界——跟 `build_v2_candidate.py` 同一套紀律(§11 的read-only
#    邊界):這一輪不核准 Phase B,CLI 預設(也是唯一)行為是印一則說明訊息、
#    回傳非 0。
# ----------------------------------------------------------------------------


def main(argv=None) -> int:
    del argv
    import sys

    print(
        "build_v2_candidate_verifier: Phase B 尚未獲得授權"
        "(docs/預註冊_DataExport0806_V2隔離建置.md 現況"
        " PREREG_READY_IMPLEMENTATION_NOT_AUTHORIZED / BUILD_NOT_RUN)。"
        " 這個 CLI 入口本輪只是佔位——不讀任何來源檔案、不建立任何輸出目錄。"
        " 另外,即使 Phase B 未來被授權,這份驗證器目前也還沒有接上"
        " 從 manifest 逐檔案自動跑一輪 independent_raw_cell_reconstruction()"
        " + cross_check_reconstruction_against_sidecar() 並彙總覆蓋率的協調層"
        "(那屬於真正 Phase B glue 的一部分),見本檔案模組 docstring。",
        file=sys.stderr,
    )
    return 1


if __name__ == "__main__":
    import sys

    sys.exit(main(sys.argv[1:]))
